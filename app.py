import os, json, threading
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import pdfplumber
from models import db, User, EmpresaDocumento, EmpresaInfo, EquipamentoIndustrial, EqIndComponente, EqIndDocumento, EqIndComponenteMedia, AgendaServico, AgendaRegisto, AgendaMaterial, FichaTecnica, FichaComponente, FichaDocumento, FeriasPeriodo, FeriasFeriado, UserSession, AusenciaRegisto, AusenciaSaldoAnual, EmpresaFecho, TIPOS_AUSENCIA, PeriodoSalarial, HoraExtra, ConfigHorario, PedidoCompra, LinhaPedido, Orcamento, ItemOrcamento, ArtigoPHC, AliasArtigo, FornecedorPHC, ConfigPHC, ConfigIA, ConfigReposicao, PendingMatch, Cliente, Embarcacao, ComponenteEmbarcacao, ConfigGeral, NotaArtigo, EventoCalendario, EntradaEquipamento, EntradaHistorico, EntradaDocumento, Assistencia, AssistenciaHistorico, AssistenciaDocumento

import time as _time_module
app = Flask(__name__)
# Token changes on every server restart - invalidates all salary sessions
_SERVER_START_TOKEN = str(int(_time_module.time()))
app.permanent_session_lifetime = __import__('datetime').timedelta(days=30)
app.config['SECRET_KEY'] = 'comprasnet-2024-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///compras.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
app.config['BAK_FOLDER'] = os.path.join(os.path.dirname(__file__), 'bak_uploads')
os.makedirs(app.config['BAK_FOLDER'], exist_ok=True)

db.init_app(app)
migrate = Migrate(app, db)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor faça login para aceder.'

def user_pode_aceder(endpoint):
    """Check if current user can access the given menu endpoint."""
    if not current_user.is_authenticated:
        return False
    if current_user.is_admin:
        return True
    perfil_id = getattr(current_user, 'perfil_id', None)
    if not perfil_id:
        return True  # no perfil = full access (backwards compat)
    perfil = Perfil.query.get(perfil_id)
    if not perfil:
        return True
    return perfil.pode_aceder(endpoint)


@app.context_processor
def inject_perfil():
    def pode_aceder(endpoint):
        from flask_login import current_user
        if not current_user.is_authenticated: return False
        if current_user.is_admin: return True
        pid = get_user_perfil_id(current_user)
        if not pid: return True
        p = Perfil.query.get(pid)
        if not p: return True
        return p.pode_aceder(endpoint)
    return dict(pode_aceder=pode_aceder)

@app.before_request
def refresh_session():
    """Keep session alive and update last_seen on every request."""
    from flask_login import current_user
    if current_user.is_authenticated:
        session.permanent = True
        session.modified = True
        # Update UserSession last_seen (skip static/api polling to avoid DB spam)
        if not request.path.startswith('/static') and request.path != '/api/users/status':
            try:
                us = UserSession.query.filter_by(user_id=current_user.id).first()
                if us:
                    us.last_seen = datetime.now()
                else:
                    db.session.add(UserSession(user_id=current_user.id, last_seen=datetime.now()))
                db.session.commit()
            except Exception:
                db.session.rollback()


@login_manager.unauthorized_handler
def handle_unauthorized():
    """Return JSON for API calls, redirect for normal pages."""
    if (request.is_json or 
        request.path.startswith('/api/') or
        request.method == 'POST'):
        return jsonify({'error': 'Sessão expirada. Recarregue a página e faça login.'}), 401
    from flask import redirect, url_for
    return redirect(url_for('login'))


@login_manager.user_loader
def load_user(uid): return User.query.get(int(uid))

def allowed_file(f): return '.' in f and f.rsplit('.',1)[1].lower() == 'pdf'

# ── PDF helpers ────────────────────────────────────────────────────────────────

def extract_pdf_text(filepath):
    text = ""
    try:
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t: text += t + "\n"
                for table in page.extract_tables():
                    for row in table:
                        if row: text += " | ".join([str(c) if c else "" for c in row]) + "\n"
    except Exception as e:
        text = f"Erro: {e}"
    return text

def analyze_pdf_with_claude(pdf_text, filename):
    """Wrapper: delegates to ai_provider using current ConfigIA from DB."""
    from ai_provider import analyze_pdf
    cfg = ConfigIA.query.first()
    if not cfg:
        # Auto-create default LM Studio config
        cfg = ConfigIA(); db.session.add(cfg); db.session.commit()
    return analyze_pdf(cfg, pdf_text, filename)

# ── AUTH ───────────────────────────────────────────────────────────────────────


@app.route('/favicon.ico')
def favicon():
    """Serve favicon - use company logo if available, else default."""
    cfg = ConfigGeral.query.first()
    if cfg and cfg.empresa_logo_path:
        logo_path = os.path.join(app.config['UPLOAD_FOLDER'], cfg.empresa_logo_path)
        if os.path.exists(logo_path):
            try:
                from PIL import Image
                import io
                img = Image.open(logo_path).convert('RGBA')
                # Create padded square with navy background
                size = max(img.size)
                bg = Image.new('RGBA', (size, size), '#1e3a5f')
                offset = ((size - img.width)//2, (size - img.height)//2)
                bg.paste(img, offset, img if img.mode == 'RGBA' else None)
                # Resize to 32x32 for favicon
                ico = bg.resize((32, 32), Image.LANCZOS)
                buf = io.BytesIO()
                ico.save(buf, format='ICO', sizes=[(16,16),(32,32)])
                buf.seek(0)
                from flask import Response
                return Response(buf.getvalue(), mimetype='image/vnd.microsoft.icon',
                    headers={'Cache-Control': 'public, max-age=3600'})
            except Exception:
                pass
    return send_from_directory(os.path.join(app.root_path, 'static'),
        'favicon.ico', mimetype='image/vnd.microsoft.icon')

@app.route('/icon-<int:size>.png')
def pwa_icon(size):
    valid = [72, 96, 128, 144, 152, 192, 384, 512]
    if size not in valid:
        size = 192
    fname = f'icon-{size}.png'
    resp = send_from_directory(os.path.join(app.root_path, 'static'), fname)
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


@app.route('/api/backup/manual', methods=['POST'])
@login_required
def backup_manual():
    if not current_user.is_admin:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    from backup_manager import fazer_backup_completo
    cfg = ConfigGeral.query.first()
    ok, msg = fazer_backup_completo(app, cfg)
    return jsonify({'ok': ok, 'msg': msg})


@app.route('/admin/regenerar-icons', methods=['POST'])
@login_required
def admin_regenerar_icons():
    if not current_user.is_admin:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    ok = _regenerar_pwa_icons()
    if ok:
        return jsonify({'ok': True, 'msg': 'Ícones PWA regenerados com o logotipo actual'})
    return jsonify({'ok': False, 'error': 'Sem logotipo definido ou erro ao processar'})


@app.route('/icon-debug')
def icon_debug():
    """Debug: check what icon would be served."""
    cfg = ConfigGeral.query.first()
    info = {
        'empresa_logo_path': cfg.empresa_logo_path if cfg else None,
        'upload_folder': app.config['UPLOAD_FOLDER'],
    }
    if cfg and cfg.empresa_logo_path:
        logo_path = os.path.join(app.config['UPLOAD_FOLDER'], cfg.empresa_logo_path)
        info['logo_full_path'] = logo_path
        info['logo_exists'] = os.path.exists(logo_path)
    else:
        info['logo_full_path'] = None
        info['logo_exists'] = False
    
    # List uploads folder
    try:
        info['uploads_contents'] = os.listdir(app.config['UPLOAD_FOLDER'])
    except Exception as e:
        info['uploads_contents'] = str(e)
    
    return jsonify(info)

def _regenerar_pwa_icons_auto():
    """Called on startup to restore icons from logo."""
    with app.app_context():
        _regenerar_pwa_icons()

def _regenerar_pwa_icons():
    """Pre-generate all PWA icon sizes from the company logo."""
    cfg = ConfigGeral.query.first()
    if not cfg or not cfg.empresa_logo_path:
        return False
    logo_path = os.path.join(app.config['UPLOAD_FOLDER'], cfg.empresa_logo_path)
    if not os.path.exists(logo_path):
        return False
    try:
        from PIL import Image
        logo = Image.open(logo_path).convert('RGBA')
        static_dir = os.path.join(app.root_path, 'static')
        for size in [72, 96, 128, 144, 152, 180, 192, 384, 512]:
            canvas = Image.new('RGBA', (size, size), '#1e3a5f')
            thumb = logo.copy()
            max_dim = int(size * 0.75)
            thumb.thumbnail((max_dim, max_dim), Image.LANCZOS)
            ox = (size - thumb.width) // 2
            oy = (size - thumb.height) // 2
            canvas.paste(thumb, (ox, oy), thumb)
            canvas.save(os.path.join(static_dir, f'icon-{size}.png'), 'PNG')
        # favicon
        i32 = Image.new('RGBA', (32, 32), '#1e3a5f')
        thumb = logo.copy(); thumb.thumbnail((24,24), Image.LANCZOS)
        i32.paste(thumb, ((32-thumb.width)//2,(32-thumb.height)//2), thumb)
        i32.save(os.path.join(static_dir, 'favicon.ico'), format='ICO', sizes=[(16,16),(32,32)])
        return True
    except Exception as e:
        print(f"Icon regen error: {e}")
        return False


@app.route('/icon-maskable-<int:size>.png')
def pwa_icon_maskable(size):
    if size not in [192, 512]:
        size = 192
    resp = send_from_directory(os.path.join(app.root_path, 'static'), f'icon-maskable-{size}.png')
    resp.headers['Cache-Control'] = 'no-cache'
    return resp

@app.route('/apple-touch-icon.png')
@app.route('/apple-touch-icon-precomposed.png')
def apple_touch_icon():
    return pwa_icon(180)

@app.route('/manifest.json')
def pwa_manifest():
    cfg = ConfigGeral.query.first()
    nome = (cfg.empresa_nome if cfg and cfg.empresa_nome else 'ComprasNet').strip()
    short = nome[:12] if len(nome) > 12 else nome
    from flask import request as _req
    base = _req.host_url.rstrip('/')
    import json, time
    v = str(int(time.time()))
    manifest = {
        "name": nome,
        "short_name": short,
        "description": "Plataforma de gestao interna",
        "start_url": base + "/",  # dashboard is at root
        "scope": base + "/",
        "display": "standalone",
        "display_override": ["standalone", "minimal-ui"],
        "background_color": "#1e3a5f",
        "theme_color": "#1e3a5f",
        "orientation": "portrait-primary",
        "lang": "pt-PT",
        "dir": "ltr",
        "prefer_related_applications": False,
        "icons": [
            {"src": base + "/icon-192.png?v=" + v,          "sizes": "192x192", "type": "image/png", "purpose": "any"},
            {"src": base + "/icon-512.png?v=" + v,          "sizes": "512x512", "type": "image/png", "purpose": "any"},
            {"src": base + "/icon-maskable-192.png?v=" + v, "sizes": "192x192", "type": "image/png", "purpose": "maskable"},
            {"src": base + "/icon-maskable-512.png?v=" + v, "sizes": "512x512", "type": "image/png", "purpose": "maskable"},
        ]
    }
    resp = Response(json.dumps(manifest), mimetype='application/manifest+json')
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        u = User.query.filter_by(username=request.form.get('username','')).first()
        if u and check_password_hash(u.password_hash, request.form.get('password','')):
            login_user(u, remember=True)
            if getattr(u, 'must_change_password', False):
                return redirect(url_for('alterar_password'))
            return redirect(url_for('dashboard'))
        flash('Utilizador ou palavra-passe incorretos.', 'error')
    return render_template('login.html', cfg=ConfigGeral.query.first())

@app.route('/logout')
@login_required
def logout():
    logout_user(); return redirect(url_for('login'))

# ── DASHBOARD ─────────────────────────────────────────────────────────────────

@app.route('/dashboard')
@app.route('/')
@login_required
def dashboard():
    total_pedidos    = PedidoCompra.query.count()
    pedidos_abertos  = PedidoCompra.query.filter_by(estado='aberto').count()
    pedidos_aprovados= PedidoCompra.query.filter_by(estado='aprovado').count()
    total_orcamentos = Orcamento.query.count()
    pedidos_recentes = PedidoCompra.query.filter(PedidoCompra.estado.in_(['aberto','pedido','aprovado','pendente','em_analise'])).order_by(PedidoCompra.data_criacao.desc()).limit(8).all()

    # Artigos pedidos para dashboard - pendentes/não encomendados primeiro, recebidos no fim
    from sqlalchemy import case
    artigos_pedidos = (db.session.query(LinhaPedido, PedidoCompra, User)
        .join(PedidoCompra, LinhaPedido.pedido_id == PedidoCompra.id)
        .join(User, PedidoCompra.criado_por == User.id, isouter=True)
        .filter(
            db.or_(
                PedidoCompra.estado.in_(['aberto','aprovado','pendente']),
                # Always show por_faturar and faturado even if pedido is closed
                LinhaPedido.status.in_(['por_faturar','faturado'])
            )
        )
        .filter(LinhaPedido.status != 'cancelado')
        .order_by(
            case(
                (LinhaPedido.status == 'nao_encomendado', 0),
                (LinhaPedido.status == 'pendente', 1),
                (LinhaPedido.status == 'encomendado', 2),
                (LinhaPedido.status == 'por_faturar', 3),
                (LinhaPedido.status == 'recebido', 4),
                (LinhaPedido.status == 'faturado', 5),
                else_=6
            ),
            PedidoCompra.data_criacao.desc()
        )
        .all()
    )
    # Stock baixo: stock > 0 mas abaixo de threshold (excluir negativos)
    artigos_stock_baixo = ArtigoPHC.query.filter(
        ArtigoPHC.stock_atual > 0,
        ArtigoPHC.stock_atual <= 3
    ).order_by(ArtigoPHC.stock_atual).limit(10).all()
    # Check funcionario documents expiring/expired
    try:
        from datetime import date as _dt, timedelta
        _hoje = _dt.today()
        _aviso = _hoje + timedelta(days=30)  # warn 30 days before
        # CC and passaporte expiry on Funcionario
        docs_expirados = []
        docs_a_expirar = []
        for f in Funcionario.query.all():
            if f.cc_validade:
                if f.cc_validade < _hoje:
                    docs_expirados.append({'func': f.nome, 'tipo': 'CC', 'data': f.cc_validade, 'fid': f.id})
                elif f.cc_validade <= _aviso:
                    docs_a_expirar.append({'func': f.nome, 'tipo': 'CC', 'data': f.cc_validade, 'fid': f.id})
            if f.passaporte_validade:
                if f.passaporte_validade < _hoje:
                    docs_expirados.append({'func': f.nome, 'tipo': 'Passaporte', 'data': f.passaporte_validade, 'fid': f.id})
                elif f.passaporte_validade <= _aviso:
                    docs_a_expirar.append({'func': f.nome, 'tipo': 'Passaporte', 'data': f.passaporte_validade, 'fid': f.id})
        # FuncionarioDocumento expiry
        for doc in FuncionarioDocumento.query.filter(FuncionarioDocumento.data_validade != None).all():
            if doc.data_validade < _hoje:
                f = Funcionario.query.get(doc.funcionario_id)
                docs_expirados.append({'func': f.nome if f else '?', 'tipo': doc.titulo, 'data': doc.data_validade, 'fid': doc.funcionario_id})
            elif doc.data_validade <= _aviso:
                f = Funcionario.query.get(doc.funcionario_id)
                docs_a_expirar.append({'func': f.nome if f else '?', 'tipo': doc.titulo, 'data': doc.data_validade, 'fid': doc.funcionario_id})
        docs_alert_count = len(docs_expirados) + len(docs_a_expirar)
    except: docs_expirados=[]; docs_a_expirar=[]; docs_alert_count=0

    # Check assistencias com obra concluida/comunicado > 5 dias sem faturar
    try:
        from datetime import date as _date
        _hoje = _date.today()
        assist_alerta = []
        for _a in Assistencia.query.filter(Assistencia.status.in_(['obra_concluida','comunicado'])).all():
            _ref = _a.data_obra_concluida if _a.status == 'obra_concluida' else _a.data_comunicado
            if _ref:
                _dias = (_hoje - _ref).days
                if _dias > 5:
                    _a._dias_alerta = _dias
                    assist_alerta.append(_a)
        assist_alerta_count = len(assist_alerta)
    except: assist_alerta = []; assist_alerta_count = 0

    # Check entradas in estadia
    try:
        _verificar_estadias()
        entradas_estadia_list = EntradaEquipamento.query.filter(
            EntradaEquipamento.status.in_(['orcamentado_estadia','concluido_estadia'])
        ).order_by(EntradaEquipamento.data_status).all()
        entradas_estadia = len(entradas_estadia_list)
    except: entradas_estadia = 0; entradas_estadia_list = []
    return render_template('dashboard.html',
        entradas_estadia=entradas_estadia,
        entradas_estadia_list=entradas_estadia_list,
        assist_alerta=assist_alerta,
        assist_alerta_count=assist_alerta_count,
        docs_expirados=docs_expirados,
        docs_a_expirar=docs_a_expirar,
        docs_alert_count=docs_alert_count,
        total_pedidos=total_pedidos,
        pedidos_abertos=pedidos_abertos,
        pedidos_aprovados=pedidos_aprovados,
        total_orcamentos=total_orcamentos,
        pedidos_recentes=pedidos_recentes,
        artigos_stock_baixo=artigos_stock_baixo,
        artigos_pedidos=artigos_pedidos,
    )


@app.route('/api/artigos-pedidos/pdf')
@login_required
def artigos_pedidos_pdf():
    """Generate PDF of pending items, filtered and sorted by priority."""
    from datetime import datetime as _dt
    forn_filtro   = request.args.get('fornecedor','').strip()
    cliente_filtro = request.args.get('cliente','').strip()
    item_filtro   = request.args.get('item','').strip()
    status_filtro = request.args.get('status','').strip()

    from sqlalchemy import case
    STATUS_VISIVEIS = ['nao_encomendado', 'consultado', 'pendente', 'encomendado', 'por_faturar']
    query = (db.session.query(LinhaPedido, PedidoCompra)
        .join(PedidoCompra, LinhaPedido.pedido_id == PedidoCompra.id)
        .filter(PedidoCompra.estado.notin_(['cancelado','concluido','arquivado','rejeitado']))
        .filter(db.or_(
            LinhaPedido.status.in_(STATUS_VISIVEIS),
            LinhaPedido.status == None
        )))

    artigos = query.order_by(
        case(
            (LinhaPedido.status == 'nao_encomendado', 0),
            (LinhaPedido.status == 'consultado', 1),
            (LinhaPedido.status == 'pendente', 2),
            (LinhaPedido.status == 'encomendado', 3),
            else_=4
        ), PedidoCompra.data_criacao.desc()
    ).all()

    STATUS_INFO_LOCAL = {
        "nao_encomendado": ("🔴 Não Enc.",    "#ef4444"),
        "consultado":       ("🔍 Consultado",  "#0891b2"),
        "pendente":         ("⏳ Pendente",    "#f59e0b"),
        "encomendado":      ("📦 Encomendado", "#3b6ef0"),
        "recebido":         ("✅ Recebido",    "#22c55e"),
        "por_faturar":      ("🧾 Por Faturar", "#f97316"),
        "faturado":         ("💶 Faturado",    "#6366f1"),
        "cancelado":        ("❌ Cancelado",   "#a855f7"),
    }

    rows = []
    for linha, pedido in artigos:
        try:
            cid = linha.cliente_id or pedido.cliente_id
            cliente_nome = "Stock"
            if cid:
                row = db.session.execute(db.text('SELECT nome FROM clientes WHERE id=:id'),{'id':cid}).fetchone()
                if row: cliente_nome = row[0]
        except: cliente_nome = "Stock"

        # Fornecedor from fornecedores_json
        forn_nome = ""
        try:
            import json as _jpdf
            fjs = _jpdf.loads(linha.fornecedores_json or '[]') if linha.fornecedores_json else []
            if isinstance(fjs, list):
                forn_nome = ', '.join(f.get('nome','') for f in fjs if isinstance(f,dict) and f.get('nome'))
        except: pass
        if not forn_nome and linha.fornecedor_hab:
            forn_nome = linha.fornecedor_hab

        # Apply filters
        if forn_filtro and forn_filtro.lower() not in (forn_nome or '').lower() and forn_filtro.lower() not in (linha.fornecedor_hab or '').lower(): continue
        if cliente_filtro and cliente_filtro.lower() not in cliente_nome.lower(): continue
        if item_filtro and item_filtro.lower() not in (linha.designacao or '').lower(): continue
        if status_filtro and linha.status != status_filtro: continue

        lbl, col = STATUS_INFO_LOCAL.get(linha.status or 'nao_encomendado', ("?","#888"))
        # Get last cost price
        ultimo_preco = ''
        try:
            artigo = ArtigoPHC.query.filter_by(referencia=linha.artigo_ref).first() if linha.artigo_ref else None
            if artigo and artigo.preco_custo:
                ultimo_preco = f'{artigo.preco_custo:.2f} €'
        except: pass

        rows.append({
            'ref': linha.artigo_ref or '',
            'designacao': linha.designacao or '',
            'qty': f"{linha.quantidade or 0} {linha.unidade or ''}".strip(),
            'cliente': cliente_nome,
            'fornecedor': forn_nome or '—',
            'status': lbl,
            'status_col': col,
            'pedido_ref': (pedido.titulo or '')[:30] or f'#{pedido.id}',
            'data': pedido.data_criacao.strftime('%d/%m/%Y') if pedido.data_criacao else '',
            'preco': ultimo_preco,
            'obs': linha.observacoes or '',
        })

    cfg = ConfigGeral.query.first()
    empresa = cfg.empresa_nome if cfg else 'UCN'
    filtros_str = ' | '.join(filter(None,[
        f'Fornecedor: {forn_filtro}' if forn_filtro else '',
        f'Cliente: {cliente_filtro}' if cliente_filtro else '',
        f'Item: {item_filtro}' if item_filtro else '',
        f'Estado: {status_filtro}' if status_filtro else '',
    ])) or 'Todos os artigos'

    total_q = len(artigos)
    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<title>Artigos Pedidos — {empresa}</title>
<style>
@page{{size:A4 landscape;margin:1.2cm}}
@media print{{.no-print{{display:none}}body{{-webkit-print-color-adjust:exact;print-color-adjust:exact}}}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:Helvetica,Arial,sans-serif;font-size:9px;color:#1e293b;background:white}}
.header{{border-bottom:3px solid #1e3a5f;padding-bottom:8px;margin-bottom:10px;display:flex;justify-content:space-between;align-items:flex-end}}
h1{{font-size:16px;font-weight:800;color:#1e3a5f}}
h2{{font-size:9px;color:#64748b;margin-top:2px}}
table{{width:100%;border-collapse:collapse;font-size:8.5px}}
thead tr{{background:#1e3a5f;color:white}}
th{{padding:5px 8px;text-align:left;font-size:8px;font-weight:700}}
td{{padding:4px 8px;border-bottom:1px solid #f1f5f9;vertical-align:middle}}
tr:nth-child(even){{background:#f8fafc}}
.badge{{display:inline-block;padding:2px 7px;border-radius:4px;font-weight:700;font-size:8px;white-space:nowrap}}
.footer{{margin-top:12px;font-size:8px;color:#94a3b8;display:flex;justify-content:space-between;border-top:1px solid #e2e8f0;padding-top:6px}}
.btn-print{{position:fixed;top:16px;right:16px;background:#1e3a5f;color:white;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;font-size:12px;font-weight:700;box-shadow:0 2px 8px rgba(0,0,0,.2)}}
.ref{{font-family:monospace;color:#3b6ef0;font-size:8px}}
.price{{font-family:monospace;color:#16a34a;font-weight:700}}
</style></head><body>
<button class="btn-print no-print" onclick="window.print()">🖨 Imprimir / PDF</button>
<div class="header">
  <div>
    <h1>{empresa} &mdash; Lista de Artigos Pedidos</h1>
    <h2>Filtros: {filtros_str} &nbsp;·&nbsp; {len(rows)} de {total_q} artigo(s) &nbsp;·&nbsp; Emitido em {_dt.now().strftime('%d/%m/%Y %H:%M')}</h2>
  </div>
</div>
<table>
<thead><tr>
  <th style="width:8%">Estado</th>
  <th style="width:7%">Ref.</th>
  <th style="width:24%">Designação</th>
  <th style="width:6%;text-align:center">Qtd</th>
  <th style="width:16%">Fornecedor</th>
  <th style="width:12%">Cliente</th>
  <th style="width:8%;text-align:right">Últ. Custo</th>
  <th style="width:10%">Pedido</th>
  <th style="width:9%">Data</th>
</tr></thead>
<tbody>"""
    for r in rows:
        html += f"""<tr>
<td><span class="badge" style="background:{r['status_col']}22;color:{r['status_col']};border:1px solid {r['status_col']}55">{r['status']}</span></td>
<td class="ref">{r['ref']}</td>
<td><strong>{r['designacao']}</strong>{('<br><span style="font-size:7.5px;color:#94a3b8">' + r['obs'] + '</span>') if r['obs'] else ''}</td>
<td style="text-align:center;font-weight:700">{r['qty']}</td>
<td>{r['fornecedor']}</td>
<td>{r['cliente']}</td>
<td style="text-align:right" class="price">{r['preco']}</td>
<td style="color:#3b6ef0">{r['pedido_ref']}</td>
<td style="color:#94a3b8">{r['data']}</td>
</tr>"""
    html += f"""</tbody></table>
<div class="footer">
  <span>{empresa} &nbsp;·&nbsp; Uso Interno &nbsp;·&nbsp; Lista de Artigos Pedidos</span>
  <span>Total: {len(rows)} artigos</span>
</div>
</body></html>"""

    from flask import Response
    return Response(html, mimetype='text/html')


@app.route('/api/dashboard/artigos-pedidos')
@login_required
def api_dashboard_artigos_pedidos():
    from sqlalchemy import case as sa_case
    artigos = (db.session.query(LinhaPedido, PedidoCompra, User)
        .join(PedidoCompra, LinhaPedido.pedido_id == PedidoCompra.id)
        .join(User, PedidoCompra.criado_por == User.id, isouter=True)
        .filter(PedidoCompra.estado.in_(["aberto","pedido","aprovado","pendente","em_analise"]))
        .order_by(
            sa_case(
                (LinhaPedido.status == "recebido", 2),
                (LinhaPedido.status == "cancelado", 3),
                else_=0
            ),
            PedidoCompra.data_criacao.desc()
        ).all()
    )
    STATUS_INFO = {
        "nao_encomendado": ("🔴 Não Enc.",    "#ef4444", "rgba(239,68,68,.15)"),
        "pendente":         ("⏳ Pendente",    "#f59e0b", "rgba(245,158,11,.15)"),
        "encomendado":      ("📦 Encomendado", "#3b6ef0", "rgba(59,110,240,.15)"),
        "recebido":         ("✅ Recebido",    "#22c55e", "rgba(34,197,94,.15)"),
        "por_faturar":      ("🧾 Por Faturar", "#f97316", "rgba(249,115,22,.15)"),
        "faturado":         ("💶 Faturado",    "#6366f1", "rgba(99,102,241,.15)"),
        "consultado":       ("🔍 Consultado",  "#0891b2", "rgba(8,145,178,.15)"),
        "cancelado":        ("❌ Cancelado",   "#a855f7", "rgba(168,85,247,.15)"),
    }
    rows = []
    for linha, pedido, user in artigos:
        s = linha.status or "nao_encomendado"
        lbl, col, bg = STATUS_INFO.get(s, STATUS_INFO["nao_encomendado"])
        hist = LinhaPedidoHistorico.query.filter_by(linha_id=linha.id)            .order_by(LinhaPedidoHistorico.data.desc()).first()
        # Get cliente name - use direct query bypassing cache
        cliente_nome = "Stock"
        try:
            cid = linha.cliente_id or pedido.cliente_id
            if cid:
                row = db.session.execute(
                    db.text('SELECT nome FROM clientes WHERE id = :id'), {'id': cid}
                ).fetchone()
                if row: cliente_nome = row[0]
        except: pass
        # Get fornecedores
        import json as _json
        forn_list = []
        try:
            forn_raw = linha.fornecedores_json or '[]'
            parsed = _json.loads(forn_raw) if isinstance(forn_raw, str) else forn_raw
            if isinstance(parsed, list):
                for item in parsed:
                    if isinstance(item, dict):
                        forn_list.append(item)
                    elif isinstance(item, str) and item:
                        forn_list.append({'nome': item, 'id': ''})
        except: forn_list = []
        forn_nomes = [f.get('nome','') for f in forn_list if f.get('nome')]
        forn_ids   = [str(f.get('id','')) for f in forn_list if f.get('id')]
        rows.append({
            "linha_id": linha.id,
            "pedido_id": pedido.id,
            "designacao": linha.designacao or linha.referencia or "—",
            "referencia": linha.referencia or "",
            "quantidade": int(linha.quantidade or 1),
            "unidade": linha.unidade or "un",
            "data_pedido": pedido.data_criacao.strftime("%d/%m/%Y"),
            "criado_por": user.nome[:20] if user else "—",
            "cliente": cliente_nome,
            "fornecedor": forn_nomes[0] if forn_nomes else "—",
            "fornecedores_ids": forn_ids,
            "fornecedores_nomes": forn_nomes,
            "is_stock": cliente_nome == "Stock",
            "status": s,
            "status_label": ("✅ Recebido — Stock" if cliente_nome == "Stock" else lbl) if s == "recebido" else lbl,
            "status_color": col,
            "status_bg": bg,
            "dim": s in ["recebido","faturado","cancelado"],
            "alerta_faturar": bool(s == "por_faturar" and hist and (datetime.now() - hist.data).days >= 10),
            "dias_por_faturar": int((datetime.now() - hist.data).days) if s == "por_faturar" and hist else 0,
            "data_status": hist.data.strftime("%d/%m/%Y %H:%M") if hist else "—",
            "alterado_por": hist.user_nome[:20] if hist else "—",
        })
    return jsonify(rows)


@app.route('/api/clientes/search')
@login_required
def api_clientes_search():
    q = request.args.get('q','').strip()
    clientes = Cliente.query.filter(Cliente.nome.ilike(f'%{q}%')).order_by(Cliente.nome).limit(20).all() if q else Cliente.query.order_by(Cliente.nome).limit(20).all()
    return jsonify([{'id': c.id, 'nome': c.nome} for c in clientes])

@app.route('/pedidos/<int:pid>/estado', methods=['POST'])
@login_required
def pedido_estado(pid):
    p = PedidoCompra.query.get_or_404(pid)
    data = request.get_json() or {}
    novo = data.get('estado','')
    if novo not in ['aberto','pedido','fechado','anulado','em_analise','aprovado','cancelado']:
        return jsonify({'ok': False, 'error': 'Estado inválido'})
    p.estado = novo
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/pedidos/linha/<int:lid>/editar', methods=['POST'])
@login_required
def linha_editar(lid):
    l = LinhaPedido.query.get_or_404(lid)
    data = request.get_json() or {}
    if 'designacao'  in data: l.designacao  = data['designacao'].strip()
    if 'referencia'  in data: l.referencia  = data['referencia'].strip()
    if 'quantidade'  in data:
        try: l.quantidade = float(data['quantidade'])
        except: pass
    if 'unidade'     in data: l.unidade     = data['unidade'].strip()
    if 'observacoes' in data: l.observacoes = data['observacoes'].strip()
    if 'cliente_id'  in data:
        cid = data['cliente_id']
        l.cliente_id = int(cid) if str(cid).strip() else None
    if 'fornecedores_json' in data:
        import json as _j
        fj = data['fornecedores_json']
        l.fornecedores_json = _j.dumps(fj) if isinstance(fj, list) else fj
    if 'fornecedores' in data:
        import json as _jj
        l.fornecedores_json = _jj.dumps(data['fornecedores'])
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/pedidos/linha/<int:lid>/eliminar', methods=['POST'])
@login_required
def linha_eliminar(lid):
    l = LinhaPedido.query.get_or_404(lid)
    pid = l.pedido_id
    db.session.delete(l)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/pedidos/<int:pid>/editar', methods=['POST'])
@login_required
def pedido_editar(pid):
    p = PedidoCompra.query.get_or_404(pid)
    data = request.get_json() or {}
    if 'titulo'       in data and data['titulo'].strip():
        p.titulo = data['titulo'].strip()
    if 'descricao'    in data: p.descricao    = data['descricao'].strip()
    if 'departamento' in data: p.departamento = data['departamento'].strip()
    if 'prioridade'   in data: p.prioridade   = data['prioridade']
    if 'estado'       in data: p.estado       = data['estado']
    if 'cliente_id' in data:
        cid = data['cliente_id']
        novo_cid = int(cid) if str(cid).strip() else None
        old_cid = p.cliente_id
        p.cliente_id = novo_cid
        # Also update lines that had the same cliente as the pedido header
        if old_cid != novo_cid:
            for l in p.linhas:
                if l.cliente_id == old_cid or l.cliente_id is None:
                    l.cliente_id = novo_cid
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/pedidos')
@login_required
def pedidos():
    estado = request.args.get('estado','aberto') or 'aberto'
    q = PedidoCompra.query
    if estado and estado != 'todos': q = q.filter_by(estado=estado)
    return render_template('pedidos.html', pedidos=q.order_by(PedidoCompra.data_criacao.desc()).all(), estado_filtro=estado)

@app.route('/pedidos/novo', methods=['GET','POST'])
@login_required
def novo_pedido():
    if request.method == 'POST':
        titulo = request.form.get('titulo','').strip()
        if not titulo:
            flash('O artigo a comprar é obrigatório.','error')
            clientes_list = Cliente.query.order_by(Cliente.nome).all()
            return render_template('novo_pedido.html', clientes=clientes_list)
        cliente_id = request.form.get('cliente_id','').strip()
        p = PedidoCompra(titulo=titulo,
            descricao=request.form.get('descricao','').strip(),
            prioridade=request.form.get('prioridade','normal'),
            departamento=request.form.get('departamento','Compras').strip(),
            cliente_id=int(cliente_id) if cliente_id else None,
            estado='aberto', criado_por=current_user.id,
            data_criacao=datetime.now())
        db.session.add(p); db.session.flush()

        # Save lines from form
        import json as _json
        linhas_json = request.form.get('linhas_json','[]')
        try:
            linhas = _json.loads(linhas_json)
        except: linhas = []
        for i, l in enumerate(linhas):
            ref = l.get('referencia','').strip()
            artigo = ArtigoPHC.query.filter_by(referencia=ref).first() if ref else None
            db.session.add(LinhaPedido(
                pedido_id=p.id, ordem=i,
                artigo_ref=ref if artigo else None,
                referencia=ref or l.get('designacao','')[:50],
                designacao=l.get('designacao','').strip(),
                unidade=l.get('unidade','un'),
                quantidade=float(l.get('quantidade',1)),
                stock_atual=float(artigo.stock_atual if artigo else 0),
                preco_custo_ref=float(artigo.preco_custo if artigo else 0),
                preco_pcp_ref=float(artigo.preco_custo_ponderado if artigo else 0),
                observacoes=l.get('observacoes','').strip(),
                cliente_id=int(l.get('cliente_id')) if str(l.get('cliente_id','')).strip() else None,
                fornecedores_json=_json.dumps(l.get('fornecedores', []))
            ))
        db.session.commit()
        flash('Pedido criado!','success')
        return redirect(url_for('pedido_detalhe', pid=p.id))
    clientes_list = Cliente.query.order_by(Cliente.nome).all()
    return render_template('novo_pedido.html', clientes=clientes_list)

def _get_fornecedores_artigo(artigo_ref):
    """Get habitual suppliers for an article from PHC data."""
    if not artigo_ref:
        return []
    try:
        cfg = ConfigPHC.query.first()
        if not cfg or not cfg.server:
            return []
        import pyodbc
        conn_str = (f"DRIVER={{SQL Server}};SERVER={cfg.server};DATABASE={cfg.database};"
                    f"UID={cfg.username};PWD={cfg.password}")
        conn = pyodbc.connect(conn_str, timeout=3)
        cur = conn.cursor()
        # Query PHC for suppliers of this article
        cur.execute("""
            SELECT DISTINCT f.nome, f.ncont, sl.epv
            FROM sl WITH(NOLOCK)
            JOIN fornecedor f WITH(NOLOCK) ON sl.no = f.no
            WHERE sl.ref = ? AND sl.ativo = 1
            ORDER BY sl.epv DESC
        """, artigo_ref)
        rows = cur.fetchall()
        conn.close()
        return [{'nome': r[0], 'nif': r[1] or '', 'preco': float(r[2] or 0)} for r in rows[:5]]
    except Exception as ex:
        return []


@app.route('/pedidos/<int:pid>')
@login_required
def pedido_detalhe(pid):
    p = PedidoCompra.query.get_or_404(pid)
    orcs = Orcamento.query.filter_by(pedido_id=pid).order_by(Orcamento.total).all()
    # Get fornecedores habituais from PHC for each line
    from models import FornecedorPHC
    linhas_json = json.dumps([{
        'id': l.id, 'referencia': l.referencia or '', 'designacao': l.designacao or '',
        'unidade': l.unidade or 'un', 'quantidade': l.quantidade,
        'stock_atual': l.stock_atual, 'preco_custo_ref': l.preco_custo_ref,
        'preco_pcp_ref': l.preco_pcp_ref, 'fornecedor_hab': l.fornecedor_hab or '',
        'observacoes': l.observacoes or '', 'artigo_ref': l.artigo_ref or '',
        'status': l.status or 'nao_encomendado',
        'cliente_id': l.cliente_id or '',
        'fornecedores_phc': _get_fornecedores_artigo(l.artigo_ref),
        'fornecedores_json': l.fornecedores_json or '[]',
    } for l in p.linhas])
    return render_template('pedido_detalhe.html', pedido=p, orcamentos=orcs,
                           melhor=orcs[0] if orcs else None, linhas_json=linhas_json)

@app.route('/pedidos/<int:pid>/linha/<int:lid>/status', methods=['POST'])
@login_required
def linha_status(pid, lid):
    l = LinhaPedido.query.filter_by(id=lid, pedido_id=pid).first_or_404()
    data = request.json or {}
    novo_status = data.get('status', 'nao_encomendado')
    notas = data.get('notas', '')
    status_ant = l.status or 'nao_encomendado'
    if novo_status != status_ant:
        hist = LinhaPedidoHistorico(
            linha_id=lid, pedido_id=pid,
            status_ant=status_ant, status_novo=novo_status,
            user_id=current_user.id, user_nome=current_user.nome,
            notas=notas, data=datetime.now()
        )
        db.session.add(hist)
        l.status = novo_status
        db.session.commit()
    return jsonify({'ok': True, 'status': l.status})

@app.route('/pedidos/<int:pid>/linha/<int:lid>/historico')
@login_required
def linha_historico(pid, lid):
    hist = LinhaPedidoHistorico.query.filter_by(linha_id=lid).order_by(
        LinhaPedidoHistorico.data.desc()).all()
    return jsonify([{
        'status_ant': h.status_ant, 'status_novo': h.status_novo,
        'user_nome': h.user_nome, 'notas': h.notas or '',
        'data': h.data.strftime('%d/%m/%Y %H:%M')
    } for h in hist])


# ── LINHAS DO PEDIDO ──────────────────────────────────────────────────────────

@app.route('/pedidos/<int:pid>/linhas', methods=['POST'])
@login_required
def salvar_linhas(pid):
    """Save/replace all lines of a purchase order (JSON body)."""
    p = PedidoCompra.query.get_or_404(pid)
    if p.estado in ['aprovado','cancelado']:
        return jsonify({'error':'Pedido fechado.'}), 400
    data = request.get_json()
    if not data or 'linhas' not in data:
        return jsonify({'error':'Dados inválidos.'}), 400

    LinhaPedido.query.filter_by(pedido_id=pid).delete()
    for i, l in enumerate(data['linhas']):
        ref = l.get('referencia','').strip()
        artigo = ArtigoPHC.query.filter_by(referencia=ref).first() if ref else None
        linha = LinhaPedido(
            pedido_id=pid, ordem=i,
            artigo_ref      = ref if artigo else None,
            referencia      = ref or l.get('designacao','')[:50],
            designacao      = l.get('designacao','').strip(),
            unidade         = l.get('unidade','un'),
            quantidade      = float(l.get('quantidade',1)),
            stock_atual     = float(artigo.stock_atual if artigo else 0),
            preco_custo_ref = float(artigo.preco_custo if artigo else 0),
            preco_pcp_ref   = float(artigo.preco_custo_ponderado if artigo else 0),
            fornecedor_hab  = l.get('fornecedor_hab','').strip(),
            observacoes     = l.get('observacoes','').strip(),
            cliente_id      = int(l.get('cliente_id')) if str(l.get('cliente_id','')).strip() else None,
            fornecedores_json = __import__('json').dumps(l.get('fornecedores',[]) or l.get('fornecedores_json',[]))
        )
        db.session.add(linha)
    db.session.commit()
    return jsonify({'ok': True, 'total_linhas': len(data['linhas'])})

@app.route('/pedidos/<int:pid>/linhas/<int:lid>', methods=['DELETE'])
@login_required
def apagar_linha(pid, lid):
    l = LinhaPedido.query.filter_by(id=lid, pedido_id=pid).first_or_404()
    db.session.delete(l); db.session.commit()
    return jsonify({'ok': True})

# ── ARTIGOS PHC API ───────────────────────────────────────────────────────────

@app.route('/api/artigos')
@login_required
def api_artigos():
    """Search PHC articles — used by the live search in pedido form."""
    q = request.args.get('q','').strip()
    if len(q) < 2:
        return jsonify([])
    like = f"%{q}%"
    rows = (ArtigoPHC.query
            .filter(db.or_(
                ArtigoPHC.referencia.ilike(like),
                ArtigoPHC.designacao.ilike(like),
                ArtigoPHC.familia.ilike(like)))
            .order_by(ArtigoPHC.referencia)
            .limit(30).all())
    return jsonify([{
        'referencia':   a.referencia,
        'designacao':   a.designacao,
        'stock_atual':  a.stock_atual,
        'unidade':      a.unidade,
        'preco_custo':  a.preco_custo,
        'pcp':          a.preco_custo_ponderado,
        'familia':      a.familia,
    } for a in rows])

@app.route('/api/artigos/<ref>')
@login_required
def api_artigo_detalhe(ref):
    a = ArtigoPHC.query.filter_by(referencia=ref).first()
    if not a: return jsonify({'error':'Não encontrado'}), 404
    # Try to get purchase history if PHC is configured
    historico = []
    try:
        cfg = ConfigPHC.query.first()
        if cfg and cfg.ultima_sync:
            from phc_sync import get_historico_compras
            historico = get_historico_compras(cfg, ref)
    except Exception:
        pass
    # Get known suppliers for this article from purchase history + FornecedorPHC
    fornecedores = []
    seen = set()
    for h in historico:
        nome = h.get('fornecedor_nome', '')
        if nome and nome not in seen:
            seen.add(nome)
            # Try to find email in local FornecedorPHC cache
            forn_db = FornecedorPHC.query.filter(
                FornecedorPHC.nome.ilike(f'%{nome}%')
            ).first()
            fornecedores.append({
                'nome':   nome,
                'no':     h.get('fornecedor_no'),
                'email':  forn_db.email if forn_db else '',
                'tel':    forn_db.telefone if forn_db else '',
                'ultimo_preco': h.get('preco', 0),
                'ultima_compra': h.get('data_compra', ''),
            })

    return jsonify({
        'referencia':  a.referencia,
        'designacao':  a.designacao,
        'stock_atual': a.stock_atual,
        'unidade':     a.unidade,
        'preco_custo': a.preco_custo,
        'pcp':         a.preco_custo_ponderado,
        'familia':     a.familia,
        'taxa_iva':    a.taxa_iva,
        'historico':   historico,
        'fornecedores': fornecedores,
    })

# ── ORÇAMENTOS ────────────────────────────────────────────────────────────────

@app.route('/pedidos/<int:pid>/upload', methods=['POST'])
@login_required
def upload_orcamento(pid):
    p = PedidoCompra.query.get_or_404(pid)
    if p.estado not in ['aberto','em_analise']:
        return jsonify({'error':'Pedido não permite orçamentos neste estado.'}), 400
    if Orcamento.query.filter_by(pedido_id=pid).count() >= 3:
        return jsonify({'error':'Já existem 3 orçamentos neste pedido.'}), 400
    if 'pdf' not in request.files:
        return jsonify({'error':'Nenhum ficheiro.'}), 400
    file = request.files['pdf']
    if not file.filename or not allowed_file(file.filename):
        return jsonify({'error':'Apenas PDFs são aceites.'}), 400

    fname = secure_filename(f"pc{pid}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
    fpath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
    file.save(fpath)

    dados, erro = analyze_pdf_with_claude(extract_pdf_text(fpath), file.filename)
    forn_manual = request.form.get('fornecedor','').strip()
    if not dados:
        dados = {'empresa': forn_manual or 'Fornecedor Desconhecido',
                 'total':0,'subtotal':0,'desconto_total':0,'desconto_percentagem':0,
                 'iva_valor':0,'numero_orcamento':None,'data_orcamento':None,
                 'validade':None,'contacto':None,'nif':None,'observacoes':erro,'items':[]}
    if forn_manual: dados['empresa'] = forn_manual

    orc = Orcamento(
        pedido_id=pid,
        empresa=dados.get('empresa','Desconhecido'),
        nif=dados.get('nif'),
        numero_orcamento=dados.get('numero_orcamento'),
        data_orcamento=datetime.strptime(dados['data_orcamento'],'%Y-%m-%d').date() if dados.get('data_orcamento') else None,
        validade=datetime.strptime(dados['validade'],'%Y-%m-%d').date() if dados.get('validade') else None,
        subtotal=float(dados.get('subtotal') or 0),
        desconto_total=float(dados.get('desconto_total') or 0),
        desconto_percentagem=float(dados.get('desconto_percentagem') or 0),
        iva_valor=float(dados.get('iva_valor') or 0),
        total=float(dados.get('total') or 0),
        moeda=dados.get('moeda','EUR'),
        contacto=dados.get('contacto'),
        observacoes=dados.get('observacoes'),
        ficheiro_pdf=fname,
        carregado_por=current_user.id,
        data_upload=datetime.now(),
        dados_brutos=json.dumps(dados))
    db.session.add(orc); db.session.flush()

    for item in dados.get('items',[]):
        db.session.add(ItemOrcamento(
            orcamento_id=orc.id,
            descricao=item.get('descricao',''),
            referencia=item.get('referencia'),
            quantidade=float(item.get('quantidade') or 1),
            unidade=item.get('unidade','un'),
            preco_unitario=float(item.get('preco_unitario') or 0),
            desconto_item=float(item.get('desconto_item') or 0),
            total_item=float(item.get('total_item') or 0)))

    if p.estado == 'aberto': p.estado = 'em_analise'
    db.session.commit()
    return jsonify({'success':True,'orcamento_id':orc.id,'empresa':orc.empresa,
                    'total':orc.total,'num_items':len(dados.get('items',[]))})


# ── ALIAS MATCHING ────────────────────────────────────────────────────────────

@app.route('/pedidos/<int:pid>/match', methods=['POST'])
@login_required
def match_orcamento(pid):
    """Run alias matching on all orcamentos of a pedido."""
    p = PedidoCompra.query.get_or_404(pid)
    aliases = AliasArtigo.query.all()
    results = []
    for orc in p.orcamentos:
        from alias_matcher import match_orcamento_items
        r = match_orcamento_items(orc, p, aliases)
        results.extend(r)
    db.session.commit()
    return jsonify({'ok': True, 'matches': len([r for r in results if r['artigo_ref']]),'total': len(results)})


@app.route('/api/alias', methods=['POST'])
@login_required
def criar_alias():
    """Manually link a supplier description to a PHC article."""
    data = request.get_json()
    artigo_ref     = data.get('artigo_ref', '').strip()
    descricao_orig = data.get('descricao_orig', '').strip()
    fornecedor     = data.get('fornecedor', '').strip()
    referencia_forn= data.get('referencia_forn', '').strip()
    item_id        = data.get('item_id')

    if not artigo_ref or not descricao_orig:
        return jsonify({'error': 'artigo_ref e descricao_orig são obrigatórios'}), 400

    artigo = ArtigoPHC.query.filter_by(referencia=artigo_ref).first()
    if not artigo:
        return jsonify({'error': f'Artigo {artigo_ref} não encontrado'}), 404

    from alias_matcher import save_alias
    save_alias(db, artigo_ref, descricao_orig, fornecedor,
               referencia_forn, current_user.id, confianca=1.0)

    # Update the item if provided
    if item_id:
        item = ItemOrcamento.query.get(item_id)
        if item:
            item.artigo_ref_match = artigo_ref
            item.match_confianca  = 1.0
            db.session.commit()

    return jsonify({'ok': True, 'artigo_ref': artigo_ref, 'designacao': artigo.designacao})


@app.route('/api/alias/<int:alias_id>', methods=['DELETE'])
@login_required
def apagar_alias(alias_id):
    alias = AliasArtigo.query.get_or_404(alias_id)
    db.session.delete(alias)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/aliases')
@login_required
def listar_aliases():
    """List all aliases, optionally filtered by artigo_ref."""
    ref = request.args.get('ref', '').strip()
    q = AliasArtigo.query
    if ref:
        q = q.filter_by(artigo_ref=ref)
    aliases = q.order_by(AliasArtigo.vezes_usado.desc()).all()
    return jsonify([{
        'id':             a.id,
        'artigo_ref':     a.artigo_ref,
        'fornecedor':     a.fornecedor,
        'descricao_orig': a.descricao_orig,
        'referencia_forn':a.referencia_forn,
        'confianca':      a.confianca,
        'vezes_usado':    a.vezes_usado,
        'data_criacao':   a.data_criacao.strftime('%d/%m/%Y') if a.data_criacao else '',
    } for a in aliases])


@app.route('/pedidos/<int:pid>/comparacao')
@login_required
def comparacao_pedido(pid):
    """Comparison matrix page: pedido lines vs orcamento items."""
    p    = PedidoCompra.query.get_or_404(pid)
    orcs = Orcamento.query.filter_by(pedido_id=pid).all()
    from alias_matcher import build_comparison_matrix
    matrix = build_comparison_matrix(p, orcs)
    return render_template('comparacao.html', pedido=p, orcamentos=orcs, matrix=matrix)


# ── PDF PREVIEW ───────────────────────────────────────────────────────────────

@app.route('/uploads/preview/<filename>')
@login_required
def preview_pdf(filename):
    """Serve PDF for inline preview."""
    return send_from_directory(
        app.config['UPLOAD_FOLDER'], filename,
        mimetype='application/pdf',
        as_attachment=False
    )

@app.route('/pedidos/<int:pid>/aprovar', methods=['POST'])
@login_required
def aprovar_pedido(pid):
    p = PedidoCompra.query.get_or_404(pid)
    oid = request.form.get('orcamento_id')
    if oid:
        for o in p.orcamentos: o.selecionado = False
        o2 = Orcamento.query.get(oid)
        if o2: o2.selecionado = True
    p.estado = 'aprovado'; p.aprovado_por = current_user.id; p.data_aprovacao = datetime.now()
    db.session.commit(); flash('Pedido aprovado!','success')
    return redirect(url_for('pedido_detalhe', pid=pid))

@app.route('/pedidos/<int:pid>/cancelar', methods=['POST'])
@login_required
def cancelar_pedido(pid):
    p = PedidoCompra.query.get_or_404(pid)
    p.estado = 'cancelado'; db.session.commit(); flash('Pedido cancelado.','info')
    return redirect(url_for('pedido_detalhe', pid=pid))

@app.route('/orcamentos/<int:oid>')
@login_required
def orcamento_detalhe(oid):
    return render_template('orcamento_detalhe.html', orc=Orcamento.query.get_or_404(oid))

@app.route('/orcamentos/<int:oid>/apagar', methods=['POST'])
@login_required
def apagar_orcamento(oid):
    orc = Orcamento.query.get_or_404(oid); pid = orc.pedido_id
    fp = os.path.join(app.config['UPLOAD_FOLDER'], orc.ficheiro_pdf)
    if os.path.exists(fp): os.remove(fp)
    ItemOrcamento.query.filter_by(orcamento_id=oid).delete()
    db.session.delete(orc); db.session.commit()
    flash('Orçamento removido.','info')
    return redirect(url_for('pedido_detalhe', pid=pid))

@app.route('/uploads/<filename>')
def download_pdf(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)



# ── REPOSIÇÃO ─────────────────────────────────────────────────────────────────

@app.route('/reposicao')
@login_required
def reposicao():
    from reposicao import CONFIG_PADRAO, analisar_todos
    cfg_phc    = ConfigPHC.query.first()
    cfg_global = ConfigReposicao.query.filter_by(artigo_ref=None).first()
    if not cfg_global:
        cfg_global = ConfigReposicao()
        db.session.add(cfg_global)
        db.session.commit()

    # Build config dict
    config = dict(CONFIG_PADRAO)
    if cfg_global:
        config.update({
            'meses_historico':             cfg_global.meses_historico or 60,
            'lead_time_dias':              cfg_global.lead_time_dias or 7,
            'fator_seguranca':             cfg_global.fator_seguranca or 1.5,
            'meses_cobertura':             cfg_global.meses_cobertura or 2,
            'custo_encomenda':             cfg_global.custo_encomenda or 25,
            'taxa_posse_anual':            cfg_global.taxa_posse_anual or 0.20,
            'quantidade_minima_encomenda': cfg_global.quantidade_minima_encomenda or 1,
            'alertar_dias_cobertura':      cfg_global.alertar_dias_cobertura or 30,
            'min_anos_historico':          getattr(cfg_global, 'min_anos_historico', 2) or 2,
            'min_meses_com_venda':         getattr(cfg_global, 'min_meses_com_venda', 3) or 3,
            'min_total_vendido':           getattr(cfg_global, 'min_total_vendido', 3) or 3,
            'ignorar_sem_movimento_anos':  getattr(cfg_global, 'ignorar_sem_movimento_anos', 3) or 3,
            'min_facturas_sugerir':        getattr(cfg_global, 'min_facturas_sugerir', 8) or 8,
            'min_facturas_sugerir':        getattr(cfg_global, 'min_facturas_sugerir', 5) or 5,
        })

    # Run analysis only if requested
    analise = request.args.get('analise') == '1'
    resultados = []
    erro = None
    if analise:
        try:
            artigos = ArtigoPHC.query.filter(ArtigoPHC.stock_atual >= 0).all()
            resultados = analisar_todos(cfg_phc, config, artigos)
        except Exception as e:
            import traceback
            erro = str(e)
            logger.error(traceback.format_exc())

    total_artigos    = ArtigoPHC.query.count()
    precisam         = [r for r in resultados if r.get('precisa_encomendar')]
    criticos         = [r for r in resultados if r.get('urgencia') == 'critico']
    sem_dados        = [r for r in resultados if not r.get('tem_dados')]

    return render_template('reposicao.html',
        cfg=cfg_global,
        config_padrao=config,
        total_artigos=total_artigos,
        resultados=resultados,
        analise_feita=analise,
        precisam=precisam,
        criticos=criticos,
        sem_dados=sem_dados,
        erro=erro,
    )


@app.route('/reposicao/config', methods=['POST'])
@login_required
def salvar_config_reposicao():
    """Save global or per-article replenishment config."""
    artigo_ref = request.form.get('artigo_ref', '').strip() or None
    cfg = ConfigReposicao.query.filter_by(artigo_ref=artigo_ref).first()
    if not cfg:
        cfg = ConfigReposicao(artigo_ref=artigo_ref)
        db.session.add(cfg)

    cfg.meses_historico             = int(request.form.get('meses_historico', 24))
    cfg.lead_time_dias              = float(request.form.get('lead_time_dias', 7))
    cfg.fator_seguranca             = float(request.form.get('fator_seguranca', 1.5))
    cfg.meses_cobertura             = int(request.form.get('meses_cobertura', 2))
    cfg.custo_encomenda             = float(request.form.get('custo_encomenda', 25))
    taxa_raw = float(request.form.get('taxa_posse_anual', 20))
    # Accept both 0.20 and 20 (percent) - if > 1, treat as percentage
    cfg.taxa_posse_anual = taxa_raw / 100 if taxa_raw > 1 else taxa_raw
    cfg.quantidade_minima_encomenda = float(request.form.get('quantidade_minima_encomenda', 1))
    cfg.alertar_dias_cobertura      = int(request.form.get('alertar_dias_cobertura', 30))
    cfg.ignorar_parados_dias        = int(request.form.get('ignorar_parados_dias', 365))
    cfg.min_anos_historico          = float(request.form.get('min_anos_historico', 2) or 2)
    cfg.min_meses_com_venda         = int(request.form.get('min_meses_com_venda', 3) or 3)
    cfg.min_total_vendido           = float(request.form.get('min_total_vendido', 3) or 3)
    cfg.ignorar_sem_movimento_anos  = float(request.form.get('ignorar_sem_movimento_anos', 3) or 3)
    cfg.min_facturas_sugerir        = int(request.form.get('min_facturas_sugerir', 8) or 8)
    cfg.min_facturas_sugerir        = int(request.form.get('min_facturas_sugerir', 5) or 5)
    cfg.atualizado_em               = datetime.now()
    cfg.atualizado_por              = current_user.id
    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True})
    flash('Configuração de reposição guardada.', 'success')
    # If requested, redirect to analysis
    if request.args.get('redirect_analise') == '1':
        return redirect(url_for('reposicao') + '?analise=1')
    return redirect(url_for('reposicao'))


@app.route('/reposicao/analisar/<ref>')
@login_required
def analisar_artigo(ref):
    """Analyse one article and show detailed breakdown page."""
    from reposicao import CONFIG_PADRAO, calcular_metricas, SQL_VENDAS_ARTIGO

    cfg_phc    = ConfigPHC.query.first()
    cfg_global = ConfigReposicao.query.filter_by(artigo_ref=None).first()
    artigo     = ArtigoPHC.query.filter_by(referencia=ref).first_or_404()

    config = dict(CONFIG_PADRAO)
    if cfg_global:
        for k in list(CONFIG_PADRAO.keys()) + ['min_anos_historico','min_meses_com_venda',
                                                'min_total_vendido','ignorar_sem_movimento_anos']:
            v = getattr(cfg_global, k, None)
            if v is not None:
                config[k] = v

    # Fetch per-month sales and diversity from PHC
    vendas_por_mes = {}
    vendas_lista   = []
    diversidade    = {}

    if cfg_phc and cfg_phc.ultima_sync:
        try:
            from phc_sync import get_phc_connection
            from reposicao import SQL_DIVERSIDADE_ARTIGO
            conn   = get_phc_connection(cfg_phc)
            cursor = conn.cursor()
            meses  = config.get('meses_historico', 60)
            cursor.execute(SQL_VENDAS_ARTIGO, ref, meses)
            for row in cursor.fetchall():
                ano, mes, total, nfat = row[0], row[1], float(row[2]), row[3]
                vendas_por_mes[(ano, mes)] = total
                vendas_lista.append({'ano':ano,'mes':mes,'total':total,'nfat':nfat})
            # Fetch invoice detail
            from reposicao import SQL_DETALHE_VENDAS
            cursor.execute(SQL_DETALHE_VENDAS, ref, meses)
            vendas_detalhe = []
            for row in cursor.fetchall():
                vendas_detalhe.append({
                    'num_fatura':   str(row[0]) if row[0] else '-',
                    'cliente_nome': (row[1] or '').strip(),
                    'data':         row[2].strftime('%d/%m/%Y') if row[2] else '',
                    'quantidade':   float(row[3]),
                    'preco_venda':  float(row[4] or 0),
                })
            # Diversity
            cursor.execute(SQL_DIVERSIDADE_ARTIGO, ref)
            row = cursor.fetchone()
            if row:
                diversidade = {'num_clientes': int(row[0] or 0),
                               'num_facturas': int(row[1] or 0)}
            conn.close()
        except Exception as e:
            app.logger.warning(f'PHC error analisar_artigo: {e}')

    stock  = artigo.stock_atual or 0
    preco  = artigo.preco_custo_ponderado or artigo.preco_custo or 0
    result = calcular_metricas(vendas_por_mes, stock, preco, config)
    result['referencia'] = ref
    result['designacao'] = artigo.designacao or ''
    result['stock_atual'] = stock
    result['preco_custo'] = preco
    result['unidade']    = artigo.unidade or 'un'

    # Compute period info for display
    periodo = {}
    if vendas_por_mes:
        chaves = sorted(vendas_por_mes.keys())
        primeiro = chaves[0]
        ultimo   = chaves[-1]
        from datetime import datetime as _dt
        meses_periodo = (_dt.now().year - primeiro[0])*12 + (_dt.now().month - primeiro[1]) + 1
        total_vendido = sum(vendas_por_mes.values())
        periodo = {
            'primeiro_mes': f"{primeiro[1]:02d}/{primeiro[0]}",
            'ultimo_mes':   f"{ultimo[1]:02d}/{ultimo[0]}",
            'meses_periodo': meses_periodo,
            'anos_periodo':  round(meses_periodo/12, 1),
            'total_vendido': total_vendido,
            'meses_com_venda': len([v for v in vendas_por_mes.values() if v>0]),
            'media_mensal':  round(total_vendido/meses_periodo, 3),
            'media_anual':   round(total_vendido/meses_periodo*12, 2),
        }

    # Sort vendas_lista by date
    vendas_lista.sort(key=lambda x: (x['ano'], x['mes']))

    # Compute relevance score
    score_info = {}
    previsoes  = {}
    if periodo:
        from reposicao import score_relevancia, calcular_todos_metodos
        previsoes = calcular_todos_metodos(vendas_por_mes, periodo.get('meses_periodo', 1))
        score_info = score_relevancia(
            total_vendido = periodo.get('total_vendido', 0),
            num_clientes  = diversidade.get('num_clientes', 0),
            num_facturas  = diversidade.get('num_facturas', 0),
            meses_periodo = periodo.get('meses_periodo', 1),
            preco_custo   = preco,
            config        = config,
        )

    return render_template('reposicao_artigo.html',
        artigo=artigo, result=result, vendas_lista=vendas_lista,
        periodo=periodo, config=config, score_info=score_info,
        diversidade=diversidade, previsoes=previsoes,
        vendas_detalhe=vendas_detalhe if 'vendas_detalhe' in dir() else [])


@app.route('/reposicao/lista')
@login_required
def lista_reposicao():
    """Return all articles needing replenishment as JSON (for table)."""
    from reposicao import analisar_artigo as _analisar, CONFIG_PADRAO

    cfg_phc    = ConfigPHC.query.first()
    cfg_global = ConfigReposicao.query.filter_by(artigo_ref=None).first()

    config = dict(CONFIG_PADRAO)
    if cfg_global:
        for k in CONFIG_PADRAO:
            v = getattr(cfg_global, k, None)
            if v is not None:
                config[k] = v

    familia = request.args.get('familia', '').strip()
    apenas_alertas = request.args.get('alertas', '0') == '1'

    query = ArtigoPHC.query
    if familia:
        query = query.filter_by(familia=familia)

    artigos = query.order_by(ArtigoPHC.referencia).limit(200).all()
    resultados = []
    for a in artigos:
        cfg_art = ConfigReposicao.query.filter_by(artigo_ref=a.referencia).first()
        cfg_final = dict(config)
        if cfg_art:
            for k in CONFIG_PADRAO:
                v = getattr(cfg_art, k, None)
                if v is not None: cfg_final[k] = v

        r = _analisar(cfg_phc, cfg_final, a.referencia,
                      stock_atual=a.stock_atual,
                      preco_custo=a.preco_custo_ponderado or a.preco_custo)

        if apenas_alertas and not r.get('precisa_encomendar'):
            continue
        # Skip completely inactive articles
        dias_sem_mov = r.get('dias_sem_movimento', 0) or 0
        if dias_sem_mov > cfg_final.get('ignorar_parados_dias', 365):
            continue

        resultados.append(r)

    # Sort: needs ordering first, then by coverage days
    resultados.sort(key=lambda x: (
        not x.get('precisa_encomendar', False),
        x.get('dias_cobertura_atual') or 9999
    ))

    return jsonify(resultados)


@app.route('/reposicao/gerar_pedido', methods=['POST'])
@login_required
def gerar_pedido_reposicao():
    """Create a PedidoCompra from selected replenishment suggestions."""
    data = request.get_json()
    artigos_sel = data.get('artigos', [])
    if not artigos_sel:
        return jsonify({'error': 'Nenhum artigo selecionado.'}), 400

    titulo = data.get('titulo') or f"Reposição de stock — {datetime.now().strftime('%d/%m/%Y')}"
    p = PedidoCompra(
        titulo=titulo,
        descricao='Gerado automaticamente pelo motor de reposição.',
        prioridade='normal',
        estado='aberto',
        criado_por=current_user.id,
        data_criacao=datetime.now()
    )
    db.session.add(p)
    db.session.flush()

    for i, a in enumerate(artigos_sel):
        artigo = ArtigoPHC.query.filter_by(referencia=a['referencia']).first()
        linha = LinhaPedido(
            pedido_id=p.id, ordem=i,
            artigo_ref=a['referencia'],
            referencia=a['referencia'],
            designacao=a.get('designacao', ''),
            unidade=a.get('unidade', 'un'),
            quantidade=float(a.get('quantidade_sugerida', 1)),
            stock_atual=float(artigo.stock_atual if artigo else 0),
            preco_custo_ref=float(artigo.preco_custo if artigo else 0),
            preco_pcp_ref=float(artigo.preco_custo_ponderado if artigo else 0),
        )
        db.session.add(linha)

    db.session.commit()
    return jsonify({'ok': True, 'pedido_id': p.id, 'titulo': titulo})

# ── BAK UPLOAD & RESTORE ──────────────────────────────────────────────────────

@app.route('/admin/phc/upload_bak', methods=['POST'])
@login_required
def upload_bak():
    if not current_user.is_admin:
        return jsonify({'error': 'Sem permissão'}), 403
    if 'bak' not in request.files:
        return jsonify({'error': 'Nenhum ficheiro enviado.'}), 400
    file = request.files['bak']
    if not file.filename.lower().endswith('.bak'):
        return jsonify({'error': 'Apenas ficheiros .bak são aceites.'}), 400
    bak_path = os.path.join(app.config['BAK_FOLDER'], 'phc_backup.bak')
    file.save(bak_path)
    cfg = ConfigPHC.query.first()
    if not cfg:
        return jsonify({'error': 'Configure primeiro a ligação ao SQL Server.'}), 400
    win_path = os.path.abspath(bak_path)
    messages = []
    def progress(msg): messages.append(msg)
    try:
        from bak_restore import restore_bak, verify_phc_database
        ok, msg = restore_bak(cfg, win_path, progress_callback=progress)
        if not ok:
            return jsonify({'error': msg, 'log': messages}), 500
        valid, vmsg, stats = verify_phc_database(cfg)
        if not valid:
            return jsonify({'error': vmsg, 'log': messages}), 500
        messages.append(vmsg)
        from phc_sync import sync_all
        sync_stats = sync_all(app, cfg)
        messages.append(
            f"Sincronização: {sync_stats['artigos']['inseridos']} artigos novos / "
            f"{sync_stats['artigos']['atualizados']} atualizados"
        )
        return jsonify({
            'success': True, 'log': messages,
            'stats': {
                'artigos': sync_stats['artigos']['inseridos'] + sync_stats['artigos']['atualizados'],
                'fornecedores': sync_stats['fornecedores']['inseridos'] + sync_stats['fornecedores']['atualizados'],
            }
        })
    except ImportError as e:
        return jsonify({'error': f'Módulo em falta: {e}. Instale pyodbc.'}), 500
    except Exception as e:
        messages.append(f'Erro: {e}')
        return jsonify({'error': str(e), 'log': messages}), 500


@app.route('/admin/phc/check_sqlserver')
@login_required
def check_sqlserver():
    if not current_user.is_admin:
        return jsonify({'error': 'Sem permissão'}), 403
    cfg = ConfigPHC.query.first()
    if not cfg:
        return jsonify({'ok': False, 'msg': 'Sem configuração SQL Server.'})
    try:
        from bak_restore import check_sqlserver_available
        ok, msg = check_sqlserver_available(cfg)
        return jsonify({'ok': ok, 'msg': msg})
    except ImportError:
        return jsonify({'ok': False, 'msg': 'pyodbc não instalado.'})

# ── PHC CONFIG & SYNC ─────────────────────────────────────────────────────────

@app.route('/admin/phc', methods=['GET','POST'])
@login_required
def admin_phc():
    if not current_user.is_admin:
        flash('Sem permissão.','error'); return redirect(url_for('dashboard'))
    cfg = ConfigPHC.query.first() or ConfigPHC()

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'save':
            cfg.servidor     = request.form.get('servidor', r'.\SQLEXPRESS').strip()
            cfg.porta        = int(request.form.get('porta', 1433))
            cfg.base_dados   = request.form.get('base_dados','PHC').strip()
            cfg.autenticacao = request.form.get('autenticacao','sql')
            cfg.utilizador   = request.form.get('utilizador','sa').strip()
            pw = request.form.get('password','')
            if pw: cfg.password = pw
            cfg.sync_auto = request.form.get('sync_auto') == 'on'
            cfg.sync_hora = request.form.get('sync_hora','06:00')
            if not cfg.id: db.session.add(cfg)
            db.session.commit()
            flash('Configuração guardada.','success')

        elif action == 'test':
            if not cfg.id:
                flash('Guarde primeiro a configuração.','error')
            else:
                try:
                    from phc_sync import test_connection
                    ok, msg = test_connection(cfg)
                    flash(msg, 'success' if ok else 'error')
                except ImportError:
                    flash('pyodbc não instalado. Instale com: pip install pyodbc','error')

        elif action == 'sync':
            if not cfg.id:
                flash('Configure e guarde a ligação primeiro.','error')
            else:
                try:
                    from phc_sync import sync_artigos, sync_fornecedores
                    ins_a, upd_a, err_a = sync_artigos(cfg)
                    ins_f, upd_f, err_f = sync_fornecedores(cfg)
                    from phc_sync import sync_clientes
                    ins_c, upd_c, err_c = sync_clientes(cfg)
                    cfg.ultima_sync = datetime.now()
                    db.session.commit()
                    msg = (f"✅ Sync concluída — "
                           f"Artigos: {ins_a}+{upd_a} | "
                           f"Fornecedores: {ins_f}+{upd_f} | "
                           f"Clientes: {ins_c}+{upd_c}")
                    if err_a or err_f or err_c:
                        msg += f" | {len(err_a+err_f+err_c)} erros"
                    flash(msg, 'success')
                except Exception as e:
                    import traceback
                    flash(f'Erro na sincronização: {e}', 'error')
                    logger.error(traceback.format_exc())

        return redirect(url_for('admin_phc'))

    total_artigos  = ArtigoPHC.query.count()
    total_forn     = FornecedorPHC.query.count()
    total_clientes = Cliente.query.count()
    return render_template('admin_phc.html', cfg=cfg,
                           total_artigos=total_artigos,
                           total_forn=total_forn,
                           total_clientes=total_clientes)

@app.route('/admin/utilizadores')
@login_required
def admin_utilizadores():
    if not current_user.is_admin:
        flash('Sem permissão.','error'); return redirect(url_for('dashboard'))
    perfis = Perfil.query.order_by(Perfil.nome).all()
    return render_template('admin_utilizadores.html', utilizadores=User.query.all(), perfis=perfis, get_user_perfil_id=get_user_perfil_id)

@app.route('/admin/utilizadores/novo', methods=['POST'])
@login_required
def novo_utilizador():
    if not current_user.is_admin: return jsonify({'error':'Sem permissão'}), 403
    username = request.form.get('username','').strip()
    if User.query.filter_by(username=username).first():
        flash('Username já existe.','error')
        return redirect(url_for('admin_utilizadores'))
    db.session.add(User(
        username=username, nome=request.form.get('nome','').strip(),
        password_hash=generate_password_hash(request.form.get('password','')),
        is_admin=request.form.get('is_admin')=='on',
        email=request.form.get('email','').strip(),
        departamento=request.form.get('departamento','').strip()))
    db.session.commit(); flash('Utilizador criado.','success')
    return redirect(url_for('admin_utilizadores'))

@app.route('/admin/utilizadores/<int:uid>/apagar', methods=['POST'])
@login_required
def apagar_utilizador(uid):
    if not current_user.is_admin: return jsonify({'error':'Sem permissão'}), 403
    if uid == current_user.id:
        flash('Não pode apagar a sua conta.','error')
        return redirect(url_for('admin_utilizadores'))
    db.session.delete(User.query.get_or_404(uid)); db.session.commit()
    flash('Utilizador removido.','info')
    return redirect(url_for('admin_utilizadores'))

# ── INIT ──────────────────────────────────────────────────────────────────────

# ── CHANGELOG ─────────────────────────────────────────────────────────────────

@app.route('/changelog')
@login_required
def changelog():
    try:
        entries = ChangelogEntry.query.order_by(ChangelogEntry.criado_em.desc()).limit(100).all()
    except Exception:
        entries = []
    return render_template('changelog.html', entries=entries)


# ── EMAIL CONSULTA ─────────────────────────────────────────────────────────────

@app.route('/api/email_consulta', methods=['POST'])
@login_required
def gerar_email_consulta():
    """Generate a quote request email text for a list of articles."""
    data       = request.get_json()
    fornecedor = data.get('fornecedor', '')
    artigos    = data.get('artigos', [])   # [{referencia, designacao, quantidade, unidade}]
    empresa    = data.get('empresa_propria', 'a nossa empresa')

    if not artigos:
        return jsonify({'error': 'Sem artigos'}), 400

    linhas_artigos = '\n'.join([
        f"  - {a.get('referencia','')} | {a.get('designacao','')} | Qtd: {a.get('quantidade',1)} {a.get('unidade','un')}"
        for a in artigos
    ])

    texto = f"""Boa tarde,

Gostaríamos de solicitar orçamento para os seguintes artigos/materiais:

{linhas_artigos}

Agradecemos o envio do vosso melhor preço, incluindo condições de pagamento, prazo de entrega e validade da proposta.

Para qualquer esclarecimento, estamos disponíveis.

Com os melhores cumprimentos,
{current_user.nome}
{empresa}"""

    return jsonify({
        'ok': True,
        'texto': texto,
        'assunto': f'Pedido de Orçamento — {len(artigos)} referência(s)',
        'para': data.get('email_fornecedor', ''),
    })


# ── PENDING MATCHES ───────────────────────────────────────────────────────────

@app.route('/pendentes')
@login_required
def pendentes():
    """All unconfirmed matches across all pedidos."""
    pedido_id = request.args.get('pedido_id', type=int)
    q = PendingMatch.query.filter_by(confirmado=False)
    if pedido_id:
        q = q.filter_by(pedido_id=pedido_id)
    pendentes = q.order_by(PendingMatch.criado_em.desc()).all()

    # Count by pedido for sidebar badge
    total = PendingMatch.query.filter_by(confirmado=False).count()
    artigos_phc = ArtigoPHC.query.order_by(ArtigoPHC.referencia).all()
    artigos_phc_json = json.dumps([{
        'referencia': a.referencia, 'designacao': a.designacao
    } for a in artigos_phc])
    return render_template('pendentes.html', pendentes=pendentes,
                           total=total, pedido_id_filtro=pedido_id,
                           artigos_phc=artigos_phc,
                           artigos_phc_json=artigos_phc_json)


@app.route('/pendentes/<int:pm_id>/confirmar', methods=['POST'])
@login_required
def confirmar_match(pm_id):
    """Confirm a pending match — saves alias and updates item."""
    pm = PendingMatch.query.get_or_404(pm_id)
    data = request.get_json()
    artigo_ref = data.get('artigo_ref', '').strip()

    if not artigo_ref:
        return jsonify({'error': 'artigo_ref obrigatório'}), 400

    # Verify article exists
    artigo = ArtigoPHC.query.filter_by(referencia=artigo_ref).first()
    if not artigo:
        return jsonify({'error': f'Artigo {artigo_ref} não encontrado'}), 404

    # Update pending match
    pm.confirmado        = True
    pm.artigo_ref_final  = artigo_ref
    pm.confirmado_por    = current_user.id
    pm.data_confirmacao  = datetime.now()

    # Update the item
    from models import ItemOrcamento
    item = ItemOrcamento.query.get(pm.item_id)
    if item:
        item.artigo_ref_match = artigo_ref
        item.match_confianca  = 1.0

    # Save alias so it learns
    from alias_matcher import save_alias
    save_alias(db, artigo_ref,
               pm.descricao_forn, pm.fornecedor,
               pm.referencia_forn, current_user.id, confianca=1.0)

    db.session.commit()
    return jsonify({'ok': True, 'artigo_ref': artigo_ref,
                    'designacao': artigo.designacao})


@app.route('/pendentes/<int:pm_id>/ignorar', methods=['POST'])
@login_required
def ignorar_match(pm_id):
    """Mark as confirmed with no match (intentionally unlinked)."""
    pm = PendingMatch.query.get_or_404(pm_id)
    pm.confirmado       = True
    pm.artigo_ref_final = None
    pm.confirmado_por   = current_user.id
    pm.data_confirmacao = datetime.now()
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/pendentes/count')
@login_required
def count_pendentes():
    """Badge count for sidebar."""
    n = PendingMatch.query.filter_by(confirmado=False).count()
    return jsonify({'count': n})


# ══════════════════════════════════════════════════════════════
#  CONFIG GERAL (tema, empresa, backup)
# ══════════════════════════════════════════════════════════════

def get_config_geral():
    cfg = ConfigGeral.query.first()
    if not cfg:
        cfg = ConfigGeral()
        db.session.add(cfg)
        db.session.commit()
    return cfg


@app.context_processor
def inject_config():
    """Make ConfigGeral and datetime available in all templates."""
    from datetime import datetime as _dt
    try:
        return {
            'cfg_geral': get_config_geral(),
            'cfg_ia': ConfigIA.query.first(),
            'now': _dt.now
        }
    except Exception:
        return {'cfg_geral': None, 'cfg_ia': None, 'now': _dt.now}

@app.route('/admin/testar-pasta', methods=['POST'])
@login_required
def admin_testar_pasta():
    if not current_user.is_admin:
        return jsonify({'ok': False, 'msg': 'Sem permissão'})
    data = request.get_json() or {}
    path = data.get('path', '').strip()
    if not path:
        return jsonify({'ok': False, 'msg': 'Caminho vazio'})
    try:
        if not os.path.isabs(path):
            path = os.path.join(os.path.dirname(os.path.abspath(__file__)), path)
        path = os.path.normpath(path)
        if os.path.isdir(path):
            backups = [f for f in os.listdir(path) if f.startswith('comprasnet_backup_')]
            try:
                import shutil as _sh
                free = f'{_sh.disk_usage(path).free // (1024**3):.1f} GB livres'
            except: free = ''
            msg = f'Pasta existe — {len(backups)} backup(s)'
            if free: msg += f' — {free}'
            return jsonify({'ok': True, 'msg': msg})
        else:
            parent = os.path.dirname(path)
            if os.path.isdir(parent):
                os.makedirs(path, exist_ok=True)
                return jsonify({'ok': True, 'msg': f'Pasta criada com sucesso'})
            return jsonify({'ok': False, 'msg': f'Caminho nao acessivel: {path}'})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)})



@app.route('/admin/config', methods=['GET', 'POST'])
@login_required
def admin_config():
    if not current_user.is_admin:
        flash('Sem permissão.', 'error')
        return redirect(url_for('dashboard'))
    # Ensure all columns exist (safe for older DBs)
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            for col, typ in [('dashboard_layouts','TEXT')]:
                try:
                    conn.execute(text(f"ALTER TABLE config_geral ADD COLUMN {col} {typ}"))
                    conn.commit()
                except Exception:
                    pass
    except Exception:
        pass
    cfg = get_config_geral()
    if request.method == 'POST':
        action = request.form.get('action', 'save')
        if action == 'save':
            cfg.empresa_nome       = request.form.get('empresa_nome', '').strip() or 'ComprasNet'
            cfg.empresa_abrev      = request.form.get('empresa_abrev', '').strip()
            cfg.empresa_nif        = request.form.get('empresa_nif', '').strip()
            cfg.empresa_morada     = request.form.get('empresa_morada', '').strip()
            cfg.empresa_tel        = request.form.get('empresa_tel', '').strip()
            cfg.empresa_email      = request.form.get('empresa_email', '').strip()
            cfg.cor_accent         = request.form.get('cor_accent', '#3b6ef0')
            cfg.cor_bg             = request.form.get('cor_bg', '#0f1117')
            cfg.cor_surface        = request.form.get('cor_surface', '#171b25')
            cfg.backup_local_path  = request.form.get('backup_local_path', 'backups').strip()
            cfg.backup_rede_path   = request.form.get('backup_rede_path', '').strip()
            cfg.backup_hora        = request.form.get('backup_hora', '02:00')
            cfg.backup_manter_dias = int(request.form.get('backup_manter_dias', 30))
            cfg.backup_auto_ativo  = request.form.get('backup_auto_ativo') == 'on'
            cfg.salarios_pin = request.form.get('salarios_senha','').strip() or ''
            cfg.claude_chat_ativo  = request.form.get('claude_chat_ativo') == 'on'
            cfg.claude_chat_sistema= request.form.get('claude_chat_sistema', '').strip()
            try:
                cfg.logo_altura    = int(request.form.get('logo_altura', 48) or 48)
                cfg.logo_largura   = int(request.form.get('logo_largura', 180) or 180)
                cfg.logo_filtro    = request.form.get('logo_filtro', '').strip()
            except Exception:
                # Columns may not exist yet - run fix_db.py
                from sqlalchemy import text
                with db.engine.connect() as conn:
                    for col, typ, default in [
                        ('logo_altura','INTEGER','48'),
                        ('logo_largura','INTEGER','180'),
                        ('logo_filtro','VARCHAR(100)',"''"),
                    ]:
                        try:
                            conn.execute(text(f"ALTER TABLE config_geral ADD COLUMN {col} {typ} DEFAULT {default}"))
                            conn.commit()
                        except Exception:
                            pass
                cfg.logo_altura    = int(request.form.get('logo_altura', 48) or 48)
                cfg.logo_largura   = int(request.form.get('logo_largura', 180) or 180)
                cfg.logo_filtro    = request.form.get('logo_filtro', '').strip()
            # SMTP
            cfg.smtp_host = request.form.get('smtp_host', '').strip()
            cfg.smtp_port = int(request.form.get('smtp_port', 587) or 587)
            cfg.smtp_user = request.form.get('smtp_user', '').strip()
            smtp_pass_input = request.form.get('smtp_pass', '').strip()
            if smtp_pass_input:  # only update if not empty
                cfg.smtp_pass = smtp_pass_input
            cfg.smtp_from = request.form.get('smtp_from', '').strip()
            cfg.smtp_tls  = 1 if request.form.get('smtp_tls') else 0
            # Salarios PIN
            try:
                if request.form.get('salarios_pin_clear'):
                    cfg.salarios_pin = ''
                elif request.form.get('salarios_pin','').strip():
                    cfg.salarios_pin = request.form.get('salarios_pin','').strip()
            except Exception:
                from sqlalchemy import text
                with db.engine.connect() as _c:
                    try: _c.execute(text("ALTER TABLE config_geral ADD COLUMN salarios_pin VARCHAR(10) DEFAULT ''")); _c.commit()
                    except: pass
                if request.form.get('salarios_pin_clear'):
                    cfg.salarios_pin = ''
                elif request.form.get('salarios_pin','').strip():
                    cfg.salarios_pin = request.form.get('salarios_pin','').strip()
            try:
                cfg.salario_dia_inicio = int(request.form.get('salario_dia_inicio', 1) or 1)
                cfg.salario_dia_fecho  = int(request.form.get('salario_dia_fecho', 27) or 27)
            except: pass
            try:
                cfg.salario_dia_inicio = int(request.form.get('salario_dia_inicio', 1) or 1)
                cfg.salario_dia_fecho  = int(request.form.get('salario_dia_fecho', 27) or 27)
            except: pass
            # Logo upload
            if 'logo' in request.files and request.files['logo'].filename:
                logo = request.files['logo']
                logo_path = os.path.join(app.config['UPLOAD_FOLDER'], 'logo_empresa.png')
                logo.save(logo_path)
                cfg.empresa_logo_path = 'logo_empresa.png'
                _regenerar_pwa_icons()
            try:
                db.session.commit()
                flash('Configurações guardadas com sucesso! ✅', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Erro ao guardar: {e}', 'error')
        elif action == 'backup':
            from backup_manager import fazer_backup
            ok, msg = fazer_backup(app, cfg)
            flash(msg, 'success' if ok else 'error')
        return redirect(url_for('admin_config'))

    from backup_manager import listar_backups
    backups = listar_backups(app, cfg)
    return render_template('admin_config.html', cfg=cfg, backups=backups)


# ══════════════════════════════════════════════════════════════
#  CLIENTES
# ══════════════════════════════════════════════════════════════

@app.route('/clientes')
@login_required
def clientes():
    q = request.args.get('q', '').strip()
    query = Cliente.query.filter_by(ativo=True)
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(
            Cliente.nome.ilike(like),
            Cliente.abreviatura.ilike(like),
            Cliente.nif.ilike(like),
        ))
    todos = query.order_by(Cliente.nome).all()
    return render_template('clientes.html', clientes=todos, q=q)


@app.route('/clientes/sync_phc', methods=['POST'])
@login_required
def sync_clientes_phc():
    """Sync clients from PHC ec table (clientes only)."""
    if not current_user.is_admin:
        return jsonify({'error': 'Sem permissão'}), 403
    cfg_phc = ConfigPHC.query.first()
    if not cfg_phc:
        return jsonify({'error': 'PHC não configurado'}), 400
    try:
        from phc_sync import get_phc_connection
        conn = get_phc_connection(cfg_phc)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT no, nome, ISNULL(abrev,'') abrev, ISNULL(ncont,'') nipc,
                   ISNULL(morada,'') morada, ISNULL(local,'') local,
                   ISNULL(codpost,'') codpost, ISNULL(pais,'Portugal') pais,
                   ISNULL(telefone,'') tel, ISNULL(tlm,'') tlm,
                   ISNULL(email,'') email, ISNULL(www,'') www
            FROM ec
            WHERE cliente=1 AND ISNULL(inactivo,0)=0
            ORDER BY nome
        """)
        rows = cursor.fetchall()
        conn.close()
        inseridos = atualizados = 0
        for r in rows:
            c = Cliente.query.filter_by(phc_no=int(r[0])).first()
            if c:
                c.nome=r[1];c.abreviatura=r[2];c.nif=r[3];c.morada=r[4]
                c.localidade=r[5];c.cod_postal=r[6];c.pais=r[7]
                c.telefone=r[8];c.telemovel=r[9];c.email=r[10];c.website=r[11]
                c.ultima_sync_phc=datetime.now(); atualizados+=1
            else:
                db.session.add(Cliente(phc_no=int(r[0]),nome=r[1],abreviatura=r[2],
                    nif=r[3],morada=r[4],localidade=r[5],cod_postal=r[6],pais=r[7],
                    telefone=r[8],telemovel=r[9],email=r[10],website=r[11],
                    ultima_sync_phc=datetime.now())); inseridos+=1
        db.session.commit()
        return jsonify({'ok':True,'inseridos':inseridos,'atualizados':atualizados})
    except Exception as e:
        return jsonify({'error':str(e)}), 500


@app.route('/clientes/novo', methods=['GET','POST'])
@login_required
def novo_cliente():
    if request.method == 'POST':
        nome = request.form.get('nome','').strip()
        if not nome:
            flash('Nome é obrigatório.','error')
            return render_template('cliente_form.html', cliente=None)
        c = Cliente(nome=nome,
            abreviatura=request.form.get('abreviatura','').strip(),
            nif=request.form.get('nif','').strip(),
            morada=request.form.get('morada','').strip(),
            localidade=request.form.get('localidade','').strip(),
            cod_postal=request.form.get('cod_postal','').strip(),
            pais=request.form.get('pais','Portugal').strip(),
            telefone=request.form.get('telefone','').strip(),
            telemovel=request.form.get('telemovel','').strip(),
            email=request.form.get('email','').strip(),
            website=request.form.get('website','').strip(),
            notas=request.form.get('notas','').strip())
        db.session.add(c); db.session.commit()
        flash(f'Cliente {nome} criado.','success')
        return redirect(url_for('cliente_detalhe', cid=c.id))
    return render_template('cliente_form.html', cliente=None)


@app.route('/clientes/<int:cid>')
@login_required
def cliente_detalhe(cid):
    c = Cliente.query.get_or_404(cid)
    # Get articles sold to this client from PHC
    consumiveis = []
    cfg_phc = ConfigPHC.query.first()
    if cfg_phc and cfg_phc.ultima_sync:
        try:
            from phc_sync import get_phc_connection
            conn = get_phc_connection(cfg_phc)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT TOP 200
                    fl.ref, fl.design, fl.qtt, fl.preco,
                    ft.data, ft.fno, ft.serie
                FROM fl
                INNER JOIN ft ON ft.ftstamp=fl.ftstamp
                INNER JOIN ec ON ec.no=ft.no
                WHERE ec.no=? AND ft.anulado=0
                  AND fl.ref IS NOT NULL AND fl.ref<>''
                ORDER BY ft.data DESC
            """, c.phc_no or -1)
            rows = cursor.fetchall()
            conn.close()
            consumiveis = [{'ref':r[0],'design':r[1],'qtt':float(r[2] or 0),
                            'preco':float(r[3] or 0),
                            'data':r[4].strftime('%d/%m/%Y') if r[4] else '—',
                            'doc':f"{r[6]}{r[5]}"} for r in rows]
        except Exception:
            pass
    return render_template('cliente_detalhe.html', cliente=c, consumiveis=consumiveis)


@app.route('/clientes/<int:cid>/editar', methods=['GET','POST'])
@login_required
def editar_cliente(cid):
    c = Cliente.query.get_or_404(cid)
    if request.method == 'POST':
        c.nome=request.form.get('nome','').strip()
        c.abreviatura=request.form.get('abreviatura','').strip()
        c.nif=request.form.get('nif','').strip()
        c.morada=request.form.get('morada','').strip()
        c.localidade=request.form.get('localidade','').strip()
        c.cod_postal=request.form.get('cod_postal','').strip()
        c.pais=request.form.get('pais','Portugal').strip()
        c.telefone=request.form.get('telefone','').strip()
        c.telemovel=request.form.get('telemovel','').strip()
        c.email=request.form.get('email','').strip()
        c.website=request.form.get('website','').strip()
        c.notas=request.form.get('notas','').strip()
        c.atualizado_em=datetime.now()
        db.session.commit()
        flash('Cliente atualizado.','success')
        return redirect(url_for('cliente_detalhe', cid=cid))
    return render_template('cliente_form.html', cliente=c)


# ── Embarcações ──────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════
# BASE DADOS EQUIPAMENTOS INDUSTRIAIS
# ══════════════════════════════════════════════════════════════════════
EQ_IND_UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'eq_industrial')
os.makedirs(EQ_IND_UPLOAD_DIR, exist_ok=True)

EQ_TIPOS = {'gerador': 'Grupo Gerador', 'grupo_hidraulico': 'Grupo Hidráulico',
             'motor': 'Motor Industrial', 'compressor': 'Compressor', 'outro': 'Outro Equipamento'}

@app.route('/equipamentos-industriais')
@login_required
def eq_industriais():
    q    = request.args.get('q','').strip()
    tipo = request.args.get('tipo','').strip()
    query = EquipamentoIndustrial.query
    if q:
        query = query.filter(db.or_(
            EquipamentoIndustrial.nome.ilike(f'%{q}%'),
            EquipamentoIndustrial.referencia_interna.ilike(f'%{q}%'),
            EquipamentoIndustrial.local_instalacao.ilike(f'%{q}%'),
        ))
    if tipo:
        query = query.filter_by(tipo=tipo)
    equipamentos = query.order_by(EquipamentoIndustrial.criado_em.desc()).all()
    clientes = Cliente.query.order_by(Cliente.nome).all()
    return render_template('eq_industriais.html',
        equipamentos=equipamentos, clientes=clientes, tipos=EQ_TIPOS, q=q, tipo=tipo)

@app.route('/equipamentos-industriais/novo', methods=['GET','POST'])
@login_required
def eq_industrial_novo():
    clientes = Cliente.query.order_by(Cliente.nome).all()
    if request.method == 'POST':
        from datetime import date as _d
        cid = request.form.get('cliente_id','').strip()
        if not cid:
            flash('Seleccione um cliente', 'error')
            clientes = Cliente.query.order_by(Cliente.nome).all()
            return render_template('eq_industrial_form.html', clientes=clientes, tipos=EQ_TIPOS, eq=None)
        eq = EquipamentoIndustrial(
            cliente_id=int(cid),
            nome=request.form.get('nome','').strip(),
            tipo=request.form.get('tipo','outro'),
            referencia_interna=request.form.get('referencia_interna','').strip() or None,
            local_instalacao=request.form.get('local_instalacao','').strip() or None,
            notas=request.form.get('notas','').strip() or None,
            estado='ativo',
        )
        di = request.form.get('data_instalacao','').strip()
        df = request.form.get('data_fabricacao','').strip()
        try:
            if di: eq.data_instalacao = _d.fromisoformat(di)
            if df: eq.data_fabricacao = _d.fromisoformat(df)
        except: pass
        db.session.add(eq)
        db.session.flush()
        # Foto upload
        f = request.files.get('foto')
        if f and f.filename:
            import re as _re
            safe = _re.sub(r'[^\w\-\.]', '_', f.filename)
            fname = f'eq{eq.id}_{safe}'
            f.save(os.path.join(EQ_IND_UPLOAD_DIR, fname))
            eq.foto_path = fname
        db.session.commit()
        flash('Equipamento criado ✅', 'success')
        return redirect(url_for('eq_industrial_detalhe', eid=eq.id))
    return render_template('eq_industrial_form.html', clientes=clientes, tipos=EQ_TIPOS, eq=None)

@app.route('/equipamentos-industriais/<int:eid>')
@login_required
def eq_industrial_detalhe(eid):
    eq = EquipamentoIndustrial.query.get_or_404(eid)
    return render_template('eq_industrial_detalhe.html', eq=eq, tipos=EQ_TIPOS)

@app.route('/equipamentos-industriais/<int:eid>/editar', methods=['POST'])
@login_required
def eq_industrial_editar(eid):
    eq = EquipamentoIndustrial.query.get_or_404(eid)
    from datetime import date as _d
    eq.nome              = request.form.get('nome', eq.nome).strip()
    eq.tipo              = request.form.get('tipo', eq.tipo)
    eq.cliente_id        = int(request.form.get('cliente_id', eq.cliente_id))
    eq.referencia_interna= request.form.get('referencia_interna','').strip() or None
    eq.local_instalacao  = request.form.get('local_instalacao','').strip() or None
    eq.estado            = request.form.get('estado', eq.estado)
    eq.notas             = request.form.get('notas','').strip() or None
    eq.atualizado_em     = datetime.now()
    di = request.form.get('data_instalacao','').strip()
    df = request.form.get('data_fabricacao','').strip()
    try:
        if di: eq.data_instalacao = _d.fromisoformat(di)
        if df: eq.data_fabricacao = _d.fromisoformat(df)
    except: pass
    f = request.files.get('foto')
    if f and f.filename:
        import re as _re
        safe = _re.sub(r'[^\w\-\.]', '_', f.filename)
        fname = f'eq{eq.id}_{safe}'
        f.save(os.path.join(EQ_IND_UPLOAD_DIR, fname))
        eq.foto_path = fname
    db.session.commit()
    flash('Equipamento actualizado ✅', 'success')
    return redirect(url_for('eq_industrial_detalhe', eid=eid))

@app.route('/equipamentos-industriais/<int:eid>/componente', methods=['POST'])
@login_required
def eq_industrial_componente_add(eid):
    eq = EquipamentoIndustrial.query.get_or_404(eid)
    data = request.get_json() or {}
    def g(k): return data.get(k,'') or None
    def gf(k):
        v = data.get(k,'')
        try: return float(v) if v else None
        except: return None
    def gi(k):
        v = data.get(k,'')
        try: return int(v) if v else None
        except: return None
    comp = EqIndComponente(
        equipamento_id=eid, tipo=g('tipo') or 'motor',
        marca_grupo=g('marca_grupo'), modelo_grupo=g('modelo_grupo'), nserie_grupo=g('nserie_grupo'), dados_grupo=g('dados_grupo'),
        marca_motor=g('marca_motor'), modelo_motor=g('modelo_motor'), nserie_motor=g('nserie_motor'),
        familia_motor=g('familia_motor'), tipo_motor=g('tipo_motor'),
        potencia_motor_kw=gf('potencia_motor_kw'), potencia_motor_cv=gf('potencia_motor_cv'),
        rpm_motor=gi('rpm_motor'), cilindros=gi('cilindros'),
        combustivel=g('combustivel'), instalacao_eletrica=g('instalacao_eletrica'), dados_motor=g('dados_motor'),
        marca_alternador=g('marca_alternador'), nserie_alternador=g('nserie_alternador'),
        potencia_kva=gf('potencia_kva'), potencia_kw_alt=gf('potencia_kw_alt'),
        tensao_saida=g('tensao_saida'), frequencia=gi('frequencia'),
        fator_potencia=g('fator_potencia'), dados_alternador=g('dados_alternador'),
        descricao=g('descricao'), dados_outros=g('dados_outros'),
    )
    db.session.add(comp)
    db.session.commit()
    return jsonify({'ok': True, 'id': comp.id})

@app.route('/equipamentos-industriais/componente/<int:cid>/editar', methods=['POST'])
@login_required
def eq_industrial_componente_edit(cid):
    comp = EqIndComponente.query.get_or_404(cid)
    data = request.get_json() or {}
    def g(k): return data.get(k,'') or None
    def gf(k):
        v = data.get(k,'')
        try: return float(v) if v else None
        except: return None
    def gi(k):
        v = data.get(k,'')
        try: return int(v) if v else None
        except: return None
    # Update all fields
    for field in ['marca_grupo','modelo_grupo','nserie_grupo','dados_grupo',
                  'marca_motor','modelo_motor','nserie_motor','familia_motor','tipo_motor',
                  'combustivel','instalacao_eletrica','dados_motor',
                  'marca_alternador','nserie_alternador','tensao_saida','fator_potencia','dados_alternador',
                  'descricao','dados_outros']:
        setattr(comp, field, g(field))
    for field in ['potencia_motor_kw','potencia_motor_cv','potencia_kva','potencia_kw_alt']:
        setattr(comp, field, gf(field))
    for field in ['rpm_motor','cilindros','frequencia']:
        setattr(comp, field, gi(field))
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/equipamentos-industriais/componente/<int:cid>/media/upload', methods=['POST'])
@login_required
def eq_ind_comp_media_upload(cid):
    comp = EqIndComponente.query.get_or_404(cid)
    f = request.files.get('ficheiro')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'Sem ficheiro'})
    import re as _re
    safe = _re.sub(r'[^\w\-\.]', '_', f.filename)
    tipo = 'pdf' if f.filename.lower().endswith('.pdf') else 'foto'
    fname = f'comp{cid}_{int(datetime.now().timestamp())}_{safe}'
    f.save(os.path.join(EQ_IND_UPLOAD_DIR, fname))
    m = EqIndComponenteMedia(
        componente_id=cid, tipo=tipo,
        ficheiro_path=fname,
        titulo=request.form.get('titulo','').strip() or f.filename,
    )
    db.session.add(m)
    db.session.commit()
    return jsonify({'ok': True, 'id': m.id, 'titulo': m.titulo, 'tipo': tipo, 'path': fname})

@app.route('/equipamentos-industriais/comp-media/<int:mid>/ver')
@login_required
def eq_ind_comp_media_ver(mid):
    m = EqIndComponenteMedia.query.get_or_404(mid)
    return send_from_directory(EQ_IND_UPLOAD_DIR, m.ficheiro_path)

@app.route('/equipamentos-industriais/comp-media/<int:mid>/apagar', methods=['POST'])
@login_required
def eq_ind_comp_media_apagar(mid):
    m = EqIndComponenteMedia.query.get_or_404(mid)
    try: os.remove(os.path.join(EQ_IND_UPLOAD_DIR, m.ficheiro_path))
    except: pass
    db.session.delete(m)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/equipamentos-industriais/componente/<int:cid>/dados')
@login_required
def eq_ind_comp_dados(cid):
    comp = EqIndComponente.query.get_or_404(cid)
    medias = [{'id': m.id, 'titulo': m.titulo, 'tipo': m.tipo} for m in comp.medias.all()]
    return jsonify({'ok': True, 'tipo': comp.tipo,
        'marca_grupo': comp.marca_grupo, 'modelo_grupo': comp.modelo_grupo, 'nserie_grupo': comp.nserie_grupo, 'dados_grupo': comp.dados_grupo,
        'marca_motor': comp.marca_motor, 'modelo_motor': comp.modelo_motor, 'nserie_motor': comp.nserie_motor,
        'familia_motor': comp.familia_motor, 'tipo_motor': comp.tipo_motor,
        'potencia_motor_kw': comp.potencia_motor_kw, 'potencia_motor_cv': comp.potencia_motor_cv,
        'rpm_motor': comp.rpm_motor, 'cilindros': comp.cilindros,
        'combustivel': comp.combustivel, 'instalacao_eletrica': comp.instalacao_eletrica, 'dados_motor': comp.dados_motor,
        'marca_alternador': comp.marca_alternador, 'nserie_alternador': comp.nserie_alternador,
        'potencia_kva': comp.potencia_kva, 'potencia_kw_alt': comp.potencia_kw_alt,
        'tensao_saida': comp.tensao_saida, 'frequencia': comp.frequencia,
        'fator_potencia': comp.fator_potencia, 'dados_alternador': comp.dados_alternador,
        'descricao': comp.descricao, 'dados_outros': comp.dados_outros,
        'medias': medias})


@app.route('/equipamentos-industriais/componente/<int:cid>', methods=['DELETE'])
@login_required
def eq_industrial_componente_del(cid):
    comp = EqIndComponente.query.get_or_404(cid)
    db.session.delete(comp)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/equipamentos-industriais/<int:eid>/doc/upload', methods=['POST'])
@login_required
def eq_industrial_doc_upload(eid):
    eq = EquipamentoIndustrial.query.get_or_404(eid)
    f  = request.files.get('ficheiro')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'Sem ficheiro'})
    import re as _re
    safe = _re.sub(r'[^\w\-\.]', '_', f.filename)
    fname = f'eqdoc{eid}_{int(datetime.now().timestamp())}_{safe}'
    f.save(os.path.join(EQ_IND_UPLOAD_DIR, fname))
    doc = EqIndDocumento(
        equipamento_id=eid,
        tipo=request.form.get('tipo','outro'),
        titulo=request.form.get('titulo','').strip() or f.filename,
        ficheiro_path=fname,
        notas=request.form.get('notas','').strip() or None,
        user_id=current_user.id,
    )
    db.session.add(doc)
    db.session.commit()
    return jsonify({'ok': True, 'id': doc.id, 'titulo': doc.titulo})

@app.route('/equipamentos-industriais/doc/<int:did>/ver')
@login_required
def eq_industrial_doc_ver(did):
    doc = EqIndDocumento.query.get_or_404(did)
    return send_from_directory(EQ_IND_UPLOAD_DIR, doc.ficheiro_path)

@app.route('/equipamentos-industriais/doc/<int:did>/apagar', methods=['POST'])
@login_required
def eq_industrial_doc_apagar(did):
    doc = EqIndDocumento.query.get_or_404(did)
    try: os.remove(os.path.join(EQ_IND_UPLOAD_DIR, doc.ficheiro_path))
    except: pass
    db.session.delete(doc)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/equipamentos-industriais/foto/<path:fname>')
@login_required
def eq_industrial_foto(fname):
    return send_from_directory(EQ_IND_UPLOAD_DIR, fname)


# ══════════════════════════════════════════════════════════════════════
# MANUAIS EMBARCAÇÕES — Upload / Download / Listagem
# ══════════════════════════════════════════════════════════════════════
MANUAIS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'manuais_embarcacoes')
os.makedirs(MANUAIS_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════
# AGENDA DIGITAL
# ══════════════════════════════════════════════════════════════════════
@app.route('/agenda')
@login_required
def agenda():
    from datetime import date as _d
    import calendar as _cal
    hoje = _d.today()
    ano = int(request.args.get('ano', hoje.year))
    mes = int(request.args.get('mes', hoje.month))
    mes_ini = _d(ano, mes, 1)
    mes_fim = _d(ano, mes, _cal.monthrange(ano, mes)[1])
    funcs = Funcionario.query.filter_by(ativo=True, agenda_ativo=True).order_by(Funcionario.nome).all()
    stats = {}
    for f in funcs:
        regs = AgendaRegisto.query.filter(
            AgendaRegisto.funcionario_id == f.id,
            AgendaRegisto.data >= mes_ini,
            AgendaRegisto.data <= mes_fim,
        ).all()
        stats[f.id] = {
            'horas_mes': sum(r.horas or 0 for r in regs),
            'he_mes': sum(r.he_horas or 0 for r in regs),
            'faltas_mes': sum(r.falta_horas or 0 for r in regs if r.tem_falta),
            'registos_mes': len(regs),
        }
    return render_template('agenda.html', funcs=funcs, stats=stats, hoje=hoje, ano=ano, mes=mes)

@app.route('/agenda/config-users', methods=['GET','POST'])
@login_required
def agenda_config_users():
    if not current_user.is_admin:
        return redirect(url_for('agenda'))
    if request.method == 'POST':
        funcs = Funcionario.query.filter_by(ativo=True).all()
        for u in funcs:
            u.agenda_ativo = str(u.id) in request.form.getlist('agenda_ativo')
        db.session.commit()
        flash('Configuração guardada ✅', 'success')
        return redirect(url_for('agenda'))
    funcs = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()
    return render_template('agenda_config_users.html', funcs=funcs)


@app.route('/agenda/sync-funcionarios', methods=['POST'])
@login_required
def agenda_sync_funcionarios():
    if not current_user.is_admin:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    # Mark inactive funcionarios as agenda_ativo=False
    todos = Funcionario.query.all()
    alterados = 0
    for f in todos:
        if not f.ativo and f.agenda_ativo:
            f.agenda_ativo = False
            alterados += 1
    db.session.commit()
    ativos = Funcionario.query.filter_by(ativo=True).count()
    agenda = Funcionario.query.filter_by(ativo=True, agenda_ativo=True).count()
    return jsonify({'ok': True, 'alterados': alterados, 'total_ativos': ativos, 'com_agenda': agenda})


@app.route('/agenda/funcionario/<int:fid>')
@login_required
def agenda_funcionario(fid):
    func = Funcionario.query.get_or_404(fid)
    ano  = int(request.args.get('ano', datetime.now().year))
    mes  = int(request.args.get('mes', datetime.now().month))
    from datetime import date as _d
    import calendar as _cal
    # Get registos for this month
    mes_ini = _d(ano, mes, 1)
    mes_fim = _d(ano, mes, _cal.monthrange(ano, mes)[1])
    registos = AgendaRegisto.query.filter(
        AgendaRegisto.funcionario_id == fid,
        AgendaRegisto.data >= mes_ini,
        AgendaRegisto.data <= mes_fim,
    ).all()
    # Map by date
    reg_map = {}
    for r in registos:
        if r.data not in reg_map:
            reg_map[r.data] = []
        reg_map[r.data].append(r)
    # Stats
    total_horas = sum(r.horas or 0 for r in registos)
    total_he    = sum(r.he_horas or 0 for r in registos)
    return render_template('agenda_funcionario.html',
        func=func, ano=ano, mes=mes, mes_ini=mes_ini, mes_fim=mes_fim,
        reg_map=reg_map, total_horas=total_horas, total_he=total_he,
        hoje=_d.today(), cal=_cal)

@app.route('/agenda/funcionario/<int:fid>/dia/<int:ano>/<int:mes>/<int:dia>')
@login_required
def agenda_dia(fid, ano, mes, dia):
    func = Funcionario.query.get_or_404(fid)
    from datetime import date as _d
    data = _d(ano, mes, dia)
    registos = AgendaRegisto.query.filter_by(
        funcionario_id=fid, data=data
    ).order_by(AgendaRegisto.criado_em).all()
    # Services in progress for suggestions
    servicos_curso = AgendaServico.query.filter(
        AgendaServico.estado == 'em_curso'
    ).order_by(AgendaServico.numero.desc().nullslast(), AgendaServico.atualizado_em.desc()).limit(30).all()
    clientes_obj = Cliente.query.order_by(Cliente.nome).all()
    clientes = [{'id': c.id, 'nome': c.nome} for c in clientes_obj]
    # Embarcacao suggestions
    emb_sugs = db.session.query(AgendaServico.embarcacao_nome).filter(
        AgendaServico.embarcacao_nome != None,
        AgendaServico.embarcacao_nome != ''
    ).distinct().limit(30).all()
    emb_sugs = [e[0] for e in emb_sugs if e[0]]
    # Build conflict map: {cliente_id: [sv,...], embarcacao: [sv,...]}
    import json as _json
    conflitos = {}
    for sv in servicos_curso:
        if sv.cliente_id:
            k = f'cli_{sv.cliente_id}'
            conflitos.setdefault(k, []).append({'id': sv.id, 'num': sv.numero or '', 'titulo': sv.titulo})
        if sv.embarcacao_nome:
            k = f'emb_{sv.embarcacao_nome.lower().strip()}'
            conflitos.setdefault(k, []).append({'id': sv.id, 'num': sv.numero or '', 'titulo': sv.titulo})
    conflitos_json = _json.dumps(conflitos)

    return render_template('agenda_dia.html',
        func=func, data=data, registos=registos,
        servicos_curso=servicos_curso, clientes=clientes,
        emb_sugs=emb_sugs, conflitos_json=conflitos_json)

@app.route('/agenda/registo/novo', methods=['POST'])
@login_required
def agenda_registo_novo():
    data = request.get_json() or {}
    from datetime import date as _d
    try:
        dt = _d.fromisoformat(data.get('data',''))
    except:
        return jsonify({'ok': False, 'error': 'Data inválida'})
    fid = data.get('funcionario_id')
    if not fid:
        return jsonify({'ok': False, 'error': 'Funcionário obrigatório'})
    # Get or create servico
    servico_id = data.get('servico_id')
    if not servico_id:
        # Create new servico
        # Auto-assign service number
        ultimo_num = db.session.query(db.func.max(AgendaServico.numero)).scalar() or 0
        sv = AgendaServico(
            numero=ultimo_num + 1,
            titulo=data.get('titulo') or data.get('equipamento') or 'Serviço ' + dt.strftime('%d/%m/%Y'),
            cliente_id=int(data['cliente_id']) if data.get('cliente_id') else None,
            embarcacao_nome=data.get('embarcacao_nome','').strip() or None,
            equipamento=data.get('equipamento','').strip() or None,
            local_servico=data.get('local_servico','').strip() or None,
            tipo=data.get('tipo','cliente'),
            estado='em_curso',
            data_inicio=dt,
            descricao=data.get('descricao','').strip() or None,
            criado_por=current_user.id,
        )
        db.session.add(sv)
        db.session.flush()
        servico_id = sv.id
    else:
        sv = AgendaServico.query.get(int(servico_id))

    reg = AgendaRegisto(
        servico_id=servico_id,
        funcionario_id=int(fid),
        data=dt,
        horas=float(data.get('horas') or 0),
        hora_inicio=data.get('hora_inicio','').strip() or None,
        hora_fim=data.get('hora_fim','').strip() or None,
        tem_he=bool(data.get('tem_he')),
        he_inicio=data.get('he_inicio','').strip() or None,
        he_fim=data.get('he_fim','').strip() or None,
        he_horas=float(data.get('he_horas') or 0),
        deslocacao_viatura=bool(data.get('deslocacao_viatura')),
        viatura_tipo=data.get('viatura_tipo','propria'),
        n_viagens=int(data.get('n_viagens') or 0),
        km=float(data.get('km') or 0) or None,
        n_almoco=int(data.get('n_almoco') or 0),
        custo_refeicao=float(data.get('custo_refeicao') or 0),
        obs_refeicao=data.get('obs_refeicao','').strip() or None,
        tem_falta=bool(data.get('tem_falta')),
        falta_tipo=data.get('falta_tipo','') or None,
        falta_horas=float(data.get('falta_horas') or 0),
        falta_obs=data.get('falta_obs','').strip() or None,
        descricao_trabalho=data.get('descricao_trabalho','').strip() or None,
        estado=data.get('estado','em_curso'),
    )
    db.session.add(reg)
    db.session.flush()
    # Materiais
    for m in data.get('materiais', []):
        mat = AgendaMaterial(
            registo_id=reg.id,
            artigo_ref=m.get('ref','').strip() or None,
            descricao=m.get('descricao','').strip(),
            quantidade=float(m.get('quantidade') or 1),
            unidade=m.get('unidade','un'),
            observacoes=m.get('observacoes','').strip() or None,
            origem=m.get('origem','manual'),
        )
        db.session.add(mat)
    db.session.commit()
    return jsonify({'ok': True, 'registo_id': reg.id, 'servico_id': servico_id})

@app.route('/agenda/registo/<int:rid>', methods=['GET'])
@login_required
def agenda_registo_get(rid):
    r = AgendaRegisto.query.get_or_404(rid)
    sv = r.servico
    return jsonify({'ok': True,
        'id': r.id, 'data': r.data.isoformat(),
        'servico_id': r.servico_id,
        'numero': sv.numero or '', 'titulo': sv.titulo, 'cliente_id': sv.cliente_id,
        'cliente_nome': sv.cliente.nome if sv.cliente else '',
        'embarcacao_nome': sv.embarcacao_nome or '',
        'equipamento': sv.equipamento or '',
        'local_servico': sv.local_servico or '',
        'tipo': sv.tipo or 'cliente',
        'horas': r.horas, 'hora_inicio': r.hora_inicio or '', 'hora_fim': r.hora_fim or '',
        'tem_he': r.tem_he, 'he_inicio': r.he_inicio or '', 'he_fim': r.he_fim or '', 'he_horas': r.he_horas,
        'deslocacao_viatura': r.deslocacao_viatura, 'viatura_tipo': r.viatura_tipo or 'propria',
        'n_viagens': r.n_viagens, 'km': r.km or '',
        'n_almoco': r.n_almoco, 'custo_refeicao': r.custo_refeicao, 'obs_refeicao': r.obs_refeicao or '',
        'estado': r.estado or 'em_curso',
        'tem_falta': r.tem_falta or False,
        'falta_tipo': r.falta_tipo or 'dia_completo',
        'falta_horas': r.falta_horas or 0,
        'falta_obs': r.falta_obs or '',
        'descricao_trabalho': r.descricao_trabalho or '',
        'materiais': [{'ref': m.artigo_ref or '', 'descricao': m.descricao, 'quantidade': m.quantidade,
                       'unidade': m.unidade, 'observacoes': m.observacoes or '', 'origem': m.origem} for m in r.materiais.all()],
    })

@app.route('/agenda/registo/<int:rid>/editar', methods=['POST'])
@login_required
def agenda_registo_editar(rid):
    r = AgendaRegisto.query.get_or_404(rid)
    data = request.get_json() or {}
    sv = r.servico
    # Update servico fields
    if data.get('cliente_id'): sv.cliente_id = int(data['cliente_id'])
    sv.embarcacao_nome = data.get('embarcacao_nome','').strip() or sv.embarcacao_nome
    sv.equipamento     = data.get('equipamento','').strip() or sv.equipamento
    sv.local_servico   = data.get('local_servico','').strip() or sv.local_servico
    sv.tipo            = data.get('tipo', sv.tipo)
    sv.atualizado_em   = datetime.now()
    # Update registo
    r.horas            = float(data.get('horas') or 0)
    r.hora_inicio      = data.get('hora_inicio','').strip() or None
    r.hora_fim         = data.get('hora_fim','').strip() or None
    r.tem_he           = bool(data.get('tem_he'))
    r.he_inicio        = data.get('he_inicio','').strip() or None
    r.he_fim           = data.get('he_fim','').strip() or None
    r.he_horas         = float(data.get('he_horas') or 0)
    r.deslocacao_viatura = bool(data.get('deslocacao_viatura'))
    r.viatura_tipo     = data.get('viatura_tipo','propria')
    r.n_viagens        = int(data.get('n_viagens') or 0)
    r.km               = float(data.get('km') or 0) or None
    r.n_almoco         = int(data.get('n_almoco') or 0)
    r.custo_refeicao   = float(data.get('custo_refeicao') or 0)
    r.obs_refeicao     = data.get('obs_refeicao','').strip() or None
    r.tem_falta        = bool(data.get('tem_falta'))
    r.falta_tipo       = data.get('falta_tipo','') or None
    r.falta_horas      = float(data.get('falta_horas') or 0)
    r.falta_obs        = data.get('falta_obs','').strip() or None
    r.descricao_trabalho = data.get('descricao_trabalho','').strip() or None
    r.estado           = data.get('estado', r.estado)
    r.atualizado_em    = datetime.now()
    # Replace materiais
    for m in r.materiais.all():
        db.session.delete(m)
    for m in data.get('materiais', []):
        mat = AgendaMaterial(
            registo_id=r.id,
            artigo_ref=m.get('ref','').strip() or None,
            descricao=m.get('descricao','').strip(),
            quantidade=float(m.get('quantidade') or 1),
            unidade=m.get('unidade','un'),
            observacoes=m.get('observacoes','').strip() or None,
            origem=m.get('origem','manual'),
        )
        db.session.add(mat)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/agenda/registo/<int:rid>/apagar', methods=['POST'])
@login_required
def agenda_registo_apagar(rid):
    r = AgendaRegisto.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/agenda/servico/<int:sid>/apagar', methods=['POST'])
@login_required
def agenda_servico_apagar(sid):
    sv = AgendaServico.query.get_or_404(sid)
    # Delete all registos and their materiais (cascade)
    db.session.delete(sv)
    db.session.commit()
    flash('Serviço eliminado ✅', 'success')
    return jsonify({'ok': True})


@app.route('/agenda/servico/<int:sid>/concluir', methods=['POST'])
@login_required
def agenda_servico_concluir(sid):
    sv = AgendaServico.query.get_or_404(sid)
    from datetime import date as _d
    sv.estado = 'concluido'
    sv.data_fim = _d.today()
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/agenda/servicos')
@login_required
def agenda_servicos():
    estado = request.args.get('estado','em_curso')
    q      = request.args.get('q','').strip()
    query  = AgendaServico.query
    if estado: query = query.filter_by(estado=estado)
    if q:
        query = query.filter(db.or_(
            AgendaServico.titulo.ilike(f'%{q}%'),
            AgendaServico.embarcacao_nome.ilike(f'%{q}%'),
        ))
    servicos = query.order_by(AgendaServico.atualizado_em.desc()).all()
    return render_template('agenda_servicos.html', servicos=servicos, estado=estado, q=q)

@app.route('/agenda/servico/<int:sid>/pdf')
@login_required
def agenda_servico_pdf(sid):
    sv = AgendaServico.query.get_or_404(sid)
    registos = AgendaRegisto.query.filter_by(servico_id=sid).order_by(AgendaRegisto.data, AgendaRegisto.funcionario_id).all()
    cfg = ConfigGeral.query.first()
    return render_template('agenda_servico_pdf.html', sv=sv, registos=registos, cfg=cfg, now=datetime.now())

@app.route('/agenda/estatisticas')
@login_required
def agenda_estatisticas():
    from datetime import date as _d
    import calendar as _cal

    ano = int(request.args.get('ano', _d.today().year))
    mes = int(request.args.get('mes', _d.today().month))
    mes_ini = _d(ano, mes, 1)
    mes_fim = _d(ano, mes, _cal.monthrange(ano, mes)[1])

    # Calcular dias úteis do mês (excl. fds e feriados nacionais)
    feriados = set()
    try:
        for f in FeriasFeriado.query.filter(
            FeriasFeriado.data >= mes_ini,
            FeriasFeriado.data <= mes_fim
        ).all():
            feriados.add(f.data)
    except Exception:
        pass

    # Also exclude empresa_fechos
    try:
        for f in EmpresaFecho.query.filter(
            EmpresaFecho.data >= mes_ini,
            EmpresaFecho.data <= mes_fim
        ).all():
            feriados.add(f.data)
    except Exception:
        pass

    dias_uteis = 0
    cur = mes_ini
    from datetime import timedelta
    while cur <= mes_fim:
        if cur.weekday() < 5 and cur not in feriados:
            dias_uteis += 1
        cur += timedelta(days=1)
    horas_teoricas = dias_uteis * 8

    funcs = Funcionario.query.filter_by(ativo=True, agenda_ativo=True).order_by(Funcionario.nome).all()

    # Per-funcionario stats with tipo breakdown
    stats = []
    chart_labels = []
    chart_cliente = []
    chart_casa = []
    chart_stock = []

    for f in funcs:
        regs = AgendaRegisto.query.filter(
            AgendaRegisto.funcionario_id == f.id,
            AgendaRegisto.data >= mes_ini,
            AgendaRegisto.data <= mes_fim,
        ).all()
        if not regs:
            continue

        servico_ids = list(set(r.servico_id for r in regs))
        servicos = {sv.id: sv for sv in AgendaServico.query.filter(AgendaServico.id.in_(servico_ids)).all()}

        h_cli   = round(sum(r.horas or 0 for r in regs if servicos.get(r.servico_id) and servicos[r.servico_id].tipo == 'cliente'), 1)
        h_casa  = round(sum(r.horas or 0 for r in regs if servicos.get(r.servico_id) and servicos[r.servico_id].tipo == 'casa'), 1)
        h_stock = round(sum(r.horas or 0 for r in regs if servicos.get(r.servico_id) and servicos[r.servico_id].tipo == 'stock'), 1)
        h_total = round(h_cli + h_casa + h_stock, 1)
        he      = round(sum(r.he_horas or 0 for r in regs), 1)

        stats.append({
            'func': f,
            'horas': h_total,
            'he': he,
            'dias': len(set(r.data for r in regs)),
            'almoco': sum(r.n_almoco or 0 for r in regs),
            'custo_ref': round(sum(r.custo_refeicao or 0 for r in regs), 2),
            'h_cliente': h_cli,
            'h_casa': h_casa,
            'h_stock': h_stock,
            'n_servicos': len(set(r.servico_id for r in regs)),
        })

        # Chart data - use first name only for readability
        nome_curto = f.nome.split()[0] + ' ' + f.nome.split()[-1] if len(f.nome.split()) > 1 else f.nome
        chart_labels.append(nome_curto)
        chart_cliente.append(h_cli)
        chart_casa.append(h_casa)
        chart_stock.append(h_stock)

    stats.sort(key=lambda x: x['horas'], reverse=True)

    # Top clients
    top_cli = db.session.query(
        Cliente.nome,
        db.func.count(db.distinct(AgendaServico.id)).label('n_serv'),
        db.func.sum(AgendaRegisto.horas).label('h_total')
    ).join(AgendaServico, AgendaServico.cliente_id == Cliente.id
    ).join(AgendaRegisto, AgendaRegisto.servico_id == AgendaServico.id
    ).filter(AgendaRegisto.data >= mes_ini, AgendaRegisto.data <= mes_fim
    ).group_by(Cliente.nome).order_by(db.text('h_total desc')).limit(10).all()

    # Services stats
    sv_em_curso  = AgendaServico.query.filter_by(estado='em_curso').count()
    sv_concluido = AgendaServico.query.filter_by(estado='concluido').count()

    # Top equipamentos
    top_equip = db.session.query(
        AgendaServico.equipamento,
        db.func.count(db.distinct(AgendaServico.id)).label('n')
    ).filter(
        AgendaServico.equipamento != None,
        AgendaServico.equipamento != ''
    ).group_by(AgendaServico.equipamento).order_by(db.text('n desc')).limit(8).all()

    # Monthly horas (last 6 months)
    meses_hist = []
    for i in range(5, -1, -1):
        m2 = mes - i; a2 = ano
        while m2 <= 0: m2 += 12; a2 -= 1
        m_ini = _d(a2, m2, 1)
        m_fim = _d(a2, m2, _cal.monthrange(a2, m2)[1])
        h = db.session.query(db.func.sum(AgendaRegisto.horas)).filter(
            AgendaRegisto.data >= m_ini, AgendaRegisto.data <= m_fim
        ).scalar() or 0
        meses_hist.append({'label': f'{m2:02d}/{a2}', 'horas': round(float(h), 1)})

    import json
    chart_data = json.dumps({
        'labels': chart_labels,
        'cliente': chart_cliente,
        'casa': chart_casa,
        'stock': chart_stock,
        'teorico': horas_teoricas,
    })

    return render_template('agenda_estatisticas.html',
        stats=stats, top_cli=top_cli, top_equip=top_equip,
        sv_em_curso=sv_em_curso, sv_concluido=sv_concluido,
        meses_hist=meses_hist, chart_data=chart_data,
        dias_uteis=dias_uteis, horas_teoricas=horas_teoricas,
        ano=ano, mes=mes, mes_ini=mes_ini, mes_fim=mes_fim)


@app.route('/api/agenda/embarcacoes')
@login_required
def api_agenda_embarcacoes():
    q = request.args.get('q','').strip()
    sugs = db.session.query(AgendaServico.embarcacao_nome).filter(
        AgendaServico.embarcacao_nome.ilike(f'%{q}%')
    ).distinct().limit(15).all()
    return jsonify([s[0] for s in sugs if s[0]])


# ══════════════════════════════════════════════════════════════════════
# EMPRESA
# ══════════════════════════════════════════════════════════════════════
EMPRESA_UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'empresa')
os.makedirs(EMPRESA_UPLOAD_DIR, exist_ok=True)

EMPRESA_TIPOS = {
    'certidao':            {'label': '📜 Certidão Permanente',           'icon': '📜', 'cor': '#3b6ef0'},
    'nao_divida_financas': {'label': '🧾 Declaração Não Dívida Finanças', 'icon': '🧾', 'cor': '#f59e0b'},
    'nao_divida_ss':       {'label': '🏛 Declaração Não Dívida Seg. Social','icon': '🏛', 'cor': '#8b5cf6'},
    'iban':                {'label': '🏦 Comprovativo Bancário / IBAN',    'icon': '🏦', 'cor': '#10b981'},
    'outro':               {'label': '📁 Outro Documento',                 'icon': '📁', 'cor': '#64748b'},
}

@app.route('/empresa')
@login_required
def empresa():
    info = EmpresaInfo.query.first()
    docs = EmpresaDocumento.query.order_by(EmpresaDocumento.tipo, EmpresaDocumento.data_upload.desc()).all()
    from datetime import date, timedelta
    hoje = date.today()
    return render_template('empresa.html',
        info=info, docs=docs, tipos=EMPRESA_TIPOS,
        hoje=hoje, timedelta=timedelta)

@app.route('/empresa/info', methods=['POST'])
@login_required
def empresa_info_save():
    if not current_user.is_admin:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    info = EmpresaInfo.query.first()
    if not info:
        info = EmpresaInfo()
        db.session.add(info)
    campos = ['nome_completo','nome_comercial','nif','nipc','cae','forma_juridica',
              'morada','codigo_postal','localidade','telefone','email','website',
              'banco','iban','swift','conservatoria','num_registo','capital_social',
              'seguradora','apolice']
    for c in campos:
        setattr(info, c, request.form.get(c, '').strip() or None)
    # seguro_validade
    sv = request.form.get('seguro_validade','').strip()
    if sv:
        from datetime import date as _d
        try: info.seguro_validade = _d.fromisoformat(sv)
        except: pass
    else:
        info.seguro_validade = None
    db.session.commit()
    flash('Informações da empresa guardadas ✅', 'success')
    return redirect(url_for('empresa'))

@app.route('/empresa/doc/novo', methods=['POST'])
@login_required
def empresa_doc_novo():
    tipo   = request.form.get('tipo', 'outro')
    titulo = request.form.get('titulo', '').strip()
    if not titulo:
        titulo = EMPRESA_TIPOS.get(tipo, {}).get('label', 'Documento')
    doc = EmpresaDocumento(
        tipo=tipo, titulo=titulo,
        numero_acesso=request.form.get('numero_acesso','').strip() or None,
        notas=request.form.get('notas','').strip() or None,
        user_id=current_user.id,
    )
    de = request.form.get('data_emissao','').strip()
    dv = request.form.get('data_validade','').strip()
    from datetime import date as _d
    try:
        if de: doc.data_emissao = _d.fromisoformat(de)
        if dv: doc.data_validade = _d.fromisoformat(dv)
    except: pass
    db.session.add(doc)
    db.session.flush()
    # PDF upload
    f = request.files.get('pdf')
    if f and f.filename.lower().endswith('.pdf'):
        import re as _re
        safe = _re.sub(r'[^\w\-\.]', '_', f.filename)
        fname = f'{doc.id}_{safe}'
        f.save(os.path.join(EMPRESA_UPLOAD_DIR, fname))
        doc.pdf_path = fname
    db.session.commit()
    flash('Documento adicionado ✅', 'success')
    return redirect(url_for('empresa'))

@app.route('/empresa/doc/<int:did>/ver')
@login_required
def empresa_doc_ver(did):
    doc = EmpresaDocumento.query.get_or_404(did)
    if not doc.pdf_path:
        return 'Sem PDF', 404
    return send_from_directory(EMPRESA_UPLOAD_DIR, doc.pdf_path, mimetype='application/pdf')

@app.route('/empresa/doc/<int:did>/download')
@login_required
def empresa_doc_download(did):
    doc = EmpresaDocumento.query.get_or_404(did)
    if not doc.pdf_path:
        return 'Sem PDF', 404
    return send_from_directory(EMPRESA_UPLOAD_DIR, doc.pdf_path, as_attachment=True, download_name=doc.titulo+'.pdf')

@app.route('/empresa/doc/<int:did>/apagar', methods=['POST'])
@login_required
def empresa_doc_apagar(did):
    if not current_user.is_admin:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    doc = EmpresaDocumento.query.get_or_404(did)
    if doc.pdf_path:
        try: os.remove(os.path.join(EMPRESA_UPLOAD_DIR, doc.pdf_path))
        except: pass
    db.session.delete(doc)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/manuais-embarcacoes')
@login_required
def manuais_embarcacoes_list():
    """Return list of uploaded manuals as JSON."""
    files = []
    try:
        for fname in sorted(os.listdir(MANUAIS_DIR)):
            if fname.lower().endswith('.pdf'):
                path = os.path.join(MANUAIS_DIR, fname)
                size = os.path.getsize(path)
                mtime = os.path.getmtime(path)
                from datetime import datetime as _dt
                files.append({
                    'nome': fname,
                    'tamanho': f'{size/1024:.0f} KB' if size < 1024*1024 else f'{size/1024/1024:.1f} MB',
                    'data': _dt.fromtimestamp(mtime).strftime('%d/%m/%Y'),
                })
    except Exception as e:
        pass
    return jsonify({'ok': True, 'manuais': files})

@app.route('/manuais-embarcacoes/upload', methods=['POST'])
@login_required
def manuais_embarcacoes_upload():
    """Upload a PDF manual."""
    if 'pdf' not in request.files:
        return jsonify({'ok': False, 'error': 'Nenhum ficheiro enviado'})
    f = request.files['pdf']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'ok': False, 'error': 'Apenas ficheiros PDF são permitidos'})
    import re as _re
    safe = _re.sub(r'[^\w\-\.\s]', '_', f.filename)
    dest = os.path.join(MANUAIS_DIR, safe)
    f.save(dest)
    return jsonify({'ok': True, 'nome': safe})

@app.route('/manuais-embarcacoes/download/<path:nome>')
@login_required
def manuais_embarcacoes_download(nome):
    return send_from_directory(MANUAIS_DIR, nome, as_attachment=True)

@app.route('/manuais-embarcacoes/view/<path:nome>')
@login_required
def manuais_embarcacoes_view(nome):
    return send_from_directory(MANUAIS_DIR, nome, mimetype='application/pdf')

@app.route('/manuais-embarcacoes/delete/<path:nome>', methods=['POST'])
@login_required
def manuais_embarcacoes_delete(nome):
    if not current_user.is_admin:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    path = os.path.join(MANUAIS_DIR, nome)
    if os.path.exists(path):
        os.remove(path)
    return jsonify({'ok': True})


@app.route('/clientes/<int:cid>/embarcacoes/nova', methods=['GET','POST'])
@login_required
def nova_embarcacao(cid):
    cliente = Cliente.query.get_or_404(cid)
    if request.method == 'POST':
        e = Embarcacao(
            cliente_id=cid,
            nome=request.form.get('nome','').strip(),
            matricula=request.form.get('matricula','').strip(),
            tipo=request.form.get('tipo','').strip(),
            ano_construcao=request.form.get('ano_construcao') or None,
            comprimento=request.form.get('comprimento') or None,
            largura=request.form.get('largura') or None,
            notas=request.form.get('notas','').strip()
        )
        db.session.add(e); db.session.flush()
        # Handle foto upload
        foto = request.files.get('foto')
        if foto and foto.filename:
            import uuid
            ext = os.path.splitext(foto.filename)[1].lower()
            fname = f'emb_{e.id}_{uuid.uuid4().hex[:8]}{ext}'
            fdir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'embarcacoes')
            os.makedirs(fdir, exist_ok=True)
            foto.save(os.path.join(fdir, fname))
            e.foto_path = fname
        # Default components
        for cat in ['Motor Propulsor','Caixa Inversora/Redutora','Veio Propulsor','Hélice']:
            db.session.add(ComponenteEmbarcacao(
                embarcacao_id=e.id, categoria=cat,
                label=cat, campos_extra='[]', ordem=['Motor Propulsor','Caixa Inversora/Redutora','Veio Propulsor','Hélice'].index(cat)
            ))
        db.session.commit()
        flash(f'Embarcação {e.nome} criada.','success')
        return redirect(url_for('embarcacao_detalhe', cid=cid, eid=e.id))
    return render_template('embarcacao_form.html', cliente=cliente, embarcacao=None)


@app.route('/clientes/<int:cid>/embarcacoes/<int:eid>')
@login_required
def embarcacao_detalhe(cid, eid):
    cliente = Cliente.query.get_or_404(cid)
    emb = Embarcacao.query.filter_by(id=eid, cliente_id=cid).first_or_404()
    return render_template('embarcacao_detalhe.html', cliente=cliente, embarcacao=emb)


@app.route('/clientes/<int:cid>/embarcacoes/<int:eid>/componente', methods=['POST'])
@login_required
def salvar_componente(cid, eid):
    data = request.get_json()
    comp_id = data.get('id')
    if comp_id:
        comp = ComponenteEmbarcacao.query.get_or_404(comp_id)
    else:
        comp = ComponenteEmbarcacao(embarcacao_id=eid)
        db.session.add(comp)
    comp.categoria   = data.get('categoria','').strip()
    comp.label       = data.get('label','').strip()
    comp.marca       = data.get('marca','').strip()
    comp.modelo      = data.get('modelo','').strip()
    comp.num_serie   = data.get('num_serie','').strip()
    comp.ano         = data.get('ano') or None
    comp.campos_extra= json.dumps(data.get('campos_extra', []))
    comp.notas       = data.get('notas','').strip()
    comp.atualizado_em = datetime.now()
    db.session.commit()
    return jsonify({'ok': True, 'id': comp.id})


@app.route('/componentes/<int:comp_id>', methods=['DELETE'])
@login_required
def apagar_componente(comp_id):
    comp = ComponenteEmbarcacao.query.get_or_404(comp_id)
    db.session.delete(comp); db.session.commit()
    return jsonify({'ok': True})


# ── PHC Consumíveis search ───────────────────────────────────

@app.route('/api/clientes/<int:cid>/consumiveis')
@login_required
def api_consumiveis_cliente(cid):
    cliente = Cliente.query.get_or_404(cid)
    q = request.args.get('q','').strip()
    cfg_phc = ConfigPHC.query.first()
    if not cfg_phc or not cfg_phc.ultima_sync or not cliente.phc_no:
        return jsonify([])
    try:
        from phc_sync import get_phc_connection
        conn = get_phc_connection(cfg_phc)
        cursor = conn.cursor()
        like = f'%{q}%' if q else '%'
        cursor.execute("""
            SELECT TOP 100
                fl.ref, fl.design, fl.qtt, fl.preco, ft.data, ft.fno, ft.serie
            FROM fl
            INNER JOIN ft ON ft.ftstamp=fl.ftstamp
            WHERE ft.no=? AND ft.anulado=0
              AND (fl.ref LIKE ? OR fl.design LIKE ?)
              AND fl.ref IS NOT NULL AND fl.ref<>''
            ORDER BY ft.data DESC
        """, cliente.phc_no, like, like)
        rows = cursor.fetchall()
        conn.close()
        return jsonify([{
            'ref':r[0],'design':r[1],'qtt':float(r[2] or 0),
            'preco':float(r[3] or 0),
            'data':r[4].strftime('%d/%m/%Y') if r[4] else '—',
            'doc':f"{r[6]}{r[5]}"
        } for r in rows])
    except Exception as e:
        return jsonify({'error':str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  CLAUDE CHAT
# ══════════════════════════════════════════════════════════════

@app.route('/chat')
@login_required
def chat_claude():
    cfg = get_config_geral()
    cfg_ia = ConfigIA.query.first()
    return render_template('chat_claude.html', cfg=cfg, cfg_ia=cfg_ia)


@app.route('/api/chat', methods=['POST'])
@login_required
def api_chat():
    """Chat using the configured AI provider (Gemini, Claude, LM Studio, Ollama)."""
    data       = request.get_json()
    # Support both 'messages' array and single 'message' string
    single_msg = data.get('message')
    messages   = data.get('messages', [])
    if single_msg and not messages:
        messages = [{'role': 'user', 'content': single_msg}]
    imagem_b64 = data.get('imagem_b64')
    imagem_tipo= data.get('imagem_tipo', 'image/jpeg')

    cfg_ia  = ConfigIA.query.first()
    cfg     = get_config_geral()
    # Allow caller to override system prompt (e.g. from backlog AI agent)
    _sys_override = data.get('system', '').strip()
    if _sys_override:
        sistema = _sys_override
    else:
        sistema = (cfg.claude_chat_sistema if cfg else None) or \
                  'Es um assistente tecnico especializado em equipamentos e compras industriais. Responde sempre em portugues.'

    if not cfg_ia:
        return jsonify({'error': 'IA não configurada. Configure em Admin → Provedor IA.'}), 400

    provider = cfg_ia.provider

    try:
        # ── Gemini ────────────────────────────────────────────────────────────
        if provider == 'gemini':
            from ai_provider import _get_best_gemini_model
            api_key = cfg_ia.gemini_api_key or ''
            if not api_key:
                return jsonify({'error': 'Chave API Gemini não configurada.'}), 400
            model = cfg_ia.gemini_model or _get_best_gemini_model(api_key)
            # Build conversation history for Gemini
            contents = []
            if sistema:
                contents.append({'role':'user','parts':[{'text': sistema}]})
                contents.append({'role':'model','parts':[{'text': 'Entendido. Estou pronto para ajudar.'}]})
            for m in messages:
                role = 'user' if m['role'] == 'user' else 'model'
                parts = []
                if imagem_b64 and m == messages[-1]:
                    parts.append({'inline_data':{'mime_type': imagem_tipo, 'data': imagem_b64}})
                parts.append({'text': m['content']})
                contents.append({'role': role, 'parts': parts})
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
            payload = json.dumps({
                'contents': contents,
                'generationConfig': {'temperature': 0.7, 'maxOutputTokens': 1500}
            }).encode()
            import urllib.request
            req = urllib.request.Request(url, data=payload,
                headers={'Content-Type':'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=120) as r:
                resp = json.loads(r.read())
            text = resp['candidates'][0]['content']['parts'][0]['text']
            return jsonify({'ok': True, 'text': text, 'provider': 'gemini', 'model': model})

        # ── Claude API ────────────────────────────────────────────────────────
        elif provider == 'claude':
            import anthropic
            client = anthropic.Anthropic(api_key=cfg_ia.claude_api_key or '')
            last = messages[-1] if messages else {'role':'user','content':'Olá'}
            content = []
            if imagem_b64:
                content.append({'type':'image','source':{
                    'type':'base64','media_type':imagem_tipo,'data':imagem_b64}})
            content.append({'type':'text','text':last.get('content','')})
            api_msgs = [{'role':m['role'],'content':m['content']} for m in messages[:-1]]
            api_msgs.append({'role':last['role'],'content':content})
            resp = client.messages.create(
                model='claude-sonnet-4-20250514', max_tokens=1500,
                system=sistema, messages=api_msgs)
            return jsonify({'ok':True,'text':resp.content[0].text, 'provider':'claude'})

        # ── LM Studio / Ollama ────────────────────────────────────────────────
        elif provider in ('lmstudio', 'ollama'):
            import urllib.request
            base = f"http://{cfg_ia.lm_host}:{cfg_ia.lm_port}"
            api_msgs = [{'role':'system','content':sistema}] + [
                {'role':m['role'],'content':m['content']} for m in messages]
            payload = json.dumps({
                'model': cfg_ia.lm_model or 'default',
                'messages': api_msgs,
                'temperature': 0.7, 'max_tokens': 1500, 'stream': False
            }).encode()
            req = urllib.request.Request(base + '/v1/chat/completions', data=payload,
                headers={'Content-Type':'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=60) as r:
                resp = json.loads(r.read())
            return jsonify({'ok':True,'text':resp['choices'][0]['message']['content'],
                           'provider':provider})

        else:
            return jsonify({'error': f'Provedor "{provider}" não suportado no chat.'}), 400

    except Exception as e:
        err = str(e)
        # Friendlier error messages
        if 'Connection refused' in err or 'recusou' in err.lower():
            prov = cfg_ia.provider if cfg_ia else 'local'
            msg = f'Não foi possível ligar ao servidor {prov.upper()}. Verifique se está a correr em {cfg_ia.lm_host}:{cfg_ia.lm_port}.'
        elif 'API_KEY_INVALID' in err or 'invalid' in err.lower():
            msg = 'Chave API inválida. Verifique em Admin → Provedor IA.'
        else:
            msg = err
        return jsonify({'ok': False, 'error': msg})


# ── AUTO-UPDATE ────────────────────────────────────────────────────────────────

@app.route('/admin/update', methods=['GET', 'POST'])
@login_required
def admin_update():
    if not current_user.is_admin:
        flash('Sem permissão.', 'error')
        return redirect(url_for('dashboard'))
    return render_template('admin_update.html')


@app.route('/api/update/check')
@login_required
def update_check():
    """Check if there are updates available on GitHub."""
    if not current_user.is_admin:
        return jsonify({'error': 'Sem permissão'}), 403
    import subprocess
    cwd = os.path.dirname(os.path.abspath(__file__))

    def run(cmd):
        return subprocess.run(cmd, capture_output=True, text=True,
                              timeout=30, cwd=cwd)

    try:
        # Init git if not present
        if not os.path.exists(os.path.join(cwd, '.git')):
            run(['git', 'init'])
            run(['git', 'remote', 'add', 'origin',
                 'https://github.com/luismiguelsilvauni-cpu/comprasnet.git'])
        else:
            # Ensure remote exists
            r = run(['git', 'remote', 'get-url', 'origin'])
            if r.returncode != 0:
                run(['git', 'remote', 'add', 'origin',
                     'https://github.com/luismiguelsilvauni-cpu/comprasnet.git'])

        # Fetch remote without merging
        result = run(['git', 'fetch', 'origin', 'main'])
        if result.returncode != 0:
            return jsonify({'error': f'Erro ao verificar: {result.stderr.strip()}'}), 500

        # Compare local vs remote
        local = subprocess.run(
            ['git', 'rev-parse', 'HEAD'],
            capture_output=True, text=True,
            cwd=os.path.dirname(os.path.abspath(__file__))
        ).stdout.strip()

        remote = subprocess.run(
            ['git', 'rev-parse', 'origin/main'],
            capture_output=True, text=True,
            cwd=os.path.dirname(os.path.abspath(__file__))
        ).stdout.strip()

        # Get list of new commits
        log = subprocess.run(
            ['git', 'log', f'HEAD..origin/main', '--oneline'],
            capture_output=True, text=True,
            cwd=os.path.dirname(os.path.abspath(__file__))
        ).stdout.strip()

        tem_updates = local != remote
        commits = [l for l in log.split('\n') if l.strip()] if log else []

        return jsonify({
            'tem_updates': tem_updates,
            'commits_pendentes': len(commits),
            'commits': commits[:10],
            'versao_local':  local[:8],
            'versao_remota': remote[:8],
        })
    except FileNotFoundError:
        return jsonify({'error': 'Git não encontrado. Instale o Git.'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/update/apply', methods=['POST'])
@login_required
def update_apply():
    """Apply updates from GitHub (git pull + db upgrade)."""
    if not current_user.is_admin:
        return jsonify({'error': 'Sem permissão'}), 403
    import subprocess
    cwd = os.path.dirname(os.path.abspath(__file__))

    def run(cmd):
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=60, cwd=cwd)
        return r.returncode == 0, r.stdout + r.stderr

    steps = []

    # 1. Backup first
    try:
        from backup_manager import fazer_backup
        ok, msg = fazer_backup(app)
        steps.append({'passo': 'Backup', 'ok': ok, 'msg': msg.split("\n")[0]})
    except Exception as e:
        steps.append({'passo': 'Backup', 'ok': False, 'msg': str(e)})

    # 0. Init git if not present
    git_dir = os.path.join(cwd, '.git')
    if not os.path.exists(git_dir):
        ok0, out0 = run(['git', 'init'])
        run(['git', 'remote', 'add', 'origin',
             'https://github.com/luismiguelsilvauni-cpu/comprasnet.git'])
        steps.append({'passo': 'Git inicializado', 'ok': ok0, 'msg': 'Repositório Git criado'})
    else:
        # Ensure remote exists
        ok_r, out_r = run(['git', 'remote', 'get-url', 'origin'])
        if not ok_r:
            run(['git', 'remote', 'add', 'origin',
                 'https://github.com/luismiguelsilvauni-cpu/comprasnet.git'])
            steps.append({'passo': 'Remote configurado', 'ok': True, 'msg': 'github.com/luismiguelsilvauni-cpu/comprasnet'})

    # 2. Git fetch
    ok, out = run(['git', 'fetch', 'origin', 'main'])
    steps.append({'passo': 'Verificar actualizações (git fetch)', 'ok': ok, 'msg': 'OK' if ok else out[:200]})

    # Force checkout all tracked files from origin/main (no merge conflicts)
    ok, out = run(['git', 'checkout', 'origin/main', '--', '.'])
    steps.append({'passo': 'Aplicar ficheiros actualizados', 'ok': ok, 'msg': 'OK' if ok else out[:200]})
    if not ok:
        # Fallback: try pull with unrelated histories
        ok, out = run(['git', 'pull', 'origin', 'main', '--allow-unrelated-histories'])
        steps.append({'passo': 'Download (fallback pull)', 'ok': ok, 'msg': out[:200]})
        if not ok:
            return jsonify({'ok': False, 'steps': steps, 'msg': 'Falhou na actualização'})

    # Auto-register commits in changelog and backlog
    try:
        log_r = subprocess.run(
            ['git', 'log', 'ORIG_HEAD..HEAD', '--oneline', '--no-merges'],
            capture_output=True, text=True, cwd=cwd
        )
        for line in log_r.stdout.strip().splitlines():
            if ' ' in line:
                commit_msg = line.split(' ', 1)[1].strip()
                auto_registar_commit(commit_msg)
    except Exception as e:
        app.logger.warning(f"Auto-register error: {e}")

    # 3. pip install
    import sys
    ok, out = run([sys.executable, '-m', 'pip', 'install', '-r', 'requirements.txt', '-q'])
    steps.append({'passo': 'Dependências Python', 'ok': ok, 'msg': 'OK' if ok else out[:200]})

    # 4. DB migrations
    ok, out = run([sys.executable, '-m', 'flask', 'db', 'upgrade'])
    steps.append({'passo': 'Migração base de dados', 'ok': ok, 'msg': 'OK' if ok else out[:200]})

    # 5. Restart — try service first, fallback to process restart
    steps.append({'passo': 'Reiniciar servidor', 'ok': True, 'msg': 'A reiniciar em 3 segundos...'})

    def restart():
        import time, subprocess, sys
        time.sleep(3)
        # Try NSSM service restart first
        nssm = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'nssm.exe')
        if os.path.exists(nssm):
            subprocess.Popen([nssm, 'restart', 'ComprasNet'],
                           creationflags=subprocess.CREATE_NO_WINDOW)
        else:
            # Restart as standalone process
            subprocess.Popen([sys.executable, os.path.abspath(__file__)],
                           cwd=os.path.dirname(os.path.abspath(__file__)),
                           creationflags=subprocess.CREATE_NO_WINDOW
                           if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
            import signal
            os.kill(os.getpid(), signal.SIGTERM)

    import threading
    threading.Thread(target=restart, daemon=True).start()

    return jsonify({'ok': True, 'steps': steps,
                    'msg': 'Actualização aplicada. O servidor irá reiniciar automaticamente.'})


# ── TUNNEL STATUS ─────────────────────────────────────────────────────────────

# Global variable to store tunnel URL
_tunnel_url = None

def set_tunnel_url(url: str):
    global _tunnel_url
    _tunnel_url = url

def detect_tunnel_url() -> str | None:
    """Try to detect active Cloudflare Tunnel URL from cloudflared process."""
    global _tunnel_url
    if _tunnel_url:
        return _tunnel_url
    try:
        import subprocess, re
        # Try to read cloudflared log from temp file if arrancar_externo.bat writes it
        log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'tunnel.log')
        if os.path.exists(log_path):
            with open(log_path, 'r', errors='ignore') as f:
                content = f.read()
            match = re.search(r'https://[a-z0-9-]+\.trycloudflare\.com', content)
            if match:
                _tunnel_url = match.group(0)
                return _tunnel_url
        # Try cloudflared metrics API
        import urllib.request
        with urllib.request.urlopen('http://localhost:20241/metrics', timeout=2) as r:
            metrics = r.read().decode()
        # Parse hostname from metrics
        match = re.search(r'cloudflared_tunnel_user_hostnames_counts\{[^}]*hostname="([^"]+)"', metrics)
        if match:
            hostname = match.group(1)
            if hostname and '.' in hostname:
                _tunnel_url = f'https://{hostname}'
                return _tunnel_url
    except Exception:
        pass
    return _tunnel_url


@app.route('/conectividade')
@login_required
def conectividade():
    tunnel_url = None
    try:
        import subprocess
        r = subprocess.run(['curl', '-s', 'http://localhost:4040/api/tunnels'],
            capture_output=True, text=True, timeout=2)
        import json as _j
        data = _j.loads(r.stdout)
        tunnels = data.get('tunnels', [])
        if tunnels:
            tunnel_url = tunnels[0].get('public_url', '')
    except: pass
    return render_template('conectividade.html', tunnel_url=tunnel_url)

@app.route('/acesso-externo')
@login_required
def acesso_externo():
    tunnel_url = detect_tunnel_url()
    return render_template('acesso_externo.html', tunnel_url=tunnel_url)


@app.route('/api/tunnel/url', methods=['GET', 'POST'])
@login_required
def api_tunnel_url():
    global _tunnel_url
    if request.method == 'POST':
        data = request.get_json()
        url = data.get('url', '').strip()
        if url.startswith('https://') and 'trycloudflare.com' in url:
            _tunnel_url = url
            return jsonify({'ok': True, 'url': url})
        return jsonify({'error': 'URL inválido'}), 400
    return jsonify({'url': _tunnel_url, 'ativo': _tunnel_url is not None})


@app.route('/api/gemini/models')
@login_required
def api_gemini_models():
    """Fetch available Gemini models from Google API in real-time."""
    cfg_ia = ConfigIA.query.first()
    if not cfg_ia or not cfg_ia.gemini_api_key:
        return jsonify({'error': 'Chave API Gemini não configurada'}), 400
    try:
        import urllib.request, json as _json
        url = f"https://generativelanguage.googleapis.com/v1beta/models?key={cfg_ia.gemini_api_key}"
        req = urllib.request.Request(url, method="GET")
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read())
        # Filter only models that support generateContent
        models = []
        for m in data.get('models', []):
            methods = m.get('supportedGenerationMethods', [])
            if 'generateContent' in methods:
                name = m['name'].replace('models/', '')
                # Only include flash/pro variants (skip embedding etc)
                if any(x in name for x in ['flash', 'pro', 'ultra']):
                    models.append({
                        'id': name,
                        'label': m.get('displayName', name),
                        'description': m.get('description', '')[:80],
                    })
        # Always include known stable models even if not in list
        known = [
            {'id':'gemini-2.0-flash',            'label':'Gemini 2.0 Flash',              'description':'Rápido e eficiente'},
            {'id':'gemini-2.0-flash-lite',        'label':'Gemini 2.0 Flash Lite',         'description':'Mais leve'},
            {'id':'gemini-1.5-flash',             'label':'Gemini 1.5 Flash',              'description':'Estável'},
            {'id':'gemini-1.5-flash-latest',      'label':'Gemini 1.5 Flash (latest)',     'description':'Versão mais recente 1.5'},
            {'id':'gemini-1.5-pro',               'label':'Gemini 1.5 Pro',               'description':'Mais capaz, mais lento'},
            {'id':'gemini-2.0-flash-exp',         'label':'Gemini 2.0 Flash Exp',          'description':'Experimental'},
        ]
        existing_ids = {m['id'] for m in models}
        for k in known:
            if k['id'] not in existing_ids:
                models.insert(0, k)
        models.sort(key=lambda x: (0 if 'flash' in x['id'] else 1, x['id']))
        return jsonify({'models': models})
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        # Even on error, return the known stable models
        return jsonify({'models': [
            {'id':'gemini-2.0-flash',        'label':'Gemini 2.0 Flash',        'description':'Recomendado'},
            {'id':'gemini-1.5-flash',        'label':'Gemini 1.5 Flash',        'description':'Estável'},
            {'id':'gemini-1.5-pro',          'label':'Gemini 1.5 Pro',          'description':'Mais capaz'},
            {'id':'gemini-2.0-flash-lite',   'label':'Gemini 2.0 Flash Lite',   'description':'Mais rápido'},
        ], 'warning': f'API retornou {e.code} — a mostrar modelos conhecidos'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── PWA + MOBILE ──────────────────────────────────────────────────────────────

@app.route('/manifest-static.json')
def manifest_static():
    return send_from_directory('static', 'manifest.json',
                               mimetype='application/manifest+json')



@app.route('/versao')
def versao():
    return 'VERSAO-f3effc2-OK-PERIODOS-FECHO'

@app.route('/pwa-test')
def pwa_test():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'pwa-test.html')


@app.route('/sw.js')
def service_worker():
    return send_from_directory('static', 'sw.js',
                               mimetype='application/javascript')

@app.route('/mobile')
@login_required
def mobile_home():
    total_artigos  = ArtigoPHC.query.count()
    pedidos_abertos = PedidoCompra.query.filter_by(estado='aberto').count()
    total_clientes = Cliente.query.filter_by(ativo=True).count()
    return render_template('mobile/home.html',
        total_artigos=total_artigos,
        pedidos_abertos=pedidos_abertos,
        total_clientes=total_clientes)

@app.route('/mobile/artigos')
@login_required
def mobile_artigos():
    q = request.args.get('q', '').strip()
    artigos = []
    if q and len(q) >= 2:
        like = f'%{q}%'
        artigos = ArtigoPHC.query.filter(db.or_(
            ArtigoPHC.referencia.ilike(like),
            ArtigoPHC.designacao.ilike(like),
        )).order_by(ArtigoPHC.referencia).limit(30).all()
    return render_template('mobile/artigos.html', artigos=artigos, q=q)

@app.route('/mobile/artigos/<ref>')
@login_required
def mobile_artigo_detalhe(ref):
    artigo = ArtigoPHC.query.filter_by(referencia=ref).first_or_404()
    historico = []
    try:
        cfg_phc = ConfigPHC.query.first()
        if cfg_phc and cfg_phc.ultima_sync:
            from phc_sync import get_historico_compras
            historico = get_historico_compras(cfg_phc, ref)[:5]
    except Exception:
        pass
    return render_template('mobile/artigo_detalhe.html', artigo=artigo, historico=historico)

@app.route('/mobile/pedidos')
@login_required
def mobile_pedidos():
    estado = request.args.get('estado', '')
    q = PedidoCompra.query
    if estado:
        q = q.filter_by(estado=estado)
    pedidos = q.order_by(PedidoCompra.data_criacao.desc()).limit(30).all()
    return render_template('mobile/pedidos.html', pedidos=pedidos, estado=estado)

@app.route('/mobile/clientes')
@login_required
def mobile_clientes():
    q_str = request.args.get('q', '').strip()
    query = Cliente.query.filter_by(ativo=True)
    if q_str:
        like = f'%{q_str}%'
        query = query.filter(db.or_(
            Cliente.nome.ilike(like),
            Cliente.nif.ilike(like),
        ))
    clientes = query.order_by(Cliente.nome).limit(30).all()
    return render_template('mobile/clientes.html', clientes=clientes, q=q_str)

@app.route('/mobile/clientes/<int:cid>')
@login_required
def mobile_cliente_detalhe(cid):
    cliente = Cliente.query.get_or_404(cid)
    return render_template('mobile/cliente_detalhe.html', cliente=cliente)

@app.route('/mobile/chat')
@login_required
def mobile_chat():
    cfg_ia = ConfigIA.query.first()
    return render_template('mobile/chat.html', cfg_ia=cfg_ia)


# ══════════════════════════════════════════════════════════════
#  STOCK
# ══════════════════════════════════════════════════════════════

@app.route('/stock')
@login_required
def stock():
    q = request.args.get('q', '').strip()
    if q and len(q) >= 2:
        like = f'%{q}%'
        artigos = ArtigoPHC.query.filter(db.or_(
            ArtigoPHC.referencia.ilike(like),
            ArtigoPHC.designacao.ilike(like),
            ArtigoPHC.familia.ilike(like),
        )).order_by(ArtigoPHC.referencia).limit(500).all()
    else:
        artigos = ArtigoPHC.query.order_by(ArtigoPHC.referencia).limit(200).all()
    return render_template('stock.html', artigos=artigos, q=q, total=ArtigoPHC.query.count())


@app.route('/stock/<ref>')
@login_required
def stock_artigo(ref):
    artigo = ArtigoPHC.query.filter_by(referencia=ref).first_or_404()
    historico = []
    vendas = []
    try:
        cfg_phc = ConfigPHC.query.first()
        if cfg_phc and cfg_phc.ultima_sync:
            from phc_sync import get_historico_compras, get_phc_connection
            historico = get_historico_compras(cfg_phc, ref)
            # Get sales history
            try:
                conn = get_phc_connection(cfg_phc)
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT TOP 30 fl.preco, fl.qtt, ft.data, ec.nome
                    FROM fl
                    INNER JOIN ft ON ft.ftstamp=fl.ftstamp
                    LEFT JOIN ec ON ec.no=ft.no
                    WHERE fl.ref=? AND ft.anulado=0
                    ORDER BY ft.data DESC
                """, ref)
                rows = cursor.fetchall()
                conn.close()
                vendas = [{'preco': float(r[0] or 0), 'qtt': float(r[1] or 0),
                           'data': r[2].strftime('%Y-%m-%d') if r[2] else '',
                           'cliente': r[3] or ''} for r in rows]
            except Exception:
                pass
    except Exception:
        pass

    # Get notes for this article
    from models import NotaArtigo
    notas = NotaArtigo.query.filter_by(artigo_ref=ref).order_by(NotaArtigo.criado_em.desc()).all()
    return render_template('stock_artigo.html', artigo=artigo,
                           historico=historico, vendas=vendas, notas=notas)


@app.route('/stock/<ref>/nota', methods=['POST'])
@login_required
def stock_artigo_nota(ref):
    artigo = ArtigoPHC.query.filter_by(referencia=ref).first_or_404()
    from models import NotaArtigo
    texto = request.form.get('texto', '').strip()
    if texto:
        nota = NotaArtigo(artigo_ref=ref, texto=texto,
                          criado_por=current_user.id)
        db.session.add(nota)
        db.session.commit()
        flash('Nota guardada.', 'success')
    return redirect(url_for('stock_artigo', ref=ref))


@app.route('/stock/<ref>/nota/<int:nid>/apagar', methods=['POST'])
@login_required
def stock_apagar_nota(ref, nid):
    from models import NotaArtigo
    nota = NotaArtigo.query.get_or_404(nid)
    db.session.delete(nota); db.session.commit()
    return jsonify({'ok': True})




# ── DASHBOARD WIDGETS ─────────────────────────────────────────────────────────

@app.route('/api/dashboard/layout', methods=['GET'])
@login_required
def get_dashboard_layout():
    """Get saved widget layout for current user."""
    from models import ConfigGeral
    cfg = get_config_geral()
    key = f'dashboard_layout_{current_user.id}'
    layout_json = getattr(cfg, 'dashboard_layouts', None)
    # Store per-user layouts in a simple JSON field
    try:
        import json as _json
        layouts = _json.loads(cfg.dashboard_layouts or '{}') if hasattr(cfg, 'dashboard_layouts') else {}
        return jsonify(layouts.get(str(current_user.id), []))
    except Exception:
        return jsonify([])

@app.route('/api/dashboard/layout', methods=['POST'])
@login_required
def save_dashboard_layout():
    """Save widget layout for current user."""
    data = request.get_json()
    try:
        from models import ConfigGeral
        cfg = get_config_geral()
        if not hasattr(cfg, 'dashboard_layouts') or cfg.dashboard_layouts is None:
            cfg.dashboard_layouts = '{}'
        layouts = json.loads(cfg.dashboard_layouts)
        layouts[str(current_user.id)] = data
        cfg.dashboard_layouts = json.dumps(layouts)
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  CALENDÁRIO
# ══════════════════════════════════════════════════════════════

@app.route('/api/calendario/eventos')
@login_required
def api_calendario_eventos():
    """Return events for a given month/range."""
    ano  = request.args.get('ano',  type=int, default=datetime.now().year)
    mes  = request.args.get('mes',  type=int, default=datetime.now().month)
    from datetime import date
    inicio = date(ano, mes, 1)
    # Last day of month
    import calendar as cal
    ultimo = cal.monthrange(ano, mes)[1]
    fim    = date(ano, mes, ultimo)

    eventos = EventoCalendario.query.filter(
        EventoCalendario.data_inicio >= inicio,
        EventoCalendario.data_inicio <= fim
    ).order_by(EventoCalendario.data_inicio).all()

    return jsonify([{
        'id':          e.id,
        'titulo':      e.titulo,
        'tipo':        e.tipo,
        'data_inicio': e.data_inicio.strftime('%Y-%m-%d'),
        'data_fim':    e.data_fim.strftime('%Y-%m-%d') if e.data_fim else None,
        'hora':        e.hora,
        'descricao':   e.descricao,
        'artigos':     json.loads(e.artigos_json or '[]'),
        'pedido_id':   e.pedido_id,
        'fornecedor':  e.fornecedor,
        'concluido':   e.concluido,
        'criado_por':  e.criado_por,
        'criado_nome': User.query.get(e.criado_por).nome if e.criado_por else '—',
    } for e in eventos])


@app.route('/api/calendario/eventos', methods=['POST'])
@login_required
def api_criar_evento():
    if not current_user.is_authenticated:
        return jsonify({'error': 'Sessão expirada'}), 401
    if not current_user.is_authenticated:
        return jsonify({'error': 'Sessão expirada — recarregue a página'}), 401
    data = request.get_json()
    from datetime import date
    def parse_date(s):
        return datetime.strptime(s, '%Y-%m-%d').date() if s else None

    e = EventoCalendario(
        titulo       = data.get('titulo','').strip() or 'Sem título',
        tipo         = data.get('tipo','manual'),
        data_inicio  = parse_date(data.get('data_inicio')),
        data_fim     = parse_date(data.get('data_fim')),
        hora         = (data.get('hora') or '').strip() or None,
        descricao    = data.get('descricao','').strip(),
        artigos_json = json.dumps(data.get('artigos',[])),
        pedido_id    = data.get('pedido_id'),
        fornecedor   = data.get('fornecedor','').strip(),
        concluido    = data.get('concluido', False),
        criado_por   = current_user.id,
    )
    db.session.add(e)
    db.session.commit()
    return jsonify({'ok': True, 'id': e.id})


@app.route('/api/calendario/eventos/<int:eid>', methods=['PUT'])
@login_required
def api_atualizar_evento(eid):
    e    = EventoCalendario.query.get_or_404(eid)
    data = request.get_json()
    from datetime import date
    def parse_date(s):
        return datetime.strptime(s, '%Y-%m-%d').date() if s else None

    if 'titulo'      in data: e.titulo       = data['titulo']
    if 'tipo'        in data: e.tipo         = data['tipo']
    if 'data_inicio' in data: e.data_inicio  = parse_date(data['data_inicio'])
    if 'data_fim'    in data: e.data_fim     = parse_date(data['data_fim'])
    if 'hora'        in data: e.hora         = data['hora'] or None
    if 'descricao'   in data: e.descricao    = data['descricao']
    if 'artigos'     in data: e.artigos_json = json.dumps(data['artigos'])
    if 'fornecedor'  in data: e.fornecedor   = data['fornecedor']
    if 'concluido'   in data: e.concluido    = data['concluido']
    if 'pedido_id'   in data: e.pedido_id    = data['pedido_id']
    e.atualizado_em = datetime.now()
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/calendario/eventos/<int:eid>', methods=['DELETE'])
@login_required
def api_apagar_evento(eid):
    e = EventoCalendario.query.get_or_404(eid)
    db.session.delete(e)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/calendario/pedido/<int:pid>/criar_eventos', methods=['POST'])
@login_required
def api_criar_eventos_pedido(pid):
    """Create purchase + delivery events from an approved pedido."""
    pedido = PedidoCompra.query.get_or_404(pid)
    data   = request.get_json()
    from datetime import date, timedelta

    artigos = [{'ref': l.artigo_ref, 'design': l.designacao, 'qtt': l.quantidade}
               for l in pedido.linhas if l.artigo_ref]

    data_compra = datetime.strptime(data.get('data_compra', datetime.now().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
    dias_entrega = int(data.get('dias_entrega', 7))
    data_entrega = data_compra + timedelta(days=dias_entrega)
    fornecedor   = data.get('fornecedor', '')

    # Create compra event
    e1 = EventoCalendario(
        titulo       = f'Compra: PC-{pid:04d} — {pedido.titulo[:40]}',
        tipo         = 'compra',
        data_inicio  = data_compra,
        artigos_json = json.dumps(artigos),
        pedido_id    = pid,
        fornecedor   = fornecedor,
        criado_por   = current_user.id,
    )
    # Create delivery event
    e2 = EventoCalendario(
        titulo       = f'Entrega prevista: PC-{pid:04d} — {pedido.titulo[:30]}',
        tipo         = 'entrega_prevista',
        data_inicio  = data_entrega,
        artigos_json = json.dumps(artigos),
        pedido_id    = pid,
        fornecedor   = fornecedor,
        criado_por   = current_user.id,
    )
    db.session.add(e1); db.session.add(e2)
    db.session.commit()
    return jsonify({'ok': True, 'id_compra': e1.id, 'id_entrega': e2.id,
                    'data_entrega': data_entrega.strftime('%Y-%m-%d')})


# ── HEALTH CHECK ──────────────────────────────────────────────────────────────

@app.route('/admin/health')
@login_required
def admin_health():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    from health_check import startup_check
    result = startup_check(app, auto_fix=True)
    return render_template('admin_health.html', result=result)


@app.route('/admin/config/remove-logo', methods=['POST'])
@login_required
def remove_logo():
    if not current_user.is_admin:
        return jsonify({'error': 'Sem permissão'}), 403
    cfg = get_config_geral()
    if cfg.empresa_logo_path:
        logo_path = os.path.join(app.config['UPLOAD_FOLDER'], cfg.empresa_logo_path)
        try:
            if os.path.exists(logo_path):
                os.remove(logo_path)
        except Exception:
            pass
        cfg.empresa_logo_path = None
        db.session.commit()
    return jsonify({'ok': True})


@app.route('/reposicao/por-fornecedor')
@login_required
def reposicao_por_fornecedor():
    """Group articles needing replenishment by their usual supplier."""
    from reposicao import CONFIG_PADRAO, analisar_todos
    import json as _json

    cfg_phc    = ConfigPHC.query.first()
    cfg_global = ConfigReposicao.query.filter_by(artigo_ref=None).first()

    config = dict(CONFIG_PADRAO)
    if cfg_global:
        config.update({
            'meses_historico':             cfg_global.meses_historico or 60,
            'lead_time_dias':              cfg_global.lead_time_dias or 7,
            'fator_seguranca':             cfg_global.fator_seguranca or 1.5,
            'meses_cobertura':             cfg_global.meses_cobertura or 2,
            'custo_encomenda':             cfg_global.custo_encomenda or 25,
            'taxa_posse_anual':            cfg_global.taxa_posse_anual or 0.20,
            'quantidade_minima_encomenda': cfg_global.quantidade_minima_encomenda or 1,
            'alertar_dias_cobertura':      cfg_global.alertar_dias_cobertura or 30,
            'min_anos_historico':          getattr(cfg_global,'min_anos_historico',2) or 2,
            'min_meses_com_venda':         getattr(cfg_global,'min_meses_com_venda',3) or 3,
            'min_total_vendido':           getattr(cfg_global,'min_total_vendido',3) or 3,
            'ignorar_sem_movimento_anos':  getattr(cfg_global,'ignorar_sem_movimento_anos',3) or 3,
            'min_facturas_sugerir':        getattr(cfg_global,'min_facturas_sugerir',8) or 8,
        })

    # SQL: for each article, find most frequent supplier from purchase invoices (fo table)
    SQL_FORNECEDOR_ARTIGO = (
        "SELECT fi.ref, fo.nome AS fornecedor, fo.no AS fornecedor_no,"
        " COUNT(*) AS n_compras"
        " FROM fi"
        " INNER JOIN ft ON ft.ftstamp = fi.ftstamp"
        " INNER JOIN fo ON fo.no = ft.no"
        " WHERE fi.ref IS NOT NULL AND fi.ref <> ''"
        " AND fi.qtt > 0"
        " GROUP BY fi.ref, fo.nome, fo.no"
    )

    # Run bulk analysis using local DB only (no PHC connection needed here)
    artigos = ArtigoPHC.query.filter(ArtigoPHC.stock_atual >= 0).all()
    erro = None
    try:
        resultados = analisar_todos(cfg_phc, config, artigos)
    except Exception as e:
        import traceback
        erro = str(e)
        logger.error(traceback.format_exc())
        resultados = []

    app.logger.info(f"Total resultados: {len(resultados)}")
    precisam = [r for r in resultados if r.get('quantidade_sugerida', 0) > 0
                and r.get('recomendacao') != 'nao_fazer_stock']
    app.logger.info(f"Precisam encomendar: {len(precisam)}")

    # Fetch supplier per article from PHC - always try regardless of analysis
    fornecedor_por_ref = {}
    print(f"[DEBUG] A buscar fornecedores PHC, precisam={len(precisam)}, cfg_phc={cfg_phc is not None}")
    if cfg_phc:
        try:
            from phc_sync import get_phc_connection
            conn = get_phc_connection(cfg_phc)
            cursor = conn.cursor()
            # Simpler query: last purchase invoice per article
            SQL_FORN_SIMPLES = (
                "SELECT RTRIM(fn.ref) AS ref, fo.nome, fo.no"
                " FROM fn"
                " INNER JOIN fo ON fo.fostamp = fn.fostamp"
                " WHERE fn.ref IS NOT NULL AND RTRIM(fn.ref) <> ''"
                " AND fn.qtt > 0"
            )
            cursor.execute(SQL_FORN_SIMPLES)
            contagens = {}
            rows = cursor.fetchall()
            print(f'[DEBUG] Supplier rows fetched: {len(rows)}')
            for ref, forn_nome, forn_no in rows:
                ref = (ref or '').strip()
                if not ref: continue
                if ref not in contagens:
                    contagens[ref] = {}
                key = str(forn_no)
                contagens[ref][key] = contagens[ref].get(key, {'nome': forn_nome, 'no': forn_no, 'n': 0})
                contagens[ref][key]['n'] += 1
            # Pick most frequent supplier per article
            for ref, forns in contagens.items():
                best = max(forns.values(), key=lambda x: x['n'])
                fornecedor_por_ref[ref] = {'nome': best['nome'], 'no': best['no']}
            print(f'[DEBUG] Suppliers identified: {len(fornecedor_por_ref)}')
            # Debug: check if precisam refs match
            if precisam:
                sample = [r['referencia'] for r in precisam[:5]]
                print(f'[DEBUG] Sample precisam refs: {sample}')
                # Check first key in fornecedor_por_ref
                if fornecedor_por_ref:
                    first_key = next(iter(fornecedor_por_ref))
                    print(f'[DEBUG] First key in dict: {repr(first_key)}')
                for ref in sample:
                    forn = fornecedor_por_ref.get(ref)
                    forn2 = fornecedor_por_ref.get(ref.strip())
                    forn3 = fornecedor_por_ref.get(ref.upper())
                    print(f'[DEBUG]   {repr(ref)} -> direct={forn} strip={forn2} upper={forn3}')
            conn.close()
        except Exception as e:
            app.logger.warning(f"PHC supplier fetch error: {e}")

    # Group by supplier
    por_fornecedor = {}
    sem_fornecedor = []
    for r in precisam:
        ref_key = r['referencia'].strip()
        forn = fornecedor_por_ref.get(ref_key)
        if forn:
            key = forn['nome']
            if key not in por_fornecedor:
                por_fornecedor[key] = {'nome': key, 'no': forn['no'], 'artigos': []}
            por_fornecedor[key]['artigos'].append(r)
        else:
            sem_fornecedor.append(r)

    # Sort by supplier name
    fornecedores = sorted(por_fornecedor.values(), key=lambda x: x['nome'])
    if sem_fornecedor:
        fornecedores.append({'nome': 'Sem fornecedor identificado', 'no': None,
                             'artigos': sem_fornecedor})

    return render_template('reposicao_fornecedor.html',
        fornecedores=fornecedores,
        total_artigos=sum(len(f['artigos']) for f in fornecedores),
        config=config, erro=erro if 'erro' in dir() else None)


@app.route('/reposicao/criar-pedido', methods=['POST'])
@login_required
def reposicao_criar_pedido():
    """Create purchase order from replenishment suggestions for one supplier."""
    data         = request.get_json()
    fornecedor   = data.get('fornecedor', 'Fornecedor')
    artigos      = data.get('artigos', [])

    if not artigos:
        return jsonify({'error': 'Sem artigos seleccionados'}), 400

    titulo = f"Reposicao - {fornecedor} - {datetime.now().strftime('%d/%m/%Y')}"
    p = PedidoCompra(
        titulo       = titulo,
        descricao    = f"Pedido gerado automaticamente pela analise de reposicao de stock.",
        prioridade   = 'normal',
        estado       = 'aberto',
        criado_por   = current_user.id,
        data_criacao = datetime.now(),
    )
    db.session.add(p)
    db.session.flush()

    for i, a in enumerate(artigos):
        ref    = a.get('referencia', '')
        artigo = ArtigoPHC.query.filter_by(referencia=ref).first()
        linha  = LinhaPedido(
            pedido_id       = p.id,
            ordem           = i,
            artigo_ref      = ref,
            referencia      = ref,
            designacao      = a.get('designacao', ''),
            unidade         = a.get('unidade', 'un'),
            quantidade      = int(a.get('quantidade_sugerida', 1)),
            stock_atual     = a.get('stock_atual', 0),
            preco_custo_ref = artigo.preco_custo if artigo else 0,
            preco_pcp_ref   = artigo.preco_custo_ponderado if artigo else 0,
            fornecedor_hab  = fornecedor,
            observacoes     = f"EOQ sugerido pela analise de reposicao. Score: {a.get('score','-')}",
        )
        db.session.add(linha)

    db.session.commit()
    return jsonify({'ok': True, 'pedido_id': p.id,
                    'url': url_for('pedido_detalhe', pid=p.id)})


@app.route('/api/reposicao/resultados')
@login_required
def api_reposicao_resultados():
    """Return paginated replenishment results as JSON."""
    from reposicao import CONFIG_PADRAO, analisar_todos
    cfg_phc    = ConfigPHC.query.first()
    cfg_global = ConfigReposicao.query.filter_by(artigo_ref=None).first()
    config = dict(CONFIG_PADRAO)
    if cfg_global:
        config.update({
            'meses_historico':            cfg_global.meses_historico or 60,
            'lead_time_dias':             cfg_global.lead_time_dias or 7,
            'fator_seguranca':            cfg_global.fator_seguranca or 1.5,
            'meses_cobertura':            cfg_global.meses_cobertura or 2,
            'custo_encomenda':            cfg_global.custo_encomenda or 25,
            'taxa_posse_anual':           cfg_global.taxa_posse_anual or 0.20,
            'quantidade_minima_encomenda':cfg_global.quantidade_minima_encomenda or 1,
            'alertar_dias_cobertura':     cfg_global.alertar_dias_cobertura or 30,
            'min_anos_historico':         getattr(cfg_global,'min_anos_historico',2) or 2,
            'min_meses_com_venda':        getattr(cfg_global,'min_meses_com_venda',3) or 3,
            'min_total_vendido':          getattr(cfg_global,'min_total_vendido',3) or 3,
            'ignorar_sem_movimento_anos': getattr(cfg_global,'ignorar_sem_movimento_anos',3) or 3,
            'min_facturas_sugerir':       getattr(cfg_global,'min_facturas_sugerir',8) or 8,
        })

    urgencia  = request.args.get('urgencia', 'necessarios')
    pesquisa  = request.args.get('q', '').strip().lower()
    sort_col  = request.args.get('sort', '')
    sort_dir  = request.args.get('dir', 'asc')
    page      = int(request.args.get('page', 1))
    per_page  = 50

    artigos = ArtigoPHC.query.filter(ArtigoPHC.stock_atual >= 0).all()
    try:
        resultados = analisar_todos(cfg_phc, config, artigos)
    except Exception as e:
        return jsonify({'error': str(e), 'rows': [], 'total': 0})

    # Filter by urgencia
    SEM = {'parado','pouco_historico','irrelevante','sem_dados','nao_recomendado'}
    PRECISAM = {'critico','urgente','necessario'}
    def match_urg(r):
        u = r.get('urgencia','')
        if urgencia == 'todos':       return True
        if urgencia == 'necessarios': return u in PRECISAM
        if urgencia == 'sem_relevancia': return u in SEM
        return u == urgencia

    rows = [r for r in resultados if match_urg(r)]

    # Filter by search
    if pesquisa:
        rows = [r for r in rows if pesquisa in (r.get('referencia','') or '').lower()
                or pesquisa in (r.get('designacao','') or '').lower()]

    total = len(rows)

    # Sort
    if sort_col:
        rev = sort_dir == 'desc'
        key_map = {
            'referencia':   lambda r: r.get('referencia','') or '',
            'designacao':   lambda r: r.get('designacao','') or '',
            'stock':        lambda r: float(r.get('stock_atual',0) or 0),
            'media_ano':    lambda r: float(r.get('consumo_medio_mensal',0) or 0)*12,
            'score':        lambda r: float(r.get('score',0) or 0),
            'dias_cob':     lambda r: float(r.get('dias_cobertura_atual',0) or 0),
            'qtd':          lambda r: float(r.get('quantidade_sugerida',0) or 0),
        }
        fn = key_map.get(sort_col)
        if fn:
            rows.sort(key=fn, reverse=rev)

    # Paginate
    start = (page-1) * per_page
    rows_page = rows[start:start+per_page]

    return jsonify({
        'rows': rows_page,
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': max(1, (total + per_page - 1) // per_page),
    })


@app.route('/api/clientes')
@login_required
def api_clientes():
    q = request.args.get('q','').strip()
    if not q or len(q) < 2:
        return jsonify([])
    # Try local Cliente table first
    try:
        clientes = Cliente.query.filter(Cliente.nome.ilike(f'%{q}%')).limit(10).all()
        if clientes:
            return jsonify([{'no': c.no if hasattr(c,'no') else c.id, 'nome': c.nome} for c in clientes])
    except Exception:
        pass
    # Query PHC directly
    try:
        cfg_phc = ConfigPHC.query.first()
        if cfg_phc:
            from phc_sync import get_phc_connection
            conn = get_phc_connection(cfg_phc)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT no, nome FROM cl WHERE nome LIKE ? AND ISNULL(inactivo,0)=0 ORDER BY nome",
                (f'%{q}%',)
            )
            rows = [{'no': r[0], 'nome': (r[1] or '').strip()} for r in cursor.fetchmany(10)]
            conn.close()
            return jsonify(rows)
    except Exception as e:
        app.logger.warning(f"api_clientes PHC error: {e}")
    return jsonify([])

@app.route('/roadmap')
@login_required
def roadmap():
    return render_template('roadmap.html')


class ChangelogEntry(db.Model):
    __tablename__ = 'changelog_entry'
    id          = db.Column(db.Integer, primary_key=True)
    versao      = db.Column(db.String(20), nullable=False)
    descricao   = db.Column(db.Text, default='')
    tipo        = db.Column(db.String(20), default='feat')  # feat/fix/chore
    commit_msg  = db.Column(db.Text, default='')
    criado_em   = db.Column(db.DateTime, default=datetime.now)


class Equipamento(db.Model):
    __tablename__ = 'equipamento'
    id              = db.Column(db.Integer, primary_key=True)
    # Cliente / Embarcacao
    cliente_nome    = db.Column(db.String(200))
    embarcacao      = db.Column(db.String(200))
    # Motor
    motor_marca     = db.Column(db.String(100))
    motor_modelo    = db.Column(db.String(200))
    motor_potencia  = db.Column(db.String(50))
    serial_number   = db.Column(db.String(100), index=True)
    catalogo        = db.Column(db.String(100))
    base_code       = db.Column(db.String(100))
    manufactured_date = db.Column(db.String(50))
    # Caixa Redutora
    caixa_modelo    = db.Column(db.String(200))
    caixa_ratio     = db.Column(db.String(50))
    caixa_serial    = db.Column(db.String(100))
    # Legacy / outros
    catalogo           = db.Column(db.String(100))
    material           = db.Column(db.String(200))
    manufacturing_date = db.Column(db.String(50))
    base_engine_pt     = db.Column(db.String(300))
    base_engine_eng    = db.Column(db.String(300))
    fuel_system_pt     = db.Column(db.String(300))
    fuel_system_eng    = db.Column(db.String(300))
    tipo_motor      = db.Column(db.String(20), default='principal')
    ativo           = db.Column(db.Boolean, default=True)  # principal / auxiliar
    notas           = db.Column(db.Text, default='')
    criado_em       = db.Column(db.DateTime, default=datetime.now)
    opcoes          = db.relationship('EquipamentoOpcao', backref='equipamento', lazy=True, cascade='all, delete-orphan')
    consumiveis     = db.relationship('EquipamentoConsumivel', backref='equipamento', lazy=True, cascade='all, delete-orphan')
    documentos      = db.relationship('EquipamentoDocumento', backref='equipamento', lazy=True, cascade='all, delete-orphan')
    motores_aux     = db.relationship('EquipamentoMotorAux', backref='equipamento', lazy=True, cascade='all, delete-orphan', order_by='EquipamentoMotorAux.numero')

class EquipamentoConsumivel(db.Model):
    __tablename__ = 'equipamento_consumivel'
    id              = db.Column(db.Integer, primary_key=True)
    equipamento_id  = db.Column(db.Integer, db.ForeignKey('equipamento.id'), nullable=False)
    ref             = db.Column(db.String(50))
    designacao      = db.Column(db.String(300))
    unidade         = db.Column(db.String(20))
    quantidade      = db.Column(db.Float, default=1)
    notas           = db.Column(db.String(300))
    ordem           = db.Column(db.Integer, default=0)


# ── PERFIS E PERMISSÕES ──────────────────────────────────────────────────────

# ── MAPA DE FÉRIAS ─────────────────────────────────────────────────────────────

# Palette of distinct colors for employees
FERIAS_CORES = [
    '#3b6ef0','#22c55e','#f59e0b','#ef4444','#a855f7',
    '#06b6d4','#ec4899','#10b981','#f97316','#6366f1',
    '#84cc16','#14b8a6','#e11d48','#7c3aed','#0891b2',
]

def _ferias_cor(funcionario_id):
    """Assign a stable color to each employee."""
    return FERIAS_CORES[(funcionario_id - 1) % len(FERIAS_CORES)]

# Default Portuguese public holidays
PT_FERIADOS = [
    (1,1,'Ano Novo'),(4,25,'25 de Abril'),(5,1,'Dia do Trabalhador'),
    (6,10,'Dia de Portugal'),(8,15,'Assunção de Nossa Senhora'),
    (10,5,'Implantação da República'),(11,1,'Dia de Todos os Santos'),
    (12,1,'Restauração da Independência'),(12,8,'Imaculada Conceição'),
    (12,25,'Natal'),
]

@app.route('/ferias')
@login_required
def ferias_mapa():
    # Redirect to the new unified ausencias view
    return redirect(url_for('ausencias'))

@app.route('/ferias_legacy')
@login_required
def ferias_mapa_legacy():
    from datetime import date
    ano = int(request.args.get('ano', date.today().year))
    # Get all employees
    funcs = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()         if hasattr(Funcionario, 'ativo') else         Funcionario.query.order_by(Funcionario.nome).all()
    # Get periodos for this year
    periodos = FeriasPeriodo.query.filter_by(ano=ano).all()
    # Get feriados for this year
    feriados = FeriasFeriado.query.filter_by(ano=ano).all()
    # If no feriados exist for this year, seed defaults
    if not feriados:
        for mes, dia, nome in PT_FERIADOS:
            try:
                db.session.add(FeriasFeriado(
                    ano=ano, data=date(ano,mes,dia), nome=nome, tipo='nacional'))
            except: pass
        try: db.session.commit()
        except: db.session.rollback()
        feriados = FeriasFeriado.query.filter_by(ano=ano).all()
    import json
    periodos_json = json.dumps([{
        'id': p.id,
        'funcionario_id': p.funcionario_id,
        'funcionario_nome': p.funcionario.nome if p.funcionario else '?',
        'data_inicio': p.data_inicio.isoformat(),
        'data_fim': p.data_fim.isoformat(),
        'tipo': p.tipo,
        'notas': p.notas or '',
        'cor': p.cor or _ferias_cor(p.funcionario_id),
    } for p in periodos])
    feriados_json = json.dumps([{
        'id': f.id, 'data': f.data.isoformat(),
        'nome': f.nome, 'tipo': f.tipo,
    } for f in feriados])
    return render_template('ferias.html',
        ano=ano, funcs=funcs,
        periodos_json=periodos_json,
        feriados_json=feriados_json,
        func_cores=json.dumps({f.id: _ferias_cor(f.id) for f in funcs}))

@app.route('/ferias/periodo/adicionar', methods=['POST'])
@login_required
def ferias_add():
    data = request.get_json() or {}
    from datetime import date as _d
    try:
        inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        fim    = datetime.strptime(data['data_fim'],    '%Y-%m-%d').date()
    except: return jsonify({'ok': False, 'error': 'Datas inválidas'})
    p = FeriasPeriodo(
        funcionario_id=int(data['funcionario_id']),
        ano=inicio.year, data_inicio=inicio, data_fim=fim,
        tipo=data.get('tipo','ferias'),
        notas=data.get('notas','').strip(),
        cor=data.get('cor',''),
        criado_por=current_user.id, criado_em=datetime.now()
    )
    db.session.add(p); db.session.commit()
    f = Funcionario.query.get(p.funcionario_id)
    return jsonify({'ok': True, 'id': p.id,
        'cor': p.cor or _ferias_cor(p.funcionario_id),
        'funcionario_nome': f.nome if f else '?'})

@app.route('/ferias/periodo/<int:pid>/editar', methods=['POST'])
@login_required
def ferias_editar(pid):
    p = FeriasPeriodo.query.get_or_404(pid)
    data = request.get_json() or {}
    if 'data_inicio' in data:
        p.data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
    if 'data_fim' in data:
        p.data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
    if 'notas' in data: p.notas = data['notas']
    if 'tipo'  in data: p.tipo  = data['tipo']
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/ferias/periodo/<int:pid>/eliminar', methods=['POST'])
@login_required
def ferias_eliminar(pid):
    p = FeriasPeriodo.query.get_or_404(pid)
    db.session.delete(p); db.session.commit()
    return jsonify({'ok': True})

@app.route('/ferias/feriado/adicionar', methods=['POST'])
@login_required
def feriado_add():
    data = request.get_json() or {}
    try: dt = datetime.strptime(data['data'], '%Y-%m-%d').date()
    except: return jsonify({'ok': False, 'error': 'Data inválida'})
    # Check if exists
    existing = FeriasFeriado.query.filter_by(data=dt).first()
    if existing:
        existing.nome = data.get('nome', existing.nome)
        existing.tipo = data.get('tipo', existing.tipo)
        db.session.commit()
        return jsonify({'ok': True, 'id': existing.id})
    f = FeriasFeriado(ano=dt.year, data=dt,
        nome=data.get('nome','Feriado'), tipo=data.get('tipo','nacional'))
    db.session.add(f); db.session.commit()
    return jsonify({'ok': True, 'id': f.id})

@app.route('/ferias/feriado/<int:fid>/eliminar', methods=['POST'])
@login_required
def feriado_eliminar(fid):
    f = FeriasFeriado.query.get_or_404(fid)
    db.session.delete(f); db.session.commit()
    return jsonify({'ok': True})

# ── USER STATUS ────────────────────────────────────────────────────────────────

@app.route('/api/user/heartbeat', methods=['POST'])
@login_required
def user_heartbeat():
    us = UserSession.query.filter_by(user_id=current_user.id).first()
    if us:
        us.last_seen = datetime.now()
    else:
        db.session.add(UserSession(user_id=current_user.id, last_seen=datetime.now()))
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/users/status')
@login_required
def api_users_status():
    now = datetime.now()
    sessions = {s.user_id: s for s in UserSession.query.all()}
    all_users = User.query.order_by(User.nome).all()
    result = []
    for u in all_users:
        s = sessions.get(u.id)
        if s and s.last_seen:
            diff = (now - s.last_seen).total_seconds()
            if diff < 300:    status = 'online'
            elif diff < 7200: status = 'recente'
            else:             status = 'offline'
            mins = int(diff // 60)
        else:
            status = 'offline'
            mins = 9999
        result.append({
            'id': u.id,
            'nome': u.nome,
            'status': status,
            'mins': mins,
        })
    result.sort(key=lambda x: (0 if x['status']=='online' else 1 if x['status']=='recente' else 2, x['mins']))
    return jsonify(result)


@app.route('/uploads/embarcacoes/<path:fname>')
@login_required
def embarcacao_foto_serve(fname):
    import mimetypes
    fdir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'embarcacoes')
    return send_from_directory(fdir, fname)


# ── MÓDULO FORNECEDORES ────────────────────────────────────────────────────────

@app.route('/fornecedores')
@login_required
def fornecedores():
    q     = request.args.get('q','').strip()
    marca = request.args.get('marca','').strip()
    query = FornecedorPHC.query
    if q:
        query = query.filter(
            FornecedorPHC.nome.ilike(f'%{q}%') |
            FornecedorPHC.nif.ilike(f'%{q}%') |
            FornecedorPHC.email.ilike(f'%{q}%') |
            FornecedorPHC.localidade.ilike(f'%{q}%') |
            FornecedorPHC.telefone.ilike(f'%{q}%')
        )
    if marca:
        query = query.filter(FornecedorPHC.marcas.ilike(f'%{marca}%'))
    items = query.order_by(FornecedorPHC.nome).limit(200).all()
    # Get all unique marcas for filter dropdown
    all_marcas = set()
    for f in FornecedorPHC.query.filter(FornecedorPHC.marcas != None, FornecedorPHC.marcas != '').all():
        for m in (f.marcas or '').split(','):
            m = m.strip()
            if m: all_marcas.add(m)
    return render_template('fornecedores.html', items=items, q=q,
        marca_filtro=marca, all_marcas=sorted(all_marcas))

@app.route('/fornecedores/resync', methods=['POST'])
@login_required
def fornecedores_resync():
    if not current_user.is_admin:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    try:
        cfg = ConfigPHC.query.first()
        if not cfg:
            return jsonify({'ok': False, 'error': 'PHC não configurado'})
        from phc_sync import sync_fornecedores
        ins, upd, errs = sync_fornecedores(cfg)
        db.session.commit()
        # Clear cache
        if hasattr(app, '_forn_cache'):
            del app._forn_cache
        return jsonify({'ok': True, 'inseridos': ins, 'atualizados': upd, 'erros': len(errs)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/fornecedores/<int:fid>/editar', methods=['POST'])
@login_required
def fornecedor_editar(fid):
    f = FornecedorPHC.query.get_or_404(fid)
    data = request.get_json() or {}
    # Only touch NAV fields (PHC fields are read-only from sync)
    if 'nav_telefone' in data: f.nav_telefone = data['nav_telefone'].strip()
    if 'nav_email'    in data: f.nav_email    = data['nav_email'].strip()
    if 'nav_morada'   in data: f.nav_morada   = data['nav_morada'].strip()
    if 'nav_notas'    in data: f.nav_notas    = data['nav_notas'].strip()
    db.session.commit()
    if hasattr(app, '_forn_cache'):
        del app._forn_cache
    return jsonify({'ok': True})


@app.route('/fornecedores/<int:fid>/marcas', methods=['POST'])
@login_required
def fornecedor_marcas(fid):
    f = FornecedorPHC.query.get_or_404(fid)
    data = request.get_json() or {}
    f.marcas = data.get('marcas', '').strip()
    db.session.commit()
    # Refresh in-memory cache
    if hasattr(app, '_forn_cache'):
        del app._forn_cache
    return jsonify({'ok': True})


# ── MÓDULO FICHAS TÉCNICAS ────────────────────────────────────────────────────

UPLOAD_FICHAS = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'fichas')
os.makedirs(UPLOAD_FICHAS, exist_ok=True)

CATEGORIAS_PADRAO = [
    'Filtros', 'Correias & Correntes', 'Fluidos & Lubrificantes',
    'Juntas & Vedantes', 'Rolamentos', 'Elétrico & Sensores',
    'Arrefecimento', 'Turbocompressor', 'Injeção', 'Outro'
]

@app.route('/fichas')
@login_required
def fichas():
    q = request.args.get('q','').strip()
    query = FichaTecnica.query
    if q:
        query = query.filter(
            FichaTecnica.grupo_designacao.ilike(f'%{q}%') |
            FichaTecnica.grupo_marca.ilike(f'%{q}%') |
            FichaTecnica.motor_modelo.ilike(f'%{q}%') |
            FichaTecnica.cliente_nome.ilike(f'%{q}%') |
            FichaTecnica.grupo_serie.ilike(f'%{q}%')
        )
    items = query.order_by(FichaTecnica.numero.desc()).all()
    return render_template('fichas.html', items=items, q=q)

@app.route('/fichas/nova', methods=['GET','POST'])
@login_required
def ficha_nova():
    if request.method == 'POST':
        last = db.session.query(db.func.max(FichaTecnica.numero)).scalar() or 0
        f = FichaTecnica(
            numero=last+1,
            grupo_designacao=request.form.get('grupo_designacao','').strip(),
            grupo_marca=request.form.get('grupo_marca','').strip(),
            grupo_modelo=request.form.get('grupo_modelo','').strip(),
            grupo_serie=request.form.get('grupo_serie','').strip(),
            grupo_ano=request.form.get('grupo_ano','').strip(),
            motor_marca=request.form.get('motor_marca','').strip(),
            motor_modelo=request.form.get('motor_modelo','').strip(),
            motor_serie=request.form.get('motor_serie','').strip(),
            motor_potencia=request.form.get('motor_potencia','').strip(),
            motor_cilindros=request.form.get('motor_cilindros','').strip(),
            cliente_nome=request.form.get('cliente_nome','').strip(),
            observacoes=request.form.get('observacoes','').strip(),
            criado_por=current_user.id,
            criado_em=datetime.now(), atualizado_em=datetime.now()
        )
        db.session.add(f); db.session.commit()
        flash(f'Ficha Técnica #{f.numero:04d} criada!', 'success')
        return redirect(url_for('ficha_detalhe', fid=f.id))
    return render_template('ficha_form.html', f=None, categorias=CATEGORIAS_PADRAO)

@app.route('/fichas/<int:fid>')
@login_required
def ficha_detalhe(fid):
    f = FichaTecnica.query.get_or_404(fid)
    return render_template('ficha_detalhe.html', f=f, categorias=CATEGORIAS_PADRAO)

@app.route('/fichas/<int:fid>/editar', methods=['GET','POST'])
@login_required
def ficha_editar(fid):
    f = FichaTecnica.query.get_or_404(fid)
    if request.method == 'POST':
        f.grupo_designacao = request.form.get('grupo_designacao','').strip()
        f.grupo_marca      = request.form.get('grupo_marca','').strip()
        f.grupo_modelo     = request.form.get('grupo_modelo','').strip()
        f.grupo_serie      = request.form.get('grupo_serie','').strip()
        f.grupo_ano        = request.form.get('grupo_ano','').strip()
        f.motor_marca      = request.form.get('motor_marca','').strip()
        f.motor_modelo     = request.form.get('motor_modelo','').strip()
        f.motor_serie      = request.form.get('motor_serie','').strip()
        f.motor_potencia   = request.form.get('motor_potencia','').strip()
        f.motor_cilindros  = request.form.get('motor_cilindros','').strip()
        f.cliente_nome     = request.form.get('cliente_nome','').strip()
        f.observacoes      = request.form.get('observacoes','').strip()
        f.atualizado_em    = datetime.now()
        db.session.commit()
        flash('Ficha actualizada.', 'success')
        return redirect(url_for('ficha_detalhe', fid=fid))
    return render_template('ficha_form.html', f=f, categorias=CATEGORIAS_PADRAO)

@app.route('/fichas/<int:fid>/eliminar', methods=['POST'])
@login_required
def ficha_eliminar(fid):
    f = FichaTecnica.query.get_or_404(fid)
    for doc in f.documentos:
        try: os.remove(os.path.join(UPLOAD_FICHAS, doc.nome_ficheiro))
        except: pass
    db.session.delete(f); db.session.commit()
    flash('Ficha eliminada.', 'success')
    return redirect(url_for('fichas'))

# ── Componentes ───────────────────────────────────────────────────────────────
@app.route('/fichas/<int:fid>/componentes/adicionar', methods=['POST'])
@login_required
def ficha_comp_adicionar(fid):
    FichaTecnica.query.get_or_404(fid)
    data = request.get_json() or {}
    c = FichaComponente(
        ficha_id=fid,
        categoria=data.get('categoria','Outro'),
        designacao=data.get('designacao','').strip(),
        part_number=data.get('part_number','').strip(),
        marca=data.get('marca','').strip(),
        referencia_equiv=data.get('referencia_equiv','').strip(),
        quantidade=data.get('quantidade','1'),
        unidade=data.get('unidade','un'),
        intervalo=data.get('intervalo','').strip(),
        notas=data.get('notas','').strip(),
        ordem=FichaComponente.query.filter_by(ficha_id=fid).count()
    )
    db.session.add(c); db.session.commit()
    return jsonify({'ok': True, 'id': c.id})

@app.route('/fichas/<int:fid>/componentes/<int:cid>/editar', methods=['POST'])
@login_required
def ficha_comp_editar(fid, cid):
    c = FichaComponente.query.filter_by(id=cid, ficha_id=fid).first_or_404()
    data = request.get_json() or {}
    for f in ['categoria','designacao','part_number','marca','referencia_equiv','quantidade','unidade','intervalo','notas']:
        if f in data: setattr(c, f, data[f])
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/fichas/<int:fid>/componentes/<int:cid>/eliminar', methods=['POST'])
@login_required
def ficha_comp_eliminar(fid, cid):
    c = FichaComponente.query.filter_by(id=cid, ficha_id=fid).first_or_404()
    db.session.delete(c); db.session.commit()
    return jsonify({'ok': True})

# ── Documentos / Fotos ────────────────────────────────────────────────────────
@app.route('/fichas/<int:fid>/documentos/upload', methods=['POST'])
@login_required
def ficha_doc_upload(fid):
    FichaTecnica.query.get_or_404(fid)
    file = request.files.get('ficheiro')
    if not file or not file.filename:
        return jsonify({'ok': False, 'error': 'Sem ficheiro'})
    import uuid, mimetypes
    ext = os.path.splitext(file.filename)[1].lower()
    nome_unico = f'ft{fid}_{uuid.uuid4().hex[:10]}{ext}'
    fpath = os.path.join(UPLOAD_FICHAS, nome_unico)
    file.save(fpath)
    mime = mimetypes.guess_type(file.filename)[0] or 'application/octet-stream'
    tipo = 'foto' if mime.startswith('image/') else 'documento'
    doc = FichaDocumento(
        ficha_id=fid, tipo=tipo,
        nome_original=file.filename, nome_ficheiro=nome_unico,
        descricao=request.form.get('descricao','').strip(),
        tamanho=os.path.getsize(fpath), mime=mime,
        criado_por=current_user.id, criado_em=datetime.now()
    )
    db.session.add(doc); db.session.commit()
    return jsonify({'ok': True, 'id': doc.id, 'nome': doc.nome_original,
        'tipo': doc.tipo, 'mime': doc.mime,
        'data': doc.criado_em.strftime('%d/%m/%Y %H:%M'),
        'uploader': current_user.nome})

@app.route('/fichas/<int:fid>/documentos/<int:did>/preview')
@login_required
def ficha_doc_preview(fid, did):
    doc = FichaDocumento.query.filter_by(id=did, ficha_id=fid).first_or_404()
    return send_from_directory(UPLOAD_FICHAS, doc.nome_ficheiro, as_attachment=False)

@app.route('/fichas/<int:fid>/documentos/<int:did>/download')
@login_required
def ficha_doc_download(fid, did):
    doc = FichaDocumento.query.filter_by(id=did, ficha_id=fid).first_or_404()
    return send_from_directory(UPLOAD_FICHAS, doc.nome_ficheiro,
        as_attachment=True, download_name=doc.nome_original)

@app.route('/fichas/<int:fid>/documentos/<int:did>/apagar', methods=['POST'])
@login_required
def ficha_doc_apagar(fid, did):
    doc = FichaDocumento.query.filter_by(id=did, ficha_id=fid).first_or_404()
    try: os.remove(os.path.join(UPLOAD_FICHAS, doc.nome_ficheiro))
    except: pass
    db.session.delete(doc); db.session.commit()
    return jsonify({'ok': True})


MENUS_DISPONIVEIS = [
    ('dashboard',          '🏠 Dashboard'),
    ('pedidos',            '📋 Pedidos de Compra'),
    ('reposicao',          '📦 Reposição'),
    ('inventario',         '📊 Inventário'),
    ('clientes',           '👥 Clientes'),
    ('tecnico',            '🔧 Técnico'),
    ('biblioteca_modelos', '📚 Biblioteca PDF'),
    ('funcionarios',       '👤 Funcionários'),
    ('salarios',           '💶 Salários'),
    ('ferias',             '📅 Mapa Férias / Faltas'),
    ('fornecedores',       '🏭 Fornecedores'),
    ('fichas',             '📋 Fichas Técnicas'),
    ('assistencias',       '🔧 Assistências'),
    ('entradas',           '📥 Entradas'),
    ('partilha',           '📁 Partilha'),
    ('empresa',            '🏢 Empresa'),
    ('eq_industriais',     '⚙️ Equipamentos Industriais'),
    ('agenda',             '📅 Agenda Digital'),
    ('conectividade',      '🌐 Conectividade'),
    ('roadmap',            '🗺️ Roadmap'),
    ('changelog',          '📝 Changelog'),
    ('admin_config',       '⚙️ Configurações'),
    ('admin_utilizadores', '👥 Utilizadores'),
]

# ── MÓDULO ASSISTÊNCIAS ───────────────────────────────────────────────────────

UPLOAD_ASSIST = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'assistencias')
os.makedirs(UPLOAD_ASSIST, exist_ok=True)

ASSIST_STATUS = [
    ('rececionado',      '📥 Rececionado'),
    ('em_execucao',      '🔧 Em Execução'),
    ('obra_concluida',   '✅ Obra Concluída'),
    ('comunicado',       '📨 Comunicado'),
    ('faturado_fechado', '🧾 Faturado – Fechado'),
]
ASSIST_STATUS_DICT = dict(ASSIST_STATUS)
ASSIST_COLORS = {
    'rececionado':      '#3b6ef0',
    'em_execucao':      '#f59e0b',
    'obra_concluida':   '#22c55e',
    'comunicado':       '#06b6d4',
    'faturado_fechado': '#6b7280',
}

def _recalc_assistencia_dates(a):
    """Recalculate all date fields from history entries (most recent per status_novo)."""
    hist = AssistenciaHistorico.query.filter_by(assist_id=a.id).order_by(
        AssistenciaHistorico.criado_em.asc()).all()
    # Get most recent real date per status
    dates = {}
    for h in hist:
        if h.data_real:
            dates[h.status_novo] = h.data_real
    # Apply
    if 'rececionado'      in dates: a.data_rececionado    = dates['rececionado']
    if 'em_execucao'      in dates: a.data_em_execucao    = dates['em_execucao']
    if 'obra_concluida'   in dates: a.data_obra_concluida = dates['obra_concluida']
    if 'comunicado'       in dates: a.data_comunicado     = dates['comunicado']
    if 'faturado_fechado' in dates: a.data_faturado       = dates['faturado_fechado']
    # Most recent status = current status
    if hist:
        a.status = hist[-1].status_novo
    _assist_calc_durations(a)


def _assist_calc_durations(a):
    """Recalculate duration fields. Negative values stored as-is but flagged."""
    if a.data_rececionado and a.data_obra_concluida:
        a.dias_recepcao_conclusao = (a.data_obra_concluida - a.data_rececionado).days
    if a.data_obra_concluida and a.data_comunicado:
        a.dias_conclusao_comunicado = (a.data_comunicado - a.data_obra_concluida).days
    if a.data_obra_concluida and a.data_faturado:
        a.dias_conclusao_faturado = (a.data_faturado - a.data_obra_concluida).days

@app.route('/assistencias')
@login_required
def assistencias():
    status_f = request.args.get('status','')
    q = Assistencia.query
    if status_f:
        q = q.filter_by(status=status_f)
    from datetime import date as _dt
    _hoje = _dt.today()
    items = q.order_by(Assistencia.numero.desc()).all()
    for _a in items:
        # Dias alerta (obra concluida/comunicado sem faturar > 5 dias)
        if _a.status in ('obra_concluida','comunicado'):
            _ref = _a.data_obra_concluida if _a.status == 'obra_concluida' else _a.data_comunicado
            if _ref:
                _d = (_hoje - _ref).days
                _a._dias_alerta = _d if _d > 5 else None
            else: _a._dias_alerta = None
        else: _a._dias_alerta = None
    return render_template('assistencias.html', items=items,
        status_list=ASSIST_STATUS, status_filtro=status_f,
        colors=ASSIST_COLORS)

@app.route('/assistencias/estatisticas')
@login_required
def assistencias_stats():
    from datetime import date as _dt
    from sqlalchemy import func, extract
    _hoje = _dt.today()

    # All closed assistencias with duration data
    todas = Assistencia.query.all()
    fechadas = [a for a in todas if a.status == 'faturado_fechado']
    abertas = [a for a in todas if a.status != 'faturado_fechado']

    # KPIs
    total = len(todas)
    n_fechadas = len(fechadas)
    n_abertas = len(abertas)

    duracoes_obra  = [a.dias_recepcao_conclusao  for a in fechadas if a.dias_recepcao_conclusao  is not None]
    duracoes_com   = [a.dias_conclusao_comunicado for a in todas    if a.dias_conclusao_comunicado is not None]
    duracoes_fat   = [a.dias_conclusao_faturado   for a in fechadas if a.dias_conclusao_faturado   is not None]

    avg_obra = round(sum(duracoes_obra)/len(duracoes_obra), 1) if duracoes_obra else 0
    avg_com  = round(sum(duracoes_com) /len(duracoes_com),  1) if duracoes_com  else 0
    avg_fat  = round(sum(duracoes_fat) /len(duracoes_fat),  1) if duracoes_fat  else 0
    max_obra = max(duracoes_obra) if duracoes_obra else 0
    max_fat  = max(duracoes_fat)  if duracoes_fat  else 0
    min_obra = min(duracoes_obra) if duracoes_obra else 0
    max_com  = max(duracoes_com)  if duracoes_com  else 0

    # Por ano/mês — agrupado por ano de data_rececionado
    from collections import defaultdict
    by_year  = defaultdict(lambda: {'total':0,'fechadas':0,'sum_obra':0,'n_obra':0,'sum_com':0,'n_com':0,'sum_fat':0,'n_fat':0})
    by_month = defaultdict(lambda: {'total':0,'fechadas':0,'sum_obra':0,'n_obra':0,'sum_com':0,'n_com':0})
    by_status = defaultdict(int)

    for a in todas:
        by_status[a.status] += 1
        yr = a.data_rececionado.year if a.data_rececionado else (a.criado_em.year if a.criado_em else _hoje.year)
        mo = f"{yr}-{a.data_rececionado.month:02d}" if a.data_rececionado else str(yr)
        by_year[yr]['total'] += 1
        by_month[mo]['total'] += 1
        if a.status == 'faturado_fechado':
            by_year[yr]['fechadas'] += 1
            by_month[mo]['fechadas'] += 1
        if a.dias_recepcao_conclusao is not None:
            by_year[yr]['sum_obra'] += a.dias_recepcao_conclusao
            by_year[yr]['n_obra']   += 1
            by_month[mo]['sum_obra'] += a.dias_recepcao_conclusao
            by_month[mo]['n_obra']   += 1
        if a.dias_conclusao_comunicado is not None:
            by_year[yr]['sum_com']  += a.dias_conclusao_comunicado
            by_year[yr]['n_com']    += 1
            by_month[mo]['sum_com'] += a.dias_conclusao_comunicado
            by_month[mo]['n_com']   += 1
        if a.dias_conclusao_faturado is not None:
            by_year[yr]['sum_fat'] += a.dias_conclusao_faturado
            by_year[yr]['n_fat']   += 1

    # Compute averages
    years_data = []
    for yr in sorted(by_year.keys()):
        d = by_year[yr]
        years_data.append({
            'year': yr,
            'total': d['total'],
            'fechadas': d['fechadas'],
            'avg_obra': round(d['sum_obra']/d['n_obra'],1) if d['n_obra'] else None,
            'avg_com':  round(d['sum_com'] /d['n_com'], 1) if d['n_com']  else None,
            'avg_fat':  round(d['sum_fat'] /d['n_fat'], 1) if d['n_fat']  else None,
        })

    months_data = []
    for mo in sorted(by_month.keys())[-18:]:  # last 18 months
        d = by_month[mo]
        months_data.append({
            'month': mo,
            'total': d['total'],
            'fechadas': d['fechadas'],
            'avg_obra': round(d['sum_obra']/d['n_obra'],1) if d['n_obra'] else None,
            'avg_com':  round(d['sum_com'] /d['n_com'], 1) if d['n_com']  else None,
        })

    # Top requerentes
    req_count = defaultdict(int)
    for a in todas:
        req_count[a.requerente_nome] += 1
    top_req = sorted(req_count.items(), key=lambda x: -x[1])[:10]

    # Abertas por estado com dias em aberto
    abertas_detail = []
    for a in abertas:
        ref = a.data_rececionado or (a.criado_em.date() if a.criado_em else _hoje)
        dias = (_hoje - ref).days
        abertas_detail.append({'status': a.status, 'dias': dias})

    import json
    return render_template('assistencias_stats.html',
        total=total, n_fechadas=n_fechadas, n_abertas=n_abertas,
        avg_obra=avg_obra, avg_com=avg_com, avg_fat=avg_fat,
        max_obra=max_obra, max_fat=max_fat, min_obra=min_obra, max_com=max_com,
        years_data=json.dumps(years_data),
        months_data=json.dumps(months_data),
        by_status=json.dumps(dict(by_status)),
        top_req=json.dumps(top_req),
        status_dict=json.dumps(ASSIST_STATUS_DICT),
        colors_json=json.dumps(ASSIST_COLORS),
    )

@app.route('/assistencias/nova', methods=['GET','POST'])
@login_required
def assistencia_nova():
    if request.method == 'POST':
        last = db.session.query(db.func.max(Assistencia.numero)).scalar() or 0
        from datetime import date
        data_rec_str = request.form.get('data_rececionado','')
        data_rec = datetime.strptime(data_rec_str, '%Y-%m-%d').date() if data_rec_str else date.today()
        a = Assistencia(
            numero=last+1,
            requerente_nome=request.form.get('requerente_nome','').strip(),
            requerente_nif=request.form.get('requerente_nif','').strip(),
            num_requisicao=request.form.get('num_requisicao','').strip(),
            local_obra=request.form.get('local_obra','').strip(),
            observacoes=request.form.get('observacoes','').strip(),
            status='rececionado',
            data_rececionado=data_rec,
            criado_por=current_user.id,
            criado_em=datetime.now(), atualizado_em=datetime.now()
        )
        db.session.add(a); db.session.flush()
        db.session.add(AssistenciaHistorico(
            assist_id=a.id, status_ant=None, status_novo='rececionado',
            data_real=data_rec, user_nome=current_user.nome,
            notas='Criado', criado_em=datetime.now()
        ))
        db.session.commit()
        flash(f'Assistência #{a.numero:04d} criada!', 'success')
        return redirect(url_for('assistencia_detalhe', aid=a.id))
    from datetime import date
    return render_template('assistencia_form.html', a=None,
        hoje=date.today().strftime('%Y-%m-%d'))

@app.route('/assistencias/<int:aid>')
@login_required
def assistencia_detalhe(aid):
    a = Assistencia.query.get_or_404(aid)
    docs = AssistenciaDocumento.query.filter_by(assist_id=aid).order_by(
        AssistenciaDocumento.criado_em.desc()).all()
    return render_template('assistencia_detalhe.html', a=a, docs=docs,
        status_list=ASSIST_STATUS, status_dict=ASSIST_STATUS_DICT,
        colors=ASSIST_COLORS)

@app.route('/assistencias/<int:aid>/editar', methods=['GET','POST'])
@login_required
def assistencia_editar(aid):
    a = Assistencia.query.get_or_404(aid)
    if request.method == 'POST':
        a.requerente_nome = request.form.get('requerente_nome','').strip()
        a.requerente_nif  = request.form.get('requerente_nif','').strip()
        a.num_requisicao  = request.form.get('num_requisicao','').strip()
        a.local_obra      = request.form.get('local_obra','').strip()
        a.observacoes     = request.form.get('observacoes','').strip()
        a.atualizado_em   = datetime.now()
        db.session.commit()
        flash('Actualizado.', 'success')
        return redirect(url_for('assistencia_detalhe', aid=aid))
    from datetime import date
    return render_template('assistencia_form.html', a=a,
        hoje=date.today().strftime('%Y-%m-%d'))

@app.route('/assistencias/<int:aid>/status', methods=['POST'])
@login_required
def assistencia_status(aid):
    a = Assistencia.query.get_or_404(aid)
    data = request.get_json() or {}
    novo = data.get('status')
    notas = data.get('notas','')
    data_real_str = data.get('data_real','')
    if novo not in ASSIST_STATUS_DICT:
        return jsonify({'ok': False, 'error': 'Status inválido'})
    from datetime import date as dt
    data_real = datetime.strptime(data_real_str, '%Y-%m-%d').date() if data_real_str else dt.today()
    ant = a.status
    a.status = novo
    a.atualizado_em = datetime.now()
    # Save date per status
    if novo == 'rececionado':      a.data_rececionado    = data_real
    elif novo == 'em_execucao':    a.data_em_execucao    = data_real
    elif novo == 'obra_concluida': a.data_obra_concluida = data_real
    elif novo == 'comunicado':     a.data_comunicado     = data_real
    elif novo == 'faturado_fechado': a.data_faturado     = data_real
    _assist_calc_durations(a)
    db.session.add(AssistenciaHistorico(
        assist_id=aid, status_ant=ant, status_novo=novo,
        data_real=data_real, user_nome=current_user.nome,
        notas=notas, criado_em=datetime.now()
    ))
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/assistencias/<int:aid>/historico/<int:hid>/editar', methods=['POST'])
@login_required
def assistencia_hist_editar(aid, hid):
    h = AssistenciaHistorico.query.filter_by(id=hid, assist_id=aid).first_or_404()
    data = request.get_json() or {}
    # Update data_real
    if 'data_real' in data:
        try: h.data_real = datetime.strptime(data['data_real'], '%Y-%m-%d').date() if data['data_real'] else None
        except: pass
    if 'notas' in data:
        h.notas = data['notas']
    # Update status_ant and status_novo if provided
    if 'status_ant' in data:
        h.status_ant = data['status_ant'] or None
    if 'status_novo' in data and data['status_novo']:
        h.status_novo = data['status_novo']
    # Recalculate all assistencia dates from history (most recent per status)
    a = Assistencia.query.get(aid)
    if a:
        _recalc_assistencia_dates(a)
        a.atualizado_em = datetime.now()
    db.session.commit()
    # Check for negative durations and return warnings
    warnings = []
    if a:
        if a.dias_recepcao_conclusao is not None and a.dias_recepcao_conclusao < 0:
            warnings.append(f'⚠️ Data de Obra Concluída ({a.data_obra_concluida}) é anterior à Receção ({a.data_rececionado}). Verifique as datas.')
        if a.dias_conclusao_comunicado is not None and a.dias_conclusao_comunicado < 0:
            warnings.append(f'⚠️ Data de Comunicado ({a.data_comunicado}) é anterior à Obra Concluída ({a.data_obra_concluida}). Verifique as datas.')
        if a.dias_conclusao_faturado is not None and a.dias_conclusao_faturado < 0:
            warnings.append(f'⚠️ Data de Faturação ({a.data_faturado}) é anterior à Obra Concluída ({a.data_obra_concluida}). Verifique as datas.')
    return jsonify({'ok': True, 'warnings': warnings})

@app.route('/assistencias/<int:aid>/historico/<int:hid>/eliminar', methods=['POST'])
@login_required
def assistencia_hist_eliminar(aid, hid):
    h = AssistenciaHistorico.query.filter_by(id=hid, assist_id=aid).first_or_404()
    if not current_user.is_admin and h.user_nome != current_user.nome:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    db.session.delete(h)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/assistencias/<int:aid>/eliminar', methods=['POST'])
@login_required
def assistencia_eliminar(aid):
    a = Assistencia.query.get_or_404(aid)
    if not current_user.is_admin and a.criado_por != current_user.id:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    for doc in a.documentos:
        try: os.remove(os.path.join(UPLOAD_ASSIST, doc.nome_ficheiro))
        except: pass
    db.session.delete(a); db.session.commit()
    return jsonify({'ok': True})

@app.route('/assistencias/<int:aid>/documentos/upload', methods=['POST'])
@login_required
def assistencia_doc_upload(aid):
    Assistencia.query.get_or_404(aid)
    f = request.files.get('ficheiro')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'Sem ficheiro'})
    import uuid, mimetypes
    ext = os.path.splitext(f.filename)[1].lower()
    nome_unico = f'a{aid}_{uuid.uuid4().hex[:10]}{ext}'
    fpath = os.path.join(UPLOAD_ASSIST, nome_unico)
    f.save(fpath)
    mime = mimetypes.guess_type(f.filename)[0] or 'application/octet-stream'
    tipo = request.form.get('tipo','documento')  # 'email' or 'documento'
    doc = AssistenciaDocumento(
        assist_id=aid, tipo=tipo,
        nome_original=f.filename, nome_ficheiro=nome_unico,
        descricao=request.form.get('descricao','').strip(),
        tamanho=os.path.getsize(fpath), mime=mime,
        criado_por=current_user.id, criado_em=datetime.now()
    )
    db.session.add(doc); db.session.commit()
    return jsonify({'ok': True, 'id': doc.id, 'nome': doc.nome_original,
        'tipo': doc.tipo, 'data': doc.criado_em.strftime('%d/%m/%Y %H:%M'),
        'uploader': current_user.nome, 'descricao': doc.descricao})

@app.route('/assistencias/<int:aid>/documentos/<int:did>/email-view')
@login_required
def assistencia_email_view(aid, did):
    """Render .eml/.msg as readable HTML without download."""
    doc = AssistenciaDocumento.query.filter_by(id=did, assist_id=aid).first_or_404()
    fpath = os.path.join(UPLOAD_ASSIST, doc.nome_ficheiro)
    ext = os.path.splitext(doc.nome_original)[1].lower()
    try:
        if ext == '.eml':
            import email as _email
            with open(fpath, 'rb') as f:
                msg = _email.message_from_bytes(f.read())
            subject = msg.get('Subject', '(sem assunto)')
            from_addr = msg.get('From', '')
            to_addr = msg.get('To', '')
            date_h = msg.get('Date', '')
            body = ''
            if msg.is_multipart():
                for part in msg.walk():
                    ct = part.get_content_type()
                    if ct == 'text/html':
                        body = part.get_payload(decode=True).decode('utf-8','replace'); break
                    elif ct == 'text/plain' and not body:
                        body = '<pre style="white-space:pre-wrap;font-family:Arial">' + part.get_payload(decode=True).decode('utf-8','replace') + '</pre>'
            else:
                payload = msg.get_payload(decode=True)
                body = '<pre style="white-space:pre-wrap;font-family:Arial">' + (payload.decode('utf-8','replace') if payload else '') + '</pre>'
            html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><style>
body{{font-family:Arial,sans-serif;max-width:900px;margin:20px auto;padding:0 16px;color:#333}}
.header{{background:#f5f5f5;border-radius:8px;padding:14px 16px;margin-bottom:16px;font-size:13px}}
.label{{font-weight:700;color:#555;min-width:60px;display:inline-block}}
</style></head><body>
<div class="header">
<div><span class="label">Assunto:</span> {subject}</div>
<div><span class="label">De:</span> {from_addr}</div>
<div><span class="label">Para:</span> {to_addr}</div>
<div><span class="label">Data:</span> {date_h}</div>
</div>
<div>{body}</div>
</body></html>"""
            from flask import make_response
            resp = make_response(html)
            resp.headers['Content-Type'] = 'text/html; charset=utf-8'
            return resp
        else:
            # For .msg or other, fallback to plain text
            with open(fpath, 'rb') as f:
                content = f.read().decode('utf-8','replace')
            html = f'<!DOCTYPE html><html><head><meta charset="UTF-8"></head><body><pre style="white-space:pre-wrap;font-family:monospace;font-size:12px">{content}</pre></body></html>'
            from flask import make_response
            resp = make_response(html)
            resp.headers['Content-Type'] = 'text/html; charset=utf-8'
            return resp
    except Exception as ex:
        return f'<html><body><p>Erro ao abrir email: {ex}</p></body></html>', 500


@app.route('/assistencias/<int:aid>/documentos/<int:did>/preview')
@login_required
def assistencia_doc_preview(aid, did):
    doc = AssistenciaDocumento.query.filter_by(id=did, assist_id=aid).first_or_404()
    return send_from_directory(UPLOAD_ASSIST, doc.nome_ficheiro, as_attachment=False)

@app.route('/assistencias/<int:aid>/documentos/<int:did>/download')
@login_required
def assistencia_doc_download(aid, did):
    doc = AssistenciaDocumento.query.filter_by(id=did, assist_id=aid).first_or_404()
    return send_from_directory(UPLOAD_ASSIST, doc.nome_ficheiro,
        as_attachment=True, download_name=doc.nome_original)

@app.route('/assistencias/<int:aid>/documentos/<int:did>/apagar', methods=['POST'])
@login_required
def assistencia_doc_apagar(aid, did):
    doc = AssistenciaDocumento.query.filter_by(id=did, assist_id=aid).first_or_404()
    if not current_user.is_admin and doc.criado_por != current_user.id:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    try: os.remove(os.path.join(UPLOAD_ASSIST, doc.nome_ficheiro))
    except: pass
    db.session.delete(doc); db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/assistencias/fornecedores')
@login_required
def api_assist_fornecedores():
    q = request.args.get('q','').strip()
    if len(q) < 2: return jsonify([])
    # Try PHC fornecedores
    try:
        cfg_phc = ConfigPHC.query.first()
        if cfg_phc and cfg_phc.server:
            import pyodbc
            conn = pyodbc.connect(
                f"DRIVER={{SQL Server}};SERVER={cfg_phc.server};DATABASE={cfg_phc.database};"
                f"UID={cfg_phc.username};PWD={cfg_phc.password}", timeout=3)
            cur = conn.cursor()
            cur.execute("SELECT TOP 15 nome, ncont FROM fornecedor WHERE nome LIKE ? AND ativo=1 ORDER BY nome", f'%{q}%')
            rows = [{'nome': r[0], 'nif': r[1] or ''} for r in cur.fetchall()]
            conn.close()
            if rows: return jsonify(rows)
    except: pass
    # Try local FornecedorPHC table
    try:
        forn = FornecedorPHC.query.filter(FornecedorPHC.nome.ilike(f'%{q}%')).limit(15).all()
        if forn:
            return jsonify([{'nome': f.nome, 'nif': getattr(f,'ncont','') or ''} for f in forn])
    except: pass
    # Fallback: previous requerentes in assistencias
    prev = db.session.query(Assistencia.requerente_nome, Assistencia.requerente_nif)\
        .filter(Assistencia.requerente_nome.ilike(f'%{q}%'))\
        .distinct().limit(10).all()
    return jsonify([{'nome': r[0], 'nif': r[1] or ''} for r in prev])


# ── MÓDULO ENTRADAS ───────────────────────────────────────────────────────────

ENTRADAS_STATUS = [
    ('rececionado',            '📥 Rececionado'),
    ('pre_orcamento',          '📝 Pré-Orçamento'),
    ('orcamentado',            '📋 Orçamentado'),
    ('material_pedido',        '📦 Material Pedido'),
    ('material_stock',         '🏪 Material em Stock'),
    ('em_reparacao',           '🔧 Em Reparação'),
    ('reparacao_concluida',    '✅ Reparação Concluída'),
    ('faturado',               '🧾 Faturado'),
    ('orcamentado_estadia',    '⏳ Orçamentado – Em Estadia'),
    ('concluido_estadia',      '🏁 Concluído – Em Estadia'),
    ('concluido_fechado',      '✔️ Concluído – Fechado'),
]
ENTRADAS_STATUS_DICT = dict(ENTRADAS_STATUS)

# States that trigger day counting and auto-escalation
# ── ENTRADAS: DOCUMENTOS ─────────────────────────────────────────────────────
UPLOAD_ENTRADAS = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'entradas')
os.makedirs(UPLOAD_ENTRADAS, exist_ok=True)

@app.route('/entradas/<int:eid>/documentos/upload', methods=['POST'])
@login_required
def entrada_doc_upload(eid):
    EntradaEquipamento.query.get_or_404(eid)
    f = request.files.get('ficheiro')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'Sem ficheiro'})
    import uuid, mimetypes
    ext = os.path.splitext(f.filename)[1].lower()
    nome_unico = f'e{eid}_{uuid.uuid4().hex[:10]}{ext}'
    fpath = os.path.join(UPLOAD_ENTRADAS, nome_unico)
    f.save(fpath)
    mime = mimetypes.guess_type(f.filename)[0] or 'application/octet-stream'
    doc = EntradaDocumento(
        entrada_id=eid, nome_original=f.filename,
        nome_ficheiro=nome_unico,
        descricao=request.form.get('descricao','').strip(),
        tamanho=os.path.getsize(fpath), mime=mime,
        criado_por=current_user.id, criado_em=datetime.now()
    )
    db.session.add(doc)
    db.session.commit()
    return jsonify({'ok': True, 'id': doc.id,
        'nome': doc.nome_original, 'mime': doc.mime,
        'tamanho': doc.tamanho,
        'uploader': current_user.nome,
        'data': doc.criado_em.strftime('%d/%m/%Y %H:%M')})

@app.route('/entradas/<int:eid>/documentos/<int:did>/download')
@login_required
def entrada_doc_download(eid, did):
    doc = EntradaDocumento.query.filter_by(id=did, entrada_id=eid).first_or_404()
    return send_from_directory(UPLOAD_ENTRADAS, doc.nome_ficheiro,
        as_attachment=True, download_name=doc.nome_original)

@app.route('/entradas/<int:eid>/documentos/<int:did>/preview')
@login_required
def entrada_doc_preview(eid, did):
    doc = EntradaDocumento.query.filter_by(id=did, entrada_id=eid).first_or_404()
    return send_from_directory(UPLOAD_ENTRADAS, doc.nome_ficheiro, as_attachment=False)

@app.route('/entradas/<int:eid>/documentos/<int:did>/apagar', methods=['POST'])
@login_required
def entrada_doc_apagar(eid, did):
    doc = EntradaDocumento.query.filter_by(id=did, entrada_id=eid).first_or_404()
    if not current_user.is_admin and doc.criado_por != current_user.id:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    try: os.remove(os.path.join(UPLOAD_ENTRADAS, doc.nome_ficheiro))
    except: pass
    db.session.delete(doc)
    db.session.commit()
    return jsonify({'ok': True})


def _recalc_entrada_dates(e):
    """Recalculate all date fields and durations for an entrada from its history."""
    hist = EntradaHistorico.query.filter_by(entrada_id=e.id).order_by(
        EntradaHistorico.criado_em.asc()).all()
    dates = {}
    for h in hist:
        if h.data_real:
            dates[h.status_novo] = h.data_real
    if 'pre_orcamento'       in dates: e.data_pre_orcamento      = dates['pre_orcamento']
    if 'orcamentado'          in dates: e.data_orcamento           = dates['orcamentado']
    # Also preserve manually-set data_orcamento if history has none
    # (set via inline date editor — don't overwrite with None)
    if 'material_pedido'      in dates: e.data_material_pedido     = dates['material_pedido']
    if 'material_stock'       in dates: e.data_material_stock      = dates['material_stock']
    if 'em_reparacao'         in dates: e.data_em_reparacao        = dates['em_reparacao']
    if 'reparacao_concluida'  in dates: e.data_reparacao_concluida = dates['reparacao_concluida']
    if 'faturado'             in dates: e.data_faturado            = dates['faturado']
    if 'concluido_fechado'    in dates:
        e.data_fecho = dates['concluido_fechado']
        if e.data_rececao:
            e.dias_total = (e.data_fecho - e.data_rececao).days
    # Also use inline-edited data_orcamento if history has none
    # (data_orcamento may be set directly via the inline field)
    # Durations
    if e.data_rececao and e.data_orcamento:
        e.dias_rec_orcamento = (e.data_orcamento - e.data_rececao).days
    if e.data_rececao and e.data_faturado:
        e.dias_rec_faturado = (e.data_faturado - e.data_rececao).days
    if e.data_orcamento and e.data_em_reparacao:
        e.dias_orc_reparacao = (e.data_em_reparacao - e.data_orcamento).days
    if e.data_material_pedido and e.data_material_stock:
        e.dias_mat_stock = (e.data_material_stock - e.data_material_pedido).days
    if e.data_material_stock and e.data_em_reparacao:
        e.dias_stock_reparacao = (e.data_em_reparacao - e.data_material_stock).days
    if e.data_material_pedido and e.data_em_reparacao:
        e.dias_mat_reparacao = (e.data_em_reparacao - e.data_material_pedido).days
    if e.data_em_reparacao and e.data_reparacao_concluida:
        e.dias_reparacao_concluida = (e.data_reparacao_concluida - e.data_em_reparacao).days
    if e.data_em_reparacao and e.data_faturado:
        e.dias_reparacao_fat = (e.data_faturado - e.data_em_reparacao).days


ESTADIA_RULES = {
    'orcamentado':  {'dias': 10, 'escalate': 'orcamentado_estadia'},
    'faturado':     {'dias': 5,  'escalate': 'concluido_estadia'},
}

def _dias_uteis(data_inicio, data_fim):
    """Count working days between two dates (Mon-Fri)."""
    from datetime import date, timedelta
    count = 0
    cur = data_inicio
    while cur < data_fim:
        cur += timedelta(days=1)
        if cur.weekday() < 5:  # Mon=0 ... Fri=4
            count += 1
    return count

def _verificar_estadias():
    """Check all entries and escalate status if working days exceeded."""
    from datetime import date
    changed = 0
    for e in EntradaEquipamento.query.all():
        rule = ESTADIA_RULES.get(e.status)
        if not rule:
            continue
        # Use real date if set, otherwise registration date
        ref_date = e.data_status_real or (e.data_status.date() if e.data_status else None)
        if not ref_date:
            continue
        dias = _dias_uteis(ref_date, date.today())
        if dias >= rule['dias']:
            status_ant = e.status
            e.status = rule['escalate']
            e.data_status = datetime.now()
            hist = EntradaHistorico(
                entrada_id=e.id, status_ant=status_ant,
                status_novo=e.status, user_nome='Sistema',
                notas=f'Automático: {dias} dias úteis atingidos',
                criado_em=datetime.now()
            )
            db.session.add(hist)
            changed += 1
    if changed:
        db.session.commit()
    return changed

@app.route('/entradas/estatisticas')
@login_required
def entradas_stats():
    from datetime import date as _dt
    from collections import defaultdict
    _hoje = _dt.today()
    todas = EntradaEquipamento.query.all()
    fechadas = [e for e in todas if e.status == 'concluido_fechado']
    abertas  = [e for e in todas if e.status != 'concluido_fechado']

    def avg(lst): return round(sum(lst)/len(lst),1) if lst else 0
    def safe(lst): return [x for x in lst if x is not None]

    g = lambda e, col: getattr(e, col, None)
    d_total   = safe([g(e,'dias_total')               for e in fechadas])
    d_rorc    = safe([g(e,'dias_rec_orcamento')        for e in todas])
    d_rf      = safe([g(e,'dias_rec_faturado')         for e in todas])
    d_orcrp   = safe([g(e,'dias_orc_reparacao')        for e in todas])
    d_cmrp    = safe([g(e,'dias_mat_reparacao')         for e in todas])
    d_cms     = safe([g(e,'dias_mat_stock')             for e in todas])
    d_srp     = safe([g(e,'dias_stock_reparacao')       for e in todas])
    d_rpc     = safe([g(e,'dias_reparacao_concluida')   for e in todas])
    d_rpf     = safe([g(e,'dias_reparacao_fat')         for e in todas])

    # By status count
    by_status = defaultdict(int)
    for e in todas: by_status[e.status] += 1

    # By year
    by_year = defaultdict(lambda:{'total':0,'fechadas':0,'sum_total':0,'n_total':0,
                                   'sum_rf':0,'n_rf':0,'sum_cmrp':0,'n_cmrp':0,
                                   'sum_cms':0,'n_cms':0,'sum_srp':0,'n_srp':0,
                                   'sum_rpc':0,'n_rpc':0,'sum_rpf':0,'n_rpf':0})
    for e in todas:
        yr = e.data_rececao.year if e.data_rececao else (e.criado_em.year if e.criado_em else _hoje.year)
        by_year[yr]['total'] += 1
        if e.status=='concluido_fechado': by_year[yr]['fechadas'] += 1
        if e.dias_total        is not None: by_year[yr]['sum_total']+=e.dias_total;  by_year[yr]['n_total']+=1
        if e.dias_rec_faturado is not None: by_year[yr]['sum_rf']  +=e.dias_rec_faturado; by_year[yr]['n_rf']+=1
        for _col,_key in [('dias_rec_faturado','rf'),('dias_mat_reparacao','cmrp'),
                             ('dias_mat_stock','cms'),('dias_stock_reparacao','srp'),
                             ('dias_reparacao_concluida','rpc'),('dias_reparacao_fat','rpf')]:
            _v = g(e, _col)
            if _v is not None: by_year[yr]['sum_'+_key]+=_v; by_year[yr]['n_'+_key]+=1

    years_data = []
    for yr in sorted(by_year.keys()):
        d = by_year[yr]
        years_data.append({'year':yr,'total':d['total'],'fechadas':d['fechadas'],
            'avg_total': round(d['sum_total']/d['n_total'],1) if d['n_total'] else None,
            'avg_rf':    round(d['sum_rf']   /d['n_rf'],   1) if d['n_rf']    else None,
            'avg_cmrp':  round(d['sum_cmrp'] /d['n_cmrp'], 1) if d['n_cmrp']  else None,
            'avg_cms':   round(d['sum_cms']  /d['n_cms'],  1) if d['n_cms']   else None,
            'avg_srp':   round(d['sum_srp']  /d['n_srp'],  1) if d['n_srp']   else None,
            'avg_rpc':   round(d['sum_rpc']  /d['n_rpc'],  1) if d['n_rpc']   else None,
            'avg_rpf':   round(d['sum_rpf']  /d['n_rpf'],  1) if d['n_rpf']   else None,
        })

    # Top clientes
    from collections import Counter
    top_cli = Counter(e.cliente_nome for e in todas).most_common(10)

    # By month (last 18)
    by_month = defaultdict(lambda:{'total':0,'sum_rf':0,'n_rf':0})
    for e in todas:
        if e.data_rececao:
            mo = f"{e.data_rececao.year}-{e.data_rececao.month:02d}"
            by_month[mo]['total'] += 1
            if e.dias_rec_faturado is not None:
                by_month[mo]['sum_rf'] += e.dias_rec_faturado; by_month[mo]['n_rf'] += 1
    months_data = [{'month':mo,'total':d['total'],
        'avg_rf':round(d['sum_rf']/d['n_rf'],1) if d['n_rf'] else None}
        for mo,d in sorted(by_month.items())[-18:]]

    import json
    return render_template('entradas_stats.html',
        total=len(todas), n_fechadas=len(fechadas), n_abertas=len(abertas),
        avg_total=avg(d_total), avg_rorc=avg(d_rorc), avg_rf=avg(d_rf),
        avg_cmrp=avg(d_cmrp),
        avg_cms=avg(d_cms), avg_srp=avg(d_srp), avg_rpc=avg(d_rpc), avg_rpf=avg(d_rpf),
        max_total=max(d_total) if d_total else 0,
        min_total=min(d_total) if d_total else 0,
        years_data=json.dumps(years_data),
        months_data=json.dumps(months_data),
        by_status=json.dumps(dict(by_status)),
        top_cli=json.dumps(top_cli),
        status_dict=json.dumps({k:v for k,v in [
            ('rececionado','📥 Rececionado'),('pre_orcamento','📝 Pré-Orç.'),
            ('orcamentado','📋 Orçamentado'),('material_pedido','📦 Mat. Pedido'),
            ('material_stock','🏪 Mat. Stock'),('em_reparacao','🔧 Em Reparação'),
            ('reparacao_concluida','✅ Rep. Concluída'),
            ('faturado','🧾 Faturado'),('orcamentado_estadia','⏳ Orç.-Estadia'),
            ('concluido_estadia','🏁 Conc.-Estadia'),('concluido_fechado','✔️ Fechado')]}),
        colors_json=json.dumps({'rececionado':'#3b6ef0','pre_orcamento':'#94a3b8',
            'orcamentado':'#f59e0b','material_pedido':'#6366f1','material_stock':'#0ea5e9',
            'em_reparacao':'#22c55e','reparacao_concluida':'#10b981',
            'faturado':'#06b6d4','orcamentado_estadia':'#ef4444',
            'concluido_estadia':'#a855f7','concluido_fechado':'#6b7280'}),
    )

@app.route('/entradas')
@login_required
def entradas():
    _verificar_estadias()
    status_f = request.args.get('status', '')
    q = EntradaEquipamento.query
    if status_f:
        q = q.filter_by(status=status_f)
    entradas_list = q.order_by(EntradaEquipamento.numero.desc()).all()
    from datetime import date
    _hoje = date.today()
    # Ensure dias are calculated from date fields (in case DB column is NULL)
    for e in entradas_list:
        def _calc(d1, d2):
            if d1 and d2 and d2 >= d1: return (d2 - d1).days
            return None
        if e.dias_rec_orcamento is None and e.data_rececao and e.data_orcamento:
            e.dias_rec_orcamento = _calc(e.data_rececao, e.data_orcamento)
        if e.dias_rec_faturado is None and e.data_rececao and e.data_faturado:
            e.dias_rec_faturado = _calc(e.data_rececao, e.data_faturado)
        if e.dias_orc_reparacao is None and e.data_orcamento and e.data_em_reparacao:
            e.dias_orc_reparacao = _calc(e.data_orcamento, e.data_em_reparacao)
        if e.dias_mat_reparacao is None and e.data_material_pedido and e.data_em_reparacao:
            e.dias_mat_reparacao = _calc(e.data_material_pedido, e.data_em_reparacao)
        if e.dias_reparacao_fat is None and e.data_em_reparacao and e.data_faturado:
            e.dias_reparacao_fat = _calc(e.data_em_reparacao, e.data_faturado)
        if e.dias_total is None and e.data_rececao and e.data_fecho:
            e.dias_total = _calc(e.data_rececao, e.data_fecho)
    # Compute dias for each entry (days in current status)
    for e in entradas_list:
        try:
            ref = getattr(e,'data_status_real',None) or (e.data_status.date() if e.data_status else None)
            if ref:
                e._dias_contagem = (_hoje - ref).days
                rule = ESTADIA_RULES.get(e.status)
                e._dias_limite = rule['dias'] if rule else 9999
            else:
                e._dias_contagem = None
                e._dias_limite = 9999
        except:
            e._dias_contagem = None
            e._dias_limite = 9999
    em_estadia = sum(1 for e in entradas_list if e.status in ('orcamentado_estadia','concluido_estadia'))
    return render_template('entradas.html',
        entradas=entradas_list, status_list=ENTRADAS_STATUS,
        status_filtro=status_f, em_estadia=em_estadia)

@app.route('/entradas/nova', methods=['GET','POST'])
@login_required
def entrada_nova():
    if request.method == 'POST':
        from datetime import date
        # Auto increment number
        last = db.session.query(db.func.max(EntradaEquipamento.numero)).scalar() or 0
        e = EntradaEquipamento(
            numero=last+1,
            data_rececao=datetime.strptime(request.form.get('data_rececao', datetime.now().strftime('%Y-%m-%d')), '%Y-%m-%d').date(),
            cliente_nome=request.form.get('cliente_nome','').strip(),
            marca=request.form.get('marca','').strip(),
            marca_grupo=request.form.get('marca_grupo','').strip(),
            modelo=request.form.get('modelo','').strip(),
            modelo_grupo=request.form.get('modelo_grupo','').strip(),
            num_serie=request.form.get('num_serie','').strip(),
            num_serie_grupo=request.form.get('num_serie_grupo','').strip(),
            observacoes=request.form.get('observacoes','').strip(),
            status='rececionado',
            data_status=datetime.now(),
            criado_por=current_user.id,
            criado_em=datetime.now(),
            atualizado_em=datetime.now(),
        )
        db.session.add(e)
        db.session.flush()
        db.session.add(EntradaHistorico(
            entrada_id=e.id, status_ant=None, status_novo='rececionado',
            user_id=current_user.id, user_nome=current_user.nome,
            notas='Entrada criada', criado_em=datetime.now()
        ))
        db.session.commit()
        flash(f'Entrada #{e.numero} criada com sucesso!', 'success')
        return redirect(url_for('entrada_detalhe', eid=e.id))
    from datetime import date
    return render_template('entrada_form.html', entrada=None,
        hoje=date.today().strftime('%Y-%m-%d'))

@app.route('/entradas/<int:eid>')
@login_required
def entrada_detalhe(eid):
    e = EntradaEquipamento.query.get_or_404(eid)
    from datetime import date
    rule = ESTADIA_RULES.get(e.status)
    if rule:
        ref = e.data_status_real or (e.data_status.date() if e.data_status else None)
        dias_contagem = _dias_uteis(ref, date.today()) if ref else None
    else:
        dias_contagem = None
    docs = EntradaDocumento.query.filter_by(entrada_id=eid).order_by(EntradaDocumento.criado_em.desc()).all()
    return render_template('entrada_detalhe.html', e=e,
        status_list=ENTRADAS_STATUS, dias_contagem=dias_contagem,
        status_dict=ENTRADAS_STATUS_DICT, docs=docs)

@app.route('/entradas/<int:eid>/editar', methods=['GET','POST'])
@login_required
def entrada_editar(eid):
    e = EntradaEquipamento.query.get_or_404(eid)
    if request.method == 'POST':
        e.data_rececao  = datetime.strptime(request.form['data_rececao'], '%Y-%m-%d').date()
        e.cliente_nome  = request.form.get('cliente_nome','').strip()
        e.marca         = request.form.get('marca','').strip()
        e.modelo        = request.form.get('modelo','').strip()
        e.num_serie     = request.form.get('num_serie','').strip()
        e.observacoes   = request.form.get('observacoes','').strip()
        e.atualizado_em = datetime.now()
        db.session.commit()
        flash('Entrada actualizada.', 'success')
        return redirect(url_for('entrada_detalhe', eid=eid))
    return render_template('entrada_form.html', entrada=e,
        hoje=e.data_rececao.strftime('%Y-%m-%d'))

@app.route('/entradas/<int:eid>/status', methods=['POST'])
@login_required
def entrada_status(eid):
    e = EntradaEquipamento.query.get_or_404(eid)
    data = request.get_json() or {}
    novo = data.get('status')
    notas = data.get('notas','')
    data_real_str = data.get('data_real','')
    if novo not in ENTRADAS_STATUS_DICT:
        return jsonify({'ok': False, 'error': 'Status inválido'})
    # Parse real date
    from datetime import date as date_type
    data_real = None
    if data_real_str:
        try:
            data_real = datetime.strptime(data_real_str, '%Y-%m-%d').date()
        except: pass
    if not data_real:
        data_real = date_type.today()
    ant = e.status
    e.status = novo
    e.data_status = datetime.now()
    e.data_status_real = data_real
    e.atualizado_em = datetime.now()
    # Store date per status
    if novo == 'pre_orcamento':          e.data_pre_orcamento      = data_real
    elif novo == 'orcamentado':          e.data_orcamento           = data_real
    elif novo == 'material_pedido':      e.data_material_pedido     = data_real
    elif novo == 'material_stock':       e.data_material_stock      = data_real
    elif novo == 'em_reparacao':         e.data_em_reparacao        = data_real
    elif novo == 'reparacao_concluida':  e.data_reparacao_concluida = data_real
    elif novo == 'faturado':             e.data_faturado            = data_real
    elif novo == 'concluido_fechado':
        e.data_fecho = data_real
        if e.data_rececao:
            e.dias_total = (data_real - e.data_rececao).days
    db.session.add(EntradaHistorico(
        entrada_id=eid, status_ant=ant, status_novo=novo,
        user_id=current_user.id, user_nome=current_user.nome,
        notas=notas, data_real=data_real, criado_em=datetime.now()
    ))
    _recalc_entrada_dates(e)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/entradas/<int:eid>/pdf')
@login_required
def entrada_pdf(eid):
    e = EntradaEquipamento.query.get_or_404(eid)
    cfg = ConfigGeral.query.first()
    empresa_nome = cfg.empresa_nome if cfg else 'NavTech'
    html = render_template('entrada_pdf.html', e=e, cfg=cfg, empresa_nome=empresa_nome,
        upload_url=request.host_url.rstrip('/') + '/uploads',
        status_dict=ENTRADAS_STATUS_DICT)
    # Try wkhtmltopdf
    for wk_path in [r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe',
                    r'C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe']:
        if os.path.exists(wk_path):
            try:
                import pdfkit
                pdf_bytes = pdfkit.from_string(html, False,
                    options={'page-size':'A4','margin-top':'12mm','margin-bottom':'12mm',
                             'margin-left':'12mm','margin-right':'12mm','encoding':'UTF-8','quiet':''},
                    configuration=pdfkit.configuration(wkhtmltopdf=wk_path))
                from flask import Response
                return Response(pdf_bytes, mimetype='application/pdf',
                    headers={'Content-Disposition': f'inline; filename="Entrada_{e.numero}.pdf"'})
            except: pass
    # Fallback: browser print
    from flask import make_response
    resp = make_response(html)
    resp.headers['Content-Type'] = 'text/html; charset=utf-8'
    return resp

@app.route('/entradas/<int:eid>/eliminar', methods=['POST'])
@login_required
def entrada_eliminar(eid):
    e = EntradaEquipamento.query.get_or_404(eid)
    if not current_user.is_admin and e.criado_por != current_user.id:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    # Delete history first
    EntradaHistorico.query.filter_by(entrada_id=eid).delete()
    db.session.delete(e)
    db.session.commit()
    flash(f'Entrada #{e.numero:04d} eliminada.', 'success')
    return jsonify({'ok': True})


@app.route('/entradas/<int:eid>/historico/<int:hid>/editar', methods=['POST'])
@login_required
def entrada_hist_editar(eid, hid):
    h = EntradaHistorico.query.filter_by(id=hid, entrada_id=eid).first_or_404()
    data = request.get_json() or {}
    if 'data_real' in data:
        try: h.data_real = datetime.strptime(data['data_real'], '%Y-%m-%d').date() if data['data_real'] else None
        except: pass
    if 'notas'       in data: h.notas      = data['notas']
    if 'status_ant'  in data: h.status_ant = data['status_ant'] or None
    if 'status_novo' in data and data['status_novo']: h.status_novo = data['status_novo']
    e = EntradaEquipamento.query.get(eid)
    if e:
        _recalc_entrada_dates(e)
        e.atualizado_em = datetime.now()
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/entradas/<int:eid>/historico/<int:hid>/eliminar', methods=['POST'])
@login_required
def entrada_hist_eliminar(eid, hid):
    h = EntradaHistorico.query.filter_by(id=hid, entrada_id=eid).first_or_404()
    db.session.delete(h)
    e = EntradaEquipamento.query.get(eid)
    if e:
        _recalc_entrada_dates(e)
        e.atualizado_em = datetime.now()
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/entradas/<int:eid>/data-orcamento', methods=['POST'])
@login_required
def entrada_data_orcamento(eid):
    e = EntradaEquipamento.query.get_or_404(eid)
    data = request.get_json() or {}
    ds = data.get('data_orcamento','')
    try:
        e.data_orcamento = datetime.strptime(ds, '%Y-%m-%d').date() if ds else None
        # If status is orcamentado, also update data_status_real for day counting
        if e.status == 'orcamentado' and e.data_orcamento:
            e.data_status_real = e.data_orcamento
        e.atualizado_em = datetime.now()
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as ex:
        return jsonify({'ok': False, 'error': str(ex)})


@app.route('/api/entradas/estadia-count')
@login_required
def api_entradas_estadia_count():
    _verificar_estadias()
    count = EntradaEquipamento.query.filter(
        EntradaEquipamento.status.in_(['orcamentado_estadia','concluido_estadia'])
    ).count()
    return jsonify({'count': count})


# ── MÓDULO PARTILHA ───────────────────────────────────────────────────────────

UPLOAD_PARTILHA = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'partilha')
os.makedirs(UPLOAD_PARTILHA, exist_ok=True)
PARTILHA_MAX_GB = 5
PARTILHA_EXTENSOES_PREVIEW = {'.pdf','.png','.jpg','.jpeg','.gif','.webp','.svg','.txt','.mp4','.mp3'}

def _tamanho_total_partilha():
    from sqlalchemy import func
    r = db.session.query(func.sum(PartilhaFicheiro.tamanho)).scalar()
    return r or 0

@app.route('/partilha')
@login_required
def partilha():
    if current_user.is_admin:
        ficheiros = PartilhaFicheiro.query.order_by(PartilhaFicheiro.criado_em.desc()).all()
    else:
        ficheiros = PartilhaFicheiro.query.filter(
            db.or_(
                PartilhaFicheiro.visibilidade == 'todos',
                PartilhaFicheiro.criado_por == current_user.id,
                PartilhaFicheiro.destinatario_id == current_user.id
            )
        ).order_by(PartilhaFicheiro.criado_em.desc()).all()
    utilizadores = User.query.order_by(User.nome).all()
    total_bytes = _tamanho_total_partilha()
    total_gb = total_bytes / (1024**3)
    return render_template('partilha.html',
        ficheiros=ficheiros, utilizadores=utilizadores,
        total_gb=total_gb, max_gb=PARTILHA_MAX_GB)

@app.route('/partilha/upload', methods=['POST'])
@login_required
def partilha_upload():
    if _tamanho_total_partilha() >= PARTILHA_MAX_GB * 1024**3:
        return jsonify({'ok': False, 'error': 'Limite de 5GB atingido'})
    f = request.files.get('ficheiro')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'Sem ficheiro'})
    import uuid, mimetypes
    ext = os.path.splitext(f.filename)[1].lower()
    nome_unico = str(uuid.uuid4()) + ext
    fpath = os.path.join(UPLOAD_PARTILHA, nome_unico)
    f.save(fpath)
    tamanho = os.path.getsize(fpath)
    mime = mimetypes.guess_type(f.filename)[0] or 'application/octet-stream'
    visib = request.form.get('visibilidade', 'todos')
    dest_id = request.form.get('destinatario_id') or None
    if dest_id: dest_id = int(dest_id)
    pf = PartilhaFicheiro(
        nome=nome_unico, nome_original=f.filename,
        descricao=request.form.get('descricao','').strip(),
        tamanho=tamanho, mime=mime, path=fpath,
        criado_por=current_user.id,
        visibilidade=visib, destinatario_id=dest_id,
        criado_em=datetime.now()
    )
    db.session.add(pf); db.session.commit()
    return jsonify({'ok': True, 'id': pf.id})

@app.route('/partilha/<int:fid>/download')
@login_required
def partilha_download(fid):
    pf = PartilhaFicheiro.query.get_or_404(fid)
    # Check access
    if not current_user.is_admin:
        if pf.visibilidade != 'todos' and pf.criado_por != current_user.id and pf.destinatario_id != current_user.id:
            return 'Sem acesso', 403
    return send_from_directory(UPLOAD_PARTILHA, pf.nome, as_attachment=True, download_name=pf.nome_original)

@app.route('/partilha/<int:fid>/preview')
@login_required
def partilha_preview(fid):
    pf = PartilhaFicheiro.query.get_or_404(fid)
    if not current_user.is_admin:
        if pf.visibilidade != 'todos' and pf.criado_por != current_user.id and pf.destinatario_id != current_user.id:
            return 'Sem acesso', 403
    ext = os.path.splitext(pf.nome_original)[1].lower()
    if ext not in PARTILHA_EXTENSOES_PREVIEW:
        return redirect(url_for('partilha_download', fid=fid))
    return send_from_directory(UPLOAD_PARTILHA, pf.nome, as_attachment=False)

@app.route('/partilha/<int:fid>/apagar', methods=['POST'])
@login_required
def partilha_apagar(fid):
    pf = PartilhaFicheiro.query.get_or_404(fid)
    if not current_user.is_admin and pf.criado_por != current_user.id:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    try: os.remove(pf.path)
    except: pass
    db.session.delete(pf); db.session.commit()
    return jsonify({'ok': True})


# ── MÓDULO FUNCIONÁRIOS ───────────────────────────────────────────────────────

UPLOAD_RH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'rh')
os.makedirs(UPLOAD_RH, exist_ok=True)

class Funcionario(db.Model):
    __tablename__ = 'funcionario'
    id                  = db.Column(db.Integer, primary_key=True)
    numero              = db.Column(db.String(20), unique=True, nullable=False)
    nome                = db.Column(db.String(200), nullable=False)
    categoria           = db.Column(db.String(100))
    ativo               = db.Column(db.Boolean, default=True)
    agenda_ativo        = db.Column(db.Boolean, default=False)  # aparece na Agenda Digital
    # Identificação
    data_admissao       = db.Column(db.Date)
    data_nascimento     = db.Column(db.Date)
    num_cc              = db.Column(db.String(50))
    cc_validade         = db.Column(db.Date, nullable=True)
    nif                 = db.Column(db.String(20))
    num_passaporte      = db.Column(db.String(50))
    passaporte_validade = db.Column(db.Date, nullable=True)
    filiacao_pai        = db.Column(db.String(200))
    filiacao_mae        = db.Column(db.String(200))
    situacao_militar    = db.Column(db.String(100))
    estado_civil        = db.Column(db.String(50))
    conjuge             = db.Column(db.String(200))
    titulares_rendimento= db.Column(db.Integer, default=1)
    num_dependentes     = db.Column(db.Integer, default=0)
    natural_freguesia   = db.Column(db.String(100))
    natural_concelho    = db.Column(db.String(100))
    socio_numero        = db.Column(db.String(50))
    sindicato           = db.Column(db.String(100))
    num_seg_social      = db.Column(db.String(30))
    carta_conducao      = db.Column(db.Boolean, default=False)
    # Contacto
    morada              = db.Column(db.Text)
    telemovel           = db.Column(db.String(30))
    email               = db.Column(db.String(200))
    contacto_emergencia = db.Column(db.String(200))
    nome_emergencia     = db.Column(db.String(200))
    # Financeiro
    iban                = db.Column(db.String(50))
    seguro_companhia    = db.Column(db.String(200))
    seguro_apolice      = db.Column(db.String(100))
    # Misc
    agregado_familiar   = db.Column(db.Integer, default=0)
    notas               = db.Column(db.Text)
    obs                 = db.Column(db.Text)
    criado_em           = db.Column(db.DateTime, default=datetime.now)
    atualizado_em       = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    documentos          = db.relationship('FuncionarioDocumento', backref='funcionario', lazy=True, cascade='all, delete-orphan')
    formacoes           = db.relationship('FuncionarioFormacao', backref='funcionario', lazy=True, cascade='all, delete-orphan')
    situacoes_prof      = db.relationship('FuncionarioSituacaoProf', backref='funcionario', lazy=True, cascade='all, delete-orphan', order_by='FuncionarioSituacaoProf.data.desc()')
    faltas              = db.relationship('FuncionarioFalta', backref='funcionario', lazy=True, cascade='all, delete-orphan')

    @property
    def idade(self):
        if not self.data_nascimento: return None
        today = datetime.now().date()
        d = self.data_nascimento
        return today.year - d.year - ((today.month, today.day) < (d.month, d.day))

class FuncionarioSituacaoProf(db.Model):
    __tablename__ = 'funcionario_situacao_prof'
    id              = db.Column(db.Integer, primary_key=True)
    funcionario_id  = db.Column(db.Integer, db.ForeignKey('funcionario.id'), nullable=False)
    data            = db.Column(db.Date, nullable=False)
    categoria_prof  = db.Column(db.String(200))
    vencimento      = db.Column(db.Numeric(10,2))
    refeicao        = db.Column(db.Numeric(10,2))
    premios_outros  = db.Column(db.Numeric(10,2))
    notas           = db.Column(db.Text)
    criado_em       = db.Column(db.DateTime, default=datetime.now)

class FuncionarioFalta(db.Model):
    __tablename__ = 'funcionario_falta'
    id              = db.Column(db.Integer, primary_key=True)
    funcionario_id  = db.Column(db.Integer, db.ForeignKey('funcionario.id'), nullable=False)
    ano             = db.Column(db.Integer, nullable=False)
    mes             = db.Column(db.Integer, nullable=False)  # 1-12
    dias_falta      = db.Column(db.Numeric(5,1), default=0)
    horas_falta     = db.Column(db.Numeric(5,1), default=0)
    tipo            = db.Column(db.String(50), default='injustificada')  # justificada/injustificada/baixa/ferias
    notas           = db.Column(db.Text)
    criado_em       = db.Column(db.DateTime, default=datetime.now)

class FuncionarioDocumento(db.Model):
    __tablename__ = 'funcionario_documento'
    id              = db.Column(db.Integer, primary_key=True)
    funcionario_id  = db.Column(db.Integer, db.ForeignKey('funcionario.id'), nullable=False)
    tipo            = db.Column(db.String(100), nullable=False)  # cc, passaporte, morada, higiene, aptidao, outro
    titulo          = db.Column(db.String(300), nullable=False)
    pdf_filename    = db.Column(db.String(500))
    pdf_path        = db.Column(db.String(1000))
    data_validade   = db.Column(db.Date, nullable=True)
    criado_em       = db.Column(db.DateTime, default=datetime.now)

class FuncionarioFormacao(db.Model):
    __tablename__ = 'funcionario_formacao'
    id              = db.Column(db.Integer, primary_key=True)
    funcionario_id  = db.Column(db.Integer, db.ForeignKey('funcionario.id'), nullable=False)
    titulo          = db.Column(db.String(300), nullable=False)
    entidade        = db.Column(db.String(200))
    data_inicio     = db.Column(db.Date)
    data_fim        = db.Column(db.Date)
    horas           = db.Column(db.Integer)
    pdf_filename    = db.Column(db.String(500))
    pdf_path        = db.Column(db.String(1000))
    criado_em       = db.Column(db.DateTime, default=datetime.now)


# ── MÓDULO SALÁRIOS ───────────────────────────────────────────────────────────

UPLOAD_SALARIOS = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'salarios')
os.makedirs(UPLOAD_SALARIOS, exist_ok=True)

class TabelaIRS(db.Model):
    __tablename__ = 'tabela_irs'
    id          = db.Column(db.Integer, primary_key=True)
    ano         = db.Column(db.Integer, nullable=False)
    descricao   = db.Column(db.String(200))
    pdf_filename= db.Column(db.String(500))
    pdf_path    = db.Column(db.String(1000))
    dados_json  = db.Column(db.Text, default='{}')  # parsed IRS brackets
    criado_em   = db.Column(db.DateTime, default=datetime.now)

class DocContabilistico(db.Model):
    __tablename__ = 'doc_contabilistico'
    id          = db.Column(db.Integer, primary_key=True)
    titulo      = db.Column(db.String(300), nullable=False)
    tipo        = db.Column(db.String(100))  # tabela_irs, outro
    ano         = db.Column(db.Integer)
    pdf_filename= db.Column(db.String(500))
    pdf_path    = db.Column(db.String(1000))
    criado_em   = db.Column(db.DateTime, default=datetime.now)

class ReciboSalario(db.Model):
    __tablename__ = 'recibo_salario'
    id              = db.Column(db.Integer, primary_key=True)
    funcionario_id  = db.Column(db.Integer, db.ForeignKey('funcionario.id'), nullable=False)
    ano             = db.Column(db.Integer, nullable=False)
    mes             = db.Column(db.Integer, nullable=False)   # 1-12, 13=ferias, 14=natal, 15+=extra
    mes_label       = db.Column(db.String(50))                # "Janeiro", "Subsídio Férias", etc.
    estado          = db.Column(db.String(20), default='rascunho')  # rascunho/processado/pago
    # Valores base
    vencimento_base     = db.Column(db.Numeric(10,2), default=0)
    vencimento_base_rht = db.Column(db.Numeric(10,4), default=0)   # F9
    vencimento_base_g   = db.Column(db.Numeric(10,2), default=0)   # G9 BASE
    subsidio_refeicao   = db.Column(db.Numeric(10,2), default=0)
    sub_refeicao_dias   = db.Column(db.Numeric(5,1), default=0)    # D14
    sub_refeicao_vdia   = db.Column(db.Numeric(8,4), default=0)    # F14
    horas_extra         = db.Column(db.Numeric(10,2), default=0)   # E13 horas
    horas_extra_rht     = db.Column(db.Numeric(10,4), default=0)   # F13 RHT
    premios             = db.Column(db.Numeric(10,2), default=0)   # H10
    outros_abonos       = db.Column(db.Numeric(10,2), default=0)
    faltas_dias         = db.Column(db.Numeric(5,1), default=0)    # D11
    faltas_horas        = db.Column(db.Numeric(5,1), default=0)    # E12
    # Deduções
    irs_retencao        = db.Column(db.Numeric(10,2), default=0)   # H19
    irs_taxa            = db.Column(db.Numeric(8,6), default=0)    # B19
    irs_parcela_abater  = db.Column(db.Numeric(10,2), default=0)  # C19
    irs_taxa_efetiva    = db.Column(db.Numeric(8,6), default=0)   # D19
    irs_base            = db.Column(db.Numeric(10,2), default=0)   # E19
    seg_social_func     = db.Column(db.Numeric(10,2), default=0)   # H18 11%
    seg_social_taxa     = db.Column(db.Numeric(8,4), default=0)    # D18
    seg_social_base     = db.Column(db.Numeric(10,2), default=0)   # E18
    seg_social_emp      = db.Column(db.Numeric(10,2), default=0)   # 23.75%
    outros_descontos    = db.Column(db.Numeric(10,2), default=0)
    faltas_valor        = db.Column(db.Numeric(10,2), default=0)
    # Totais calculados
    total_abonos    = db.Column(db.Numeric(10,2), default=0)
    total_descontos = db.Column(db.Numeric(10,2), default=0)
    liquido         = db.Column(db.Numeric(10,2), default=0)
    # Extra info
    notas           = db.Column(db.Text)
    dados_json      = db.Column(db.Text, default='{}')  # extra fields
    pdf_filename    = db.Column(db.String(500))
    pdf_path        = db.Column(db.String(1000))
    criado_em       = db.Column(db.DateTime, default=datetime.now)
    atualizado_em   = db.Column(db.DateTime, default=datetime.now)

    funcionario     = db.relationship('Funcionario', backref='recibos')

MESES_LABELS = {
    1:'Janeiro', 2:'Fevereiro', 3:'Março', 4:'Abril',
    5:'Maio', 6:'Junho', 7:'Julho', 8:'Agosto',
    9:'Setembro', 10:'Outubro', 11:'Novembro', 12:'Dezembro',
    13:'Subsídio de Férias', 14:'Subsídio de Natal'
}


class LinhaPedidoHistorico(db.Model):
    __tablename__ = 'linha_pedido_historico'
    id          = db.Column(db.Integer, primary_key=True)
    linha_id    = db.Column(db.Integer, db.ForeignKey('linhas_pedido.id'), nullable=False)
    pedido_id   = db.Column(db.Integer, nullable=False)
    status_ant  = db.Column(db.String(30))
    status_novo = db.Column(db.String(30), nullable=False)
    user_id     = db.Column(db.Integer, db.ForeignKey('users.id'))
    user_nome   = db.Column(db.String(120))
    data        = db.Column(db.DateTime, default=datetime.now)
    notas       = db.Column(db.String(300))


class PartilhaFicheiro(db.Model):
    __tablename__ = 'partilha_ficheiros'
    id            = db.Column(db.Integer, primary_key=True)
    nome          = db.Column(db.String(255), nullable=False)
    nome_original = db.Column(db.String(255), nullable=False)
    descricao     = db.Column(db.String(500), default='')
    tamanho       = db.Column(db.BigInteger, default=0)
    mime          = db.Column(db.String(100), default='')
    path          = db.Column(db.String(500), nullable=False)
    criado_por    = db.Column(db.Integer, db.ForeignKey('users.id'))
    criado_em     = db.Column(db.DateTime, default=datetime.now)
    visibilidade  = db.Column(db.String(20), default='todos')  # 'todos' ou user id
    destinatario_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    uploader      = db.relationship('User', foreign_keys=[criado_por])
    destinatario  = db.relationship('User', foreign_keys=[destinatario_id])


class RegistoPendente(db.Model):
    __tablename__ = 'registo_pendente'
    id            = db.Column(db.Integer, primary_key=True)
    nome          = db.Column(db.String(120), nullable=False)
    username      = db.Column(db.String(80), nullable=False)
    email         = db.Column(db.String(200), nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    criado_em     = db.Column(db.DateTime, default=datetime.now)
    estado        = db.Column(db.String(20), default='pendente')  # pendente/aceite/recusado


class Perfil(db.Model):
    __tablename__ = 'perfil'
    id        = db.Column(db.Integer, primary_key=True)
    nome      = db.Column(db.String(100), unique=True, nullable=False)
    descricao = db.Column(db.String(300), default='')
    menus     = db.Column(db.Text, default='')  # JSON list of allowed endpoints
    criado_em = db.Column(db.DateTime, default=datetime.now)

    def get_menus(self):
        try: return json.loads(self.menus) if self.menus else []
        except: return []

    def set_menus(self, lst):
        self.menus = json.dumps(lst)

    def pode_aceder(self, endpoint):
        return endpoint in self.get_menus()


class EquipamentoDocumento(db.Model):
    __tablename__ = 'equipamento_documento'
    id              = db.Column(db.Integer, primary_key=True)
    equipamento_id  = db.Column(db.Integer, db.ForeignKey('equipamento.id'), nullable=False)
    componente      = db.Column(db.String(100), nullable=False)  # 'motor', 'caixa', 'aux_1', etc
    titulo          = db.Column(db.String(300), nullable=False)
    pdf_filename    = db.Column(db.String(500))
    pdf_path        = db.Column(db.String(1000))
    notas           = db.Column(db.Text, default='')
    criado_em       = db.Column(db.DateTime, default=datetime.now)


class EquipamentoMotorAux(db.Model):
    __tablename__ = 'equipamento_motor_aux'
    id              = db.Column(db.Integer, primary_key=True)
    equipamento_id  = db.Column(db.Integer, db.ForeignKey('equipamento.id'), nullable=False)
    numero          = db.Column(db.Integer, default=1)   # Motor Aux 1, 2, 3...
    marca           = db.Column(db.String(100))
    modelo          = db.Column(db.String(200))
    serial_number   = db.Column(db.String(100))
    potencia        = db.Column(db.String(50))
    rpm             = db.Column(db.String(50))
    ano             = db.Column(db.String(20))
    notas           = db.Column(db.Text, default='')
    ordem           = db.Column(db.Integer, default=0)


class ModeloPDF(db.Model):
    """Central library of PDFs shared by component model (caixa TM170, motor D13K, etc.)"""
    __tablename__ = 'modelo_pdf'
    id              = db.Column(db.Integer, primary_key=True)
    tipo_componente = db.Column(db.String(50), nullable=False)  # 'caixa', 'motor', 'aux', 'factory_code'
    modelo_codigo   = db.Column(db.String(100), nullable=False, index=True)  # e.g. 'TM170', '1152'
    titulo          = db.Column(db.String(300), nullable=False)
    pdf_filename    = db.Column(db.String(500))
    pdf_path        = db.Column(db.String(1000))
    thumb_path      = db.Column(db.String(1000))
    criado_em       = db.Column(db.DateTime, default=datetime.now)


class CampoTecnicoModelo(db.Model):
    """Custom technical fields shared by component model (e.g. all MG5085 caixas)"""
    __tablename__ = 'campo_tecnico_modelo'
    id              = db.Column(db.Integer, primary_key=True)
    tipo_componente = db.Column(db.String(50), nullable=False)  # 'caixa', 'motor', 'aux'
    modelo_codigo   = db.Column(db.String(100), nullable=False, index=True)
    titulo          = db.Column(db.String(200), nullable=False)
    valor           = db.Column(db.Text, default='')
    ordem           = db.Column(db.Integer, default=0)
    pdf_filename    = db.Column(db.String(500))
    pdf_path        = db.Column(db.String(1000))
    criado_em       = db.Column(db.DateTime, default=datetime.now)


class FactoryCodePDF(db.Model):
    __tablename__ = 'factory_code_pdf'
    id           = db.Column(db.Integer, primary_key=True)
    factory_code = db.Column(db.String(50), nullable=False, index=True)
    titulo       = db.Column(db.String(300))
    pdf_filename = db.Column(db.String(500))
    pdf_path     = db.Column(db.String(1000))
    criado_em    = db.Column(db.DateTime, default=datetime.now)


class EquipamentoOpcao(db.Model):
    __tablename__ = 'equipamento_opcao'
    id              = db.Column(db.Integer, primary_key=True)
    equipamento_id  = db.Column(db.Integer, db.ForeignKey('equipamento.id'), nullable=False)
    option_name     = db.Column(db.String(200), nullable=False)
    ordered         = db.Column(db.String(200))
    factory         = db.Column(db.String(200))
    distributor     = db.Column(db.String(200))
    pdf_filename    = db.Column(db.String(500))
    pdf_path        = db.Column(db.String(1000))
    ordem           = db.Column(db.Integer, default=0)


class BacklogItem(db.Model):
    __tablename__ = 'backlog_item'
    id         = db.Column(db.Integer, primary_key=True)
    titulo     = db.Column(db.String(200), nullable=False)
    descricao  = db.Column(db.Text, default='')
    tipo       = db.Column(db.String(20), default='medium')  # bug/small/medium/large/epic
    estado     = db.Column(db.String(20), default='pending') # pending/in_progress/done
    prioridade = db.Column(db.Integer, default=10)
    notas      = db.Column(db.Text, default='')
    criado_em  = db.Column(db.DateTime, default=datetime.now)
    atualizado_em = db.Column(db.DateTime, default=datetime.now)


@app.route('/api/backlog', methods=['GET'])
@login_required
def api_backlog_list():
    items = BacklogItem.query.order_by(BacklogItem.prioridade, BacklogItem.id).all()
    return jsonify([{
        'id': i.id, 'titulo': i.titulo, 'descricao': i.descricao,
        'notas': i.notas, 'tipo': i.tipo, 'estado': i.estado, 'prioridade': i.prioridade
    } for i in items])

@app.route('/api/backlog', methods=['POST'])
@login_required
def api_backlog_criar():
    d = request.get_json() or {}
    item = BacklogItem(
        titulo=d.get('titulo','').strip() or 'Sem titulo',
        descricao=d.get('descricao','').strip(),
        notas=d.get('notas','').strip(),
        tipo=d.get('tipo','medium'),
        estado=d.get('estado','pending'),
        prioridade=int(d.get('prioridade',10)),
    )
    db.session.add(item)
    db.session.commit()
    return jsonify({'ok': True, 'id': item.id})

@app.route('/api/backlog/<int:bid>', methods=['PUT'])
@login_required
def api_backlog_atualizar(bid):
    item = BacklogItem.query.get_or_404(bid)
    d = request.get_json() or {}
    if 'titulo'     in d: item.titulo     = d['titulo']
    if 'descricao'  in d: item.descricao  = d['descricao']
    if 'notas'      in d: item.notas      = d['notas']
    if 'tipo'       in d: item.tipo       = d['tipo']
    if 'estado'     in d: item.estado     = d['estado']
    if 'prioridade' in d: item.prioridade = int(d['prioridade'])
    item.atualizado_em = datetime.now()
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/backlog/<int:bid>', methods=['DELETE'])
@login_required
def api_backlog_apagar(bid):
    item = BacklogItem.query.get_or_404(bid)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'ok': True})


def auto_registar_commit(commit_msg, versao=None):
    """Called after git pull to auto-update backlog and changelog."""
    try:
        import re
        from datetime import datetime as dt
        
        # Detect type from conventional commit prefix
        tipo = 'chore'
        if commit_msg.startswith('feat:'): tipo = 'feat'
        elif commit_msg.startswith('fix:'): tipo = 'fix'
        elif commit_msg.startswith('refactor:'): tipo = 'refactor'
        
        # Clean description
        desc = re.sub(r'^(feat|fix|chore|refactor|docs|style|test):\s*', '', commit_msg).strip()
        
        # Auto-version based on date + count
        if not versao:
            today = dt.now().strftime('%Y.%m.%d')
            existing = ChangelogEntry.query.filter(
                ChangelogEntry.versao.like(f'{today}%')
            ).count()
            versao = f'{today}.{existing+1}'
        
        # Add to changelog
        entry = ChangelogEntry(
            versao=versao,
            descricao=desc,
            tipo=tipo,
            commit_msg=commit_msg,
        )
        db.session.add(entry)
        
        # Try to match backlog items and mark done
        if tipo in ('feat', 'fix'):
            desc_lower = desc.lower()
            items = BacklogItem.query.filter(BacklogItem.estado != 'done').all()
            for item in items:
                title_lower = item.titulo.lower()
                # Simple keyword match - if 3+ words from title appear in commit
                words = [w for w in title_lower.split() if len(w) > 4]
                matches = sum(1 for w in words if w in desc_lower)
                if matches >= 2:
                    item.estado = 'done'
                    item.atualizado_em = dt.now()
                    app.logger.info(f"Auto-marked backlog item done: {item.titulo}")
        
        db.session.commit()
        app.logger.info(f"Auto-registered commit: {versao} - {desc}")
        return versao
    except Exception as e:
        app.logger.warning(f"auto_registar_commit error: {e}")
        return None


@app.route('/api/admin/registar-commit', methods=['POST'])
@login_required  
def api_registar_commit():
    """Called by update script after git pull."""
    d = request.get_json() or {}
    commit_msg = d.get('commit_msg', '').strip()
    versao = d.get('versao', '')
    if not commit_msg:
        return jsonify({'error': 'commit_msg required'}), 400
    with app.app_context():
        v = auto_registar_commit(commit_msg, versao or None)
    return jsonify({'ok': True, 'versao': v})


@app.route('/api/fornecedores')
@login_required
def api_fornecedores():
    q = request.args.get('q','').strip().lower()
    
    # 1. Use in-memory cache (populated once per server start)
    if not hasattr(app, '_forn_cache'):
        app._forn_cache = []
        try:
            rows = FornecedorPHC.query.order_by(FornecedorPHC.nome).limit(2000).all()
            app._forn_cache = [{'id': str(f.numero), 'no': str(f.numero), 'nome': f.nome, 'ncont': f.nif or ''} for f in rows]
        except: pass
    
    if app._forn_cache:
        if q:
            results = [f for f in app._forn_cache if q in f['nome'].lower()][:15]
        else:
            results = app._forn_cache[:15]
        if results:
            return jsonify(results)
    
    # 2. Fallback: previous fornecedores from pedido lines (no network)
    try:
        import json as _jf
        seen = set()
        results = []
        for l in LinhaPedido.query.filter(LinhaPedido.fornecedores_json.isnot(None)).limit(300).all():
            try:
                for f in _jf.loads(l.fornecedores_json or '[]'):
                    if isinstance(f, dict) and f.get('nome'):
                        nome = f['nome']
                        if (not q or q in nome.lower()) and nome not in seen:
                            seen.add(nome)
                            results.append({'id': f.get('id',''), 'no': f.get('id',''), 'nome': nome, 'ncont': ''})
            except: pass
        # Also from fornecedor_hab field
        for l in LinhaPedido.query.filter(LinhaPedido.fornecedor_hab.isnot(None)).filter(LinhaPedido.fornecedor_hab != '').limit(200).all():
            nome = l.fornecedor_hab.strip()
            if nome and (not q or q in nome.lower()) and nome not in seen:
                seen.add(nome)
                results.append({'id': '', 'no': '', 'nome': nome, 'ncont': ''})
        if results:
            return jsonify(sorted(results, key=lambda x: x['nome'])[:15])
    except: pass
    
    # 3. Last resort: PHC direct query with short timeout
    try:
        cfg_phc = ConfigPHC.query.first()
        if cfg_phc:
            from phc_sync import get_phc_connection
            conn = get_phc_connection(cfg_phc)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT TOP 15 no, nome FROM PHC_Uniao..fo WHERE nome LIKE ? AND nome IS NOT NULL ORDER BY nome",
                (f'%{q}%',)
            )
            rows = [{'id': r[0], 'no': r[0], 'nome': (r[1] or '').strip(), 'ncont': ''} for r in cursor.fetchall()]
            conn.close()
            # Add to cache
            if not hasattr(app, '_forn_cache'): app._forn_cache = []
            for r in rows:
                if r['nome'] and r not in app._forn_cache:
                    app._forn_cache.append(r)
            return jsonify(rows)
    except: pass
    
    return jsonify([])

# Endpoint to refresh the fornecedor cache
@app.route('/api/fornecedores/refresh', methods=['POST'])
@login_required
def api_fornecedores_refresh():
    if hasattr(app, '_forn_cache'):
        del app._forn_cache
    return jsonify({'ok': True})


@app.route('/admin/ia', methods=['GET', 'POST'])
@login_required
def admin_ia():
    if not current_user.is_admin:
        flash('Sem permissão.', 'error'); return redirect(url_for('dashboard'))

    cfg = ConfigIA.query.first()
    if not cfg:
        cfg = ConfigIA(); db.session.add(cfg); db.session.commit()

    from ai_provider import RECOMMENDED_MODELS, test_provider

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'save':
            cfg.provider     = request.form.get('provider', 'lmstudio')
            cfg.lm_host      = request.form.get('lm_host', 'localhost').strip()
            cfg.lm_port      = int(request.form.get('lm_port', 1234))
            cfg.lm_model     = request.form.get('lm_model', '').strip()
            key = request.form.get('claude_api_key', '').strip()
            if key: cfg.claude_api_key = key
            gkey = request.form.get('gemini_api_key', '').strip()
            if gkey: cfg.gemini_api_key = gkey
            cfg.gemini_model = request.form.get('gemini_model', 'gemini-1.5-flash').strip()
            db.session.commit()
            flash('Configuração de IA guardada.', 'success')

        elif action == 'test':
            ok, msg = test_provider(cfg)
            cfg.teste_ok = ok
            cfg.ultimo_teste = datetime.now()
            db.session.commit()
            flash(msg, 'success' if ok else 'error')

        elif action == 'test_pdf':
            # Quick test with dummy text
            from ai_provider import analyze_pdf
            sample = """Empresa: Test Fornecedor Lda  NIF: 123456789
Orçamento Nº: ORC-2024-001   Data: 15/01/2024
Artigo: Cilindro Hidráulico CH-50   Qtd: 2 un   Preço: 150,00 €   Total: 300,00 €
Subtotal: 300,00 €   IVA 23%: 69,00 €   TOTAL: 369,00 €"""
            dados, erro = analyze_pdf(cfg, sample, "teste.pdf")
            if dados:
                flash(f'✅ Extração OK — Empresa: {dados.get("empresa")} · Total: {dados.get("total")} €', 'success')
            else:
                flash(f'❌ Falha na extração: {erro}', 'error')

        return redirect(url_for('admin_ia'))

    return render_template('admin_ia.html', cfg=cfg,
                           modelos=RECOMMENDED_MODELS.get('lmstudio', []))

def ensure_sqlserver_running():
    """Start SQL Server Express if not running."""
    import subprocess, time, logging
    _log = logging.getLogger(__name__)
    SERVICE = 'MSSQL$SQLEXPRESS'
    try:
        r = subprocess.run(['sc', 'query', SERVICE],
                          capture_output=True, text=True, timeout=5)
        if 'RUNNING' in r.stdout:
            _log.info("✅ SQL Server já está a correr")
            return True
        if 'STOPPED' in r.stdout:
            _log.info("🔄 A iniciar SQL Server Express...")
            start = subprocess.run(['net', 'start', SERVICE],
                                  capture_output=True, text=True, timeout=30)
            if start.returncode == 0:
                time.sleep(3)
                _log.info("✅ SQL Server iniciado com sucesso")
                return True
            else:
                _log.warning(f"⚠️  Não foi possível iniciar SQL Server: {start.stdout}")
                return False
    except Exception as e:
        _log.warning(f"⚠️  Erro ao verificar SQL Server: {e}")
        return False


@app.route('/api/backlog/count')
@login_required
def api_backlog_count():
    try:
        pending = BacklogItem.query.filter(BacklogItem.estado != 'done').count()
        return jsonify({'pending': pending})
    except Exception:
        return jsonify({'pending': 0})


@app.route('/inventario')
@login_required
def inventario():
    return render_template('inventario.html')


@app.route('/api/inventario/kpis')
@login_required
def api_inventario_kpis():
    """Calculate inventory KPIs from PHC data."""
    try:
        cfg_phc = ConfigPHC.query.first()
        if not cfg_phc:
            return jsonify({'error': 'PHC nao configurado'}), 400
        from phc_sync import get_phc_connection
        conn = get_phc_connection(cfg_phc)
        cur = conn.cursor()

        # --- Base stock data ---
        cur.execute("""
            SELECT st.ref, st.design, ISNULL(st.stock,0) AS stock,
                   ISNULL(st.epcusto,0) AS custo, ISNULL(st.epcpond,0) AS custo_pond,
                   ISNULL(st.epv1,0) AS pvp, ISNULL(st.familia,'') AS familia,
                   ISNULL(st.unidade,'un') AS unidade
            FROM st
            WHERE ISNULL(st.inactivo,0)=0 AND st.ref IS NOT NULL AND st.ref <> ''
        """)
        artigos = cur.fetchall()

        # --- Sales last 12 months ---
        cur.execute("""
            SELECT fi.ref, SUM(fi.qtt) AS total_vendido,
                   SUM(fi.qtt * ISNULL(fi.epv,0)) AS total_faturado,
                   COUNT(DISTINCT ft.ftstamp) AS num_facturas,
                   MAX(ft.fdata) AS ultima_venda
            FROM fi
            INNER JOIN ft ON ft.ftstamp = fi.ftstamp
            WHERE ft.tipodoc = 1 AND ft.anulado = 0
              AND ft.fdata >= DATEADD(month, -12, GETDATE())
              AND fi.qtt > 0
            GROUP BY fi.ref
        """)
        vendas = {r[0].strip(): {
            'vendido': float(r[1] or 0),
            'faturado': float(r[2] or 0),
            'facturas': int(r[3] or 0),
            'ultima_venda': r[4].strftime('%Y-%m-%d') if r[4] else None
        } for r in cur.fetchall()}

        # --- Purchases last 12 months ---
        cur.execute("""
            SELECT RTRIM(fn.ref), SUM(fn.qtt) AS total_comprado,
                   0 AS total_compras
            FROM fn
            INNER JOIN fo ON fo.fostamp = fn.fostamp
            WHERE fo.data >= DATEADD(month, -12, GETDATE())
              AND fn.qtt > 0
            GROUP BY RTRIM(fn.ref)
        """)
        compras = {r[0]: {'comprado': float(r[1] or 0), 'valor_compras': float(r[2] or 0)}
                   for r in cur.fetchall()}

        conn.close()

        # --- Calculate metrics per article ---
        items = []
        total_valor_custo = 0
        total_valor_pvp = 0
        total_margem = 0
        sem_movimento = 0
        stock_negativo = 0
        total_artigos = 0

        for r in artigos:
            ref = (r[0] or '').strip()
            design = (r[1] or '').strip()
            stock = float(r[2])
            custo = float(r[3]) or float(r[4])  # epcusto or epcpond
            pvp = float(r[5])
            familia = (r[6] or '').strip()

            v = vendas.get(ref, {})
            c = compras.get(ref, {})

            vendido_12m = v.get('vendido', 0)
            faturado_12m = v.get('faturado', 0)
            ultima_venda = v.get('ultima_venda')

            valor_stock_custo = stock * custo
            valor_stock_pvp = stock * pvp
            margem_pct = ((pvp - custo) / pvp * 100) if pvp > 0 else 0
            margem_valor = (pvp - custo) * vendido_12m if vendido_12m > 0 else 0

            # Days without movement
            dias_sem_venda = None
            if ultima_venda:
                from datetime import datetime
                dias_sem_venda = (datetime.now() - datetime.strptime(ultima_venda, '%Y-%m-%d')).days

            total_valor_custo += max(0, valor_stock_custo)
            total_valor_pvp += max(0, valor_stock_pvp)
            total_margem += margem_valor
            total_artigos += 1

            if stock < 0: stock_negativo += 1
            if not ultima_venda or (dias_sem_venda and dias_sem_venda > 180): sem_movimento += 1

            items.append({
                'ref': ref,
                'design': design,
                'stock': stock,
                'custo': custo,
                'pvp': pvp,
                'familia': familia,
                'valor_custo': valor_stock_custo,
                'valor_pvp': valor_stock_pvp,
                'margem_pct': round(margem_pct, 1),
                'vendido_12m': vendido_12m,
                'faturado_12m': faturado_12m,
                'margem_valor': round(margem_valor, 2),
                'ultima_venda': ultima_venda,
                'dias_sem_venda': dias_sem_venda,
                'num_facturas': v.get('facturas', 0),
            })

        # --- ABC Classification by revenue ---
        items_com_venda = sorted([i for i in items if i['faturado_12m'] > 0],
                                  key=lambda x: x['faturado_12m'], reverse=True)
        total_fat = sum(i['faturado_12m'] for i in items_com_venda)
        acum = 0
        abc_counts = {'A': 0, 'B': 0, 'C': 0}
        for item in items_com_venda:
            acum += item['faturado_12m']
            pct = acum / total_fat if total_fat else 0
            if pct <= 0.7:
                item['abc'] = 'A'
                abc_counts['A'] += 1
            elif pct <= 0.9:
                item['abc'] = 'B'
                abc_counts['B'] += 1
            else:
                item['abc'] = 'C'
                abc_counts['C'] += 1
        for item in items:
            if 'abc' not in item:
                item['abc'] = '-'

        # --- GMROI ---
        gmroi = (total_margem / total_valor_custo * 100) if total_valor_custo > 0 else 0

        # --- Top performers ---
        def rot(x): return (x['stock'] / (x['vendido_12m']/12)) if x['vendido_12m'] > 0 else 0
        top_faturacao = sorted(items, key=lambda x: x['faturado_12m'], reverse=True)[:10]
        top_margem    = sorted([a for a in items if a['margem_valor'] > 0], key=lambda x: x['margem_valor'], reverse=True)[:10]
        sem_vendas    = sorted([a for a in items if a['vendido_12m'] == 0 and a['stock'] > 0], key=lambda x: x['valor_custo'], reverse=True)[:10]
        excesso       = sorted([a for a in items if a['vendido_12m'] > 0 and a['stock'] > 0], key=rot, reverse=True)[:10]

        # --- Alerts ---
        alertas = []
        for i in items:
            if i['stock'] < 0:
                alertas.append({'tipo': 'negativo', 'ref': i['ref'], 'design': i['design'][:50], 'valor': i['stock']})
            elif i['stock'] > 0 and i['dias_sem_venda'] and i['dias_sem_venda'] > 365:
                alertas.append({'tipo': 'obsoleto', 'ref': i['ref'], 'design': i['design'][:50], 'valor': i['dias_sem_venda']})
            elif i['margem_pct'] < 0 and i['vendido_12m'] > 0:
                alertas.append({'tipo': 'margem_neg', 'ref': i['ref'], 'design': i['design'][:50], 'valor': i['margem_pct']})

        # --- Families breakdown ---
        familias = {}
        for i in items:
            f = i['familia'] or 'Sem familia'
            if f not in familias:
                familias[f] = {'artigos': 0, 'valor_custo': 0, 'faturado': 0}
            familias[f]['artigos'] += 1
            familias[f]['valor_custo'] += max(0, i['valor_custo'])
            familias[f]['faturado'] += i['faturado_12m']
        top_familias = sorted(familias.items(), key=lambda x: x[1]['faturado'], reverse=True)[:8]

        return jsonify({
            'kpis': {
                'total_artigos': total_artigos,
                'artigos_com_stock': len([i for i in items if i['stock'] > 0]),
                'artigos_sem_stock': len([i for i in items if i['stock'] <= 0]),
                'stock_negativo': stock_negativo,
                'sem_movimento_180d': sem_movimento,
                'valor_stock_custo': round(total_valor_custo, 2),
                'valor_stock_pvp': round(total_valor_pvp, 2),
                'margem_total': round(total_margem, 2),
                'gmroi': round(gmroi, 1),
                'faturacao_12m': round(sum(i['faturado_12m'] for i in items), 2),
            },
            'abc': abc_counts,
            'alertas': alertas[:20],
            'top_faturacao': top_faturacao,
            'top_margem': top_margem,
            'sem_vendas': sem_vendas,
            'familias': [{'familia': k, **v} for k, v in top_familias],
        })

    except Exception as e:
        app.logger.error(f"inventario KPIs error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/tecnico/<int:eid>/consumiveis')
@login_required
def tecnico_consumiveis(eid):
    e = Equipamento.query.get_or_404(eid)
    consumiveis = EquipamentoConsumivel.query.filter_by(equipamento_id=eid).order_by(EquipamentoConsumivel.ordem).all()
    q = request.args.get('q', '').strip()
    artigos_sugeridos = []
    if q and len(q) >= 2:
        artigos_sugeridos = ArtigoPHC.query.filter(
            db.or_(ArtigoPHC.referencia.ilike(f'%{q}%'), ArtigoPHC.designacao.ilike(f'%{q}%'))
        ).limit(10).all()
    return render_template('tecnico_consumiveis.html', equipamento=e, consumiveis=consumiveis,
                           artigos_sugeridos=artigos_sugeridos, q=q)

@app.route('/tecnico/<int:eid>/consumiveis/adicionar', methods=['POST'])
@login_required
def tecnico_consumivel_adicionar(eid):
    Equipamento.query.get_or_404(eid)
    c = EquipamentoConsumivel(
        equipamento_id=eid,
        ref=request.form.get('ref','').strip(),
        designacao=request.form.get('designacao','').strip(),
        unidade=request.form.get('unidade','un').strip(),
        quantidade=float(request.form.get('quantidade', 1) or 1),
        notas=request.form.get('notas','').strip(),
        ordem=EquipamentoConsumivel.query.filter_by(equipamento_id=eid).count(),
    )
    db.session.add(c)
    db.session.commit()
    return redirect(url_for('tecnico_consumiveis', eid=eid))

@app.route('/tecnico/consumivel/<int:cid>/apagar', methods=['POST'])
@login_required
def tecnico_consumivel_apagar(cid):
    c = EquipamentoConsumivel.query.get_or_404(cid)
    eid = c.equipamento_id
    db.session.delete(c)
    db.session.commit()
    return redirect(url_for('tecnico_consumiveis', eid=eid))

@app.route('/tecnico/consumivel/<int:cid>/editar', methods=['POST'])
@login_required
def tecnico_consumivel_editar(cid):
    c = EquipamentoConsumivel.query.get_or_404(cid)
    c.ref        = request.form.get('ref', c.ref)
    c.designacao = request.form.get('designacao', c.designacao)
    c.unidade    = request.form.get('unidade', c.unidade)
    c.quantidade = float(request.form.get('quantidade', c.quantidade) or 1)
    c.notas      = request.form.get('notas', c.notas)
    db.session.commit()
    return redirect(url_for('tecnico_consumiveis', eid=c.equipamento_id))


# ── MÓDULO TÉCNICO — GESTÃO DE EQUIPAMENTOS ────────────────────────────────

UPLOAD_TECNICO = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'tecnico')

def ensure_upload_dir():
    os.makedirs(UPLOAD_TECNICO, exist_ok=True)

@app.route('/tecnico')
@login_required
def tecnico():
    equipamentos = Equipamento.query.order_by(Equipamento.embarcacao).all()
    return render_template('tecnico.html', equipamentos=equipamentos)

@app.route('/tecnico/novo', methods=['GET', 'POST'])
@login_required
def tecnico_novo():
    if request.method == 'POST':
        e = Equipamento(
            cliente_nome=request.form.get('cliente_nome','').strip(),
            embarcacao=request.form.get('embarcacao','').strip(),
            motor_marca=request.form.get('motor_marca','').strip(),
            motor_modelo=request.form.get('motor_modelo','').strip(),
            motor_potencia=request.form.get('motor_potencia','').strip(),
            serial_number=request.form.get('serial_number','').strip(),
            base_code=request.form.get('base_code','').strip(),
            manufactured_date=request.form.get('manufactured_date','').strip(),
            material=request.form.get('material','').strip(),
            caixa_modelo=request.form.get('caixa_modelo','').strip(),
            caixa_ratio=request.form.get('caixa_ratio','').strip(),
            caixa_serial=request.form.get('caixa_serial','').strip(),
            notas=request.form.get('notas','').strip(),
        )
        db.session.add(e)
        db.session.commit()
        return redirect(url_for('tecnico_detalhe', eid=e.id))
    return render_template('tecnico_form.html', equipamento=None)

@app.route('/tecnico/<int:eid>')
@login_required
def tecnico_detalhe(eid):
    e = Equipamento.query.get_or_404(eid)
    opcoes = EquipamentoOpcao.query.filter_by(equipamento_id=eid).order_by(EquipamentoOpcao.ordem).all()
    motores_aux = EquipamentoMotorAux.query.filter_by(equipamento_id=eid).order_by(EquipamentoMotorAux.numero).all()
    documentos = EquipamentoDocumento.query.filter_by(equipamento_id=eid).order_by(EquipamentoDocumento.componente, EquipamentoDocumento.criado_em).all()
    return render_template('tecnico_detalhe.html', equipamento=e, opcoes=opcoes, motores_aux=motores_aux, documentos=documentos)

@app.route('/tecnico/<int:eid>/editar', methods=['GET', 'POST'])
@login_required
def tecnico_editar(eid):
    e = Equipamento.query.get_or_404(eid)
    if request.method == 'POST':
        e.cliente_nome      = request.form.get('cliente_nome','').strip()
        e.embarcacao        = request.form.get('embarcacao','').strip()
        e.motor_marca       = request.form.get('motor_marca','').strip()
        e.motor_modelo      = request.form.get('motor_modelo','').strip()
        e.motor_potencia    = request.form.get('motor_potencia','').strip()
        e.serial_number     = request.form.get('serial_number','').strip()
        e.base_code         = request.form.get('base_code','').strip()
        e.manufactured_date = request.form.get('manufactured_date','').strip()
        e.catalogo          = request.form.get('catalogo','').strip()
        e.material          = request.form.get('material','').strip()
        e.caixa_modelo      = request.form.get('caixa_modelo','').strip()
        e.caixa_ratio       = request.form.get('caixa_ratio','').strip()
        e.caixa_serial      = request.form.get('caixa_serial','').strip()
        e.tipo_motor        = request.form.get('tipo_motor','principal')
        e.ativo             = request.form.get('ativo') == '1'
        e.catalogo          = request.form.get('catalogo','').strip()
        e.material          = request.form.get('material','').strip()
        e.manufacturing_date = request.form.get('manufacturing_date','').strip()
        e.base_engine_pt    = request.form.get('base_engine_pt','').strip()
        e.base_engine_eng   = request.form.get('base_engine_eng','').strip()
        e.fuel_system_pt    = request.form.get('fuel_system_pt','').strip()
        e.fuel_system_eng   = request.form.get('fuel_system_eng','').strip()
        e.notas             = request.form.get('notas','').strip()
        db.session.commit()
        return redirect(url_for('tecnico_detalhe', eid=eid))
    return render_template('tecnico_form.html', equipamento=e)

@app.route('/tecnico/<int:eid>/apagar', methods=['POST'])
@login_required
def tecnico_apagar(eid):
    e = Equipamento.query.get_or_404(eid)
    db.session.delete(e)
    db.session.commit()
    return redirect(url_for('tecnico'))

@app.route('/tecnico/<int:eid>/opcao', methods=['POST'])
@login_required
def tecnico_opcao_criar(eid):
    Equipamento.query.get_or_404(eid)
    o = EquipamentoOpcao(
        equipamento_id=eid,
        option_name=request.form.get('option_name','').strip(),
        ordered=request.form.get('ordered','').strip(),
        factory=request.form.get('factory','').strip(),
        distributor=request.form.get('distributor','').strip(),
        ordem=EquipamentoOpcao.query.filter_by(equipamento_id=eid).count(),
    )
    db.session.add(o)
    db.session.commit()
    return redirect(url_for('tecnico_detalhe', eid=eid))

@app.route('/tecnico/opcao/<int:oid>/editar', methods=['POST'])
@login_required
def tecnico_opcao_editar(oid):
    o = EquipamentoOpcao.query.get_or_404(oid)
    o.option_name  = request.form.get('option_name', o.option_name)
    o.ordered      = request.form.get('ordered', o.ordered)
    o.factory      = request.form.get('factory', o.factory)
    o.distributor  = request.form.get('distributor', o.distributor)
    db.session.commit()
    return redirect(url_for('tecnico_detalhe', eid=o.equipamento_id))

@app.route('/tecnico/opcao/<int:oid>/apagar', methods=['POST'])
@login_required
def tecnico_opcao_apagar(oid):
    o = EquipamentoOpcao.query.get_or_404(oid)
    eid = o.equipamento_id
    db.session.delete(o)
    db.session.commit()
    return redirect(url_for('tecnico_detalhe', eid=eid))

@app.route('/tecnico/opcao/<int:oid>/upload', methods=['POST'])
@login_required
def tecnico_opcao_upload(oid):
    o = EquipamentoOpcao.query.get_or_404(oid)
    f = request.files.get('pdf')
    if not f or not f.filename.lower().endswith('.pdf'):
        flash('Ficheiro inválido — apenas PDF.', 'error')
        return redirect(url_for('tecnico_detalhe', eid=o.equipamento_id))
    ensure_upload_dir()
    safe = f"{oid}_{f.filename.replace(' ','_')}"
    path = os.path.join(UPLOAD_TECNICO, safe)
    f.save(path)
    o.pdf_filename = f.filename
    o.pdf_path = safe

    # Determine reference code by priority: distributor > factory > ordered
    dist = (o.distributor or '').strip().replace('*','').strip()
    fac  = (o.factory    or '').strip().replace('*','').strip()
    ord_ = (o.ordered    or '').strip().replace('*','').strip()
    ref_code = dist or fac or ord_
    ref_type = 'distributor' if dist else ('factory' if fac else 'ordered' if ord_ else None)

    # Get catalog of this equipment - REQUIRED for propagation
    eq = Equipamento.query.get(o.equipamento_id)
    catalogo = (eq.catalogo or '').strip().upper() if eq else None

    propagated = 0
    if ref_code and catalogo:
        # Only propagate if equipment has a catalog defined
        eq_ids = [e.id for e in Equipamento.query.filter(
            db.func.upper(Equipamento.catalogo) == catalogo
        ).all()]

        def match_others(col):
            return EquipamentoOpcao.query.filter(
                EquipamentoOpcao.id != oid,
                EquipamentoOpcao.equipamento_id.in_(eq_ids),
                db.func.replace(db.func.replace(col,'*',''),' ','') == ref_code
            ).all()

        if dist:   others = match_others(EquipamentoOpcao.distributor)
        elif fac:  others = match_others(EquipamentoOpcao.factory)
        else:      others = match_others(EquipamentoOpcao.ordered)

        for other in others:
            other.pdf_filename = f.filename
            other.pdf_path = safe
            propagated += 1

    db.session.commit()

    if propagated:
        flash(f'✅ PDF propagado para {propagated} linhas com catálogo={catalogo} e {ref_type}={ref_code}.', 'success')
    elif ref_code and not catalogo:
        flash('PDF associado. ⚠️ Sem catálogo definido — não foi propagado a outras fichas.', 'warning')
    else:
        flash('PDF associado.', 'success')

    return redirect(url_for('tecnico_detalhe', eid=o.equipamento_id))

@app.route('/tecnico/opcao/<int:oid>/pdf')
@login_required
def tecnico_opcao_pdf(oid):
    o = EquipamentoOpcao.query.get_or_404(oid)
    if not o.pdf_path:
        return 'Sem PDF', 404
    return send_from_directory(UPLOAD_TECNICO, o.pdf_path,
                               as_attachment=False, download_name=o.pdf_filename)

@app.route('/tecnico/opcao/<int:oid>/pdf/download')
@login_required
def tecnico_opcao_pdf_download(oid):
    o = EquipamentoOpcao.query.get_or_404(oid)
    if not o.pdf_path:
        return 'Sem PDF', 404
    return send_from_directory(UPLOAD_TECNICO, o.pdf_path,
                               as_attachment=True, download_name=o.pdf_filename)


@app.route('/tecnico/<int:eid>/importar-html', methods=['POST'])
@login_required
def tecnico_importar_html(eid):
    """Parse John Deere JDPS HTML - table#optioninfo."""
    Equipamento.query.get_or_404(eid)
    f = request.files.get('html_file')
    if not f:
        flash('Nenhum ficheiro enviado.', 'error')
        return redirect(url_for('tecnico_detalhe', eid=eid))
    try:
        raw = f.read()
        for enc in ('utf-8', 'latin-1', 'cp1252'):
            try: content = raw.decode(enc); break
            except Exception: content = raw.decode('latin-1', errors='replace')

        from html.parser import HTMLParser

        class JDPSParser(HTMLParser):
            def __init__(self):
                super().__init__()
                self.rows = []
                self.in_table = False
                self.table_depth = 0
                self.in_tr = False
                self.cells = []
                self.in_td = False
                self.cur_attrs = {}
                self.cur_text = ''

            def handle_starttag(self, tag, attrs):
                a = dict(attrs)
                if tag == 'table':
                    if a.get('id') == 'optioninfo':
                        self.in_table = True
                        self.table_depth = 1
                    elif self.in_table:
                        self.table_depth += 1
                if self.in_table and self.table_depth == 1:
                    if tag == 'tr':
                        self.in_tr = True
                        self.cells = []
                    elif tag == 'td' and self.in_tr:
                        self.in_td = True
                        self.cur_attrs = a
                        self.cur_text = ''

            def handle_endtag(self, tag):
                if tag == 'table' and self.in_table:
                    self.table_depth -= 1
                    if self.table_depth == 0:
                        self.in_table = False
                if self.in_table and self.table_depth == 1:
                    if tag == 'td' and self.in_td:
                        self.in_td = False
                        self.cells.append((dict(self.cur_attrs), self.cur_text.strip()))
                    elif tag == 'tr' and self.in_tr:
                        self.in_tr = False
                        self._process_row()

            def handle_data(self, data):
                if self.in_td:
                    self.cur_text += data

            def _process_row(self):
                # Structure: trans=en_US=Option | trans=other=skip |
                #            plain[0]=Ordered | plain[1]=Factory |
                #            name2=viewDistribOption=Distributor | rest=skip
                option_name = ordered = factory = distributor = ''
                plain_cells = []
                for a, t in self.cells:
                    raw = t.strip()  # keep * as-is
                    if a.get('trans') == 'en_US':
                        option_name = raw
                    elif a.get('trans') == 'other':
                        continue
                    elif a.get('name2') == 'viewDistribOption':
                        distributor = raw
                    elif a.get('name2') == 'editDistribOption':
                        continue
                    elif a.get('width') == '100px' and a.get('align') == 'center':
                        continue
                    elif not a.get('trans') and not a.get('name2') and option_name:
                        plain_cells.append(raw)

                # Cell 2 (plain[0]) = Ordered (e.g. * or code)
                # Cell 3 (plain[1]) = Factory (e.g. 1102, 1299)
                if len(plain_cells) >= 1:
                    ordered = plain_cells[0]   # * or actual code
                if len(plain_cells) >= 2:
                    factory = plain_cells[1]   # factory code number

                if option_name:
                    self.rows.append({'option': option_name, 'ordered': ordered,
                                      'factory': factory, 'distributor': distributor})

        parser = JDPSParser()
        parser.feed(content)
        app.logger.info(f"JDPS: {len(parser.rows)} rows")

        added = 0
        ordem_start = EquipamentoOpcao.query.filter_by(equipamento_id=eid).count()
        for row in parser.rows:
            if not row['option']: continue
            o = EquipamentoOpcao(
                equipamento_id=eid,
                option_name=row['option'],
                ordered=row['ordered'],
                factory=row['factory'],
                distributor=row['distributor'],
                ordem=ordem_start + added,
            )
            db.session.add(o)
            added += 1
        db.session.commit()
        flash(f'Importados {added} option codes com sucesso.', 'success')
    except Exception as e:
        app.logger.error(f"JDPS import error: {e}")
        flash(f'Erro: {e}', 'error')
    return redirect(url_for('tecnico_detalhe', eid=eid))


@app.route('/tecnico/<int:eid>/opcoes/apagar-bulk', methods=['POST'])
@login_required
def tecnico_opcoes_apagar_bulk(eid):
    Equipamento.query.get_or_404(eid)
    ids = request.form.getlist('opcao_ids')
    if ids:
        EquipamentoOpcao.query.filter(
            EquipamentoOpcao.id.in_([int(i) for i in ids]),
            EquipamentoOpcao.equipamento_id == eid
        ).delete(synchronize_session=False)
        db.session.commit()
        flash(f'{len(ids)} linhas eliminadas.', 'success')
    return redirect(url_for('tecnico_detalhe', eid=eid))


@app.route('/tecnico/importar-excel', methods=['GET', 'POST'])
@login_required
def tecnico_importar_excel():
    if request.method == 'GET':
        return render_template('tecnico_importar_excel.html')
    
    f = request.files.get('excel_file')
    if not f:
        flash('Nenhum ficheiro enviado.', 'error')
        return redirect(url_for('tecnico_importar_excel'))
    
    try:
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
        ws = wb.active
        
        # Find the actual header row and starting column (skip empty rows/cols)
        header_row = 1
        col_offset = 0
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            vals = [str(v or '').strip() for v in row]
            non_empty = [(i, v) for i, v in enumerate(vals) if v]
            if len(non_empty) >= 3:  # found header row
                header_row = list(ws.iter_rows(min_row=1, max_row=10, values_only=True)).index(row) + 1
                col_offset = non_empty[0][0]
                break
        
        header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        headers = [str(v or '').strip().lower() for v in header_cells]
        app.logger.info(f"Header row={header_row} offset={col_offset} headers={headers}")

        def find_col(names):
            for name in names:
                for i, h in enumerate(headers):
                    # normalise: remove accents roughly
                    hn = h.replace('ã','a').replace('á','a').replace('â','a').replace('é','e').replace('ê','e').replace('í','i').replace('ó','o').replace('ô','o').replace('ú','u').replace('ç','c').replace('º','').replace('ª','').replace('°','').strip()
                    nm = name.replace('ã','a').replace('á','a').replace('â','a').replace('é','e').replace('ê','e').replace('í','i').replace('ó','o').replace('ô','o').replace('ú','u').replace('ç','c').replace('º','').replace('ª','').strip()
                    if nm in hn:
                        return i
            return None

        # Try header-based detection first
        col_emb      = find_col(['embarca'])
        col_mod      = find_col(['modelo motor','motor mod','modelo m'])
        col_pot      = find_col(['potencia','pot','cv','kw'])
        col_rpm      = find_col(['rpm'])
        col_serie    = find_col(['serie motor','n serie motor','n. serie motor','serie m','n serie m'])
        col_base     = find_col(['base code','base cod','eq'])
        col_data_fab = find_col(['data fab','data de fab','fab'])
        col_marca_cx = find_col(['marca caixa','marca cai'])
        col_mod_cx   = find_col(['modelo caixa','caixa mod','cai mod','modelo cai'])
        col_ratio    = find_col(['ratio'])
        col_serie_cx = find_col(['serie caixa','n serie caixa','serie cai','n serie cai'])
        col_obs      = find_col(['obs','nota'])

        # Fallback: use positional order if headers not detected
        # Order: Embarcação Modelo Motor Potência RPM Nº Série Base Code Data Fab. Caixa Red. Ratio Série Caixa Obs
        ncols = len(headers)
        o = col_offset  # column offset
        if col_emb is None:      col_emb      = o + 0
        if col_mod is None:      col_mod      = o + 1
        if col_pot is None:      col_pot      = o + 2
        if col_rpm is None:      col_rpm      = o + 3
        if col_serie is None:    col_serie    = o + 4
        col_base = None  # not used
        col_data_fab = None  # not used
        if col_mod_cx is None:   col_mod_cx   = o + 8
        if col_ratio is None:    col_ratio    = o + 9
        if col_serie_cx is None: col_serie_cx = o + 10
        if col_obs is None:      col_obs      = o + 12

        app.logger.info(f"Cols mapped: emb={col_emb} mod={col_mod} pot={col_pot} serie={col_serie} base={col_base} cx={col_mod_cx} ratio={col_ratio}")
        
        # Log first data row for debug
        first_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
        app.logger.info(f"First data row: {first_row}")
        app.logger.info(f"Total rows: {ws.max_row}")

        added = 0
        duplicados = 0
        erros = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
            if not any(v for v in row if v is not None):
                continue
            
            def get(col):
                if col is None or col >= len(row): return ''
                v = row[col]
                if v is None: return ''
                return str(v).strip()
            
            embarcacao = get(col_emb)
            if not embarcacao or embarcacao.lower() in ('none','nan','','null'):
                app.logger.info(f'Row {row_idx}: skipped (empty embarcacao)')
                continue
            app.logger.info(f'Row {row_idx}: embarcacao={embarcacao}')

            serial = get(col_serie)
            
            # Check duplicate
            dup = Equipamento.query.filter_by(embarcacao=embarcacao)
            if serial:
                dup = dup.filter_by(serial_number=serial)
            if dup.first():
                duplicados += 1
                erros.append(f"L{row_idx}: {embarcacao}")
                continue
            
            pot = get(col_pot)
            if pot and 'cv' not in pot.lower() and 'kw' not in pot.lower() and pot.replace('.','').replace(',','').isdigit():
                pot = pot + ' CV'
            
            e = Equipamento(
                embarcacao      = embarcacao,
                motor_modelo    = get(col_mod),
                motor_potencia  = pot,
                serial_number   = serial,
                caixa_modelo    = get(col_mod_cx),
                caixa_ratio     = get(col_ratio),
                caixa_serial    = get(col_serie_cx),
                notas           = get(col_obs),
            )
            db.session.add(e)
            added += 1
        
        db.session.commit()
        
        msg = f'✅ {added} equipamentos importados.'
        if duplicados:
            msg += f' {duplicados} duplicados ignorados ({", ".join(erros[:3])}{"..." if len(erros)>3 else ""}).'
        flash(msg, 'success' if added > 0 else 'warning')
        return redirect(url_for('tecnico'))
    
    except Exception as e:
        app.logger.error(f"Excel import error: {e}")
        flash(f'Erro: {e}', 'error')
        return redirect(url_for('tecnico_importar_excel'))


@app.route('/tecnico/<int:eid>/motor-aux', methods=['POST'])
@login_required
def tecnico_motor_aux_criar(eid):
    Equipamento.query.get_or_404(eid)
    n = EquipamentoMotorAux.query.filter_by(equipamento_id=eid).count() + 1
    m = EquipamentoMotorAux(
        equipamento_id=eid,
        numero=int(request.form.get('numero', n)),
        marca=request.form.get('marca','').strip(),
        modelo=request.form.get('modelo','').strip(),
        serial_number=request.form.get('serial_number','').strip(),
        potencia=request.form.get('potencia','').strip(),
        rpm=request.form.get('rpm','').strip(),
        ano=request.form.get('ano','').strip(),
        notas=request.form.get('notas','').strip(),
        ordem=n,
    )
    db.session.add(m)
    db.session.commit()
    return redirect(url_for('tecnico_detalhe', eid=eid) + '#motores-aux')

@app.route('/tecnico/motor-aux/<int:mid>/editar', methods=['POST'])
@login_required
def tecnico_motor_aux_editar(mid):
    m = EquipamentoMotorAux.query.get_or_404(mid)
    m.numero        = int(request.form.get('numero', m.numero))
    m.marca         = request.form.get('marca', m.marca or '').strip()
    m.modelo        = request.form.get('modelo', m.modelo or '').strip()
    m.serial_number = request.form.get('serial_number', m.serial_number or '').strip()
    m.potencia      = request.form.get('potencia', m.potencia or '').strip()
    m.rpm           = request.form.get('rpm', m.rpm or '').strip()
    m.ano           = request.form.get('ano', m.ano or '').strip()
    m.notas         = request.form.get('notas', m.notas or '').strip()
    db.session.commit()
    return redirect(url_for('tecnico_detalhe', eid=m.equipamento_id) + '#motores-aux')

@app.route('/tecnico/motor-aux/<int:mid>/apagar', methods=['POST'])
@login_required
def tecnico_motor_aux_apagar(mid):
    m = EquipamentoMotorAux.query.get_or_404(mid)
    eid = m.equipamento_id
    db.session.delete(m)
    db.session.commit()
    return redirect(url_for('tecnico_detalhe', eid=eid) + '#motores-aux')


@app.route('/tecnico/<int:eid>/documento/upload', methods=['POST'])
@login_required
def tecnico_doc_upload(eid):
    e = Equipamento.query.get_or_404(eid)
    f = request.files.get('doc_file')
    componente = request.form.get('componente', '').strip()
    titulo = request.form.get('titulo', '').strip()
    notas = request.form.get('notas', '').strip()
    if not f or not componente or not titulo:
        flash('Preencha todos os campos e seleccione um ficheiro.', 'error')
        return redirect(url_for('tecnico_detalhe', eid=eid))
    ensure_upload_dir()
    safe = f"{eid}_{componente}_{f.filename.replace(' ','_')}"
    path = os.path.join(UPLOAD_TECNICO, safe)
    f.save(path)

    # Link to this equipment
    doc = EquipamentoDocumento(
        equipamento_id=eid,
        componente=componente,
        titulo=titulo,
        pdf_filename=f.filename,
        pdf_path=safe,
        notas=notas,
    )
    db.session.add(doc)

    # Determine model key for sharing
    tipo_comp = None
    modelo_val = None
    if componente == 'caixa' and e.caixa_modelo:
        tipo_comp = 'caixa'
        modelo_val = e.caixa_modelo.strip()
    elif componente == 'motor' and e.motor_modelo:
        tipo_comp = 'motor'
        modelo_val = e.motor_modelo.strip()
    elif componente.startswith('aux_'):
        # find the aux motor model
        num = int(componente.split('_')[1]) if '_' in componente else 1
        aux = EquipamentoMotorAux.query.filter_by(equipamento_id=eid, numero=num).first()
        if aux and aux.modelo:
            tipo_comp = 'aux'
            modelo_val = aux.modelo.strip()

    # Save to shared library if model identified
    if tipo_comp and modelo_val:
        existing = ModeloPDF.query.filter_by(tipo_componente=tipo_comp,
                                             modelo_codigo=modelo_val,
                                             titulo=titulo).first()
        if not existing:
            mp = ModeloPDF(
                tipo_componente=tipo_comp,
                modelo_codigo=modelo_val,
                titulo=titulo,
                pdf_filename=f.filename,
                pdf_path=safe,
            )
            db.session.add(mp)

    db.session.commit()

    msg = f'Documento adicionado.'
    if tipo_comp and modelo_val:
        count = Equipamento.query
        if tipo_comp == 'caixa':
            count = count.filter_by(caixa_modelo=modelo_val).count()
        elif tipo_comp == 'motor':
            count = count.filter_by(motor_modelo=modelo_val).count()
        else:
            count = 1
        if count > 1:
            msg = f'Documento guardado e partilhado com {count} equipamentos com {modelo_val}.'
    flash(msg, 'success')
    return redirect(url_for('tecnico_detalhe', eid=eid))

@app.route('/tecnico/documento/<int:did>/ver')
@login_required
def tecnico_doc_ver(did):
    d = EquipamentoDocumento.query.get_or_404(did)
    import mimetypes
    mime = mimetypes.guess_type(d.pdf_filename or d.pdf_path or '')[0] or 'application/octet-stream'
    if mime.startswith('image/'):
        return send_from_directory(UPLOAD_TECNICO, d.pdf_path, as_attachment=False)
    return send_from_directory(UPLOAD_TECNICO, d.pdf_path,
                               as_attachment=False, download_name=d.pdf_filename)

@app.route('/tecnico/documento/<int:did>/download')
@login_required
def tecnico_doc_download(did):
    d = EquipamentoDocumento.query.get_or_404(did)
    return send_from_directory(UPLOAD_TECNICO, d.pdf_path,
                               as_attachment=True, download_name=d.pdf_filename)

@app.route('/tecnico/documento/<int:did>/apagar', methods=['POST'])
@login_required
def tecnico_doc_apagar(did):
    d = EquipamentoDocumento.query.get_or_404(did)
    eid = d.equipamento_id
    try:
        path = os.path.join(UPLOAD_TECNICO, d.pdf_path)
        if os.path.exists(path): os.remove(path)
    except Exception: pass
    db.session.delete(d)
    db.session.commit()
    return redirect(url_for('tecnico_detalhe', eid=eid))


# ── BIBLIOTECA DE FACTORY CODE PDFs ─────────────────────────────────────────

@app.route('/api/factory-code/<code>/pdf')
@login_required
def api_factory_code_pdf(code):
    """Return PDF info for a factory code from the central library."""
    pdf = FactoryCodePDF.query.filter_by(factory_code=code).order_by(FactoryCodePDF.criado_em.desc()).first()
    if not pdf:
        return jsonify({'found': False})
    return jsonify({
        'found': True,
        'id': pdf.id,
        'titulo': pdf.titulo,
        'pdf_filename': pdf.pdf_filename,
        'ver_url': url_for('factory_code_pdf_ver', pid=pdf.id),
        'download_url': url_for('factory_code_pdf_download', pid=pdf.id),
    })

@app.route('/factory-code/pdf/<int:pid>/ver')
@login_required
def factory_code_pdf_ver(pid):
    p = FactoryCodePDF.query.get_or_404(pid)
    return send_from_directory(UPLOAD_TECNICO, p.pdf_path,
                               as_attachment=False, download_name=p.pdf_filename)

@app.route('/factory-code/pdf/<int:pid>/download')
@login_required
def factory_code_pdf_download(pid):
    p = FactoryCodePDF.query.get_or_404(pid)
    return send_from_directory(UPLOAD_TECNICO, p.pdf_path,
                               as_attachment=True, download_name=p.pdf_filename)

@app.route('/tecnico/opcao/<int:oid>/upload-factory', methods=['POST'])
@login_required
def tecnico_opcao_upload_factory(oid):
    """Upload PDF for an option code — saves to central library by factory code."""
    o = EquipamentoOpcao.query.get_or_404(oid)
    f = request.files.get('pdf')
    if not f or not f.filename:
        flash('Seleccione um ficheiro.', 'error')
        return redirect(url_for('tecnico_detalhe', eid=o.equipamento_id))

    ensure_upload_dir()
    factory_code = (o.factory or o.ordered or '').strip()
    safe = f"factory_{factory_code}_{f.filename.replace(' ','_')}"
    path = os.path.join(UPLOAD_TECNICO, safe)
    f.save(path)

    # Save/update central library entry
    if factory_code:
        existing = FactoryCodePDF.query.filter_by(factory_code=factory_code).first()
        if existing:
            # Update existing — replace file
            try:
                old = os.path.join(UPLOAD_TECNICO, existing.pdf_path)
                if os.path.exists(old) and old != path: os.remove(old)
            except Exception: pass
            existing.pdf_filename = f.filename
            existing.pdf_path = safe
            existing.titulo = f.filename
        else:
            db.session.add(FactoryCodePDF(
                factory_code=factory_code,
                titulo=f.filename,
                pdf_filename=f.filename,
                pdf_path=safe,
            ))

    # Also link directly to this option
    o.pdf_filename = f.filename
    o.pdf_path = safe
    db.session.commit()

    # Count how many other options share this factory code
    if factory_code:
        others = EquipamentoOpcao.query.filter(
            EquipamentoOpcao.factory == factory_code,
            EquipamentoOpcao.id != oid
        ).count()
        if others:
            flash(f'PDF guardado e disponível automaticamente para {others} outras linhas com factory code {factory_code}.', 'success')
        else:
            flash('PDF guardado na biblioteca central.', 'success')
    else:
        flash('PDF guardado.', 'success')

    return redirect(url_for('tecnico_detalhe', eid=o.equipamento_id))

@app.route('/biblioteca-factory')
@login_required
def biblioteca_factory():
    """List all factory code PDFs in the central library."""
    pdfs = FactoryCodePDF.query.order_by(FactoryCodePDF.factory_code).all()
    equipamentos = Equipamento.query.options(
        db.joinedload(Equipamento.opcoes)
    ).all()
    return render_template('biblioteca_factory.html', pdfs=pdfs, equipamentos=equipamentos)

@app.route('/biblioteca-factory/<int:pid>/apagar', methods=['POST'])
@login_required
def biblioteca_factory_apagar(pid):
    p = FactoryCodePDF.query.get_or_404(pid)
    try:
        path = os.path.join(UPLOAD_TECNICO, p.pdf_path)
        if os.path.exists(path): os.remove(path)
    except Exception: pass
    db.session.delete(p)
    db.session.commit()
    flash('PDF removido da biblioteca.', 'success')
    return redirect(url_for('biblioteca_factory'))


# ── BIBLIOTECA DE MODELO PDFs (partilhados por modelo de componente) ─────────

@app.route('/api/modelo-pdf/<tipo>/<path:modelo>')
@login_required
def api_modelo_pdf_list(tipo, modelo):
    """Return all PDFs for a component model."""
    pdfs = ModeloPDF.query.filter_by(tipo_componente=tipo, modelo_codigo=modelo).order_by(ModeloPDF.titulo).all()
    return jsonify([{
        'id': p.id, 'titulo': p.titulo, 'pdf_filename': p.pdf_filename,
        'ver_url': url_for('modelo_pdf_ver', pid=p.id),
        'download_url': url_for('modelo_pdf_download', pid=p.id),
        'criado_em': p.criado_em.strftime('%d/%m/%Y'),
    } for p in pdfs])

@app.route('/modelo-pdf/<int:pid>/ver')
@login_required
def modelo_pdf_ver(pid):
    p = ModeloPDF.query.get_or_404(pid)
    return send_from_directory(UPLOAD_TECNICO, p.pdf_path,
                               as_attachment=False, download_name=p.pdf_filename)

@app.route('/modelo-pdf/<int:pid>/download')
@login_required
def modelo_pdf_download(pid):
    p = ModeloPDF.query.get_or_404(pid)
    return send_from_directory(UPLOAD_TECNICO, p.pdf_path,
                               as_attachment=True, download_name=p.pdf_filename)

@app.route('/modelo-pdf/<int:pid>/editar', methods=['POST'])
@login_required
def modelo_pdf_editar(pid):
    p = ModeloPDF.query.get_or_404(pid)
    novo_titulo = request.form.get('titulo', '').strip()
    if novo_titulo:
        p.titulo = novo_titulo
        db.session.commit()
    return redirect(request.referrer or url_for('biblioteca_modelos'))

@app.route('/modelo-pdf/<int:pid>/apagar', methods=['POST'])
@login_required
def modelo_pdf_apagar(pid):
    p = ModeloPDF.query.get_or_404(pid)
    try:
        path = os.path.join(UPLOAD_TECNICO, p.pdf_path)
        if os.path.exists(path): os.remove(path)
    except Exception: pass
    db.session.delete(p)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/tecnico/<int:eid>/upload-modelo', methods=['POST'])
@login_required
def tecnico_upload_modelo(eid):
    """Upload PDF to central model library (for caixa/motor model sharing)."""
    e = Equipamento.query.get_or_404(eid)
    f = request.files.get('pdf')
    tipo = request.form.get('tipo', '').strip()
    modelo = request.form.get('modelo', '').strip()
    titulo = request.form.get('titulo', '').strip()

    if not f or not tipo or not modelo or not titulo:
        return jsonify({'ok': False, 'error': 'Campos em falta'}), 400

    ensure_upload_dir()
    safe = f"modelo_{tipo}_{modelo}_{f.filename.replace(' ','_')}"
    path = os.path.join(UPLOAD_TECNICO, safe)
    f.save(path)

    p = ModeloPDF(
        tipo_componente=tipo,
        modelo_codigo=modelo,
        titulo=titulo,
        pdf_filename=f.filename,
        pdf_path=safe,
    )
    db.session.add(p)
    db.session.commit()

    # Count how many equipamentos share this model
    if tipo == 'caixa':
        count = Equipamento.query.filter_by(caixa_modelo=modelo).count()
    elif tipo == 'motor':
        count = Equipamento.query.filter_by(motor_modelo=modelo).count()
    else:
        count = 0

    return jsonify({'ok': True, 'id': p.id, 'count': count, 'titulo': titulo})

@app.route('/biblioteca-modelos', methods=['GET', 'POST'])
@login_required
def biblioteca_modelos():
    q = request.args.get('q', '').strip()
    
    if request.method == 'POST':
        f = request.files.get('pdf')
        titulo = request.form.get('titulo', '').strip()
        tipo_comp = request.form.get('tipo_componente', 'geral').strip()
        modelo_cod = request.form.get('modelo_codigo', '').strip()
        
        if not f or not titulo:
            flash('Título e ficheiro são obrigatórios.', 'error')
            return redirect(url_for('biblioteca_modelos'))
        
        ensure_upload_dir()
        safe = f"bib_{tipo_comp}_{modelo_cod}_{f.filename.replace(' ','_')}".replace('/','_')
        path = os.path.join(UPLOAD_TECNICO, safe)
        f.save(path)
        
        # Generate thumbnail of first page
        thumb = None
        try:
            from pdf2image import convert_from_path
            pages = convert_from_path(path, first_page=1, last_page=1, dpi=80)
            if pages:
                thumb_name = safe.replace('.pdf', '_thumb.jpg')
                thumb_path = os.path.join(UPLOAD_TECNICO, thumb_name)
                pages[0].save(thumb_path, 'JPEG', quality=70)
                thumb = thumb_name
        except Exception as ex:
            app.logger.warning(f"Thumbnail error: {ex}")
        
        p = ModeloPDF(
            tipo_componente=tipo_comp,
            modelo_codigo=modelo_cod,
            titulo=titulo,
            pdf_filename=f.filename,
            pdf_path=safe,
        )
        # Store thumb in notas field temporarily — we'll add a column
        if thumb:
            p.pdf_path = safe  # keep path
        db.session.add(p)
        db.session.commit()
        
        # Store thumbnail path
        if thumb:
            try:
                from sqlalchemy import text
                with db.engine.begin() as conn:
                    conn.execute(text("UPDATE modelo_pdf SET thumb_path=:t WHERE id=:id"),
                                {'t': thumb, 'id': p.id})
            except Exception:
                pass
        
        flash(f'"{titulo}" adicionado à biblioteca.', 'success')
        return redirect(url_for('biblioteca_modelos'))
    
    query = ModeloPDF.query
    if q:
        query = query.filter(
            db.or_(
                ModeloPDF.titulo.ilike(f'%{q}%'),
                ModeloPDF.modelo_codigo.ilike(f'%{q}%'),
                ModeloPDF.tipo_componente.ilike(f'%{q}%'),
            )
        )
    pdfs = query.order_by(ModeloPDF.tipo_componente, ModeloPDF.modelo_codigo, ModeloPDF.titulo).all()
    
    # Collect unique models for suggestions
    caixas = db.session.query(Equipamento.caixa_modelo).filter(
        Equipamento.caixa_modelo.isnot(None)).distinct().all()
    motores = db.session.query(Equipamento.motor_modelo).filter(
        Equipamento.motor_modelo.isnot(None)).distinct().all()
    
    return render_template('biblioteca_modelos.html', pdfs=pdfs, q=q,
                           caixas=[c[0] for c in caixas if c[0]],
                           motores=[m[0] for m in motores if m[0]])


@app.route('/tecnico/documento/<int:did>/editar-titulo', methods=['POST'])
@login_required
def tecnico_doc_editar_titulo(did):
    d = EquipamentoDocumento.query.get_or_404(did)
    novo = request.form.get('titulo', '').strip()
    if novo:
        d.titulo = novo
        db.session.commit()
    return redirect(url_for('tecnico_detalhe', eid=d.equipamento_id))


@app.route('/api/tecnico/<int:eid>/tipo', methods=['POST'])
@login_required
def api_tecnico_tipo(eid):
    e = Equipamento.query.get_or_404(eid)
    novo = request.get_json().get('tipo', 'principal')
    e.tipo_motor = novo
    db.session.commit()
    return jsonify({'ok': True, 'tipo': novo})


@app.route('/api/tecnico/<int:eid>/ativo', methods=['POST'])
@login_required
def api_tecnico_ativo(eid):
    e = Equipamento.query.get_or_404(eid)
    e.ativo = not e.ativo
    db.session.commit()
    return jsonify({'ok': True, 'ativo': e.ativo})


import re as _re

def extrair_factory_code(filename):
    """Extract 4-5 char factory/ordered code from filename.
    e.g. PC12295_1150_ST617227 → code='1150', catalog='PC12295'
    """
    name = filename.replace('.pdf','').replace('.PDF','')
    parts = _re.split(r'[_\-]', name)
    # Catalog: first part that looks like PC##### (letters + digits)
    catalog = None
    for p in parts:
        if _re.match(r'^[A-Z]{1,3}[0-9]{4,6}$', p, _re.IGNORECASE):
            catalog = p.upper()
            break
    # Code: 4-5 char alphanumeric that is NOT the catalog and NOT a long serial
    code = None
    for p in parts:
        if p.upper() == (catalog or ''):
            continue
        if _re.match(r'^[A-Z0-9]{4,5}$', p, _re.IGNORECASE) and not _re.match(r'^ST', p, _re.IGNORECASE):
            code = p.upper()
            break
    return code, catalog

@app.route('/tecnico/<int:eid>/foto-galeria', methods=['POST'])
@login_required
def tecnico_foto_galeria(eid):
    Equipamento.query.get_or_404(eid)
    f = request.files.get('foto')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'Sem ficheiro'})
    import uuid, mimetypes
    ext = os.path.splitext(f.filename)[1].lower()
    nome = f'galeria_{eid}_{uuid.uuid4().hex[:8]}{ext}'
    ensure_upload_dir()
    f.save(os.path.join(UPLOAD_TECNICO, nome))
    doc = EquipamentoDocumento(
        equipamento_id=eid, componente='foto_galeria',
        titulo=f.filename, pdf_filename=f.filename, pdf_path=nome)
    db.session.add(doc); db.session.commit()
    return jsonify({'ok': True, 'id': doc.id,
        'url': '/tecnico/documento/' + str(doc.id) + '/ver',
        'nome': doc.pdf_filename})


@app.route('/tecnico/<int:eid>/foto', methods=['POST'])
@login_required
def tecnico_foto(eid):
    e = Equipamento.query.get_or_404(eid)
    f = request.files.get('foto')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'Sem ficheiro'})
    import uuid
    ext = os.path.splitext(f.filename)[1].lower()
    nome = f'foto_{eid}_{uuid.uuid4().hex[:8]}{ext}'
    ensure_upload_dir()
    f.save(os.path.join(UPLOAD_TECNICO, nome))
    # Remove old principal photo doc
    old = EquipamentoDocumento.query.filter_by(equipamento_id=eid, componente='foto_principal').first()
    if old:
        try: os.remove(os.path.join(UPLOAD_TECNICO, old.pdf_path or ''))
        except: pass
        db.session.delete(old)
    doc = EquipamentoDocumento(
        equipamento_id=eid, componente='foto_principal',
        titulo='Foto Principal', pdf_filename=f.filename, pdf_path=nome)
    db.session.add(doc)
    db.session.commit()
    return jsonify({'ok': True, 'url': '/tecnico/documento/' + str(doc.id) + '/ver'})


@app.route('/tecnico/<int:eid>/upload-bulk', methods=['GET', 'POST'])
@login_required
def tecnico_upload_bulk(eid):
    e = Equipamento.query.get_or_404(eid)
    opcoes = EquipamentoOpcao.query.filter_by(equipamento_id=eid).order_by(EquipamentoOpcao.ordem).all()
    
    if request.method == 'GET':
        outros_docs = EquipamentoDocumento.query.filter_by(
            equipamento_id=eid, componente='outros_bulk').order_by(
            EquipamentoDocumento.criado_em.desc()).all()
        app.logger.warning(f"BULK GET eid={eid} outros_docs={len(outros_docs)}")
        # Also check all docs for this equipment
        all_docs = EquipamentoDocumento.query.filter_by(equipamento_id=eid).all()
        app.logger.warning(f"ALL DOCS for eid={eid}: {[(d.id, d.componente, d.titulo) for d in all_docs]}")
        return render_template('tecnico_upload_bulk.html', equipamento=e, opcoes=opcoes, outros_docs=outros_docs)
    
    files = request.files.getlist('pdfs')
    if not files:
        flash('Nenhum ficheiro seleccionado.', 'error')
        return redirect(url_for('tecnico_upload_bulk', eid=eid))
    
    ensure_upload_dir()
    matched = 0
    unmatched = []
    
    for f in files:
        if not f or not f.filename:
            continue
        code, catalog = extrair_factory_code(f.filename)
        opcao = None
        
        if code:
            # Match against factory OR ordered columns for this equipment
            opcao = EquipamentoOpcao.query.filter_by(equipamento_id=eid).filter(
                db.or_(
                    db.func.replace(db.func.replace(EquipamentoOpcao.factory,'*',''),' ','') == code,
                    db.func.replace(db.func.replace(EquipamentoOpcao.ordered,'*',''),' ','') == code,
                    db.func.replace(db.func.replace(EquipamentoOpcao.distributor,'*',''),' ','') == code,
                )
            ).first()
        
        if opcao:
            safe = f"{eid}_opcao_{opcao.id}_{f.filename.replace(' ','_')}"
            path = os.path.join(UPLOAD_TECNICO, safe)
            f.save(path)
            opcao.pdf_filename = f.filename
            opcao.pdf_path = safe
            matched += 1
        else:
            # Save as "outros" document linked to equipment
            safe = f"outros_{eid}_{f.filename.replace(' ','_')}"
            path = os.path.join(UPLOAD_TECNICO, safe)
            f.save(path)
            app.logger.warning(f'SAVING outros_bulk doc: {f.filename} safe={safe}')
            doc = EquipamentoDocumento(
                equipamento_id=eid,
                componente='outros_bulk',
                titulo=f.filename,
                pdf_filename=f.filename,
                pdf_path=safe,
                notas=f'Código detectado: {code or "nenhum"} — sem correspondência',
            )
            db.session.add(doc)
            unmatched.append(f.filename)
    
    db.session.commit()
    
    if matched:
        flash(f'✅ {matched} PDFs associados automaticamente.', 'success')
    if unmatched:
        flash(f'⚠️ {len(unmatched)} ficheiro(s) guardados em "Outros Ficheiros".', 'warning')
    
    return redirect(url_for('tecnico_upload_bulk', eid=eid))

@app.route('/tecnico/<int:eid>/upload-bulk/assign', methods=['POST'])
@login_required
def tecnico_upload_bulk_assign(eid):
    """Manually assign an outros_bulk doc to an option code."""
    did = request.form.get('doc_id','')
    oid = request.form.get('opcao_id','')
    if not did or not oid:
        return redirect(url_for('tecnico_upload_bulk', eid=eid))
    
    doc = EquipamentoDocumento.query.get(int(did))
    opcao = EquipamentoOpcao.query.get(int(oid))
    
    if doc and opcao and opcao.equipamento_id == eid:
        opcao.pdf_filename = doc.pdf_filename
        opcao.pdf_path = doc.pdf_path
        db.session.delete(doc)
        db.session.commit()
        flash(f'PDF "{opcao.pdf_filename}" associado a "{opcao.option_name}".', 'success')
    
    return redirect(url_for('tecnico_upload_bulk', eid=eid))


@app.route('/api/tecnico/<int:eid>/outros-docs')
@login_required
def api_tecnico_outros_docs(eid):
    docs = EquipamentoDocumento.query.filter_by(
        equipamento_id=eid, componente='outros_bulk').order_by(
        EquipamentoDocumento.criado_em.asc()).all()

    # Detect duplicates by pdf_filename
    seen = {}
    for d in docs:
        key = (d.pdf_filename or '').strip().upper()
        if key not in seen:
            seen[key] = []
        seen[key].append(d.id)

    result = []
    for d in docs:
        key = (d.pdf_filename or '').strip().upper()
        is_dup = len(seen[key]) > 1
        is_kept = seen[key][0] == d.id  # keep first occurrence
        result.append({
            'id': d.id,
            'titulo': d.titulo,
            'pdf_filename': d.pdf_filename,
            'notas': d.notas or '',
            'ver_url': url_for('tecnico_doc_ver', did=d.id),
            'download_url': url_for('tecnico_doc_download', did=d.id),
            'is_duplicate': is_dup and not is_kept,
            'dup_count': len(seen[key]),
        })
    return jsonify(result)

@app.route('/api/tecnico/<int:eid>/outros-dedup', methods=['POST'])
@login_required
def api_tecnico_outros_dedup(eid):
    """Remove duplicate outros_bulk docs, keeping the first occurrence of each filename."""
    docs = EquipamentoDocumento.query.filter_by(
        equipamento_id=eid, componente='outros_bulk').order_by(
        EquipamentoDocumento.criado_em.asc()).all()

    seen = {}
    removed = 0
    for d in docs:
        key = (d.pdf_filename or '').strip().upper()
        if key in seen:
            # duplicate — delete
            try:
                path = os.path.join(UPLOAD_TECNICO, d.pdf_path) if d.pdf_path else None
                # Don't delete the file if the kept copy uses the same path
                if path and path != os.path.join(UPLOAD_TECNICO, seen[key].pdf_path or ''):
                    if os.path.exists(path): os.remove(path)
            except Exception: pass
            db.session.delete(d)
            removed += 1
        else:
            seen[key] = d

    db.session.commit()
    return jsonify({'ok': True, 'removed': removed})

@app.route('/api/tecnico/<int:eid>/associar-outro', methods=['POST'])
@login_required
def api_tecnico_associar_outro(eid):
    data = request.get_json()
    did = data.get('doc_id')
    oid = data.get('opcao_id')
    doc = EquipamentoDocumento.query.get(int(did))
    opcao = EquipamentoOpcao.query.get(int(oid))
    if not doc or not opcao or opcao.equipamento_id != eid:
        return jsonify({'ok': False, 'error': 'Não encontrado'})
    opcao.pdf_filename = doc.pdf_filename
    opcao.pdf_path = doc.pdf_path
    db.session.delete(doc)
    db.session.commit()
    return jsonify({'ok': True})


# ── CAMPOS TÉCNICOS POR MODELO ───────────────────────────────────────────────

@app.route('/api/campos-tecnicos/<tipo>/<path:modelo>')
@login_required
def api_campos_tecnicos(tipo, modelo):
    campos = CampoTecnicoModelo.query.filter_by(
        tipo_componente=tipo, modelo_codigo=modelo
    ).order_by(CampoTecnicoModelo.ordem, CampoTecnicoModelo.id).all()
    return jsonify([{
        'id': c.id, 'titulo': c.titulo, 'valor': c.valor, 'ordem': c.ordem,
        'pdf_filename': c.pdf_filename, 'pdf_path': c.pdf_path
    } for c in campos])

@app.route('/api/campos-tecnicos/<tipo>/<path:modelo>', methods=['POST'])
@login_required
def api_campos_tecnicos_criar(tipo, modelo):
    data = request.get_json()
    titulo = (data.get('titulo') or '').strip()
    valor  = (data.get('valor') or '').strip()
    if not titulo:
        return jsonify({'ok': False, 'error': 'Título obrigatório'})
    count = CampoTecnicoModelo.query.filter_by(tipo_componente=tipo, modelo_codigo=modelo).count()
    c = CampoTecnicoModelo(
        tipo_componente=tipo, modelo_codigo=modelo,
        titulo=titulo, valor=valor, ordem=count
    )
    db.session.add(c)
    db.session.commit()
    return jsonify({'ok': True, 'id': c.id, 'titulo': c.titulo, 'valor': c.valor})

@app.route('/api/campos-tecnicos/<int:cid>', methods=['PUT'])
@login_required
def api_campos_tecnicos_editar(cid):
    c = CampoTecnicoModelo.query.get_or_404(cid)
    data = request.get_json()
    c.titulo = (data.get('titulo') or c.titulo).strip()
    c.valor  = (data.get('valor')  or '').strip()
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/campos-tecnicos/<int:cid>', methods=['DELETE'])
@login_required
def api_campos_tecnicos_apagar(cid):
    c = CampoTecnicoModelo.query.get_or_404(cid)
    db.session.delete(c)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/campos-tecnicos/<int:cid>/upload', methods=['POST'])
@login_required
def api_campos_tecnicos_upload(cid):
    c = CampoTecnicoModelo.query.get_or_404(cid)
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'Nenhum ficheiro'})
    ensure_upload_dir()
    safe = f"campo_{cid}_{f.filename.replace(' ','_')}"
    path = os.path.join(UPLOAD_TECNICO, safe)
    f.save(path)
    c.pdf_filename = f.filename
    c.pdf_path = safe
    db.session.commit()
    return jsonify({'ok': True, 'filename': f.filename,
                    'ver_url': url_for('api_campos_tecnicos_ver', cid=cid),
                    'download_url': url_for('api_campos_tecnicos_download', cid=cid)})

@app.route('/api/campos-tecnicos/<int:cid>/ver')
@login_required
def api_campos_tecnicos_ver(cid):
    c = CampoTecnicoModelo.query.get_or_404(cid)
    if not c.pdf_path:
        return '', 404
    return send_from_directory(UPLOAD_TECNICO, c.pdf_path,
                               as_attachment=False, download_name=c.pdf_filename)

@app.route('/api/campos-tecnicos/<int:cid>/download')
@login_required
def api_campos_tecnicos_download(cid):
    c = CampoTecnicoModelo.query.get_or_404(cid)
    if not c.pdf_path:
        return '', 404
    return send_from_directory(UPLOAD_TECNICO, c.pdf_path,
                               as_attachment=True, download_name=c.pdf_filename)

@app.route('/api/campos-tecnicos/<int:cid>/remover-ficheiro', methods=['POST'])
@login_required
def api_campos_tecnicos_remover_ficheiro(cid):
    c = CampoTecnicoModelo.query.get_or_404(cid)
    try:
        if c.pdf_path:
            path = os.path.join(UPLOAD_TECNICO, c.pdf_path)
            if os.path.exists(path): os.remove(path)
    except Exception: pass
    c.pdf_filename = None
    c.pdf_path = None
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/tecnico/<int:eid>/sync-pdfs', methods=['POST'])
@login_required
def api_tecnico_sync_pdfs(eid):
    """Sync PDFs: fill missing + validate existing ones using catalog+code."""
    eq = Equipamento.query.get_or_404(eid)
    catalogo = (eq.catalogo or '').strip().upper()

    if not catalogo:
        return jsonify({
            'ok': False,
            'error': 'Este equipamento não tem Catálogo definido. Preencha o campo Catálogo no Editar Equipamento.'
        })

    # All equipamentos with same catalog
    cat_eq_ids = [e.id for e in Equipamento.query.filter(
        db.func.upper(Equipamento.catalogo) == catalogo).all()]

    all_opcoes = EquipamentoOpcao.query.filter_by(equipamento_id=eid).all()
    matched = 0
    corrected = 0
    missing = 0
    details_new = []
    details_fixed = []
    details_missing = []

    def get_ref(o):
        dist = (o.distributor or '').strip().replace('*','').strip()
        fac  = (o.factory    or '').strip().replace('*','').strip()
        ord_ = (o.ordered    or '').strip().replace('*','').strip()
        code = dist or fac or ord_
        typ  = 'distributor' if dist else ('factory' if fac else ('ordered' if ord_ else None))
        return code, typ

    def find_donor(o, code):
        """Find best PDF donor within same catalog."""
        for col in [EquipamentoOpcao.distributor, EquipamentoOpcao.factory, EquipamentoOpcao.ordered]:
            donor = EquipamentoOpcao.query.filter(
                EquipamentoOpcao.id != o.id,
                EquipamentoOpcao.pdf_path.isnot(None),
                EquipamentoOpcao.equipamento_id.in_(cat_eq_ids),
                db.func.replace(db.func.replace(col,'*',''),' ','') == code
            ).first()
            if donor:
                return donor
        return None

    for o in all_opcoes:
        code, typ = get_ref(o)
        if not code:
            continue

        donor = find_donor(o, code)

        def filename_matches(fname, cat, cod):
            """Check if filename contains both catalog and code."""
            if not fname: return False
            n = fname.upper()
            return cat.upper() in n and cod.upper() in n

        if not o.pdf_path:
            # Missing PDF — try to fill with donor that has correct catalog in filename
            if donor and filename_matches(donor.pdf_filename, catalogo, code):
                o.pdf_filename = donor.pdf_filename
                o.pdf_path = donor.pdf_path
                matched += 1
                details_new.append(f'✅ {o.option_name[:35]} ({typ}={code}) ← {donor.pdf_filename}')
            else:
                missing += 1
                details_missing.append(f'❌ {o.option_name[:35]} ({typ}={code})')
        else:
            # Has PDF — verify filename contains correct catalog AND code
            if not filename_matches(o.pdf_filename, catalogo, code):
                # Wrong — remove the PDF association
                old_name = o.pdf_filename
                o.pdf_filename = None
                o.pdf_path = None
                corrected += 1
                details_fixed.append(
                    f'🗑 {o.option_name[:30]} ({typ}={code}): '
                    f'"{old_name}" removido — não contém catálogo {catalogo} + código {code}'
                )
                # Immediately try to find a correct replacement
                if donor and filename_matches(donor.pdf_filename, catalogo, code):
                    o.pdf_filename = donor.pdf_filename
                    o.pdf_path = donor.pdf_path
                    details_fixed[-1] += f' → substituído por {donor.pdf_filename}'

    db.session.commit()

    return jsonify({
        'ok': True,
        'catalogo': catalogo,
        'matched': matched,
        'corrected': corrected,
        'missing': missing,
        'total': len(all_opcoes),
        'details_new': details_new,
        'details_fixed': details_fixed,
        'details_missing': details_missing,
    })


# ── GESTÃO DE PERFIS ─────────────────────────────────────────────────────────

@app.route('/admin/perfis')
@login_required
def admin_perfis():
    if not current_user.is_admin:
        flash('Sem permissão.','error'); return redirect(url_for('dashboard'))
    perfis = Perfil.query.order_by(Perfil.nome).all()
    return render_template('admin_perfis.html', perfis=perfis, menus=MENUS_DISPONIVEIS)

@app.route('/admin/perfis/novo', methods=['POST'])
@login_required
def admin_perfis_novo():
    if not current_user.is_admin: return redirect(url_for('dashboard'))
    nome = request.form.get('nome','').strip()
    if not nome or Perfil.query.filter_by(nome=nome).first():
        flash('Nome inválido ou já existe.','error')
        return redirect(url_for('admin_perfis'))
    menus = request.form.getlist('menus')
    p = Perfil(nome=nome, descricao=request.form.get('descricao','').strip())
    p.set_menus(menus)
    db.session.add(p)
    db.session.commit()
    flash(f'Perfil "{nome}" criado.','success')
    return redirect(url_for('admin_perfis'))

@app.route('/admin/perfis/<int:pid>/editar', methods=['POST'])
@login_required
def admin_perfis_editar(pid):
    if not current_user.is_admin: return redirect(url_for('dashboard'))
    p = Perfil.query.get_or_404(pid)
    p.nome = request.form.get('nome', p.nome).strip()
    p.descricao = request.form.get('descricao','').strip()
    p.set_menus(request.form.getlist('menus'))
    db.session.commit()
    flash(f'Perfil "{p.nome}" actualizado.','success')
    return redirect(url_for('admin_perfis'))

@app.route('/admin/perfis/<int:pid>/apagar', methods=['POST'])
@login_required
def admin_perfis_apagar(pid):
    if not current_user.is_admin: return redirect(url_for('dashboard'))
    p = Perfil.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    flash('Perfil removido.','info')
    return redirect(url_for('admin_perfis'))

@app.route('/admin/utilizadores/<int:uid>/perfil', methods=['POST'])
@login_required
def admin_utilizador_perfil(uid):
    if not current_user.is_admin: return redirect(url_for('dashboard'))
    u = User.query.get_or_404(uid)
    perfil_id = request.form.get('perfil_id','')
    # Store perfil_id in departamento field as JSON for now
    import json as _json
    try:
        meta = _json.loads(u.departamento or '{}')
    except:
        meta = {'dept': u.departamento or ''}
    meta['perfil_id'] = int(perfil_id) if perfil_id else None
    u.departamento = _json.dumps(meta)
    u.is_admin = request.form.get('is_admin') == 'on'
    db.session.commit()
    flash(f'Utilizador {u.nome} actualizado.','success')
    return redirect(url_for('admin_utilizadores'))

def get_user_perfil_id(user):
    """Get perfil_id from user.departamento JSON."""
    try:
        meta = json.loads(user.departamento or '{}')
        return meta.get('perfil_id')
    except:
        return None


# ── REGISTO DE UTILIZADORES ───────────────────────────────────────────────────

@app.route('/registo', methods=['GET', 'POST'])
def registo():
    if request.method == 'POST':
        nome     = request.form.get('nome','').strip()
        username = request.form.get('username','').strip()
        email    = request.form.get('email','').strip()
        password = request.form.get('password','')

        if not all([nome, username, email, password]):
            flash('Preencha todos os campos.', 'error')
            return redirect(url_for('registo'))

        if User.query.filter_by(username=username).first():
            flash('Username já existe.', 'error')
            return redirect(url_for('registo'))

        if RegistoPendente.query.filter_by(username=username, estado='pendente').first():
            flash('Já existe um pedido de registo pendente para este username.', 'error')
            return redirect(url_for('registo'))

        r = RegistoPendente(
            nome=nome, username=username, email=email,
            password_hash=generate_password_hash(password)
        )
        db.session.add(r)
        db.session.commit()

        # Notify admins by email if SMTP configured
        try:
            _notificar_admins_registo(nome, username, email)
        except Exception as ex:
            app.logger.warning(f"Email notify error: {ex}")

        return redirect(url_for('registo_aguarda'))
    return render_template('registo.html', cfg=ConfigGeral.query.first())

@app.route('/registo/aguarda')
def registo_aguarda():
    return render_template('registo_aguarda.html', cfg=ConfigGeral.query.first())

@app.route('/admin/registos-pendentes')
@login_required
def admin_registos_pendentes():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    pendentes = RegistoPendente.query.filter_by(estado='pendente').order_by(RegistoPendente.criado_em).all()
    return render_template('admin_registos.html', pendentes=pendentes)

@app.route('/admin/registos/<int:rid>/aceitar', methods=['POST'])
@login_required
def admin_registo_aceitar(rid):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    r = RegistoPendente.query.get_or_404(rid)
    if User.query.filter_by(username=r.username).first():
        flash('Username já existe.', 'error')
        return redirect(url_for('admin_registos_pendentes'))
    u = User(nome=r.nome, username=r.username,
              password_hash=r.password_hash, is_admin=False,
              email=r.email, must_change_password=False)
    db.session.add(u)
    r.estado = 'aceite'
    db.session.commit()
    try:
        _enviar_email_aprovacao(r.email, r.nome, r.username)
        flash(f'Utilizador {r.nome} aprovado. Email enviado para {r.email}.', 'success')
    except Exception as ex:
        app.logger.warning(f"Email aprovacao error: {ex}")
        flash(f'Utilizador {r.nome} aprovado. ⚠️ Email não enviado: {ex}', 'warning')
    return redirect(url_for('admin_registos_pendentes'))

@app.route('/admin/registos/<int:rid>/recusar', methods=['POST'])
@login_required
def admin_registo_recusar(rid):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    r = RegistoPendente.query.get_or_404(rid)
    r.estado = 'recusado'
    db.session.commit()
    flash(f'Pedido de {r.nome} recusado.', 'info')
    return redirect(url_for('admin_registos_pendentes'))

def _notificar_admins_registo(nome, username, email):
    """Send email to all admins about new registration request."""
    admins = User.query.filter_by(is_admin=True).all()
    cfg = ConfigGeral.query.first()
    smtp_host = getattr(cfg, 'smtp_host', None) if cfg else None
    if not smtp_host:
        return
    import smtplib
    from email.mime.text import MIMEText
    msg = MIMEText(f"""
Novo pedido de registo na plataforma:

Nome: {nome}
Username: {username}
Email: {email}

Aceda ao menu Admin → Registos Pendentes para aprovar ou recusar.
""")
    msg['Subject'] = f'[NavTech] Novo pedido de registo: {nome}'
    msg['From'] = getattr(cfg, 'smtp_from', 'noreply@navtech.pt')
    msg['To'] = ', '.join([a.username for a in admins if '@' in (a.username or '')])
    if not msg['To']:
        return
    with smtplib.SMTP(smtp_host, getattr(cfg, 'smtp_port', 587)) as s:
        s.starttls()
        if getattr(cfg, 'smtp_user', None):
            s.login(cfg.smtp_user, cfg.smtp_pass or '')
        s.send_message(msg)

def _enviar_email_aprovacao(email, nome, username):
    """Send approval email to new user."""
    cfg = ConfigGeral.query.first()
    smtp_host = getattr(cfg, 'smtp_host', None) if cfg else None
    if not smtp_host or not email:
        return
    import smtplib
    from email.mime.text import MIMEText
    empresa = getattr(cfg, 'empresa_nome', 'NavTech') if cfg else 'NavTech'
    msg = MIMEText(f"""
Olá {nome},

O seu registo na plataforma {empresa} foi aprovado.

Pode agora aceder com o username: {username}

Bem-vindo(a)!
""")
    msg['Subject'] = f'[{empresa}] Acesso aprovado'
    msg['From'] = getattr(cfg, 'smtp_from', 'noreply@navtech.pt')
    msg['To'] = email
    with smtplib.SMTP(smtp_host, getattr(cfg, 'smtp_port', 587)) as s:
        s.starttls()
        if getattr(cfg, 'smtp_user', None):
            s.login(cfg.smtp_user, cfg.smtp_pass or '')
        s.send_message(msg)


@app.route('/api/admin/registos-count')
@login_required
def api_admin_registos_count():
    if not current_user.is_admin:
        return jsonify({'count': 0})
    return jsonify({'count': RegistoPendente.query.filter_by(estado='pendente').count()})


@app.route('/api/admin/testar-smtp', methods=['POST'])
@login_required
def api_testar_smtp():
    if not current_user.is_admin:
        return jsonify({'ok': False, 'error': 'Sem permissão'})
    cfg = ConfigGeral.query.first()
    if not cfg or not cfg.smtp_host:
        return jsonify({'ok': False, 'error': 'SMTP não configurado. Guarde primeiro.'})
    try:
        import smtplib
        from email.mime.text import MIMEText
        msg = MIMEText('Teste de email da plataforma NavTech.')
        msg['Subject'] = '[NavTech] Teste de email'
        msg['From'] = cfg.smtp_from or cfg.smtp_user
        msg['To'] = cfg.smtp_user
        port = cfg.smtp_port or 587
        if port == 465:
            # SSL directo
            import ssl
            ctx = ssl.create_default_context()
            with smtplib.SMTP_SSL(cfg.smtp_host, port, timeout=15, context=ctx) as s:
                if cfg.smtp_user and cfg.smtp_pass:
                    s.login(cfg.smtp_user, cfg.smtp_pass)
                s.send_message(msg)
        else:
            # STARTTLS (587 ou 25)
            with smtplib.SMTP(cfg.smtp_host, port, timeout=15) as s:
                s.ehlo()
                if cfg.smtp_tls:
                    s.starttls()
                    s.ehlo()
                if cfg.smtp_user and cfg.smtp_pass:
                    s.login(cfg.smtp_user, cfg.smtp_pass)
                s.send_message(msg)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


import secrets
import string

def gerar_password_aleatoria(length=10):
    chars = string.ascii_letters + string.digits + '!@#$%'
    return ''.join(secrets.choice(chars) for _ in range(length))

def enviar_email(para, assunto, corpo):
    """Generic email sender using SMTP config."""
    cfg = ConfigGeral.query.first()
    if not cfg or not getattr(cfg, 'smtp_host', None):
        raise Exception('SMTP não configurado')
    import smtplib
    from email.mime.text import MIMEText
    msg = MIMEText(corpo, 'plain', 'utf-8')
    msg['Subject'] = assunto
    msg['From'] = getattr(cfg, 'smtp_from', None) or getattr(cfg, 'smtp_user', '')
    msg['To'] = para
    port = getattr(cfg, 'smtp_port', 587) or 587
    if port == 465:
        import ssl
        with smtplib.SMTP_SSL(cfg.smtp_host, port, timeout=15, context=ssl.create_default_context()) as s:
            if cfg.smtp_user and cfg.smtp_pass:
                s.login(cfg.smtp_user, cfg.smtp_pass)
            s.send_message(msg)
    else:
        with smtplib.SMTP(cfg.smtp_host, port, timeout=15) as s:
            s.ehlo()
            if getattr(cfg, 'smtp_tls', 1): s.starttls(); s.ehlo()
            if cfg.smtp_user and cfg.smtp_pass:
                s.login(cfg.smtp_user, cfg.smtp_pass)
            s.send_message(msg)

@app.route('/recuperar-password', methods=['GET', 'POST'])
def recuperar_password():
    cfg = ConfigGeral.query.first()
    if request.method == 'POST':
        email_ou_user = request.form.get('email_ou_user','').strip()
        u = User.query.filter(
            (User.email == email_ou_user) | (User.username == email_ou_user)
        ).first()
        if u and u.email:
            nova = gerar_password_aleatoria()
            u.password_hash = generate_password_hash(nova)
            u.must_change_password = True
            db.session.commit()
            try:
                empresa = cfg.empresa_nome if cfg and cfg.empresa_nome else 'NavTech'
                enviar_email(u.email,
                    f'[{empresa}] Reset de Password',
                    f'Olá {u.nome},\n\nA sua password foi redefinida.\n\nPassword temporária: {nova}\n\nAo fazer login será pedido que defina uma nova password.\n\nSe não pediu este reset, contacte o administrador.')
                flash('Email enviado com a nova password temporária.', 'success')
            except Exception as ex:
                flash(f'Password alterada mas email não enviado: {ex}', 'warning')
        else:
            flash('Email ou utilizador não encontrado.', 'error')
        return redirect(url_for('login'))
    return render_template('recuperar_password.html', cfg=cfg)

@app.route('/alterar-password', methods=['GET', 'POST'])
@login_required
def alterar_password():
    if request.method == 'POST':
        atual = request.form.get('atual','')
        nova  = request.form.get('nova','').strip()
        conf  = request.form.get('confirmar','').strip()
        if not check_password_hash(current_user.password_hash, atual):
            flash('Password actual incorrecta.', 'error')
            return redirect(url_for('alterar_password'))
        if nova != conf:
            flash('As passwords não coincidem.', 'error')
            return redirect(url_for('alterar_password'))
        if len(nova) < 6:
            flash('A password deve ter pelo menos 6 caracteres.', 'error')
            return redirect(url_for('alterar_password'))
        current_user.password_hash = generate_password_hash(nova)
        current_user.must_change_password = False
        db.session.commit()
        flash('Password alterada com sucesso.', 'success')
        return redirect(url_for('dashboard'))
    return render_template('alterar_password.html')

@app.route('/admin/utilizadores/<int:uid>/reset-password', methods=['POST'])
@login_required
def admin_reset_password(uid):
    if not current_user.is_admin: return redirect(url_for('dashboard'))
    u = User.query.get_or_404(uid)
    nova = gerar_password_aleatoria()
    u.password_hash = generate_password_hash(nova)
    u.must_change_password = True
    db.session.commit()
    if u.email:
        try:
            cfg = ConfigGeral.query.first()
            empresa = cfg.empresa_nome if cfg else 'NavTech'
            enviar_email(u.email,
                f'[{empresa}] Reset de Password',
                f'Olá {u.nome},\n\nA sua password foi redefinida pelo administrador.\n\nPassword temporária: {nova}\n\nAo fazer login será pedido que defina uma nova password.')
            flash(f'Password de {u.nome} redefinida e email enviado para {u.email}.', 'success')
        except Exception as ex:
            flash(f'Password redefinida: {nova} (email não enviado: {ex})', 'warning')
    else:
        flash(f'Password de {u.nome} redefinida: {nova} (sem email registado)', 'warning')
    return redirect(url_for('admin_utilizadores'))

@app.route('/admin/utilizadores/<int:uid>/editar-email', methods=['POST'])
@login_required
def admin_editar_email(uid):
    if not current_user.is_admin: return redirect(url_for('dashboard'))
    u = User.query.get_or_404(uid)
    u.email = request.form.get('email','').strip()
    db.session.commit()
    flash(f'Email de {u.nome} actualizado.', 'success')
    return redirect(url_for('admin_utilizadores'))


# ── FUNCIONÁRIOS ROUTES ───────────────────────────────────────────────────────

@app.route('/funcionarios')
@login_required
def funcionarios():
    q      = request.args.get('q','').strip()
    filtro = request.args.get('filtro','todos')
    cat    = request.args.get('categoria','')
    page   = int(request.args.get('page', 1))
    per    = 20

    query = Funcionario.query
    if q:
        query = query.filter(db.or_(
            Funcionario.nome.ilike(f'%{q}%'),
            Funcionario.numero.ilike(f'%{q}%'),
        ))
    if filtro == 'ativo':   query = query.filter_by(ativo=True)
    if filtro == 'inativo': query = query.filter_by(ativo=False)
    if cat: query = query.filter_by(categoria=cat)

    total = query.count()
    items = query.order_by(Funcionario.numero).offset((page-1)*per).limit(per).all()
    categorias = [r[0] for r in db.session.query(Funcionario.categoria).distinct() if r[0]]

    return render_template('funcionarios.html',
        funcionarios=items, total=total, page=page, per=per,
        q=q, filtro=filtro, categoria=cat, categorias=categorias,
        pages=((total-1)//per+1) if total else 1)

@app.route('/funcionarios/novo', methods=['GET','POST'])
@login_required
def funcionario_novo():
    if request.method == 'POST':
        from datetime import date
        def d(f): return request.form.get(f,'').strip() or None
        def di(f):
            v = d(f)
            try: return date.fromisoformat(v) if v else None
            except: return None

        # Auto-generate numero if empty
        num = d('numero')
        if not num:
            last = Funcionario.query.order_by(Funcionario.id.desc()).first()
            num = str((int(last.numero) + 1) if last and last.numero.isdigit() else 1).zfill(4)

        f = Funcionario(
            numero=num, nome=d('nome'), categoria=d('categoria'),
            ativo=request.form.get('ativo')=='1',
            data_nascimento=di('data_nascimento'),
            morada=d('morada'), nif=d('nif'),
            num_cc=d('num_cc'), cc_validade=datetime.strptime(d('cc_validade'), '%Y-%m-%d').date() if d('cc_validade') else None,
            num_passaporte=d('num_passaporte'), passaporte_validade=datetime.strptime(d('passaporte_validade'), '%Y-%m-%d').date() if d('passaporte_validade') else None,
            agregado_familiar=int(request.form.get('agregado_familiar',0) or 0),
            telemovel=d('telemovel'), email=d('email'),
            contacto_emergencia=d('contacto_emergencia'),
            nome_emergencia=d('nome_emergencia'),
            iban=d('iban'), num_seg_social=d('num_seg_social'),
            seguro_companhia=d('seguro_companhia'),
            seguro_apolice=d('seguro_apolice'), notas=d('notas'),
        )
        db.session.add(f)
        db.session.flush()
        _rh_upload_docs(f.id, request.files)
        db.session.commit()
        flash(f'Funcionário {f.nome} criado.', 'success')
        return redirect(url_for('funcionario_detalhe', fid=f.id))
    from datetime import date as _d
    return render_template('funcionario_form.html', f=None, hoje=_d.today())

@app.route('/funcionarios/<int:fid>')
@login_required
def funcionario_detalhe(fid):
    f = Funcionario.query.get_or_404(fid)
    return render_template('funcionario_detalhe.html', f=f)

@app.route('/funcionarios/<int:fid>/editar', methods=['GET','POST'])
@login_required
def funcionario_editar(fid):
    func = Funcionario.query.get_or_404(fid)
    if request.method == 'POST':
        from datetime import date
        def d(field): return request.form.get(field,'').strip() or None
        def di(field):
            v = d(field)
            try: return date.fromisoformat(v) if v else None
            except: return None
        # Update numero if changed and not duplicate
        novo_numero = d('numero')
        if novo_numero and novo_numero != func.numero:
            existente = Funcionario.query.filter_by(numero=novo_numero).first()
            if existente and existente.id != fid:
                flash(f'Nº {novo_numero} já está atribuído a {existente.nome}.', 'error')
                return redirect(url_for('funcionario_editar', fid=fid))
            func.numero = novo_numero
        func.nome=d('nome'); func.categoria=d('categoria')
        func.ativo=request.form.get('ativo')=='1'
        func.data_nascimento=di('data_nascimento')
        func.data_admissao=di('data_admissao')
        func.morada=d('morada'); func.nif=d('nif')
        func.num_cc=d('num_cc')
        func.cc_validade=datetime.strptime(d('cc_validade'),'%Y-%m-%d').date() if d('cc_validade') else None
        func.num_passaporte=d('num_passaporte')
        func.passaporte_validade=datetime.strptime(d('passaporte_validade'),'%Y-%m-%d').date() if d('passaporte_validade') else None
        func.filiacao_pai=d('filiacao_pai'); func.filiacao_mae=d('filiacao_mae')
        func.situacao_militar=d('situacao_militar'); func.estado_civil=d('estado_civil')
        func.conjuge=d('conjuge')
        func.titulares_rendimento=int(request.form.get('titulares_rendimento',1) or 1)
        func.num_dependentes=int(request.form.get('num_dependentes',0) or 0)
        func.natural_freguesia=d('natural_freguesia'); func.natural_concelho=d('natural_concelho')
        func.socio_numero=d('socio_numero'); func.sindicato=d('sindicato')
        func.carta_conducao=request.form.get('carta_conducao')=='1'
        func.agregado_familiar=int(request.form.get('agregado_familiar',0) or 0)
        func.telemovel=d('telemovel'); func.email=d('email')
        func.contacto_emergencia=d('contacto_emergencia')
        func.nome_emergencia=d('nome_emergencia')
        func.iban=d('iban'); func.num_seg_social=d('num_seg_social')
        func.seguro_companhia=d('seguro_companhia')
        func.seguro_apolice=d('seguro_apolice')
        func.notas=d('notas'); func.obs=d('obs')
        _rh_upload_docs(fid, request.files)
        db.session.commit()
        flash('Dados actualizados.', 'success')
        return redirect(url_for('funcionario_detalhe', fid=fid))
    from datetime import date as _d
    return render_template('funcionario_form.html', f=func, hoje=_d.today())

@app.route('/funcionarios/<int:fid>/apagar', methods=['POST'])
@login_required
def funcionario_apagar(fid):
    f = Funcionario.query.get_or_404(fid)
    db.session.delete(f)
    db.session.commit()
    flash('Funcionário eliminado.', 'info')
    return redirect(url_for('funcionarios'))

@app.route('/funcionarios/<int:fid>/toggle-ativo', methods=['POST'])
@login_required
def funcionario_toggle_ativo(fid):
    f = Funcionario.query.get_or_404(fid)
    f.ativo = not f.ativo
    db.session.commit()
    return jsonify({'ok': True, 'ativo': f.ativo})

@app.route('/funcionarios/<int:fid>/doc/upload', methods=['POST'])
@login_required
def funcionario_doc_upload(fid):
    Funcionario.query.get_or_404(fid)
    file = request.files.get('file')
    tipo = request.form.get('tipo','outro')
    titulo = request.form.get('titulo','').strip()
    if not file or not titulo:
        flash('Título e ficheiro obrigatórios.', 'error')
        return redirect(url_for('funcionario_detalhe', fid=fid))
    safe = f"rh_{fid}_{tipo}_{file.filename.replace(' ','_')}"
    file.save(os.path.join(UPLOAD_RH, safe))
    db.session.add(FuncionarioDocumento(
        funcionario_id=fid, tipo=tipo, titulo=titulo,
        pdf_filename=file.filename, pdf_path=safe))
    db.session.commit()
    flash('Documento adicionado.', 'success')
    return redirect(url_for('funcionario_detalhe', fid=fid))

@app.route('/funcionarios/doc/<int:did>/ver')
@login_required
def funcionario_doc_ver(did):
    d = FuncionarioDocumento.query.get_or_404(did)
    return send_from_directory(UPLOAD_RH, d.pdf_path, as_attachment=False, download_name=d.pdf_filename)

@app.route('/funcionarios/doc/<int:did>/download')
@login_required
def funcionario_doc_download(did):
    d = FuncionarioDocumento.query.get_or_404(did)
    return send_from_directory(UPLOAD_RH, d.pdf_path, as_attachment=True, download_name=d.pdf_filename)

@app.route('/funcionarios/doc/<int:did>/apagar', methods=['POST'])
@login_required
def funcionario_doc_apagar(did):
    d = FuncionarioDocumento.query.get_or_404(did)
    fid = d.funcionario_id
    try:
        p = os.path.join(UPLOAD_RH, d.pdf_path)
        if os.path.exists(p): os.remove(p)
    except: pass
    db.session.delete(d)
    db.session.commit()
    return redirect(url_for('funcionario_detalhe', fid=fid))

@app.route('/funcionarios/<int:fid>/formacao/adicionar', methods=['POST'])
@login_required
def funcionario_formacao_add(fid):
    Funcionario.query.get_or_404(fid)
    from datetime import date
    def d(f): return request.form.get(f,'').strip() or None
    def di(f):
        v = d(f)
        try: return date.fromisoformat(v) if v else None
        except: return None
    fm = FuncionarioFormacao(
        funcionario_id=fid, titulo=d('titulo'), entidade=d('entidade'),
        data_inicio=di('data_inicio'), data_fim=di('data_fim'),
        horas=int(d('horas') or 0) if d('horas') else None)
    db.session.add(fm)
    db.session.flush()
    cert = request.files.get('certificado')
    if cert and cert.filename:
        safe = f"rh_{fid}_form_{fm.id}_{cert.filename.replace(' ','_')}"
        cert.save(os.path.join(UPLOAD_RH, safe))
        fm.pdf_filename = cert.filename
        fm.pdf_path = safe
    db.session.commit()
    flash('Formação adicionada.', 'success')
    return redirect(url_for('funcionario_detalhe', fid=fid))

@app.route('/funcionarios/formacao/<int:fmid>/apagar', methods=['POST'])
@login_required
def funcionario_formacao_apagar(fmid):
    fm = FuncionarioFormacao.query.get_or_404(fmid)
    fid = fm.funcionario_id
    try:
        if fm.pdf_path:
            p = os.path.join(UPLOAD_RH, fm.pdf_path)
            if os.path.exists(p): os.remove(p)
    except: pass
    db.session.delete(fm)
    db.session.commit()
    return redirect(url_for('funcionario_detalhe', fid=fid))

@app.route('/funcionarios/formacao/<int:fmid>/certificado')
@login_required
def funcionario_formacao_cert(fmid):
    fm = FuncionarioFormacao.query.get_or_404(fmid)
    return send_from_directory(UPLOAD_RH, fm.pdf_path, as_attachment=False, download_name=fm.pdf_filename)

@app.route('/funcionarios/exportar-excel')
@login_required
def funcionarios_exportar_excel():
    import openpyxl
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Funcionários'
    headers = ['Nº','Nome','Categoria','Ativo','Data Nasc.','Idade','NIF','Nº CC','Nº Passaporte',
               'Agregado Familiar','Telemóvel','Email','NIF','IBAN','Nº Seg. Social',
               'Seguro Companhia','Seguro Apólice','Morada']
    ws.append(headers)
    for f in Funcionario.query.order_by(Funcionario.numero).all():
        ws.append([f.numero, f.nome, f.categoria or '',
                   'Sim' if f.ativo else 'Não',
                   f.data_nascimento.strftime('%d/%m/%Y') if f.data_nascimento else '',
                   f.idade or '', f.nif or '', f.num_cc or '',
                   f.num_passaporte or '', f.agregado_familiar or 0,
                   f.telemovel or '', f.email or '', f.nif or '',
                   f.iban or '', f.num_seg_social or '',
                   f.seguro_companhia or '', f.seguro_apolice or '',
                   f.morada or ''])
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import send_file
    return send_file(buf, as_attachment=True, download_name='funcionarios.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def _rh_upload_docs(fid, files):
    """Upload inline document files (cc, passaporte, morada) from form."""
    for tipo in ['cc','passaporte','morada']:
        f = files.get(f'doc_{tipo}')
        if f and f.filename:
            safe = f"rh_{fid}_{tipo}_{f.filename.replace(' ','_')}"
            f.save(os.path.join(UPLOAD_RH, safe))
            # Replace existing doc of same type
            existing = FuncionarioDocumento.query.filter_by(funcionario_id=fid, tipo=tipo).first()
            if existing:
                existing.pdf_filename = f.filename; existing.pdf_path = safe
            else:
                db.session.add(FuncionarioDocumento(
                    funcionario_id=fid, tipo=tipo,
                    titulo={'cc':'Cartão de Cidadão','passaporte':'Passaporte','morada':'Comprovativo de Morada'}[tipo],
                    data_validade=datetime.strptime(request.form.get(f'{tipo}_validade',''),'%Y-%m-%d').date() if request.form.get(f'{tipo}_validade') else None,
                    pdf_filename=f.filename, pdf_path=safe))


@app.route('/admin/perfis/sync-menus', methods=['POST'])
@login_required
def admin_perfis_sync_menus():
    """Show all available menus — doesn't change permissions, just refreshes the list."""
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    flash('Lista de menus actualizada. Edite cada perfil para atribuir os novos menus.', 'success')
    return redirect(url_for('admin_perfis'))


@app.route('/funcionarios/<int:fid>/situacao-prof/adicionar', methods=['POST'])
@login_required
def funcionario_situacao_prof_add(fid):
    if not current_user.is_admin:
        flash('Sem permissão.', 'error')
        return redirect(url_for('funcionario_detalhe', fid=fid))
    Funcionario.query.get_or_404(fid)
    from datetime import date
    def d(f): return request.form.get(f,'').strip() or None
    def di(f):
        v = d(f)
        try: return date.fromisoformat(v) if v else None
        except: return None
    def dn(f):
        v = d(f)
        if not v: return None
        try: return float(v.replace(',','.'))
        except: return None
    sp = FuncionarioSituacaoProf(
        funcionario_id=fid,
        data=di('data') or date.today(),
        categoria_prof=d('categoria_prof'),
        vencimento=dn('vencimento'),
        refeicao=dn('refeicao'),
        premios_outros=dn('premios_outros'),
        notas=d('notas'),
    )
    db.session.add(sp)
    db.session.commit()
    flash('Situação profissional adicionada.', 'success')
    return redirect(url_for('funcionario_detalhe', fid=fid) + '#situacao-prof')

@app.route('/funcionarios/situacao-prof/<int:sid>/apagar', methods=['POST'])
@login_required
def funcionario_situacao_prof_apagar(sid):
    if not current_user.is_admin:
        return redirect(url_for('funcionarios'))
    sp = FuncionarioSituacaoProf.query.get_or_404(sid)
    fid = sp.funcionario_id
    db.session.delete(sp)
    db.session.commit()
    return redirect(url_for('funcionario_detalhe', fid=fid) + '#situacao-prof')

@app.route('/funcionarios/<int:fid>/falta/gravar', methods=['POST'])
@login_required
def funcionario_falta_gravar(fid):
    Funcionario.query.get_or_404(fid)
    ano = int(request.form.get('ano', datetime.now().year))
    mes = int(request.form.get('mes', 1))
    dias = float(request.form.get('dias_falta', 0) or 0)
    horas = float(request.form.get('horas_falta', 0) or 0)
    tipo = request.form.get('tipo', 'injustificada')
    notas = request.form.get('notas', '').strip()

    existing = FuncionarioFalta.query.filter_by(
        funcionario_id=fid, ano=ano, mes=mes, tipo=tipo).first()
    if existing:
        existing.dias_falta = dias
        existing.horas_falta = horas
        existing.notas = notas
    else:
        db.session.add(FuncionarioFalta(
            funcionario_id=fid, ano=ano, mes=mes,
            dias_falta=dias, horas_falta=horas,
            tipo=tipo, notas=notas))
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/funcionarios/<int:fid>/faltas/<int:ano>')
@login_required
def api_funcionario_faltas(fid, ano):
    faltas = FuncionarioFalta.query.filter_by(funcionario_id=fid, ano=ano).all()
    return jsonify([{
        'id': f.id, 'mes': f.mes, 'dias_falta': float(f.dias_falta or 0),
        'horas_falta': float(f.horas_falta or 0), 'tipo': f.tipo, 'notas': f.notas or ''
    } for f in faltas])


# ── SALÁRIOS ROUTES ───────────────────────────────────────────────────────────



@app.route('/salarios/logout')
@login_required
def salarios_logout():
    session.pop('salarios_autorizado', None)
    session.pop('salarios_auth_time', None)
    return redirect(url_for('salarios'))


@app.route('/api/salarios/verificar-pin', methods=['POST'])
@login_required
def salarios_verificar_pin():
    data = request.get_json() or {}
    pin = data.get('pin', '')
    cfg = ConfigGeral.query.first()
    senha = (cfg.salarios_pin or '').strip() if cfg else None
    if not senha:
        return jsonify({'ok': True})  # No PIN set, allow access
    if pin == senha:
        from flask import session as _sess
        from datetime import datetime as _dt2
        _sess['salarios_autorizado'] = True
        _sess['salarios_auth_time'] = _dt2.now().isoformat()
        _sess['salarios_server_token'] = _SERVER_START_TOKEN
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'PIN incorrecto'})


@app.route('/salarios')
@login_required
def salarios():
    cfg = ConfigGeral.query.first()
    salarios_senha = (cfg.salarios_pin or '').strip() if cfg else ''
    if salarios_senha:
        authed = False
        try:
            from datetime import datetime as _dt
            auth_time = session.get('salarios_auth_time')
            auth_token = session.get('salarios_server_token')
            if (session.get('salarios_autorizado') and auth_time
                    and auth_token == _SERVER_START_TOKEN):
                elapsed = (_dt.now() - _dt.fromisoformat(auth_time)).total_seconds()
                authed = elapsed < 1800
        except Exception:
            authed = False
        if not authed:
            session.pop('salarios_autorizado', None)
            session.pop('salarios_auth_time', None)
            session.pop('salarios_server_token', None)
            return render_template('salarios_pin.html')

    funcionarios = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()
    return render_template('salarios.html', funcionarios=funcionarios)

@app.route('/salarios/processar/<int:fid>')
@login_required
def salarios_calendario(fid):
    func = Funcionario.query.get_or_404(fid)
    ano = int(request.args.get('ano', datetime.now().year))
    # Get existing recibos for this employee/year
    recibos = {r.mes: r for r in ReciboSalario.query.filter_by(
        funcionario_id=fid, ano=ano).all()}
    # Extra months defined by user
    extra_meses = []
    for m_num, m_label in MESES_LABELS.items():
        pass
    return render_template('salarios_calendario.html',
        func=func, ano=ano, recibos=recibos,
        meses_labels=MESES_LABELS)

@app.route('/salarios/recibo/<int:fid>/<int:ano>/<int:mes>', methods=['GET','POST'])
@login_required
def salario_recibo(fid, ano, mes):
    db.session.expire_all()  # force fresh read
    func = Funcionario.query.get_or_404(fid)
    recibo = ReciboSalario.query.filter_by(
        funcionario_id=fid, ano=ano, mes=mes).first()

    # Get last situacao profissional for defaults
    ultima_sp = None
    if func.situacoes_prof:
        ultima_sp = func.situacoes_prof[0]

    if request.method == 'POST':
        def n(f): 
            v = request.form.get(f,'0').strip().replace(',','.')
            try: return float(v)
            except: return 0.0

        if not recibo:
            recibo = ReciboSalario(
                funcionario_id=fid, ano=ano, mes=mes,
                mes_label=MESES_LABELS.get(mes, f'Mês {mes}'))
            db.session.add(recibo)

        recibo.vencimento_base     = n('vencimento_base')
        recibo.vencimento_base_rht = n('vencimento_base_rht')
        recibo.vencimento_base_g   = n('vencimento_base_g')
        recibo.faltas_dias         = n('faltas_dias')
        recibo.faltas_horas        = n('faltas_horas')
        recibo.horas_extra         = n('horas_extra')        # horas
        recibo.horas_extra_rht     = n('horas_extra_rht')
        recibo.sub_refeicao_dias   = n('sub_refeicao_dias')
        recibo.sub_refeicao_vdia   = n('sub_refeicao_vdia')
        recibo.subsidio_refeicao   = n('subsidio_refeicao')
        recibo.premios             = n('premios')
        recibo.outros_abonos       = n('outros_abonos')
        recibo.irs_retencao        = n('irs_retencao')
        recibo.irs_taxa            = n('irs_taxa')
        recibo.irs_parcela_abater  = n('irs_parcela_abater')
        recibo.irs_taxa_efetiva    = n('irs_taxa_efetiva')
        recibo.irs_base            = n('irs_base')
        recibo.seg_social_func     = n('seg_social_func')
        recibo.seg_social_taxa     = n('seg_social_taxa')
        recibo.seg_social_base     = n('seg_social_base')
        recibo.seg_social_emp      = n('seg_social_emp')
        recibo.outros_descontos    = n('outros_descontos')
        recibo.faltas_valor        = n('faltas_valor')
        recibo.notas               = request.form.get('notas','').strip()
        recibo.estado              = request.form.get('estado', 'rascunho')

        # Horas extra valor = horas * RHT
        hex_valor = recibo.horas_extra * recibo.horas_extra_rht if recibo.horas_extra and recibo.horas_extra_rht else n('horas_extra_valor')

        # Calculate totals - use form values if provided (from Excel H17/H24/H26)
        form_total_iliq = n('total_iliquido')
        form_total_desc = n('total_descontos')
        form_liquido    = n('liquido')
        calc_abonos  = float(recibo.vencimento_base or 0) + float(recibo.subsidio_refeicao or 0) + hex_valor + float(recibo.premios or 0) + float(recibo.outros_abonos or 0)
        calc_descontos = float(recibo.irs_retencao or 0) + float(recibo.seg_social_func or 0) + float(recibo.outros_descontos or 0) + float(recibo.faltas_valor or 0)
        recibo.total_abonos    = form_total_iliq if form_total_iliq > 0 else calc_abonos
        recibo.total_descontos = form_total_desc if form_total_desc > 0 else calc_descontos
        recibo.liquido         = form_liquido if form_liquido > 0 else recibo.total_abonos - recibo.total_descontos
        recibo.atualizado_em   = datetime.now()
        db.session.commit()

        if request.form.get('action') == 'pdf':
            return redirect(url_for('salario_recibo_pdf', rid=recibo.id))
        flash('Recibo guardado.', 'success')
        return redirect(url_for('salario_recibo', fid=fid, ano=ano, mes=mes))

    # Get faltas for this month
    faltas_mes = FuncionarioFalta.query.filter_by(
        funcionario_id=fid, ano=ano, mes=mes).all()
    dias_falta = sum(float(f.dias_falta or 0) for f in faltas_mes)

    # Get tabelas IRS
    tabelas_irs = TabelaIRS.query.filter_by(ano=ano).order_by(TabelaIRS.id.desc()).all()

    return render_template('salario_recibo.html',
        func=func, ano=ano, mes=mes,
        mes_label=MESES_LABELS.get(mes, f'Mês {mes}'),
        recibo=recibo, ultima_sp=ultima_sp,
        dias_falta=dias_falta, tabelas_irs=tabelas_irs)

@app.route('/salarios/recibo/<int:rid>/pdf')
@login_required
def salario_recibo_pdf(rid):
    """Return printable HTML — user saves as PDF via browser print dialog."""
    db.session.expire_all()  # force fresh read from DB
    # Auto-fix: if totals are 0, recalculate from components
    _recibo_fix = ReciboSalario.query.get(rid)
    if _recibo_fix and (not _recibo_fix.total_abonos or float(_recibo_fix.total_abonos) == 0):
        try:
            _hex = float(_recibo_fix.horas_extra or 0) * float(_recibo_fix.horas_extra_rht or 0)
            _recibo_fix.total_abonos = (float(_recibo_fix.vencimento_base or 0) +
                float(_recibo_fix.subsidio_refeicao or 0) + _hex +
                float(_recibo_fix.premios or 0) + float(_recibo_fix.outros_abonos or 0))
            _recibo_fix.total_descontos = (float(_recibo_fix.irs_retencao or 0) +
                float(_recibo_fix.seg_social_func or 0) +
                float(_recibo_fix.outros_descontos or 0) + float(_recibo_fix.faltas_valor or 0))
            _recibo_fix.liquido = float(_recibo_fix.total_abonos) - float(_recibo_fix.total_descontos)
            db.session.commit()
        except Exception:
            db.session.rollback()
    r = ReciboSalario.query.get_or_404(rid)
    func = Funcionario.query.get(r.funcionario_id)
    cfg = ConfigGeral.query.first()
    empresa_nome = cfg.empresa_nome if cfg else 'União Construtora Naval Limitada'
    upload_url = request.host_url.rstrip('/') + '/uploads'
    mes_label = MESES_LABELS.get(r.mes, r.mes_label or f'Mês {r.mes}')
    html = render_template('salario_recibo_print.html',
        recibo=r, func=func, ano=r.ano,
        mes_label=mes_label,
        cfg=cfg, empresa_nome=empresa_nome,
        upload_url=upload_url)
    from flask import make_response
    resp = make_response(html)
    resp.headers['Content-Type'] = 'text/html; charset=utf-8'
    return resp

@app.route('/salarios/recibo/<int:rid>/email', methods=['POST'])
@login_required
def salario_recibo_email(rid):
    r = ReciboSalario.query.get_or_404(rid)
    func = Funcionario.query.get(r.funcionario_id)
    if not func.email:
        return jsonify({'ok': False, 'error': 'Funcionario sem email'})
    try:
        cfg = ConfigGeral.query.first()
        empresa = cfg.empresa_nome if cfg else 'Empresa'
        mes_label = MESES_LABELS.get(r.mes, f'Mes {r.mes}')
        import smtplib
        from email.message import EmailMessage
        msg = EmailMessage()
        msg['Subject'] = f'[{empresa}] Recibo Salario {mes_label} {r.ano}'
        msg['From'] = getattr(cfg,'smtp_from','') or getattr(cfg,'smtp_user','')
        msg['To'] = func.email
        msg.set_content(
            f'Ola {func.nome},\n\n'
            f'Segue em anexo o recibo de salario referente a {mes_label} de {r.ano}.\n\n'
            f'Com os melhores cumprimentos,\n{empresa}',
            charset='utf-8'
        )
        try:
            html_body = render_template('salario_recibo_print.html',
                recibo=r, func=func, ano=r.ano, mes_label=mes_label,
                cfg=cfg, empresa_nome=empresa,
                upload_url=request.host_url.rstrip('/') + '/uploads')
            html_clean = html_body.replace('window.onload', '//window.onload')
            fname = f'Recibo_{func.numero}_{r.ano}_{r.mes:02d}'
            pdf_done = False
            for wk_path in [
                r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe',
                r'C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe',
            ]:
                if os.path.exists(wk_path):
                    try:
                        import pdfkit
                        pdf_file = os.path.join(UPLOAD_SALARIOS, fname + '.pdf')
                        wk_opts = {'page-size':'A4','margin-top':'12mm','margin-bottom':'12mm',
                                   'margin-left':'12mm','margin-right':'12mm','encoding':'UTF-8','quiet':''}
                        pdfkit.from_string(html_clean, pdf_file, options=wk_opts,
                                           configuration=pdfkit.configuration(wkhtmltopdf=wk_path))
                        with open(pdf_file,'rb') as pf:
                            msg.add_attachment(pf.read(), maintype='application', subtype='pdf',
                                               filename=fname+'.pdf')
                        pdf_done = True
                    except Exception as ex_wk:
                        app.logger.warning(f'wkhtmltopdf: {ex_wk}')
                    break
            if not pdf_done:
                msg.add_attachment(html_clean.encode('utf-8'), maintype='text', subtype='html',
                                   filename=fname+'.html')
        except Exception as ex_a:
            app.logger.warning(f'Attach error: {ex_a}')
        port = getattr(cfg,'smtp_port',587) or 587
        if port == 465:
            import ssl
            with smtplib.SMTP_SSL(cfg.smtp_host, port, timeout=15,
                                   context=ssl.create_default_context()) as s:
                if cfg.smtp_user: s.login(cfg.smtp_user, cfg.smtp_pass or '')
                s.send_message(msg)
        else:
            with smtplib.SMTP(cfg.smtp_host, port, timeout=15) as s:
                s.ehlo()
                if getattr(cfg,'smtp_tls',1): s.starttls(); s.ehlo()
                if cfg.smtp_user: s.login(cfg.smtp_user, cfg.smtp_pass or '')
                s.send_message(msg)
        return jsonify({'ok': True})
    except Exception as e:
        import traceback
        app.logger.error(f'Email error: {traceback.format_exc()}')
        return jsonify({'ok': False, 'error': str(e)})


# Documentos Contabilísticos
@app.route('/salarios/documentos', methods=['GET','POST'])
@login_required
def salarios_documentos():
    if request.method == 'POST':
        f = request.files.get('file')
        titulo = request.form.get('titulo','').strip()
        tipo = request.form.get('tipo','outro')
        ano = request.form.get('ano','')
        if f and titulo:
            safe = f"contab_{tipo}_{ano}_{f.filename.replace(' ','_')}"
            f.save(os.path.join(UPLOAD_SALARIOS, safe))
            doc = DocContabilistico(titulo=titulo, tipo=tipo,
                ano=int(ano) if ano.isdigit() else None,
                pdf_filename=f.filename, pdf_path=safe)
            db.session.add(doc)
            if tipo == 'tabela_irs' and ano.isdigit():
                db.session.add(TabelaIRS(ano=int(ano), descricao=titulo,
                    pdf_filename=f.filename, pdf_path=safe))
            db.session.commit()
            flash('Documento guardado.', 'success')
        return redirect(url_for('salarios_documentos'))
    docs = DocContabilistico.query.order_by(DocContabilistico.criado_em.desc()).all()
    return render_template('salarios_documentos.html', docs=docs)

@app.route('/salarios/documentos/<int:did>/ver')
@login_required
def salario_doc_ver(did):
    d = DocContabilistico.query.get_or_404(did)
    return send_from_directory(UPLOAD_SALARIOS, d.pdf_path, as_attachment=False, download_name=d.pdf_filename)

@app.route('/salarios/documentos/<int:did>/apagar', methods=['POST'])
@login_required
def salario_doc_apagar(did):
    d = DocContabilistico.query.get_or_404(did)
    try:
        p = os.path.join(UPLOAD_SALARIOS, d.pdf_path)
        if os.path.exists(p): os.remove(p)
    except: pass
    db.session.delete(d)
    db.session.commit()
    return redirect(url_for('salarios_documentos'))

@app.route('/salarios/historico')
@login_required
def salarios_historico():
    ano = int(request.args.get('ano', datetime.now().year))
    fid = request.args.get('fid','')
    q = ReciboSalario.query.filter_by(ano=ano)
    if fid: q = q.filter_by(funcionario_id=int(fid))
    recibos = q.order_by(ReciboSalario.funcionario_id, ReciboSalario.mes).all()
    funcs = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()
    return render_template('salarios_historico.html', recibos=recibos, ano=ano, funcs=funcs, fid=fid)


@app.route('/salarios/recibo/<int:rid>/excel')
@login_required
def salario_recibo_excel(rid):
    """Download salary slip as Excel."""
    from io import BytesIO
    from flask import send_file
    r = ReciboSalario.query.get_or_404(rid)
    func = Funcionario.query.get(r.funcionario_id)
    cfg = ConfigGeral.query.first()
    empresa = cfg.empresa_nome if cfg else 'União Construtora Naval Limitada'
    mes_label = MESES_LABELS.get(r.mes, r.mes_label or f'Mês {r.mes}')

    import importlib.util, sys
    spec = importlib.util.spec_from_file_location("gerar_recibo",
        os.path.join(os.path.dirname(__file__), 'gerar_recibo.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    wb = mod.gerar_recibo_excel(func, r, mes_label, r.ano, empresa)
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"Recibo_{func.numero}_{r.ano}_{r.mes:02d}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/salarios/recibo/<int:rid>/apagar', methods=['POST'])
@login_required
def salario_recibo_apagar(rid):
    r = ReciboSalario.query.get_or_404(rid)
    fid = r.funcionario_id
    ano = r.ano
    try:
        if r.pdf_path:
            p = os.path.join(UPLOAD_SALARIOS, r.pdf_path)
            if os.path.exists(p): os.remove(p)
    except: pass
    db.session.delete(r)
    db.session.commit()
    flash('Recibo eliminado.', 'info')
    return redirect(url_for('salarios_calendario', fid=fid, ano=ano))


# ── IMPORTAÇÃO RECIBOS EXCEL ─────────────────────────────────────────────────

@app.route('/salarios/importar', methods=['GET', 'POST'])
@login_required
def salarios_importar():
    if request.method == 'POST':
        f = request.files.get('excel')
        if not f or not f.filename:
            flash('Seleccione um ficheiro Excel.', 'error')
            return redirect(url_for('salarios_importar'))
        # Save temp file
        import tempfile, os as _os
        ext = _os.path.splitext(f.filename)[1].lower() or '.xls'
        tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
        f.save(tmp.name); tmp.close()
        try:
            resultado = _parse_recibos_excel(tmp.name)
            # Strip non-serializable objects before storing in session
            resultado_safe = []
            for rec in resultado:
                r2 = {k: v for k, v in rec.items() if k != 'func_match'}
                resultado_safe.append(r2)
            session['excel_import_resultado'] = resultado_safe
            return render_template('salarios_importar_preview.html',
                resultado=resultado_safe,
                funcionarios=Funcionario.query.order_by(Funcionario.nome).all(),
                meses_labels=MESES_LABELS)
        except Exception as e:
            flash(f'Erro ao ler ficheiro: {e}', 'error')
            return redirect(url_for('salarios_importar'))
    return render_template('salarios_importar.html')

@app.route('/salarios/importar/confirmar', methods=['POST'])
@login_required
def salarios_importar_confirmar():
    """Confirm import — create ReciboSalario records from Excel data."""
    from flask import session as sess
    resultado = sess.get('excel_import_resultado', [])
    mes_global = request.form.get('mes_global', '')
    ano_global = request.form.get('ano_global', str(datetime.now().year))

    criados = 0
    erros = []
    for i, rec in enumerate(resultado):
        # Skip if not checked
        if not request.form.get(f'incluir_{i}'):
            continue
        fid = request.form.get(f'func_{i}')
        mes = request.form.get(f'mes_{i}') or mes_global
        ano = request.form.get(f'ano_{i}') or ano_global

        if not fid or not mes or not ano:
            erros.append(f"Linha {i+1} ({rec.get('sheet','?')}): sem funcionário/mês/ano")
            continue

        fid = int(fid); mes = int(mes); ano = int(ano)
        func = Funcionario.query.get(fid)
        if not func:
            erros.append(f"Linha {i+1}: funcionário não encontrado")
            continue

        # Check existing
        existing = ReciboSalario.query.filter_by(
            funcionario_id=fid, ano=ano, mes=mes).first()
        if existing:
            # Update
            r = existing
        else:
            r = ReciboSalario(funcionario_id=fid, ano=ano, mes=mes,
                mes_label=MESES_LABELS.get(mes, f'Mês {mes}'))
            db.session.add(r)

        r.vencimento_base     = rec.get('vencimento_base', 0)
        r.vencimento_base_rht = rec.get('vencimento_base_rht', 0)
        r.vencimento_base_g   = rec.get('vencimento_base_g', 0)
        r.premios             = rec.get('premios', 0)
        r.faltas_dias         = rec.get('faltas_dias', 0)
        r.faltas_horas        = rec.get('faltas_horas', 0)
        r.horas_extra         = rec.get('horas_extra_horas', 0)
        r.horas_extra_rht     = rec.get('horas_extra_rht', 0)
        r.sub_refeicao_dias   = rec.get('sub_refeicao_dias', 0)
        r.sub_refeicao_vdia   = rec.get('sub_refeicao_vdia', 0)
        r.subsidio_refeicao   = rec.get('subsidio_refeicao', 0)
        r.outros_abonos       = rec.get('outros_abonos', 0)
        r.total_abonos        = rec.get('total_iliquido', 0)
        r.seg_social_taxa     = rec.get('seg_social_taxa', 0)
        r.seg_social_base     = rec.get('seg_social_base', 0)
        r.seg_social_func     = rec.get('seg_social', 0)
        r.irs_taxa            = rec.get('irs_taxa', 0)
        r.irs_parcela_abater  = rec.get('irs_parcela_abater', 0)
        r.irs_taxa_efetiva    = rec.get('irs_taxa_efetiva', 0)
        r.irs_base            = rec.get('irs_base', 0)
        r.irs_retencao        = rec.get('irs', 0)
        r.outros_descontos    = 0
        r.total_descontos     = rec.get('total_descontos', 0)
        r.liquido             = rec.get('liquido', 0)
        r.estado              = 'processado'
        r.notas               = 'Importado de Excel: ' + rec.get('sheet','')
        if rec.get('obs'): r.notas = rec.get('obs') + ' | ' + r.notas
        r.atualizado_em       = datetime.now()
        criados += 1

    db.session.commit()
    if erros:
        for e in erros: flash(e, 'warning')
    flash(f'{criados} recibo(s) importado(s) com sucesso.', 'success')
    return redirect(url_for('salarios'))

def _parse_recibos_excel(filepath):
    """Parse UCN salary Excel format — one sheet per employee."""
    if filepath.lower().endswith('.xlsx'):
        return _parse_recibos_xlsx(filepath)

    # .xls — use xlrd
    import xlrd
    wb = xlrd.open_workbook(filepath)

    resultado = []

    # Get global month from INICIO sheet if present
    mes_global = ''
    try:
        ini = wb.sheet_by_name('INICIO')
        mes_global = str(ini.cell_value(4, 2)).strip()  # C5
    except: pass

    for sname in wb.sheet_names():
        if sname in ('INICIO', 'Transf BCP', 'inicio', 'transf'):
            continue
        ws = wb.sheet_by_name(sname)
        try:
            rec = _parse_sheet_xls(ws, sname, mes_global)
            resultado.append(rec)
        except Exception as e:
            resultado.append({'sheet': sname, 'erro': str(e), 'nome_raw': sname})

    return resultado

def _parse_sheet_xls(ws, sname, mes_global=''):
    """Parse one salary sheet. Skips non-salary sheets (dist lucros, etc).
    Finds key rows by text search within the 'original' block.
    """
    def fv(r, c):
        try: return float(ws.cell_value(r, c) or 0)
        except: return 0.0
    def sv(r, c):
        try: return str(ws.cell_value(r, c)).strip()
        except: return ''

    # Find 'original' row
    orig_row = 1
    for r in range(min(10, ws.nrows)):
        for c in range(ws.ncols):
            if sv(r, c).lower() == 'original':
                orig_row = r
                break

    O = orig_row

    # Verify this is a salary sheet (has 'Remuneração Base')
    has_rem = False
    for r in range(O, min(O+15, ws.nrows)):
        txt = sv(r, 0).upper()
        if 'REMUNERAÇÃO BASE' in txt or 'REMUNERACAO BASE' in txt:
            has_rem = True
            break
    if not has_rem:
        raise ValueError(f"Sheet '{sname}' não é um recibo de salário normal")

    # Find key rows by text in col A within the block
    rm = {}  # row_map
    for r in range(O, min(O+30, ws.nrows)):
        txt = sv(r, 0).upper()
        if not txt: continue
        if ('REMUNERAÇÃO BASE' in txt or 'REMUNERACAO BASE' in txt) and 'rem_base' not in rm:
            rm['rem_base'] = r
        elif ('GRATIFICAÇÃO' in txt or 'GRATIFICACAO' in txt or 'PRÉMIO' in txt or 'PREMIO' in txt) and 'premios' not in rm:
            rm['premios'] = r
        elif 'FALTAS' in txt and ('MÊS' in txt or 'MES' in txt or 'CORRENTE' in txt) and 'faltas_horas' not in rm:
            rm['faltas_horas'] = r
        elif 'FALTAS' in txt and 'faltas_dias' not in rm:
            rm['faltas_dias'] = r
        elif ('EXTRAORDINÁR' in txt or 'EXTRAORDINAR' in txt) and 'horas_extra' not in rm:
            rm['horas_extra'] = r
        elif ('ALIMENTAÇÃO' in txt or 'ALIMENTACAO' in txt or 'REFEIÇÃO' in txt) and 'sub_ref' not in rm:
            rm['sub_ref'] = r
        elif ('TOTAL ILÍQUIDO' in txt or 'TOTAL ILIQUIDO' in txt) and 'total_iliq' not in rm:
            rm['total_iliq'] = r
        elif 'C.R.S.S' in txt and 'crss' not in rm:
            rm['crss'] = r
        elif 'I.R.S' in txt and 'HORAS' not in txt and 'EXTRAS' not in txt and 'irs' not in rm:
            rm['irs'] = r
        elif ('TOTAL DE DESCONTOS' in txt or 'TOTAL DESCONTOS' in txt) and 'total_desc' not in rm:
            rm['total_desc'] = r
        elif ('LÍQUIDO A RECEBER' in txt or 'LIQUIDO A RECEBER' in txt) and 'liquido' not in rm:
            rm['liquido'] = r
        elif 'DISCRIMINATIVO' in txt and 'transf' not in rm:
            rm['transf'] = r

    RB  = rm.get('rem_base',  O+7)
    RPR = rm.get('premios',   O+8)
    RFD = rm.get('faltas_dias', O+9)
    RFH = rm.get('faltas_horas', O+10)
    RHE = rm.get('horas_extra', O+11)
    RSR = rm.get('sub_ref',   O+12)
    RTI = rm.get('total_iliq', O+14)
    RCS = rm.get('crss',      O+16)
    RIR = rm.get('irs',       O+17)
    RTD = rm.get('total_desc', O+21)
    RLQ = rm.get('liquido',   O+23)
    RTR = rm.get('transf',    O+25)

    # Employee name from A4 of block
    nome_raw = sv(O+2, 0).strip()
    if not nome_raw or 'NOME' in nome_raw.upper():
        nome_raw = sname

    # Match funcionario
    func_match = None
    funcs = Funcionario.query.all()
    if '-' in sname:
        parts = sname.split('-', 1)
        num = parts[0].strip()
        f_by_num = Funcionario.query.filter_by(numero=num).first()
        if f_by_num:
            func_match = f_by_num
        elif len(parts) > 1:
            nome_sheet = parts[1].strip()
            for f in funcs:
                if nome_sheet.upper() in f.nome.upper() or f.nome.upper() in nome_sheet.upper():
                    func_match = f; break
    if not func_match:
        nome_parts = [p for p in nome_raw.upper().split() if len(p) > 2]
        best = 0
        for f in funcs:
            score = sum(1 for p in nome_parts if any(p in fp for fp in f.nome.upper().split()))
            if score > best:
                best = score; func_match = f

    return {
        'sheet': sname,
        'nome_raw': nome_raw,
        'func_match_id':   func_match.id   if func_match else None,
        'func_match_nome': func_match.nome if func_match else '—',
        'mes_global': mes_global,
        'vencimento_base':     fv(RB,  7),  # H
        'vencimento_base_rht': fv(RB,  5),  # F
        'vencimento_base_g':   fv(RB,  6),  # G
        'premios':             fv(RPR, 7),  # H
        'faltas_dias':         fv(RFD, 3),  # D
        'faltas_horas':        fv(RFH, 3),  # D col index 3 = column D
        'horas_extra_horas':   fv(RHE, 4),  # E
        'horas_extra_rht':     fv(RHE, 5),  # F
        'horas_extra':         fv(RHE, 7),  # H
        'sub_refeicao_dias':   fv(RSR, 3),  # D
        'sub_refeicao_vdia':   fv(RSR, 5),  # F
        'subsidio_refeicao':   fv(RSR, 7),  # H
        'outros_abonos':       0,
        'total_iliquido':      fv(RTI, 7),  # H
        'seg_social_taxa':     fv(RCS, 3),  # D
        'seg_social_base':     fv(RCS, 4),  # E
        'seg_social':          fv(RCS, 5),  # F
        'irs_taxa':            fv(RIR, 1),  # B
        'irs_parcela_abater':  fv(RIR, 2),  # C
        'irs_taxa_efetiva':    fv(RIR, 3),  # D
        'irs_base':            fv(RIR, 4),  # E
        'irs':                 fv(RIR, 7),  # H
        'total_descontos':     fv(RTD, 7),  # H
        'liquido':             fv(RLQ, 7),  # H
        'transf_conta':        fv(RTR, 7),  # H
        'transf_refeicao':     fv(RTR+1, 7),# H next row
        'obs':                 sv(O+30, 5),  # F32
    }


def _parse_recibos_xlsx(filepath):
    """Parse .xlsx format."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    resultado = []
    for sname in wb.sheetnames:
        if sname.upper() in ('INICIO', 'TRANSF BCP'):
            continue
        ws = wb[sname]
        def fv(r, c):
            try: return float(ws.cell(row=r, column=c).value or 0)
            except: return 0.0
        def sv(r, c):
            try: return str(ws.cell(row=r, column=c).value or '').strip()
            except: return ''
        nome_raw = sv(4, 1).replace('NOME DO FUNCIONÁRIO','').replace('XPTO','').strip()
        resultado.append({
            'sheet': sname, 'nome_raw': nome_raw or sname,
            'func_match': None, 'func_match_id': None, 'func_match_nome': '—',
            'mes_global': '',
            'vencimento_base': fv(9,8), 'premios': fv(10,8),
            'faltas_dias': fv(11,4), 'faltas_horas': fv(12,4),  # D12
            'horas_extra': fv(13,8), 'subsidio_refeicao': fv(14,8),
            'outros_abonos': fv(15,8), 'total_iliquido': fv(17,8),  # H17
            'seg_social': fv(19,6), 'irs': fv(20,8),
            'total_descontos': fv(24,8), 'liquido': fv(26,8),  # H24, H26
            'transf_conta': fv(27,8), 'transf_refeicao': fv(28,8),
        })
    return resultado


def init_db():
    """Create all database tables."""
    with app.app_context():
        db.create_all()


if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=False)
# cache bust Tue Apr  7 15:34:00 UTC 2026

def _sync_ausencias_to_faltas(funcionario_id, ano):
    """Sync AusenciaRegisto records into FuncionarioFalta for a given year."""
    from datetime import date
    cfg = ConfigGeral.query.first()
    dia_inicio = cfg.salario_dia_inicio if cfg and hasattr(cfg,'salario_dia_inicio') else 1
    dia_fecho  = cfg.salario_dia_fecho  if cfg and hasattr(cfg,'salario_dia_fecho')  else 27

    # Delete existing auto-synced faltas for this year
    FuncionarioFalta.query.filter_by(
        funcionario_id=funcionario_id, ano=ano
    ).delete()

    # Get approved absences for this year (+/- spillover)
    registos = AusenciaRegisto.query.filter(
        AusenciaRegisto.funcionario_id == funcionario_id,
        AusenciaRegisto.estado.in_(["aprovado","pendente"]),
        AusenciaRegisto.data_inicio >= date(ano-1,12,1),
        AusenciaRegisto.data_fim    <= date(ano+1,1,31)
    ).all()

    # Map tipo → FuncionarioFalta.tipo
    tipo_map = {
        "ferias": "ferias", "ponte": "ferias", "fecho_empresa": "ferias",
        "falta_justificada": "justificada", "falta_injustificada": "injustificada",
        "baixa_medica": "baixa", "consulta_medica": "justificada",
        "assistencia_familia": "justificada", "licenca_sem_venc": "justificada",
        "formacao": "ferias", "teletrabalho": "ferias", "trabalho_externo": "ferias",
        "trabalho_ponte": None,  # worked on bridge day - no deduction
    }

    feriados_set = _get_feriados_set(ano)
    import calendar as _cal_mod2
    from datetime import date as _d2
    mes_ini_civil = _d2(ano, mes, 1)
    mes_fim_civil = _d2(ano, mes, _cal_mod2.monthrange(ano, mes)[1])
    dias_uteis_mes_civil = _dias_uteis_ausencia(mes_ini_civil, mes_fim_civil, feriados_set)

    # Accumulate per salary-period per tipo
    from datetime import timedelta
    from collections import defaultdict
    buckets = defaultdict(float)  # (mes_salarial, tipo_falta) → dias

    for r in registos:
        cur = r.data_inicio
        while cur <= r.data_fim:
            # Only count weekdays non-holidays (unless it's a ponte type)
            is_weekday = cur.weekday() < 5
            is_feriado = cur in feriados_set
            if r.tipo == "ponte":
                count_day = is_weekday  # pontes are weekdays only
            else:
                count_day = is_weekday and not is_feriado

            if count_day:
                # Determine salary month for this day
                if cur.day <= dia_fecho:
                    mes_sal = cur.month
                    ano_sal = cur.year
                else:
                    # After closing day → belongs to next month
                    if cur.month == 12:
                        mes_sal = 1; ano_sal = cur.year + 1
                    else:
                        mes_sal = cur.month + 1; ano_sal = cur.year

                if ano_sal == ano:
                    ft = tipo_map.get(r.tipo, "justificada")
                    val = 0.5 if r.formato in ("manha","tarde") else 1.0
                    if r.formato == "horas":
                        val = round(r.horas / 8, 2)
                    buckets[(mes_sal, ft)] += val

            cur += timedelta(days=1)

    # Write FuncionarioFalta records
    for (mes, tipo_falta), dias in buckets.items():
        if dias > 0:
            db.session.add(FuncionarioFalta(
                funcionario_id=funcionario_id,
                ano=ano, mes=mes,
                dias_falta=round(dias, 1),
                tipo=tipo_falta,
                notas=f"Auto-sync Mapa Férias/Faltas {ano}"
            ))
    db.session.commit()
    return len(buckets)


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO PERÍODOS SALARIAIS & HORAS EXTRA
# ══════════════════════════════════════════════════════════════════════════════

def _get_horario():
    """Get company work schedule or defaults."""
    cfg = ConfigHorario.query.first()
    if cfg: return cfg.hora_inicio, cfg.hora_fim, cfg.horas_dia
    return '08:30', '17:30', 8.0

def _horas_entre(h_ini, h_fim):
    """Calculate hours between two HH:MM strings."""
    def to_min(s):
        h, m = s.split(':')
        return int(h)*60 + int(m)
    return round((to_min(h_fim) - to_min(h_ini)) / 60, 2)

def _categoria_he(data, feriados_set):
    """Classify hora extra: dia_util / fds / feriado."""
    from datetime import date
    if data in feriados_set: return 'feriado'
    if data.weekday() >= 5: return 'fds'
    return 'dia_util'

def _get_or_create_periodo(ano, mes):
    """Get existing period or return None."""
    return PeriodoSalarial.query.filter_by(ano=ano, mes=mes).first()

def _resumo_periodo(periodo_id):
    """Calculate attendance summary for a salary period."""
    p = PeriodoSalarial.query.get(periodo_id)
    if not p: return {}
    from datetime import timedelta
    feriados_set = _get_feriados_set(p.ano)

    # Days in period
    dias_uteis_total = _dias_uteis_ausencia(p.data_inicio, p.data_fim, feriados_set)

    funcs = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()         if hasattr(Funcionario,'ativo') else Funcionario.query.order_by(Funcionario.nome).all()

    resumo = {}
    for f in funcs:
        ausencias = AusenciaRegisto.query.filter(
            AusenciaRegisto.funcionario_id == f.id,
            AusenciaRegisto.estado.in_(['aprovado']),
            AusenciaRegisto.data_inicio <= p.data_fim,
            AusenciaRegisto.data_fim   >= p.data_inicio
        ).all()

        r = {'nome': f.nome, 'id': f.id,
             'dias_uteis': dias_uteis_total,
             'ferias': 0.0, 'pontes': 0.0, 'faltas_just': 0.0,
             'faltas_injust': 0.0, 'baixas': 0.0, 'outros': 0.0,
             'he_util': 0.0, 'he_fds': 0.0, 'he_feriado': 0.0}

        for a in ausencias:
            # Clip to period
            ini = max(a.data_inicio, p.data_inicio)
            fim = min(a.data_fim,   p.data_fim)
            dias = _dias_uteis_ausencia(ini, fim, feriados_set, a.formato, a.horas)
            if a.tipo in ('ferias','fecho_empresa'): r['ferias'] += dias
            elif a.tipo == 'ponte': r['pontes'] += dias
            elif a.tipo == 'falta_justificada': r['faltas_just'] += dias
            elif a.tipo == 'falta_injustificada': r['faltas_injust'] += dias
            elif a.tipo == 'baixa_medica': r['baixas'] += dias
            else: r['outros'] += dias

        # Horas extra
        hes = HoraExtra.query.filter(
            HoraExtra.funcionario_id == f.id,
            HoraExtra.data >= p.data_inicio,
            HoraExtra.data <= p.data_fim,
            HoraExtra.estado.in_(['aprovado','pendente'])
        ).all()
        for he in hes:
            if he.categoria == 'dia_util': r['he_util'] += he.total_horas
            elif he.categoria == 'fds': r['he_fds'] += he.total_horas
            elif he.categoria == 'feriado': r['he_feriado'] += he.total_horas

        r['dias_trabalhados'] = round(dias_uteis_total - r['ferias'] - r['pontes'] - r['faltas_just'] - r['faltas_injust'] - r['baixas'], 1)
        r['he_total'] = round(r['he_util'] + r['he_fds'] + r['he_feriado'], 2)
        for k in ['ferias','pontes','faltas_just','faltas_injust','baixas','outros']:
            r[k] = round(r[k], 1)
        resumo[f.id] = r

    return resumo

# ── Períodos salariais ────────────────────────────────────────────────────────
@app.route('/periodos-salariais')
@login_required
def periodos_fecho():
    from datetime import date
    ano = int(request.args.get('ano', date.today().year))
    periodos = PeriodoSalarial.query.filter_by(ano=ano).order_by(PeriodoSalarial.mes).all()
    return render_template('periodos_salariais.html', periodos=periodos, ano=ano)

@app.route('/periodos-salariais/criar', methods=['POST'])
@login_required
def periodo_criar():
    data = request.get_json() or {}
    try:
        ini = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        fim = datetime.strptime(data['data_fim'],    '%Y-%m-%d').date()
    except: return jsonify({'ok': False, 'error': 'Datas inválidas'})
    mes = int(data.get('mes', ini.month))
    ano = int(data.get('ano', ini.year))
    existing = PeriodoSalarial.query.filter_by(ano=ano, mes=mes).first()
    if existing: return jsonify({'ok': False, 'error': f'Já existe período para {ano}/{mes:02d}'})
    p = PeriodoSalarial(ano=ano, mes=mes, data_inicio=ini, data_fim=fim,
        notas=data.get('notas',''), criado_por=current_user.id, criado_em=datetime.now())
    db.session.add(p); db.session.commit()
    return jsonify({'ok': True, 'id': p.id})

@app.route('/periodos-salariais/<int:pid>/editar', methods=['POST'])
@login_required
def periodo_editar(pid):
    p = PeriodoSalarial.query.get_or_404(pid)
    if p.estado == 'fechado' and not current_user.is_admin:
        return jsonify({'ok': False, 'error': 'Período fechado'})
    data = request.get_json() or {}
    if 'data_inicio' in data: p.data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
    if 'data_fim'    in data: p.data_fim    = datetime.strptime(data['data_fim'],    '%Y-%m-%d').date()
    if 'notas'       in data: p.notas       = data['notas']
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/periodos-salariais/<int:pid>/fechar', methods=['POST'])
@login_required
def periodo_fechar(pid):
    if not current_user.is_admin: return jsonify({'ok': False, 'error': 'Sem permissão'})
    p = PeriodoSalarial.query.get_or_404(pid)
    p.estado = 'fechado'; p.fechado_por = current_user.id; p.fechado_em = datetime.now()
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/periodos-salariais/<int:pid>/reabrir', methods=['POST'])
@login_required
def periodo_reabrir(pid):
    if not current_user.is_admin: return jsonify({'ok': False, 'error': 'Sem permissão'})
    p = PeriodoSalarial.query.get_or_404(pid)
    p.estado = 'aberto'; p.fechado_por = None; p.fechado_em = None
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/periodos-salariais/<int:pid>/resumo')
@login_required
def periodo_resumo(pid):
    resumo = _resumo_periodo(pid)
    p = PeriodoSalarial.query.get_or_404(pid)
    return jsonify({'ok': True, 'resumo': list(resumo.values()),
        'periodo': {'inicio': p.data_inicio.isoformat(), 'fim': p.data_fim.isoformat(),
                    'mes': p.mes, 'ano': p.ano, 'estado': p.estado}})

@app.route('/periodos-salariais/<int:pid>/exportar-pdf')
@login_required
def periodo_exportar_pdf(pid):
    p = PeriodoSalarial.query.get_or_404(pid)
    resumo = _resumo_periodo(pid)
    cfg = ConfigGeral.query.first()
    empresa = cfg.empresa_nome if cfg else 'UCN'
    from flask import Response
    import io
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
            topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles = getSampleStyleSheet()
        meses_pt = ['','Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                    'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
        elements = []
        # Title
        title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=14, spaceAfter=4)
        sub_style   = ParagraphStyle('sub',   parent=styles['Normal'],   fontSize=9,  textColor=colors.grey)
        elements.append(Paragraph(f'{empresa} — Resumo Salarial', title_style))
        elements.append(Paragraph(
            f'{meses_pt[p.mes]} {p.ano} · Período: {p.data_inicio.strftime("%d/%m/%Y")} a {p.data_fim.strftime("%d/%m/%Y")} · Estado: {p.estado.upper()}',
            sub_style))
        elements.append(Spacer(1, 0.4*cm))
        # Table
        headers = ['Funcionário','Dias Úteis','Trabalhados','Férias','Pontes',
                   'Faltas Just.','Faltas Injust.','Baixas','HE Útil (h)','HE FDS (h)','HE Feriado (h)','HE Total (h)']
        data_rows = [headers]
        for r in resumo.values():
            data_rows.append([
                r['nome'][:28],
                str(r['dias_uteis']), str(r['dias_trabalhados']),
                str(r['ferias']), str(r['pontes']),
                str(r['faltas_just']), str(r['faltas_injust']),
                str(r['baixas']),
                str(r['he_util']), str(r['he_fds']), str(r['he_feriado']), str(r['he_total'])
            ])
        tbl = Table(data_rows, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('GRID',       (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f7fa')]),
            ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ]))
        elements.append(tbl)
        doc.build(elements)
        buf.seek(0)
        fname = f'resumo_salarial_{p.ano}_{p.mes:02d}.pdf'
        return Response(buf.getvalue(), mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment;filename={fname}'})
    except ImportError:
        return jsonify({'ok': False, 'error': 'ReportLab não instalado. Use: pip install reportlab'})

# ── Horas Extra ───────────────────────────────────────────────────────────────
@app.route('/horas-extra/registar', methods=['POST'])
@login_required
def hora_extra_registar():
    data = request.get_json() or {}
    try:
        dt = datetime.strptime(data['data'], '%Y-%m-%d').date()
    except: return jsonify({'ok': False, 'error': 'Data inválida'})
    fid    = int(data['funcionario_id'])
    horas_manual = data.get('horas_manual')
    h_ini = (data.get('hora_inicio') or '').strip() or '00:00'
    h_fim = (data.get('hora_fim') or '').strip() or '00:00'
    if horas_manual and float(horas_manual) > 0:
        total = float(horas_manual)
        h_ini = '--'; h_fim = '--'
    elif h_ini != '00:00' and h_fim != '00:00':
        total = _horas_entre(h_ini, h_fim)
        if total <= 0: return jsonify({'ok': False, 'error': 'Hora fim deve ser depois do início'})
    else:
        return jsonify({'ok': False, 'error': 'Indique o período horário ou o total de horas'})
    # Check overlap
    overlap = HoraExtra.query.filter(
        HoraExtra.funcionario_id == fid,
        HoraExtra.data == dt,
        HoraExtra.estado.in_(['aprovado','pendente'])
    ).first()
    if overlap:
        return jsonify({'ok': False, 'error': f'Já existe registo de HE neste dia ({overlap.hora_inicio}–{overlap.hora_fim})'})
    # Check conflict with falta
    falta = AusenciaRegisto.query.filter(
        AusenciaRegisto.funcionario_id == fid,
        AusenciaRegisto.data_inicio <= dt,
        AusenciaRegisto.data_fim >= dt,
        AusenciaRegisto.tipo.in_(['falta_injustificada','falta_justificada','baixa_medica']),
        AusenciaRegisto.estado == 'aprovado'
    ).first()
    if falta and not data.get('forcar'):
        return jsonify({'ok': False, 'error': f'Conflito: funcionário tem {falta.tipo} neste dia', 'conflito': True})
    feriados_set = _get_feriados_set(dt.year)
    cat = _categoria_he(dt, feriados_set)
    # Find period
    p = PeriodoSalarial.query.filter(
        PeriodoSalarial.data_inicio <= dt,
        PeriodoSalarial.data_fim >= dt
    ).first()
    he = HoraExtra(
        funcionario_id=fid, data=dt,
        hora_inicio=h_ini, hora_fim=h_fim,
        total_horas=total, categoria=cat,
        periodo_id=p.id if p else None,
        observacoes=data.get('observacoes','').strip(),
        estado='aprovado',
        criado_por=current_user.id, criado_em=datetime.now()
    )
    db.session.add(he); db.session.commit()
    return jsonify({'ok': True, 'id': he.id, 'total_horas': total, 'categoria': cat})

@app.route('/horas-extra/<int:hid>/eliminar', methods=['POST'])
@login_required
def hora_extra_eliminar(hid):
    he = HoraExtra.query.get_or_404(hid)
    db.session.delete(he); db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/funcionario/<int:fid>/horas-extra')
@login_required
def api_funcionario_horas_extra(fid):
    ano = int(request.args.get('ano', datetime.now().year))
    hes = HoraExtra.query.filter(
        HoraExtra.funcionario_id == fid,
        db.extract('year', HoraExtra.data) == ano
    ).order_by(HoraExtra.data.desc()).all()
    return jsonify([{
        'id': h.id, 'data': h.data.isoformat(),
        'hora_inicio': h.hora_inicio, 'hora_fim': h.hora_fim,
        'total_horas': h.total_horas, 'categoria': h.categoria,
        'estado': h.estado, 'observacoes': h.observacoes or ''
    } for h in hes])


@app.route('/ausencias/pdf/<int:ano>/<int:mes>')
@login_required
def ausencias_pdf(ano, mes):
    """Main PDF route - always uses the professional HTML version."""
    return redirect(url_for('ausencias_pdf_html', ano=ano, mes=mes))


@app.route('/ausencias/pdf-reportlab/<int:ano>/<int:mes>')
@login_required
def ausencias_pdf_reportlab(ano, mes):
    from datetime import date, timedelta
    import io
    MESES_PT = ['','Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

    # Get salary period
    periodo = PeriodoSalarial.query.filter_by(ano=ano, mes=mes).first()
    if periodo:
        data_ini = periodo.data_inicio
        data_fim = periodo.data_fim
    else:
        # Fallback: civil month
        data_ini = date(ano, mes, 1)
        import calendar
        data_fim = date(ano, mes, calendar.monthrange(ano, mes)[1])

    feriados_set = _get_feriados_set(ano)

    # Calculate period stats
    from datetime import date as _dt_pdf
    import calendar as _cal_mod
    dias_total = (data_fim - data_ini).days + 1
    dias_uteis = _dias_uteis_ausencia(data_ini, data_fim, feriados_set)
    # Also calculate full civil month working days for reference
    mes_ini_civil = _dt_pdf(ano, mes, 1)
    mes_fim_civil = _dt_pdf(ano, mes, _cal_mod.monthrange(ano, mes)[1])
    dias_uteis_mes_civil = _dias_uteis_ausencia(mes_ini_civil, mes_fim_civil, feriados_set)
    dias_uteis_periodo = dias_uteis  # period-based
    feriados_no_periodo = [f for f in FeriasFeriado.query.filter_by(ano=ano).all()
                           if data_ini <= f.data <= data_fim]
    pontes_no_periodo = [f for f in feriados_no_periodo if f.tipo == 'ponte']
    n_feriados = len([f for f in feriados_no_periodo if f.tipo != 'ponte'])
    n_pontes   = len(pontes_no_periodo)

    funcs = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()         if hasattr(Funcionario,'ativo') else Funcionario.query.order_by(Funcionario.nome).all()

    cfg = ConfigGeral.query.first()
    empresa = cfg.empresa_nome if cfg else 'UCN'

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable, PageBreak)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
            topMargin=2*cm, bottomMargin=2*cm,
            leftMargin=2*cm, rightMargin=2*cm,
            title=f'Resumo Assiduidade {MESES_PT[mes]} {ano}')

        # Styles
        styles = getSampleStyleSheet()
        navy   = colors.HexColor('#1e3a5f')
        gold   = colors.HexColor('#e8b84b')
        light  = colors.HexColor('#f5f7fa')
        mid    = colors.HexColor('#e2e8f0')
        dark   = colors.HexColor('#334155')

        s_title  = ParagraphStyle('T', fontName='Helvetica-Bold', fontSize=18, textColor=navy, spaceAfter=4)
        s_sub    = ParagraphStyle('S', fontName='Helvetica', fontSize=10, textColor=dark, spaceAfter=2)
        s_period = ParagraphStyle('P', fontName='Helvetica-Oblique', fontSize=9, textColor=colors.grey, spaceAfter=12)
        s_sec    = ParagraphStyle('SEC', fontName='Helvetica-Bold', fontSize=12, textColor=navy, spaceBefore=14, spaceAfter=6)
        s_name   = ParagraphStyle('N', fontName='Helvetica-Bold', fontSize=11, textColor=navy, spaceBefore=10, spaceAfter=2)
        s_body   = ParagraphStyle('B', fontName='Helvetica', fontSize=9, textColor=dark, spaceAfter=2)
        s_event  = ParagraphStyle('E', fontName='Helvetica', fontSize=8.5, textColor=dark, leftIndent=12, spaceAfter=1)
        s_footer = ParagraphStyle('F', fontName='Helvetica', fontSize=8, textColor=colors.grey, alignment=TA_CENTER)

        elements = []

        # ── Header ────────────────────────────────────────────────────────────
        elements.append(Paragraph(empresa, s_title))
        elements.append(Paragraph(f'Resumo de Assiduidade – {MESES_PT[mes]} {ano}', s_sub))
        periodo_label = f'Período Salarial: {data_ini.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'
        if not periodo:
            periodo_label += '  (mês civil — sem período salarial configurado)'
        elements.append(Paragraph(periodo_label, s_period))
        elements.append(HRFlowable(width='100%', thickness=2, color=navy, spaceAfter=10))

        # ── Summary KPIs ──────────────────────────────────────────────────────
        # Aggregate totals
        total_faltas = 0; total_ferias = 0; total_he = 0
        for f in funcs:
            aus = AusenciaRegisto.query.filter(
                AusenciaRegisto.funcionario_id == f.id,
                AusenciaRegisto.estado == 'aprovado',
                AusenciaRegisto.data_inicio <= data_fim,
                AusenciaRegisto.data_fim >= data_ini
            ).all()
            for a in aus:
                ini2 = max(a.data_inicio, data_ini)
                fim2 = min(a.data_fim, data_fim)
                d = _dias_uteis_ausencia(ini2, fim2, feriados_set, a.formato, a.horas)
                if a.tipo in ('falta_justificada','falta_injustificada','baixa_medica','consulta_medica','assistencia_familia'):
                    total_faltas += d
                elif a.tipo in ('ferias','fecho_empresa','ponte'):
                    total_ferias += d
            hes = HoraExtra.query.filter(
                HoraExtra.funcionario_id == f.id,
                HoraExtra.data >= data_ini, HoraExtra.data <= data_fim,
                HoraExtra.estado.in_(['aprovado','pendente'])
            ).all()
            total_he += sum(h.total_horas for h in hes)

        kpi_data = [
            ['Mês de Referência', MESES_PT[mes] + ' ' + str(ano),
             'Dias do Período', str(dias_total)],
            ['Dias Úteis', str(int(dias_uteis)),
             'Feriados', str(len(feriados_no_periodo))],
            ['Total Faltas', f'{total_faltas:.1f} dias',
             'Total Férias', f'{total_ferias:.1f} dias'],
            ['Horas Extra', f'{total_he:.1f} h',
             'Data de Emissão', date.today().strftime('%d/%m/%Y')],
        ]
        kpi_table = Table(kpi_data, colWidths=[4*cm, 4.5*cm, 4*cm, 4.5*cm])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND',  (0,0), (-1,-1), light),
            ('BACKGROUND',  (0,0), (0,-1), mid),
            ('BACKGROUND',  (2,0), (2,-1), mid),
            ('FONTNAME',    (0,0), (-1,-1), 'Helvetica'),
            ('FONTNAME',    (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME',    (2,0), (2,-1), 'Helvetica-Bold'),
            ('FONTSIZE',    (0,0), (-1,-1), 9),
            ('PADDING',     (0,0), (-1,-1), 6),
            ('GRID',        (0,0), (-1,-1), 0.5, colors.white),
            ('ROUNDEDCORNERS', [4]),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 0.5*cm))

        # ── Per-employee sections ──────────────────────────────────────────────
        elements.append(Paragraph('Detalhe por Colaborador', s_sec))
        elements.append(HRFlowable(width='100%', thickness=1, color=mid, spaceAfter=8))

        TIPOS_LABELS = {
            'ferias':'Férias','ponte':'Ponte','fecho_empresa':'Fecho Empresa',
            'falta_justificada':'Falta Justificada','falta_injustificada':'Falta Injustificada',
            'baixa_medica':'Baixa Médica','consulta_medica':'Consulta Médica',
            'assistencia_familia':'Assistência Família','formacao':'Formação',
            'teletrabalho':'Teletrabalho','licenca_sem_venc':'Licença s/ Vencimento',
            'trabalho_externo':'Trabalho Externo',
        }

        for f in funcs:
            aus = AusenciaRegisto.query.filter(
                AusenciaRegisto.funcionario_id == f.id,
                AusenciaRegisto.estado == 'aprovado',
                AusenciaRegisto.data_inicio <= data_fim,
                AusenciaRegisto.data_fim >= data_ini
            ).order_by(AusenciaRegisto.data_inicio).all()

            hes = HoraExtra.query.filter(
                HoraExtra.funcionario_id == f.id,
                HoraExtra.data >= data_ini, HoraExtra.data <= data_fim,
                HoraExtra.estado.in_(['aprovado','pendente'])
            ).order_by(HoraExtra.data).all()

            if not aus and not hes:
                continue

            # Employee header
            elements.append(Paragraph(f.nome, s_name))
            dept = getattr(f, 'departamento', '') or getattr(f, 'cargo', '') or '—'
            elements.append(Paragraph(dept, s_body))

            # Employee summary table
            f_ferias=0; f_faltas=0; f_pontes=0; f_outros=0; f_he=0
            for a in aus:
                ini2 = max(a.data_inicio, data_ini)
                fim2 = min(a.data_fim, data_fim)
                d = _dias_uteis_ausencia(ini2, fim2, feriados_set, a.formato, a.horas)
                if a.tipo == 'ferias': f_ferias += d
                elif a.tipo in ('ponte','fecho_empresa'): f_pontes += d
                elif a.tipo in ('falta_justificada','falta_injustificada','baixa_medica',
                                'consulta_medica','assistencia_familia'): f_faltas += d
                else: f_outros += d
            for h in hes:
                f_he += h.total_horas

            emp_data = [[
                f'Dias Úteis: {int(dias_uteis)}',
                f'Férias: {f_ferias:.1f}d',
                f'Faltas: {f_faltas:.1f}d',
                f'Pontes: {f_pontes:.1f}d',
                f'HE: {f_he:.1f}h'
            ]]
            emp_table = Table(emp_data, colWidths=[3.3*cm]*5)
            emp_table.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,-1), colors.HexColor('#eef2ff')),
                ('FONTNAME',(0,0),(-1,-1),'Helvetica'),
                ('FONTSIZE',(0,0),(-1,-1),8),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),
                ('PADDING',(0,0),(-1,-1),4),
                ('GRID',(0,0),(-1,-1),0.5,colors.white),
            ]))
            elements.append(emp_table)
            elements.append(Spacer(1, 0.2*cm))

            # Events — consolidate consecutive férias
            # Group consecutive férias/pontes/fecho
            event_lines = []
            i = 0
            while i < len(aus):
                a = aus[i]
                ini2 = max(a.data_inicio, data_ini)
                fim2 = min(a.data_fim, data_fim)
                label = TIPOS_LABELS.get(a.tipo, a.tipo)
                if a.tipo in ('ferias','fecho_empresa','ponte'):
                    if ini2 == fim2:
                        event_lines.append(f'Gozou {label.lower()} no dia {ini2.strftime("%d/%m/%Y")}')
                    else:
                        event_lines.append(f'Gozou {label.lower()} de {ini2.strftime("%d/%m/%Y")} a {fim2.strftime("%d/%m/%Y")}')
                else:
                    # Faltas, baixas, etc — individual days
                    cur = ini2
                    from datetime import timedelta
                    while cur <= fim2:
                        if cur.weekday() < 5 and cur not in feriados_set:
                            event_lines.append(f'Ausente no dia {cur.strftime("%d/%m/%Y")} — {label}')
                        cur += timedelta(days=1)
                if a.observacoes:
                    event_lines[-1] += f' (obs: {a.observacoes})'
                i += 1

            for he in hes:
                event_lines.append(f'Horas extra no dia {he.data.strftime("%d/%m/%Y")}, das {he.hora_inicio} às {he.hora_fim} ({he.total_horas:.1f}h)')

            for line in event_lines:
                elements.append(Paragraph('• ' + line, s_event))

            elements.append(HRFlowable(width='100%', thickness=0.5, color=mid, spaceAfter=4))

        # ── Footer ────────────────────────────────────────────────────────────
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(
            f'Documento gerado em {date.today().strftime("%d/%m/%Y")} · {empresa} · Uso Interno',
            s_footer))

        doc.build(elements)
        buf.seek(0)
        fname = f'assiduidade_{MESES_PT[mes]}_{ano}.pdf'
        from flask import Response
        return Response(buf.getvalue(), mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment;filename={fname}'})

    except ImportError:
        # Fallback: generate printable HTML
        from flask import redirect
        return redirect(url_for('ausencias_pdf_html', ano=ano, mes=mes))
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/ausencias/pdf-html/<int:ano>/<int:mes>')
@login_required
def ausencias_pdf_html(ano, mes):
    """HTML printable version as fallback when reportlab not installed."""
    from datetime import date, timedelta
    import calendar as cal_mod
    MESES_PT = ['','Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    periodo = PeriodoSalarial.query.filter_by(ano=ano, mes=mes).first()
    if periodo:
        data_ini = periodo.data_inicio
        data_fim = periodo.data_fim
    else:
        data_ini = date(ano, mes, 1)
        data_fim = date(ano, mes, cal_mod.monthrange(ano, mes)[1])
    feriados_set = _get_feriados_set(ano)
    dias_uteis = _dias_uteis_ausencia(data_ini, data_fim, feriados_set)
    cfg = ConfigGeral.query.first()
    empresa = cfg.empresa_nome if cfg else 'UCN'
    try:
        funcs = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()
        if not funcs:  # if ativo filter returns nothing, get all
            funcs = Funcionario.query.order_by(Funcionario.nome).all()
    except Exception:
        funcs = Funcionario.query.order_by(Funcionario.nome).all()
    # Civil month calculation (base for working days)
    import calendar as _cal_m
    from datetime import date as _dm
    mes_ini_civil = _dm(ano, mes, 1)
    mes_fim_civil = _dm(ano, mes, _cal_m.monthrange(ano, mes)[1])
    dias_uteis_mes_civil = _dias_uteis_ausencia(mes_ini_civil, mes_fim_civil, set())  # excl weekends only
    feriados_mes = [f for f in FeriasFeriado.query.filter_by(ano=ano).all()
                    if mes_ini_civil <= f.data <= mes_fim_civil]
    # Only count feriados/pontes on weekdays
    n_feriados = len([f for f in feriados_mes if f.tipo != 'ponte' and f.data.weekday() < 5])
    n_pontes   = len([f for f in feriados_mes if f.tipo == 'ponte' and f.data.weekday() < 5])
    # dias_uteis_trabalho = civil month weekdays minus feriados AND pontes
    dias_uteis_trabalho = _dias_uteis_ausencia(mes_ini_civil, mes_fim_civil, {f.data for f in feriados_mes})
    # dias_uteis_total = civil month weekdays including pontes (total potential working days)
    dias_uteis_total = _dias_uteis_ausencia(mes_ini_civil, mes_fim_civil, {f.data for f in feriados_mes if f.tipo != 'ponte'})
    feriados_no_periodo = feriados_mes
    # dias_ferias_pontes = feriados + pontes that fall on weekdays
    dias_ferias_pontes = n_feriados + n_pontes
    TIPOS_LABELS = {
        'ferias':'Férias','ponte':'Ponte','fecho_empresa':'Fecho Empresa',
        'falta_justificada':'Falta Justificada','falta_injustificada':'Falta Injustificada',
        'baixa_medica':'Baixa Médica','consulta_medica':'Consulta Médica',
        'assistencia_familia':'Assistência Família/Filho','formacao':'Formação',
        'teletrabalho':'Teletrabalho','licenca_sem_venc':'Licença s/ Vencimento',
        'trabalho_externo':'Trabalho Externo',
    }
    func_data = []
    total_faltas=0; total_ferias=0; total_he=0
    for f in funcs:
        aus = AusenciaRegisto.query.filter(
            AusenciaRegisto.funcionario_id==f.id,
            AusenciaRegisto.estado=='aprovado',
            AusenciaRegisto.data_inicio<=data_fim,
            AusenciaRegisto.data_fim>=data_ini
        ).order_by(AusenciaRegisto.data_inicio).all()
        try:
            hes = HoraExtra.query.filter(
                HoraExtra.funcionario_id==f.id,
                HoraExtra.data>=data_ini, HoraExtra.data<=data_fim,
                HoraExtra.estado.in_(['aprovado','pendente'])
            ).order_by(HoraExtra.data).all()
        except Exception:
            hes = []
        f_ferias=0; f_faltas_just=0; f_faltas_injust=0; f_baixas=0; f_pontes=0
        f_he=0; f_he_util=0; f_he_fds=0
        events = []
        for a in aus:
            ini2=max(a.data_inicio,data_ini); fim2=min(a.data_fim,data_fim)
            d=_dias_uteis_ausencia(ini2,fim2,feriados_set,a.formato,a.horas)
            label=TIPOS_LABELS.get(a.tipo,a.tipo)
            cls='ferias'
            if a.tipo=='ferias': f_ferias+=d; total_ferias+=d; cls='ferias'
            elif a.tipo in('ponte','fecho_empresa'): f_pontes+=d; cls='ponte'
            elif a.tipo=='falta_justificada': f_faltas_just+=d; total_faltas+=d; cls='falta'
            elif a.tipo=='falta_injustificada': f_faltas_injust+=d; total_faltas+=d; cls='falta'
            elif a.tipo in('baixa_medica','consulta_medica','assistencia_familia'): f_baixas+=d; total_faltas+=d; cls='baixa'
            else: cls='ferias'
            # Build event text
            if a.tipo in ('falta_justificada','falta_injustificada','baixa_medica','consulta_medica','assistencia_familia'):
                # Individual days for absences
                from datetime import timedelta as _td
                cur = ini2
                while cur <= fim2:
                    if cur.weekday()<5 and cur not in feriados_set:
                        obs_part = f' — motivo: {a.observacoes}' if a.observacoes else ''
                        events.append({'txt':f'Falta em {cur.strftime("%d/%m/%Y")} ({label}){obs_part}','cls':cls,'detail':'','tipo':a.tipo})
                    cur += _td(days=1)
            else:
                if ini2==fim2:
                    txt=f'{label} no dia {ini2.strftime("%d/%m/%Y")}'
                else:
                    txt=f'{label} de {ini2.strftime("%d/%m/%Y")} a {fim2.strftime("%d/%m/%Y")}'
                detail = f'{d:.1f} dias úteis' + (f' — {a.observacoes}' if a.observacoes else '')
                events.append({'txt':txt,'cls':cls,'detail':detail,'tipo':a.tipo})
        for he in hes:
            f_he+=he.total_horas; total_he+=he.total_horas
            if he.categoria=='fds': f_he_fds+=he.total_horas
            else: f_he_util+=he.total_horas
            cat_label = {'dia_util':'dia útil','fds':'fim de semana','feriado':'feriado'}.get(he.categoria,'')
            txt=f'Horas extra dia {he.data.strftime("%d/%m/%Y")} ({he.data.strftime("%A")}), das {he.hora_inicio} às {he.hora_fim}'
            detail=f'{he.total_horas:.1f}h em {cat_label}' + (f' — {he.observacoes}' if he.observacoes else '')
            events.append({'txt':txt,'cls':'he','detail':detail,'tipo':'he'})
        # Sort events by date
        events.sort(key=lambda e: e.get('txt',''))
        f_faltas = f_faltas_just+f_faltas_injust+f_baixas
        # Add auto-pontes from FeriasFeriado (not manually registered)
        auto_pontes = sum(1 for fm in feriados_mes
            if fm.tipo == 'ponte' and fm.data.weekday() < 5
            and not any(a.data_inicio <= fm.data <= a.data_fim for a in aus))
        f_pontes_total = f_pontes + auto_pontes
        f_trabalhados = round(dias_uteis_trabalho - f_ferias - f_pontes_total - f_faltas, 1)
        dept_val = getattr(f,'departamento','') or getattr(f,'cargo','') or ''
        func_data.append({'nome':f.nome,'dept':dept_val,
            'ferias':f_ferias,'faltas_just':f_faltas_just,'faltas_injust':f_faltas_injust,
            'baixas':f_baixas,'pontes':f_pontes,'he':f_he,'he_util':f_he_util,'he_fds':f_he_fds,
            'trabalhados':max(0,f_trabalhados),'pontes':f_pontes_total,'events':events,'dias_uteis':int(dias_uteis_trabalho),
            'has_events':bool(events)})
    from datetime import date as _dt_cls

    # Build mini-calendar data
    import calendar as _cal
    MESES_NOMES = ['','Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                   'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    feriados_dict = {f.data: f for f in feriados_no_periodo}
    # Also get fecho empresa days for calendar
    try:
        fechos_empresa = AusenciaRegisto.query.filter(
            AusenciaRegisto.tipo == 'fecho_empresa',
            AusenciaRegisto.data_inicio <= data_fim,
            AusenciaRegisto.data_fim >= data_ini
        ).all()
        fechos_dias = set()
        from datetime import timedelta as _td2
        for fe in fechos_empresa:
            cur = fe.data_inicio
            while cur <= fe.data_fim:
                fechos_dias.add(cur)
                cur += _td2(days=1)
    except:
        fechos_dias = set()
    cal_months = []
    # Always show the civil month in calendar
    cur_m = mes; cur_y = ano
    end_m = mes; end_y = ano
    # Also include salary period months
    if data_ini.month != mes or data_ini.year != ano:
        cur_m = min(data_ini.month, mes); cur_y = ano
    while (cur_y, cur_m) <= (end_y, end_m):
        from datetime import date as _dtc
        nd = _cal.monthrange(cur_y, cur_m)[1]
        first_dow = _dtc(cur_y, cur_m, 1).weekday()  # 0=Mon
        weeks = []
        week = [{'day':0}] * first_dow
        for d in range(1, nd+1):
            dt = _dtc(cur_y, cur_m, d)
            dow = dt.weekday()
            fer = feriados_dict.get(dt)
            is_ref_month = (cur_y == ano and cur_m == mes)
            is_civil_weekday = dow < 5 and not fer and is_ref_month
            cell = {
                'day': d,
                'weekend': dow >= 5,
                'in_period': is_civil_weekday,  # green = reference month working days only
                'feriado': bool(fer),
                'feriado_tipo': ('ponte' if fer and fer.tipo=='ponte' else 'feriado') if fer else '',
                'feriado_nome': fer.nome if fer else '',
                'is_fecho': dt in fechos_dias,
                'border_start': dt == data_ini,
                'border_end': dt == data_fim,
                'in_fecho': data_ini <= dt <= data_fim,
            }
            week.append(cell)
            if len(week) == 7:
                weeks.append(week); week = []
        if week:
            week += [{'day':0}] * (7 - len(week))
            weeks.append(week)
        cal_months.append({'nome': MESES_NOMES[cur_m], 'ano': cur_y, 'weeks': weeks})
        if cur_m == 12: cur_y += 1; cur_m = 1
        else: cur_m += 1

    from datetime import date as _dt_cls
    return render_template('ausencias_pdf.html',
        empresa=empresa, mes=mes, ano=ano, mes_nome=MESES_PT[mes],
        data_ini=data_ini, data_fim=data_fim,
        dias_uteis=int(dias_uteis),
        dias_uteis_civil=int(dias_uteis_trabalho),
        dias_uteis_total=int(dias_uteis_total),
        dias_ferias_pontes=dias_ferias_pontes,
        total_faltas=round(total_faltas,1),
        total_ferias=round(total_ferias,1), total_he=round(total_he,1),
        func_data=func_data, hoje=date.today(),
        feriados_no_periodo=feriados_no_periodo,
        n_feriados=n_feriados, n_pontes=n_pontes,
        cal_months=cal_months, date=_dt_cls,
        periodo_fecho=periodo,
        mes_ini_civil=mes_ini_civil, mes_fim_civil=mes_fim_civil)


@app.route('/ausencias/ping')
def ausencias_ping():
    return 'AUSENCIAS_MODULE_LOADED_OK'

# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO AUSÊNCIAS — FÉRIAS / FALTAS
# ══════════════════════════════════════════════════════════════════════════════

UPLOAD_AUSENCIAS = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'ausencias')
os.makedirs(UPLOAD_AUSENCIAS, exist_ok=True)

def _dias_uteis_ausencia(data_inicio, data_fim, feriados_set=None, formato='dia', horas=0):
    """Count useful days between dates, excluding weekends and holidays."""
    from datetime import timedelta
    if formato == 'horas':
        return round(horas / 8, 2)
    if formato in ('manha', 'tarde'):
        # Check if that single day is a weekday
        if data_inicio.weekday() < 5 and (feriados_set is None or data_inicio not in feriados_set):
            return 0.5
        return 0
    count = 0
    cur = data_inicio
    while cur <= data_fim:
        if cur.weekday() < 5 and (feriados_set is None or cur not in feriados_set):
            count += 1
        cur += timedelta(days=1)
    return count

def _get_feriados_set(ano):
    from datetime import date
    feriados = FeriasFeriado.query.filter_by(ano=ano).all()
    return {f.data for f in feriados}

def _calcular_direito_ferias(funcionario, ano):
    """Calculate vacation days entitlement based on admission date."""
    from datetime import date
    admissao = funcionario.data_admissao
    if not admissao:
        return 22  # default

    # First year: 2 days per full month, max 20
    if admissao.year == ano:
        meses_completos = 0
        for m in range(admissao.month, 13):
            # Count months after admission in the same year
            if m > admissao.month or admissao.day == 1:
                meses_completos += 1
        # More precise: months from admission to end of year
        from dateutil.relativedelta import relativedelta
        end_of_year = date(ano, 12, 31)
        delta = relativedelta(end_of_year, admissao)
        meses_completos = delta.months + (delta.years * 12)
        if delta.days > 0:
            meses_completos += 1
        dias = min(meses_completos * 2, 20)
        return max(0, dias)
    
    # Year of admission + 1: proportional or full 22 days
    if admissao.year == ano - 1:
        return 22
    
    # Subsequent years: 22 days
    return 22

def _recalc_saldo_com_admissao(funcionario_id, ano):
    """Recalculate saldo considering admission date and company bridges."""
    from datetime import date, timedelta
    func = Funcionario.query.get(funcionario_id)
    if not func:
        return None
    
    saldo = AusenciaSaldoAnual.query.filter_by(funcionario_id=funcionario_id, ano=ano).first()
    if not saldo:
        saldo = AusenciaSaldoAnual(funcionario_id=funcionario_id, ano=ano)
        db.session.add(saldo)
    
    # Only set dias_direito if not manually overridden (dias_ajuste == 0 and dias_direito == 22)
    if saldo.dias_direito == 22 and saldo.dias_ajuste == 0:
        saldo.dias_direito = _calcular_direito_ferias(func, ano)
    
    # Calculate consumed: ferias + pontes + fecho_empresa
    # But only pontes/fechos AFTER admission date
    admissao = func.data_admissao or date(ano, 1, 1)
    
    consumed = db.session.query(db.func.sum(AusenciaRegisto.dias_uteis)).filter(
        AusenciaRegisto.funcionario_id == funcionario_id,
        AusenciaRegisto.ano == ano,
        AusenciaRegisto.estado.in_(['aprovado', 'pendente']),
        AusenciaRegisto.tipo.in_(['ferias', 'ponte', 'fecho_empresa']),
        AusenciaRegisto.data_inicio >= admissao  # Only after admission
    ).scalar() or 0
    
    saldo.dias_gozados = round(consumed, 2)
    saldo.dias_restantes = round((saldo.dias_direito + saldo.dias_ajuste) - consumed, 2)
    return saldo


def _recalc_saldo(funcionario_id, ano):
    """Recalculate annual leave balance for a given year."""
    saldo = AusenciaSaldoAnual.query.filter_by(funcionario_id=funcionario_id, ano=ano).first()
    if not saldo:
        saldo = AusenciaSaldoAnual(funcionario_id=funcionario_id, ano=ano, dias_direito=22)
        db.session.add(saldo)
    # Sum ferias + pontes + fecho_empresa (all consume vacation balance)
    gozados = db.session.query(db.func.sum(AusenciaRegisto.dias_uteis)).filter(
        AusenciaRegisto.funcionario_id == funcionario_id,
        AusenciaRegisto.ano == ano,
        AusenciaRegisto.estado.in_(['aprovado','pendente']),
        AusenciaRegisto.tipo.in_(['ferias','ponte','fecho_empresa'])
    ).scalar() or 0
    saldo.dias_gozados = round(gozados, 2)
    saldo.dias_restantes = round((saldo.dias_direito + saldo.dias_ajuste) - gozados, 2)
    return saldo

@app.route('/ausencias/sync-funcionario/<int:fid>/<int:ano>', methods=['POST'])
@login_required
def ausencias_sync_funcionario(fid, ano):
    try:
        n = _sync_ausencias_to_faltas(fid, ano)
        _recalc_saldo(fid, ano)
        db.session.commit()
        return jsonify({'ok': True, 'registos': n})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/ausencias/sync-all/<int:ano>', methods=['POST'])
@login_required
def ausencias_sync_all(ano):
    funcs = Funcionario.query.filter_by(ativo=True).all() if hasattr(Funcionario,'ativo') else Funcionario.query.all()
    total = 0
    for f in funcs:
        try:
            total += _sync_ausencias_to_faltas(f.id, ano)
        except: pass
    db.session.commit()
    return jsonify({'ok': True, 'total': total, 'funcionarios': len(funcs)})


@app.route('/ausencias')
@login_required
def ausencias():
    from datetime import date
    ano = int(request.args.get('ano', date.today().year))
    dept_filtro = request.args.get('dept', '')
    func_filtro = request.args.get('func', 0, type=int)
    view = request.args.get('view', 'calendario')  # calendario / lista / dashboard

    funcs = Funcionario.query.filter(Funcionario.ativo == True).order_by(Funcionario.nome).all() \
        if hasattr(Funcionario, 'ativo') else Funcionario.query.order_by(Funcionario.nome).all()
    if dept_filtro:
        funcs = [f for f in funcs if (getattr(f,'departamento','') or '') == dept_filtro]

    # Registos do ano
    try:
        query = AusenciaRegisto.query.filter_by(ano=ano)
        if func_filtro:
            query = query.filter_by(funcionario_id=func_filtro)
        registos = query.all()
    except Exception:
        registos = []

    # Feriados — seed defaults if none exist
    try:
        feriados = FeriasFeriado.query.filter_by(ano=ano).all()
        if not feriados:
            from datetime import date as _d
            PT_FERIADOS_BASE = [
                (1,1,'Ano Novo'),(4,25,'25 de Abril'),(5,1,'Dia do Trabalhador'),
                (6,10,'Dia de Portugal'),(8,15,'Assunção de Nossa Senhora'),
                (10,5,'Implantação da República'),(11,1,'Dia de Todos os Santos'),
                (12,1,'Restauração da Independência'),(12,8,'Imaculada Conceição'),
                (12,25,'Natal'),
            ]
            for mes, dia, nome in PT_FERIADOS_BASE:
                try:
                    db.session.add(FeriasFeriado(ano=ano, data=_d(ano,mes,dia), nome=nome, tipo='nacional'))
                except: pass
            try: db.session.commit()
            except: db.session.rollback()
            feriados = FeriasFeriado.query.filter_by(ano=ano).all()
    except Exception:
        feriados = []

    # Fecho empresa
    try:
        fechos = EmpresaFecho.query.filter_by(ano=ano).all()
    except Exception:
        fechos = []

    # Saldos - recalculate with admission date rules
    saldos = {}
    for f in funcs:
        try:
            s = AusenciaSaldoAnual.query.filter_by(funcionario_id=f.id, ano=ano).first()
            if not s:
                # Auto-create with correct entitlement
                try:
                    s = AusenciaSaldoAnual(
                        funcionario_id=f.id, ano=ano,
                        dias_direito=_calcular_direito_ferias(f, ano)
                    )
                    db.session.add(s)
                    db.session.commit()
                except Exception:
                    db.session.rollback()
            # Calculate breakdown by type for display
            try:
                ferias_dias = db.session.query(db.func.sum(AusenciaRegisto.dias_uteis)).filter(
                    AusenciaRegisto.funcionario_id==f.id, AusenciaRegisto.ano==ano,
                    AusenciaRegisto.estado.in_(['aprovado','pendente']),
                    AusenciaRegisto.tipo=='ferias'
                ).scalar() or 0
                pontes_dias = db.session.query(db.func.sum(AusenciaRegisto.dias_uteis)).filter(
                    AusenciaRegisto.funcionario_id==f.id, AusenciaRegisto.ano==ano,
                    AusenciaRegisto.estado.in_(['aprovado','pendente']),
                    AusenciaRegisto.tipo=='ponte'
                ).scalar() or 0
                fecho_dias = db.session.query(db.func.sum(AusenciaRegisto.dias_uteis)).filter(
                    AusenciaRegisto.funcionario_id==f.id, AusenciaRegisto.ano==ano,
                    AusenciaRegisto.estado.in_(['aprovado','pendente']),
                    AusenciaRegisto.tipo=='fecho_empresa'
                ).scalar() or 0
                if s:
                    s._ferias = round(ferias_dias, 1)
                    s._pontes = round(pontes_dias, 1)
                    s._fecho  = round(fecho_dias, 1)
            except:
                if s:
                    s._ferias = 0; s._pontes = 0; s._fecho = 0
            saldos[f.id] = s
        except Exception:
            saldos[f.id] = None

    import json
    hoje = date.today()

    registos_json = json.dumps([{
        'id': r.id,
        'funcionario_id': r.funcionario_id,
        'funcionario_nome': r.funcionario.nome if r.funcionario else '?',
        'tipo': r.tipo,
        'label': TIPOS_AUSENCIA.get(r.tipo, {}).get('label', r.tipo),
        'cor': TIPOS_AUSENCIA.get(r.tipo, {}).get('cor', '#888'),
        'icon': TIPOS_AUSENCIA.get(r.tipo, {}).get('icon', '📅'),
        'data_inicio': r.data_inicio.isoformat(),
        'data_fim': r.data_fim.isoformat(),
        'formato': r.formato,
        'horas': r.horas,
        'dias_uteis': r.dias_uteis,
        'estado': r.estado,
        'observacoes': r.observacoes or '',
    } for r in registos])

    feriados_json = json.dumps([{
        'id': f.id, 'data': f.data.isoformat(), 'nome': f.nome, 'tipo': f.tipo
    } for f in feriados])

    fechos_json = json.dumps([{
        'id': f.id, 'data_inicio': f.data_inicio.isoformat(),
        'data_fim': f.data_fim.isoformat(), 'descricao': f.descricao
    } for f in fechos])

    # Dashboard stats
    ausentes_hoje = [r for r in registos
        if r.data_inicio <= hoje <= r.data_fim and r.estado == 'aprovado']

    depts = sorted(set((getattr(f,'departamento','') or 'Geral') for f in
        (Funcionario.query.all() if not dept_filtro else funcs)))

    cfg = ConfigGeral.query.first()
    dia_inicio = cfg.salario_dia_inicio if cfg and hasattr(cfg,'salario_dia_inicio') else 1
    dia_fecho  = cfg.salario_dia_fecho  if cfg and hasattr(cfg,'salario_dia_fecho')  else 27
    # Load salary periods for calendar overlay
    import json as _json2
    try:
        periodos_sal = PeriodoSalarial.query.filter_by(ano=ano).all()
        periodos_sal_json = _json2.dumps([{
            'mes': p.mes, 'inicio': p.data_inicio.isoformat(), 'fim': p.data_fim.isoformat(),
            'estado': p.estado
        } for p in periodos_sal])
    except Exception:
        periodos_sal_json = '[]'
    try:
        from datetime import date as _dt3
        horas_extra_ano = HoraExtra.query.filter(
            db.extract('year', HoraExtra.data) == ano,
            HoraExtra.estado.in_(['aprovado','pendente'])
        ).all()
        horas_extra_json = _json2.dumps([{
            'id': h.id,
            'funcionario_id': h.funcionario_id,
            'funcionario_nome': h.funcionario.nome if h.funcionario else '?',
            'data': h.data.isoformat(),
            'hora_inicio': h.hora_inicio,
            'hora_fim': h.hora_fim,
            'total_horas': h.total_horas,
            'categoria': h.categoria,
            'estado': h.estado,
            'observacoes': h.observacoes or '',
        } for h in horas_extra_ano])
    except Exception:
        horas_extra_json = '[]'
    return render_template('ausencias.html',
        ano=ano, funcs=funcs, saldos=saldos,
        tipos=TIPOS_AUSENCIA, view=view,
        registos_json=registos_json,
        feriados_json=feriados_json,
        fechos_json=fechos_json,
        ausentes_hoje=ausentes_hoje,
        dept_filtro=dept_filtro, func_filtro=func_filtro,
        depts=depts, hoje=hoje,
        dia_inicio_sal=dia_inicio, dia_fecho_sal=dia_fecho,
        periodos_sal_json=periodos_sal_json,
        horas_extra_json=horas_extra_json)

@app.route('/ausencias/registar', methods=['POST'])
@login_required
def ausencia_registar():
    from datetime import date
    data = request.get_json() or {}
    try:
        inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        fim    = datetime.strptime(data['data_fim'],    '%Y-%m-%d').date()
    except:
        return jsonify({'ok': False, 'error': 'Datas inválidas'})
    if fim < inicio:
        return jsonify({'ok': False, 'error': 'Data fim anterior ao início'})

    fid  = int(data['funcionario_id'])
    tipo = data.get('tipo', 'ferias')
    fmt  = data.get('formato', 'dia')
    horas= float(data.get('horas', 0))

    # Check overlap
    overlap = AusenciaRegisto.query.filter(
        AusenciaRegisto.funcionario_id == fid,
        AusenciaRegisto.estado.in_(['aprovado','pendente']),
        AusenciaRegisto.data_inicio <= fim,
        AusenciaRegisto.data_fim >= inicio
    ).first()
    if overlap:
        return jsonify({'ok': False, 'error': f'Sobreposição com registo existente ({overlap.tipo} {overlap.data_inicio}–{overlap.data_fim})'})

    feriados_set = _get_feriados_set(inicio.year)
    dias = _dias_uteis_ausencia(inicio, fim, feriados_set, fmt, horas)

    # Check saldo for férias
    tipo_info = TIPOS_AUSENCIA.get(tipo, {})
    if tipo_info.get('conta_ferias'):
        saldo = AusenciaSaldoAnual.query.filter_by(funcionario_id=fid, ano=inicio.year).first()
        if saldo and saldo.dias_restantes < dias:
            if not data.get('forcar'):
                return jsonify({'ok': False, 'error': f'Saldo insuficiente ({saldo.dias_restantes:.1f} dias disponíveis, a marcar {dias:.1f})', 'saldo_insuficiente': True})

    auto_aprovar = True  # can be changed to pending based on config
    r = AusenciaRegisto(
        funcionario_id=fid, tipo=tipo, formato=fmt, horas=horas,
        data_inicio=inicio, data_fim=fim, ano=inicio.year,
        dias_uteis=dias,
        estado='aprovado' if auto_aprovar else 'pendente',
        observacoes=data.get('observacoes','').strip(),
        criado_por=current_user.id, criado_em=datetime.now()
    )
    db.session.add(r)
    db.session.flush()
    _recalc_saldo(fid, inicio.year)
    db.session.commit()
    try: _sync_ausencias_to_faltas(fid, inicio.year)
    except: pass
    return jsonify({'ok': True, 'id': r.id, 'dias_uteis': dias})

@app.route('/ausencias/<int:rid>/editar', methods=['POST'])
@login_required
def ausencia_editar(rid):
    r = AusenciaRegisto.query.get_or_404(rid)
    data = request.get_json() or {}
    old_ano = r.ano
    if 'data_inicio' in data:
        r.data_inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
    if 'data_fim' in data:
        r.data_fim = datetime.strptime(data['data_fim'], '%Y-%m-%d').date()
    if 'tipo' in data: r.tipo = data['tipo']
    if 'formato' in data: r.formato = data['formato']
    if 'horas' in data: r.horas = float(data['horas'])
    if 'estado' in data: r.estado = data['estado']
    if 'observacoes' in data: r.observacoes = data['observacoes']
    r.ano = r.data_inicio.year
    feriados_set = _get_feriados_set(r.ano)
    r.dias_uteis = _dias_uteis_ausencia(r.data_inicio, r.data_fim, feriados_set, r.formato, r.horas)
    r.alterado_por = current_user.id
    r.alterado_em  = datetime.now()
    db.session.commit()
    _recalc_saldo(r.funcionario_id, r.ano)
    if old_ano != r.ano:
        _recalc_saldo(r.funcionario_id, old_ano)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/ausencias/<int:rid>/eliminar', methods=['POST'])
@login_required
def ausencia_eliminar(rid):
    r = AusenciaRegisto.query.get_or_404(rid)
    fid, ano = r.funcionario_id, r.ano
    db.session.delete(r)
    db.session.commit()
    _recalc_saldo(fid, ano)
    db.session.commit()
    try: _sync_ausencias_to_faltas(fid, ano)
    except: pass
    return jsonify({'ok': True})

@app.route('/ausencias/saldo/<int:fid>/<int:ano>', methods=['GET', 'POST'])
@login_required
def ausencia_saldo(fid, ano):
    if request.method == 'POST':
        data = request.get_json() or {}
        saldo = AusenciaSaldoAnual.query.filter_by(funcionario_id=fid, ano=ano).first()
        if not saldo:
            saldo = AusenciaSaldoAnual(funcionario_id=fid, ano=ano)
            db.session.add(saldo)
        if 'dias_direito' in data: saldo.dias_direito = float(data['dias_direito'])
        if 'dias_ajuste'  in data: saldo.dias_ajuste  = float(data['dias_ajuste'])
        if 'notas_ajuste' in data: saldo.notas_ajuste = data['notas_ajuste'].strip()
        db.session.flush()
        _recalc_saldo(fid, ano)
        db.session.commit()
        return jsonify({'ok': True, 'restantes': saldo.dias_restantes})
    saldo = AusenciaSaldoAnual.query.filter_by(funcionario_id=fid, ano=ano).first()
    if not saldo: return jsonify({'ok': False})
    return jsonify({'ok': True, 'direito': saldo.dias_direito,
        'ajuste': saldo.dias_ajuste, 'gozados': saldo.dias_gozados, 'restantes': saldo.dias_restantes})

@app.route('/ausencias/fecho', methods=['POST'])
@login_required
def ausencia_fecho():
    data = request.get_json() or {}
    try:
        inicio = datetime.strptime(data['data_inicio'], '%Y-%m-%d').date()
        fim    = datetime.strptime(data['data_fim'],    '%Y-%m-%d').date()
    except:
        return jsonify({'ok': False, 'error': 'Datas inválidas'})
    f = EmpresaFecho(
        ano=inicio.year, data_inicio=inicio, data_fim=fim,
        descricao=data.get('descricao','').strip(),
        criado_por=current_user.id, criado_em=datetime.now()
    )
    db.session.add(f)
    # Auto-apply fecho to all active employees
    funcs = Funcionario.query.filter(Funcionario.ativo == True).all() \
        if hasattr(Funcionario, 'ativo') else Funcionario.query.all()
    feriados_set = _get_feriados_set(inicio.year)
    dias = _dias_uteis_ausencia(inicio, fim, feriados_set)
    for func in funcs:
        # Skip if overlap
        exists = AusenciaRegisto.query.filter(
            AusenciaRegisto.funcionario_id == func.id,
            AusenciaRegisto.data_inicio <= fim,
            AusenciaRegisto.data_fim >= inicio,
            AusenciaRegisto.estado.in_(['aprovado','pendente'])
        ).first()
        if not exists:
            reg = AusenciaRegisto(
                funcionario_id=func.id, tipo='fecho_empresa',
                data_inicio=inicio, data_fim=fim, ano=inicio.year,
                dias_uteis=dias, estado='aprovado',
                observacoes=data.get('descricao',''),
                criado_por=current_user.id, criado_em=datetime.now()
            )
            db.session.add(reg)
    db.session.commit()
    for func in funcs:
        _recalc_saldo(func.id, inicio.year)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/ausencias/exportar')
@login_required
def ausencias_exportar():
    import csv, io
    ano = int(request.args.get('ano', datetime.now().year))
    fmt = request.args.get('fmt', 'csv')
    fid = request.args.get('func', 0, type=int)
    query = AusenciaRegisto.query.filter_by(ano=ano)
    if fid: query = query.filter_by(funcionario_id=fid)
    registos = query.order_by(AusenciaRegisto.funcionario_id, AusenciaRegisto.data_inicio).all()
    output = io.StringIO()
    w = csv.writer(output, delimiter=';')
    w.writerow(['Funcionário','Tipo','Início','Fim','Dias Úteis','Formato','Estado','Observações'])
    for r in registos:
        w.writerow([
            r.funcionario.nome if r.funcionario else '',
            TIPOS_AUSENCIA.get(r.tipo, {}).get('label', r.tipo),
            r.data_inicio.strftime('%d/%m/%Y'),
            r.data_fim.strftime('%d/%m/%Y'),
            r.dias_uteis,
            r.formato,
            r.estado,
            r.observacoes or ''
        ])
    output.seek(0)
    from flask import Response
    return Response(
        '\ufeff' + output.getvalue(),  # BOM for Excel UTF-8
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment;filename=ausencias_{ano}.csv'}
    )

