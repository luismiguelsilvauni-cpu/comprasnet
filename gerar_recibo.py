"""Generate salary slip Excel from Funcionario and ReciboSalario data."""
import os
from openpyxl import Workbook
from openpyxl.styles import (Font, Alignment, Border, Side, PatternFill,
                               numbers)
from openpyxl.utils import get_column_letter

def thin():
    return Side(style='thin', color='000000')

def border(top=False, bottom=False, left=False, right=False):
    return Border(
        top=thin() if top else Side(style=None),
        bottom=thin() if bottom else Side(style=None),
        left=thin() if left else Side(style=None),
        right=thin() if right else Side(style=None),
    )

def gerar_recibo_excel(func, recibo, mes_label, ano, empresa_nome='União Construtora Naval Limitada'):
    """Generate salary slip matching the UCN template."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{func.numero}-{func.nome.split()[0]}"

    # Column widths matching original
    col_widths = {'A':28,'B':8,'C':8,'D':8,'E':14,'F':10,'G':12,'H':12,'I':2,'J':2,'K':2,'L':2,'M':2,'N':2}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Row heights
    for r in range(1, 36):
        ws.row_dimensions[r].height = 16

    bold = Font(bold=True, name='Arial', size=10)
    normal = Font(name='Arial', size=10)
    small = Font(name='Arial', size=9)
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='center')
    left_al = Alignment(horizontal='left', vertical='center')

    def c(cell, value, font=None, align=None, fmt=None, fill=None):
        ws[cell] = value
        if font: ws[cell].font = font
        if align: ws[cell].alignment = align
        if fmt: ws[cell].number_format = fmt
        if fill: ws[cell].fill = fill

    # ── ROW 1: Company name ──
    ws.merge_cells('A1:N1')
    c('A1', f'   {empresa_nome}', Font(bold=True, name='Arial', size=11), left_al)

    # ── ROW 3: RECIBO DE REMUNERAÇÃO ──
    ws.merge_cells('A3:D3')
    c('A3', 'RECIBO DE REMUNERAÇÃO', bold, left_al)
    ws.merge_cells('E3:G3')
    c('E3', f'Nº {func.numero}', normal, center)

    # ── ROW 4: Nome ──
    ws.merge_cells('A4:D4')
    c('A4', f'NOME DO FUNCIONÁRIO: {func.nome}', bold, left_al)
    ws.merge_cells('E4:N4')
    nif_text = f'Beneficiário CRSSP - {func.num_seg_social or "—"}'
    c('E4', nif_text, normal, left_al)

    # ── ROW 5: Morada ──
    ws.merge_cells('A5:D5')
    c('A5', f'Morada: {func.morada or "—"}', normal, left_al)
    ws.merge_cells('E5:N5')
    c('E5', f'Contribuinte - {func.nif or "—"}', normal, left_al)

    # ── ROW 6: CP / Estado Civil ──
    ws.merge_cells('A6:D6')
    c('A6', 'Código Postal:', normal, left_al)
    ws.merge_cells('F6:N6')
    dep = func.num_dependentes or 0
    titulares = func.titulares_rendimento or 1
    ec = func.estado_civil or 'Solteiro'
    c('F6', f'Situação: {ec}, {titulares} titular{"es" if titulares>1 else ""}, {dep} dependente{"s" if dep!=1 else ""}', normal, left_al)

    # ── ROW 7: Categoria / Seguro ──
    ws.merge_cells('A7:C7')
    c('A7', f'Categoria: {func.categoria or "—"}', normal, left_al)
    ws.merge_cells('D7:N7')
    c('D7', f'Abrangido pela apólice Acidentes de Trabalho nº {func.seguro_apolice or "—"} Seguradora: {func.seguro_companhia or "—"}', small, left_al)

    # ── HEADER ROW 8 ──
    ws.merge_cells('A8:C8')
    c('A8', 'PROCESSAMENTO', bold, center)
    c('D8', 'DIAS', bold, center)
    c('E8', 'HORAS', bold, center)
    c('F8', 'RHT', bold, center)
    ws.merge_cells('G8:H8')
    c('G8', 'BASE  (€)', bold, center)
    # Border bottom on row 8
    for col in ['A','B','C','D','E','F','G','H']:
        ws[f'{col}8'].border = border(bottom=True)

    # ── ROW 9: Remuneração Base ──
    venc = float(recibo.vencimento_base or 0)
    dias = 30
    horas = venc / dias / (8 if venc > 0 else 1) if venc > 0 else 0
    c('A9', 'Remuneração Base', normal, left_al)
    c('D9', dias, normal, center, '#,##0')
    c('E9', horas * dias, normal, center, '#,##0.00')
    c('F9', f'={venc}/{dias}' if dias > 0 else 0, normal, center, '#,##0.000')
    c('G9', venc, normal, right, '#,##0.00')
    c('H9', '=G9', normal, right, '#,##0.00')

    # ── ROW 10: Gratificação ──
    grat = float(recibo.premios or 0)
    c('A10', 'GRATIFICAÇÃO / PRÉMIOS', normal, left_al)
    c('H10', grat if grat else '', normal, right, '#,##0.00')

    # ── ROW 11: Faltas dias ──
    c('A11', 'Faltas/Descontos/Férias/Feriados (dias)', normal, left_al)

    # ── ROW 12: Faltas horas ──
    c('A12', 'Faltas mês corrente (horas)', normal, left_al)

    # ── ROW 13: Horas Extra ──
    hex_val = float(recibo.horas_extra or 0)
    c('A13', 'Horas Extraordinárias', normal, left_al)
    if hex_val:
        c('E13', hex_val, normal, center, '#,##0.00')
        rht_extra = venc / dias / (8 if venc > 0 else 1) * 1.5 if venc > 0 else 0
        c('F13', rht_extra, normal, center, '#,##0.000')
        c('H13', '=E13*F13', normal, right, '#,##0.00')

    # ── ROW 14: Subsídio Alimentação ──
    sub_ref = float(recibo.subsidio_refeicao or 0)
    c('A14', 'Subsídio de Alimentação', normal, left_al)
    # dias úteis estimados
    dias_ref = 22
    val_dia_ref = sub_ref / dias_ref if dias_ref > 0 and sub_ref > 0 else 10
    c('D14', dias_ref, normal, center, '#,##0')
    c('F14', val_dia_ref, normal, center, '#,##0.00')
    c('H14', sub_ref if sub_ref else f'=D14*F14', normal, right, '#,##0.00')

    # ── ROW 15: Outros Abonos ──
    outros = float(recibo.outros_abonos or 0)
    if outros:
        c('A15', 'Outros Abonos', normal, left_al)
        c('H15', outros, normal, right, '#,##0.00')

    # ── ROW 16: TOTAL ILÍQUIDO ──
    ws.merge_cells('A16:D16')
    ws.merge_cells('E16:G16')
    c('E16', 'TOTAL ILÍQUIDO >>>', bold, right)
    c('H16', '=H9+H10+H13+H14+H15', bold, right, '#,##0.00')
    ws['H16'].border = border(top=True, bottom=True)

    # ── DESCONTOS HEADER ROW 17 ──
    ws.merge_cells('A17:C17')
    c('A17', 'DESCONTOS', bold, left_al)
    c('D17', '%', bold, center)
    c('E17', 'VALOR', bold, center)
    c('H17', 'VALOR', bold, center)
    for col in ['A','B','C','D','E','F','G','H']:
        ws[f'{col}17'].border = border(bottom=True)

    # ── ROW 18: CRSS ──
    ss_rate = 0.11
    c('A18', 'C.R.S.S. Remuneração base', normal, left_al)
    stored_ss_taxa = float(getattr(recibo, 'seg_social_taxa', 0) or ss_rate)
    stored_ss_base = float(getattr(recibo, 'seg_social_base', 0) or venc)
    c('D18', stored_ss_taxa, normal, center, '0.00%')
    c('E18', stored_ss_base, normal, right, '#,##0.00')
    c('F18', f'=D18*E18', normal, right, '#,##0.00')

    # ── ROW 19: IRS ──
    irs_val = float(recibo.irs_retencao or 0)
    # Calculate effective IRS rate
    total_irs_base = venc + grat
    # Use stored IRS taxa if available, else calculate
    stored_irs_taxa = float(getattr(recibo, 'irs_taxa', 0) or 0)
    stored_irs_base = float(getattr(recibo, 'irs_base', 0) or 0)
    irs_rate = stored_irs_taxa if stored_irs_taxa else (irs_val / total_irs_base if total_irs_base > 0 and irs_val > 0 else 0)
    irs_base_show = stored_irs_base if stored_irs_base else total_irs_base
    c('A19', 'I.R.S.', normal, left_al)
    c('B19', irs_rate, normal, center, '0.0000')
    c('D19', irs_rate, normal, center, '0.0000')
    c('E19', irs_base_show, normal, right, '#,##0.00')
    c('H19', irs_val if irs_val else f'=D19*E19', normal, right, '#,##0.00')

    # ── ROW 20: IRS info ──
    c('B20', 'Tx marginal', small, center)
    c('C20', 'parcela abater', small, center)
    c('D20', 'tx efetiva', small, center)

    # ── ROW 21: IRS Horas Extra ──
    if hex_val:
        c('A21', 'I.R.S. Horas Extras', normal, left_al)
        c('D21', irs_rate, normal, center, '0.0000')

    # ── ROW 22: Outros descontos ──
    outros_desc = float(recibo.outros_descontos or 0)
    if outros_desc:
        c('A22', 'Outros Descontos', normal, left_al)
        c('H22', outros_desc, normal, right, '#,##0.00')

    # ── ROW 23: TOTAL DESCONTOS ──
    ws.merge_cells('E23:G23')
    c('E23', 'TOTAL DE DESCONTOS >>>', bold, right)
    c('H23', '=F18+H19+H22', bold, right, '#,##0.00')
    ws['H23'].border = border(top=True, bottom=True)

    # ── ROW 24: RESUMO ──
    c('A24', 'RESUMO', bold, left_al)

    # ── ROW 25: LÍQUIDO ──
    ws.merge_cells('E25:G25')
    c('E25', 'LÍQUIDO A RECEBER >>>', bold, right)
    c('H25', '=H16-H23', bold, right, '#,##0.00')
    ws['H25'].border = border(top=True, bottom=True)

    # ── ROW 26: Acertos ──
    c('D26', 'acertos >>>', normal, right)

    # ── ROWS 27-28: Transferências ──
    c('B27', 'Discriminativo:', normal, left_al)
    ws.merge_cells('C27:G27')
    c('C27', f'>>> Transferência v/ conta DO {func.iban or "—"}', normal, left_al)
    c('H27', '=H25-H28', normal, right, '#,##0.00')

    ws.merge_cells('C28:G28')
    c('C28', '>>> Transferência Subsídio de Refeição', normal, left_al)
    c('H28', '=H14', normal, right, '#,##0.00')

    # ── ROW 30: Recebi ──
    ws.merge_cells('A30:G30')
    c('A30', f'Recebi o valor deste recibo, pago pela {empresa_nome} referente ao', normal, left_al)
    c('H30', mes_label, bold, center)

    # ── ROW 31: Local/Data ──
    ws.merge_cells('A31:B31')
    c('A31', 'Vila do Conde em:', normal, left_al)
    c('C31', '___/___/______', normal, center)

    # ── ROW 32: Obs ──
    c('E32', 'obs:', small, right)

    # ── ROW 33-34: Assinatura ──
    ws.merge_cells('A33:G33')
    c('A33', '_' * 55, normal, center)
    ws.merge_cells('A34:G34')
    c('A34', '(assinatura)', normal, center)

    # Full border around recibo area
    for row in range(3, 35):
        ws[f'A{row}'].border = border(left=True)
        ws[f'N{row}'].border = border(right=True)
    for col in 'ABCDEFGHIJKLMN':
        ws[f'{col}3'].border = border(top=True)
        ws[f'{col}34'].border = border(bottom=True)

    return wb

